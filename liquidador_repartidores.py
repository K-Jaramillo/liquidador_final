#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LIQUIDADOR DE REPARTIDORES v2
- Modelo de datos centralizado (DataStore)
- Edición inline tipo Excel con Entry/Combobox flotante sincronizado
- Datos en tiempo real entre pestañas (Asignar → Liquidación → Descuentos)
"""

import tkinter as tk
import database_local as db
db.init_database()
from tkinter import ttk, messagebox, filedialog, simpledialog
import subprocess
import os
import sys
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# Intentar importar tkcalendar para selector de fecha
try:
    from tkcalendar import DateEntry
    HAS_CALENDAR = True
except ImportError:
    HAS_CALENDAR = False

# ---------------------------------------------------------------------------
# Importar TabAnotaciones para la pestaña de sticky notes
# ---------------------------------------------------------------------------
try:
    from tabs.tab_anotaciones import TabAnotaciones
    HAS_ANOTACIONES = True
except ImportError:
    HAS_ANOTACIONES = False
    print("⚠️ No se pudo cargar TabAnotaciones")

# ---------------------------------------------------------------------------
# Importar base de datos local SQLite para persistencia
# ---------------------------------------------------------------------------
try:
    import database_local as db_local
    USE_SQLITE = True
except ImportError:
    USE_SQLITE = False
    print("⚠️ No se pudo cargar database_local, usando almacenamiento en memoria")

# ---------------------------------------------------------------------------
# Funciones de acceso a datos (usan SQLite si está disponible)
# ---------------------------------------------------------------------------
def cargar_descuentos():
    """Carga todos los descuentos organizados por folio."""
    if USE_SQLITE:
        conn = db_local.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM descuentos ORDER BY folio, fecha_creacion')
        rows = cursor.fetchall()
        conn.close()
        
        result = {}
        for row in rows:
            folio_key = str(row['folio'])
            if folio_key not in result:
                result[folio_key] = {"descuentos": []}
            result[folio_key]["descuentos"].append({
                "id": row['id'],
                "tipo": row['tipo'],
                "monto": row['monto'],
                "observacion": row['observacion'] or '',
                "repartidor": row['repartidor'] or '',
                "fecha": row['fecha']
            })
        return result
    else:
        return {}

def agregar_descuento(folio, tipo, monto, observacion, repartidor):
    """Agrega un descuento a una factura."""
    if USE_SQLITE:
        from datetime import datetime as _dt
        fecha = _dt.now().strftime('%Y-%m-%d')
        db_local.agregar_descuento(fecha, int(folio), tipo, monto, repartidor, observacion)

def obtener_descuentos_factura(folio):
    """Obtiene los descuentos de una factura específica."""
    if USE_SQLITE:
        return db_local.obtener_descuentos_folio(int(folio))
    return []

def cargar_asignaciones():
    """Carga todas las asignaciones como dict {fecha_folio: repartidor}."""
    if USE_SQLITE:
        conn = db_local.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT fecha, folio, repartidor FROM asignaciones')
        rows = cursor.fetchall()
        conn.close()
        return {f"{row['fecha']}_{row['folio']}": row['repartidor'] for row in rows}
    return {}

def guardar_asignaciones(data):
    """Guarda las asignaciones (para compatibilidad, no hace nada con SQLite)."""
    pass  # Con SQLite se guarda automáticamente

def asignar_repartidor(folio, fecha, repartidor):
    """Asigna un repartidor a una factura."""
    if USE_SQLITE:
        db_local.guardar_asignacion(fecha, int(folio), repartidor)

def obtener_repartidor_factura(folio, fecha):
    """Obtiene el repartidor asignado a una factura."""
    if USE_SQLITE:
        return db_local.obtener_asignacion(fecha, int(folio)) or ''
    return ''

def limpiar_asignaciones_dia(fecha):
    """Limpia todas las asignaciones de un día."""
    if USE_SQLITE:
        db_local.limpiar_asignaciones_fecha(fecha)


# ===========================================================================
#  DATASTORE  –  Modelo de datos centralizado (única fuente de verdad)
# ===========================================================================
# WIDGET TOGGLE SWITCH PERSONALIZADO
# ===========================================================================
class ToggleSwitch(tk.Frame):
    """
    Toggle switch visual estilo iOS/Material Design compacto.
    Usa Frame + Canvas para manejar mejor el fondo con ttk.
    """
    def __init__(self, parent, variable=None, command=None, width=36, height=18,
                 bg_off='#bdbdbd', bg_on='#4caf50', frame_bg=None, **kwargs):
        # Eliminar opciones que pueden causar problemas
        kwargs.pop('bg', None)
        kwargs.pop('background', None)
        
        super().__init__(parent, **kwargs)
        
        # Configurar fondo del frame
        self._frame_bg = frame_bg or self.cget('bg')
        self.configure(bg=self._frame_bg)
        
        self.canvas = tk.Canvas(self, width=width, height=height, 
                               highlightthickness=0, bd=0, bg=self._frame_bg)
        self.canvas.pack()
        
        self.w = width
        self.h = height
        self.bg_off = bg_off
        self.bg_on = bg_on
        self.variable = variable
        self.command = command
        self.pad = 2
        self.knob_r = (height - 2 * self.pad) // 2
        
        self.is_on = variable.get() if variable else False
        self._draw()
        self.canvas.bind("<Button-1>", self._toggle)
        
        if self.variable:
            self.variable.trace_add('write', self._on_var_change)
    
    def _draw(self):
        """Dibuja el switch."""
        self.canvas.delete("all")
        
        bg = self.bg_on if self.is_on else self.bg_off
        r = self.h // 2
        
        # Píldora de fondo
        self.canvas.create_arc(0, 0, self.h, self.h, start=90, extent=180, fill=bg, outline=bg)
        self.canvas.create_arc(self.w - self.h, 0, self.w, self.h, start=-90, extent=180, fill=bg, outline=bg)
        self.canvas.create_rectangle(r, 0, self.w - r, self.h, fill=bg, outline=bg)
        
        # Knob (círculo blanco)
        cx = (self.w - self.pad - self.knob_r) if self.is_on else (self.pad + self.knob_r)
        cy = self.h // 2
        
        # Knob blanco
        self.canvas.create_oval(cx - self.knob_r, cy - self.knob_r,
                               cx + self.knob_r, cy + self.knob_r,
                               fill='white', outline='#cccccc', width=1)
    
    def _toggle(self, event=None):
        self.is_on = not self.is_on
        if self.variable:
            self.variable.set(self.is_on)
        self._draw()
        if self.command:
            self.command()
    
    def _on_var_change(self, *args):
        new_val = self.variable.get()
        if new_val != self.is_on:
            self.is_on = new_val
            self._draw()
    
    def set(self, value):
        self.is_on = bool(value)
        if self.variable:
            self.variable.set(self.is_on)
        self._draw()
    
    def set_frame_bg(self, color):
        """Actualiza el color de fondo del frame y canvas."""
        self._frame_bg = color
        self.configure(bg=color)
        self.canvas.configure(bg=color)
    
    def get(self):
        return self.is_on


# ===========================================================================
class DataStore:
    """
    Mantiene el estado global de la aplicación.
    Todas las pestañas leen/escriben aquí → sincronización automática.
    """

    def __init__(self):
        self.fecha: str = datetime.now().strftime('%Y-%m-%d')
        # Lista de dicts: {id, folio, nombre, subtotal, repartidor, cancelada, total_credito, es_credito}
        self.ventas: list = []
        # Conjunto rápido de repartidores conocidos
        self._repartidores: set = set()
        # Callbacks registrados por las pestañas
        self._listeners: list = []
        # Datos adicionales financieros
        self.devoluciones: list = []      # Lista de devoluciones del día
        self.movimientos_entrada: list = []  # Ingresos extras
        self.movimientos_salida: list = []   # Salidas

    # --- suscripción de eventos ---
    def suscribir(self, callback):
        """Registra un callback que se invoca al cambiar datos."""
        if callback not in self._listeners:
            self._listeners.append(callback)

    def _notificar(self):
        for cb in self._listeners:
            try:
                cb()
            except Exception:
                pass

    # --- ventas ---
    def set_ventas(self, ventas: list):
        self.ventas = ventas
        self._repartidores = {v['repartidor'] for v in ventas if v.get('repartidor')}
        self._notificar()

    def get_ventas(self):
        return self.ventas

    def get_total_subtotal(self) -> float:
        """Retorna el total de ventas usando TOTAL (no subtotal).
        Para facturas canceladas del mismo día: se excluyen del total vendido.
        Para facturas canceladas de otro día: NO se suman (solo informativas).
        NOTA: Se usa total_original para coincidir con el corte de caja de Firebird.
        """
        total = 0
        for v in self.ventas:
            cancelada_otro_dia = v.get('cancelada_otro_dia', False)
            cancelada = v.get('cancelada', False)
            if cancelada_otro_dia or cancelada:
                # Las canceladas no suman al total vendido
                continue
            # Usar total_original (TOTAL) para coincidir con Firebird
            total += v.get('total_original', v['subtotal'])
        return total

    def get_total_canceladas(self) -> float:
        """Retorna el total de facturas canceladas del mismo día."""
        return sum(v.get('total_original', 0) for v in self.ventas 
                   if v.get('cancelada', False) and not v.get('cancelada_otro_dia', False))

    def get_total_canceladas_otro_dia(self) -> float:
        """Retorna el total de facturas canceladas que son de otro día."""
        return sum(v.get('total_original', 0) for v in self.ventas 
                   if v.get('cancelada', False) and v.get('cancelada_otro_dia', False))

    def get_total_todas_facturas(self) -> float:
        """Retorna el total de TODAS las facturas del día (incluyendo canceladas del mismo día).
        Las canceladas de otro día NO se suman (solo son informativas).
        """
        return sum(v.get('total_original', v['subtotal']) for v in self.ventas 
                   if not v.get('cancelada_otro_dia', False))

    def get_monto_facturas_efectivo(self) -> float:
        """Retorna el MONTO FACTURAS para cuadre de caja.
        
        Solo incluye:
        - Facturas en efectivo del día (no crédito)
        - NO canceladas (del mismo día ni de otro día)
        
        Para facturas con devolución parcial (estado P), el campo TOTAL de Eleventa
        ya contiene el monto después del descuento.
        """
        return sum(
            v.get('total_original', v['subtotal']) 
            for v in self.ventas 
            if not v.get('cancelada', False)  # No canceladas
            and not v.get('cancelada_otro_dia', False)  # No canceladas de otro día
            and not v.get('es_credito', False)  # No crédito
        )

    def get_ventas_canceladas_otro_dia(self) -> list:
        """Retorna lista de ventas canceladas de otro día."""
        return [v for v in self.ventas if v.get('cancelada', False) and v.get('cancelada_otro_dia', False)]

    def get_total_credito(self) -> float:
        """Retorna el total de facturas a crédito."""
        return sum(v.get('total_credito', 0) for v in self.ventas if v.get('es_credito', False))

    def get_ventas_credito(self) -> list:
        """Retorna lista de ventas a crédito."""
        return [v for v in self.ventas if v.get('es_credito', False)]

    def get_ventas_canceladas(self) -> list:
        """Retorna lista de ventas canceladas del mismo día."""
        return [v for v in self.ventas if v.get('cancelada', False) and not v.get('cancelada_otro_dia', False)]

    # --- devoluciones ---
    def set_devoluciones(self, devoluciones: list):
        self.devoluciones = devoluciones
        self._notificar()

    def get_total_devoluciones(self) -> float:
        return sum(d.get('monto', 0) for d in self.devoluciones)

    # --- movimientos (ingresos/salidas) ---
    def set_movimientos(self, entradas: list, salidas: list):
        self.movimientos_entrada = entradas
        self.movimientos_salida = salidas
        self._notificar()

    def get_total_ingresos_extras(self) -> float:
        return sum(m.get('monto', 0) for m in self.movimientos_entrada)

    def get_total_salidas(self) -> float:
        return sum(m.get('monto', 0) for m in self.movimientos_salida)

    # --- repartidores ---
    def get_repartidores(self) -> list:
        return sorted(self._repartidores)

    def set_repartidor_factura(self, folio: int, repartidor: str):
        """Actualiza el repartidor de una factura y persiste."""
        for v in self.ventas:
            if v['folio'] == folio:
                v['repartidor'] = repartidor
                break
        if repartidor:
            self._repartidores.add(repartidor)
            asignar_repartidor(folio, self.fecha, repartidor)
        self._notificar()

    def clear_repartidor_factura(self, folio: int):
        for v in self.ventas:
            if v['folio'] == folio:
                v['repartidor'] = ''
                break
        # Eliminar de persistencia
        asignaciones = cargar_asignaciones()
        key = f"{self.fecha}_{folio}"
        if key in asignaciones:
            del asignaciones[key]
            guardar_asignaciones(asignaciones)
        self._repartidores = {v['repartidor'] for v in self.ventas if v.get('repartidor')}
        self._notificar()

    def clear_all_asignaciones(self):
        for v in self.ventas:
            v['repartidor'] = ''
        self._repartidores.clear()
        limpiar_asignaciones_dia(self.fecha)
        self._notificar()

    # --- gastos adicionales por repartidor ---
    # Estructura: list de dicts {repartidor, concepto, monto}
    # Se inicializa en __init__ junto con el resto.

    def _ensure_gastos(self):
        if not hasattr(self, 'gastos'):
            self.gastos: list = []

    def agregar_gasto(self, repartidor: str, concepto: str, monto: float, observaciones: str = ''):
        """Agrega un gasto y lo persiste en SQLite."""
        self._ensure_gastos()
        gasto = {
            'repartidor': repartidor,
            'concepto': concepto,
            'monto': monto,
            'observaciones': observaciones
        }
        # Persistir en SQLite
        if USE_SQLITE:
            gasto_id = db_local.agregar_gasto(self.fecha, repartidor, concepto, monto, observaciones)
            gasto['id'] = gasto_id
        self.gastos.append(gasto)
        self._notificar()

    def eliminar_gasto(self, index_or_id):
        """Elimina un gasto por índice o ID de SQLite."""
        self._ensure_gastos()
        if USE_SQLITE and isinstance(index_or_id, int):
            # Eliminar directamente de la BD por ID
            db_local.eliminar_gasto(index_or_id)
            # Sincronizar lista local
            self.gastos = [g for g in self.gastos if g.get('id') != index_or_id]
            self._notificar()
        elif 0 <= index_or_id < len(self.gastos):
            del self.gastos[index_or_id]
            self._notificar()

    def get_gastos(self, repartidor: str = '') -> list:
        """Obtiene los gastos (desde SQLite si está disponible)."""
        self._ensure_gastos()
        if USE_SQLITE:
            if repartidor:
                return db_local.obtener_gastos_repartidor(self.fecha, repartidor)
            return db_local.obtener_gastos_fecha(self.fecha)
        if not repartidor:
            return list(self.gastos)
        return [g for g in self.gastos if g['repartidor'] == repartidor]

    def get_total_gastos(self, repartidor: str = '') -> float:
        gastos = self.get_gastos(repartidor)
        return sum(g.get('monto', 0) for g in gastos)

    def get_total_gastos_cajero(self, repartidor: str = '') -> float:
        """Retorna el total de gastos de cajero (cajero, caja, cajera)."""
        gastos = self.get_gastos(repartidor)
        return sum(g.get('monto', 0) for g in gastos 
                   if g.get('repartidor', '').lower() in ('cajero', 'caja', 'cajera'))

    def get_total_gastos_repartidores(self, repartidor: str = '') -> float:
        """Retorna el total de gastos de repartidores (excluyendo cajero)."""
        gastos = self.get_gastos(repartidor)
        return sum(g.get('monto', 0) for g in gastos 
                   if g.get('repartidor', '').lower() not in ('cajero', 'caja', 'cajera'))

    def get_total_gastos_concepto(self, concepto: str) -> float:
        """Retorna el total de gastos de un concepto específico (ej: NÓMINA, SOCIOS)."""
        gastos = self.get_gastos()
        return sum(g.get('monto', 0) for g in gastos 
                   if g.get('concepto', '').upper() == concepto.upper())

    # --- conteo de dinero por repartidor ---
    # Estructura: dict  repartidor → {valor_int: cantidad_int}

    def _ensure_dinero(self):
        if not hasattr(self, 'dinero'):
            self.dinero: dict = {}

    def set_dinero(self, repartidor: str, conteo: dict):
        """Guarda el conteo de denominaciones para un repartidor (persiste en SQLite)."""
        self._ensure_dinero()
        self.dinero[repartidor] = dict(conteo)   # copia shallow
        # Persistir en SQLite
        if USE_SQLITE:
            db_local.guardar_conteo_dinero(self.fecha, repartidor, conteo)

    def get_dinero(self, repartidor: str) -> dict:
        """Retorna {valor_int: cantidad} para un repartidor."""
        self._ensure_dinero()
        # Primero intentar cargar de SQLite
        if USE_SQLITE:
            conteo_db = db_local.obtener_conteo_dinero(self.fecha, repartidor)
            if conteo_db:
                self.dinero[repartidor] = conteo_db
                return dict(conteo_db)
        return dict(self.dinero.get(repartidor, {}))

    def get_total_dinero(self, repartidor: str = '') -> float:
        """Suma total de dinero de los conteos múltiples. Si repartidor está vacío suma todos."""
        if USE_SQLITE:
            if repartidor:
                # Total de conteos múltiples del repartidor
                return db_local.obtener_total_conteos_repartidor(self.fecha, repartidor)
            else:
                # Total general de todos los repartidores
                return db_local.obtener_total_general_conteos_fecha(self.fecha)
        # Fallback sin SQLite
        return 0.0

    def get_total_descuentos(self, repartidor: str = '') -> float:
        """Retorna el total de descuentos de tipo 'credito' y 'devolucion'."""
        total = 0.0
        desc_todos = cargar_descuentos()
        for folio_key, datos in desc_todos.items():
            for desc in datos.get("descuentos", []):
                if desc.get("fecha", "").startswith(self.fecha):
                    if not repartidor or desc.get("repartidor") == repartidor:
                        tipo = desc.get("tipo", "").lower()
                        if tipo in ('credito', 'devolucion'):
                            total += desc.get("monto", 0)
        return total

    def get_total_ajustes(self, repartidor: str = '') -> float:
        """Retorna el total de ajustes de precios (tipo 'ajuste')."""
        total = 0.0
        desc_todos = cargar_descuentos()
        for folio_key, datos in desc_todos.items():
            for desc in datos.get("descuentos", []):
                if desc.get("fecha", "").startswith(self.fecha):
                    if not repartidor or desc.get("repartidor") == repartidor:
                        tipo = desc.get("tipo", "").lower()
                        if tipo == 'ajuste':
                            total += desc.get("monto", 0)
        return total

    # --- pagos a proveedores ---
    def agregar_pago_proveedor(self, proveedor: str, concepto: str, monto: float, repartidor: str = '', observaciones: str = ''):
        """Agrega un pago a proveedor y lo persiste en SQLite."""
        if USE_SQLITE:
            pago_id = db_local.agregar_pago_proveedor(self.fecha, proveedor, concepto, monto, repartidor, observaciones)
            self._notificar()
            return pago_id
        return -1

    def eliminar_pago_proveedor(self, pago_id: int):
        """Elimina un pago a proveedor."""
        if USE_SQLITE:
            db_local.eliminar_pago_proveedor(pago_id)
            self._notificar()

    def actualizar_pago_proveedor(self, pago_id: int, proveedor: str, concepto: str, 
                                   monto: float, repartidor: str = '', observaciones: str = ''):
        """Actualiza un pago a proveedor existente."""
        if USE_SQLITE:
            db_local.actualizar_pago_proveedor(pago_id, proveedor, concepto, monto, repartidor, observaciones)
            self._notificar()

    def get_pagos_proveedores(self, repartidor: str = '') -> list:
        """Obtiene los pagos a proveedores de la fecha actual."""
        if USE_SQLITE:
            if repartidor:
                return db_local.obtener_pagos_proveedores_repartidor(self.fecha, repartidor)
            return db_local.obtener_pagos_proveedores_fecha(self.fecha)
        return []

    def get_total_pagos_proveedores(self, repartidor: str = '') -> float:
        """Retorna el total de pagos a proveedores."""
        if USE_SQLITE:
            return db_local.obtener_total_pagos_proveedores_fecha(self.fecha, repartidor)
        return 0.0

    # --- actualizar gasto ---
    def actualizar_gasto(self, gasto_id: int, repartidor: str, concepto: str, monto: float, observaciones: str = ''):
        """Actualiza un gasto existente."""
        if USE_SQLITE:
            db_local.actualizar_gasto(gasto_id, repartidor, concepto, monto, observaciones)
            self._notificar()

    # --- préstamos ---
    def agregar_prestamo(self, repartidor: str, concepto: str, monto: float, observaciones: str = ''):
        """Agrega un préstamo y lo persiste en SQLite."""
        if USE_SQLITE:
            prestamo_id = db_local.agregar_prestamo(self.fecha, repartidor, concepto, monto, observaciones)
            self._notificar()
            return prestamo_id
        return -1

    def eliminar_prestamo(self, prestamo_id: int):
        """Elimina un préstamo."""
        if USE_SQLITE:
            db_local.eliminar_prestamo(prestamo_id)
            self._notificar()

    def actualizar_prestamo(self, prestamo_id: int, repartidor: str, concepto: str, monto: float, observaciones: str = ''):
        """Actualiza un préstamo existente."""
        if USE_SQLITE:
            db_local.actualizar_prestamo(prestamo_id, repartidor, concepto, monto, observaciones)
            self._notificar()

    def get_prestamos(self, repartidor: str = '') -> list:
        """Obtiene los préstamos de la fecha actual."""
        if USE_SQLITE:
            if repartidor:
                return db_local.obtener_prestamos_repartidor(self.fecha, repartidor)
            return db_local.obtener_prestamos_fecha(self.fecha)
        return []

    def get_total_prestamos(self, repartidor: str = '') -> float:
        """Retorna el total de préstamos."""
        if USE_SQLITE:
            return db_local.obtener_total_prestamos_fecha(self.fecha, repartidor)
        return 0.0

    # --- pagos de nómina ---
    def agregar_pago_nomina(self, empleado: str, concepto: str, monto: float, observaciones: str = ''):
        """Agrega un pago de nómina y lo persiste en SQLite."""
        if USE_SQLITE:
            pago_id = db_local.agregar_pago_nomina(self.fecha, empleado, concepto, monto, observaciones)
            self._notificar()
            return pago_id
        return -1

    def eliminar_pago_nomina(self, pago_id: int):
        """Elimina un pago de nómina."""
        if USE_SQLITE:
            db_local.eliminar_pago_nomina(pago_id)
            self._notificar()

    def actualizar_pago_nomina(self, pago_id: int, empleado: str, concepto: str, monto: float, observaciones: str = ''):
        """Actualiza un pago de nómina existente."""
        if USE_SQLITE:
            db_local.actualizar_pago_nomina(pago_id, empleado, concepto, monto, observaciones)
            self._notificar()

    def get_pagos_nomina(self, repartidor: str = '') -> list:
        """Obtiene los pagos de nómina de la fecha actual, opcionalmente filtrado por repartidor."""
        if USE_SQLITE:
            if repartidor:
                return db_local.obtener_pagos_nomina_repartidor(self.fecha, repartidor)
            return db_local.obtener_pagos_nomina_fecha(self.fecha)
        return []

    def get_total_pagos_nomina(self, repartidor: str = '') -> float:
        """Retorna el total de pagos de nómina, opcionalmente filtrado por repartidor."""
        if USE_SQLITE:
            return db_local.obtener_total_pagos_nomina_fecha(self.fecha, repartidor)
        return 0.0

    # --- pagos a socios ---
    def agregar_pago_socios(self, socio: str, concepto: str, monto: float, observaciones: str = ''):
        """Agrega un pago a socios y lo persiste en SQLite."""
        if USE_SQLITE:
            pago_id = db_local.agregar_pago_socios(self.fecha, socio, concepto, monto, observaciones)
            self._notificar()
            return pago_id
        return -1

    def eliminar_pago_socios(self, pago_id: int):
        """Elimina un pago a socios."""
        if USE_SQLITE:
            db_local.eliminar_pago_socios(pago_id)
            self._notificar()

    def actualizar_pago_socios(self, pago_id: int, socio: str, concepto: str, monto: float, observaciones: str = ''):
        """Actualiza un pago a socios existente."""
        if USE_SQLITE:
            db_local.actualizar_pago_socios(pago_id, socio, concepto, monto, observaciones)
            self._notificar()

    def get_pagos_socios(self, repartidor: str = '') -> list:
        """Obtiene los pagos a socios de la fecha actual, opcionalmente filtrado por repartidor."""
        if USE_SQLITE:
            if repartidor:
                return db_local.obtener_pagos_socios_repartidor(self.fecha, repartidor)
            return db_local.obtener_pagos_socios_fecha(self.fecha)
        return []

    def get_total_pagos_socios(self, repartidor: str = '') -> float:
        """Retorna el total de pagos a socios, opcionalmente filtrado por repartidor."""
        if USE_SQLITE:
            return db_local.obtener_total_pagos_socios_fecha(self.fecha, repartidor)
        return 0.0

    # --- transferencias ---
    def agregar_transferencia(self, destinatario: str, concepto: str, monto: float, observaciones: str = ''):
        """Agrega una transferencia y la persiste en SQLite."""
        if USE_SQLITE:
            transferencia_id = db_local.agregar_transferencia(self.fecha, destinatario, concepto, monto, observaciones)
            self._notificar()
            return transferencia_id
        return -1

    def eliminar_transferencia(self, transferencia_id: int):
        """Elimina una transferencia."""
        if USE_SQLITE:
            db_local.eliminar_transferencia(transferencia_id)
            self._notificar()

    def actualizar_transferencia(self, transferencia_id: int, destinatario: str, concepto: str, monto: float, observaciones: str = ''):
        """Actualiza una transferencia existente."""
        if USE_SQLITE:
            db_local.actualizar_transferencia(transferencia_id, destinatario, concepto, monto, observaciones)
            self._notificar()

    def get_transferencias(self, repartidor: str = '') -> list:
        """Obtiene las transferencias de la fecha actual, opcionalmente filtrado por repartidor."""
        if USE_SQLITE:
            if repartidor:
                return db_local.obtener_transferencias_repartidor(self.fecha, repartidor)
            return db_local.obtener_transferencias_fecha(self.fecha)
        return []

    def get_total_transferencias(self, repartidor: str = '') -> float:
        """Retorna el total de transferencias, opcionalmente filtrado por repartidor."""
        if USE_SQLITE:
            return db_local.obtener_total_transferencias_fecha(self.fecha, repartidor)
        return 0.0

    # --- conceptos de gastos ---
    def get_conceptos_gastos(self) -> list:
        """Obtiene la lista de conceptos guardados."""
        if USE_SQLITE:
            return db_local.obtener_conceptos_gastos()
        return []
    
    def agregar_concepto_gasto(self, concepto: str) -> bool:
        """Agrega un nuevo concepto. Retorna True si se agregó."""
        if USE_SQLITE:
            return db_local.agregar_concepto_gasto(concepto)
        return False
    
    def eliminar_concepto_gasto(self, concepto: str) -> bool:
        """Elimina un concepto."""
        if USE_SQLITE:
            return db_local.eliminar_concepto_gasto(concepto)
        return False


# ===========================================================================
#  CLASE PRINCIPAL
# ===========================================================================
class LiquidadorRepartidores:

    def __init__(self, ventana: tk.Tk):
        self.ventana = ventana
        self.ventana.title("LiquiVentas v3.0")
        self.ventana.geometry("1350x950")
        self.ventana.state('zoomed')  # Abrir maximizado
        self.ventana.minsize(1100, 800)

        # Cargar ruta FDB desde configuración
        self._cargar_configuracion_rutas()
        
        # DataStore único
        self.ds = DataStore()

        # Variable compartida para repartidor filtro en Liquidación
        self.repartidor_filtro_var = tk.StringVar()

        # Editor inline activo (referencia para destruirlo si existe)
        self._editor_activo = None

        # Crear barra de menú
        self._crear_menu_principal()
        
        self._crear_interfaz()

        # Suscribir las pestañas al DataStore
        self.ds.suscribir(self._on_data_changed)
        
        # Configurar estilos mejorados
        self._configurar_estilos()
        
        # Cargar datos de la fecha actual al iniciar (con pequeño delay para que la GUI esté lista)
        self.ventana.after(500, self._cargar_datos_inicial)
    
    def _crear_tooltip(self, widget, texto):
        """Crea un tooltip (mensaje emergente) para un widget."""
        tooltip = None
        
        def mostrar(event):
            nonlocal tooltip
            x = widget.winfo_rootx() + 20
            y = widget.winfo_rooty() + widget.winfo_height() + 5
            tooltip = tk.Toplevel(widget)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{x}+{y}")
            label = tk.Label(tooltip, text=texto, background="#333333", foreground="white",
                           relief="solid", borderwidth=1, font=("Segoe UI", 9))
            label.pack()
        
        def ocultar(event):
            nonlocal tooltip
            if tooltip:
                tooltip.destroy()
                tooltip = None
        
        widget.bind("<Enter>", mostrar)
        widget.bind("<Leave>", ocultar)
    
    def _cargar_configuracion_rutas(self):
        """Carga las rutas de BD desde la configuración guardada."""
        # Ruta por defecto para LiquiVentas
        RUTA_FDB_DEFAULT = r'D:\LiquiVentas\BDEV\PDVDATA.FDB'
        
        # Detectar automáticamente el sistema operativo
        if sys.platform == 'win32':
            # Intentar cargar ruta FDB desde configuración
            try:
                ruta_guardada = db_local.obtener_config('fdb_path')
                if ruta_guardada and os.path.exists(ruta_guardada):
                    self.ruta_fdb = ruta_guardada
                else:
                    self.ruta_fdb = RUTA_FDB_DEFAULT
            except Exception:
                self.ruta_fdb = RUTA_FDB_DEFAULT
            
            # Buscar isql en múltiples ubicaciones posibles
            posibles_rutas = [
                r'C:\Program Files\Firebird\Firebird_5_0\isql.exe',
                r'C:\Program Files\Firebird\Firebird_4_0\isql.exe',
                r'C:\Program Files\Firebird\Firebird_3_0\isql.exe',
                r'C:\Program Files\Firebird\Firebird_2_5\bin\isql.exe',
                r'C:\Program Files (x86)\Firebird\Firebird_2_5\bin\isql.exe',
                r'C:\Firebird\isql.exe',
                r'C:\Firebird\bin\isql.exe',
            ]
            self.isql_path = None
            for ruta in posibles_rutas:
                if os.path.exists(ruta):
                    self.isql_path = ruta
                    break
            if not self.isql_path:
                self.isql_path = r'C:\Program Files\Firebird\Firebird_5_0\isql.exe'
        else:
            # Linux
            self.ruta_fdb = os.path.join(os.path.dirname(__file__), 'PDVDATA.FDB')
            self.isql_path = '/opt/firebird/bin/isql'
    
    def _crear_menu_principal(self):
        """Crea la barra de menú principal."""
        menubar = tk.Menu(self.ventana)
        
        # Menú Archivo
        menu_archivo = tk.Menu(menubar, tearoff=0)
        menu_archivo.add_command(label="⚙️ Configurar Ruta BD", command=self._configurar_ruta_fdb)
        menu_archivo.add_separator()
        menu_archivo.add_command(label="📂 Abrir Carpeta de Datos", command=self._abrir_carpeta_datos)
        menu_archivo.add_separator()
        menu_archivo.add_command(label="❌ Salir", command=self.ventana.quit)
        menubar.add_cascade(label="Archivo", menu=menu_archivo)
        
        # Menú Ayuda
        menu_ayuda = tk.Menu(menubar, tearoff=0)
        menu_ayuda.add_command(label="ℹ️ Acerca de", command=self._mostrar_acerca_de)
        menubar.add_cascade(label="Ayuda", menu=menu_ayuda)
        
        self.ventana.config(menu=menubar)
    
    def _configurar_ruta_fdb(self):
        """Permite seleccionar la ruta del archivo PDVDATA.FDB."""
        from tkinter import filedialog
        
        ruta_actual = self.ruta_fdb
        nueva_ruta = filedialog.askopenfilename(
            title="Seleccionar Base de Datos Eleventa (PDVDATA.FDB)",
            initialdir=os.path.dirname(ruta_actual) if os.path.exists(os.path.dirname(ruta_actual)) else "D:\\",
            filetypes=[("Base de datos Firebird", "*.FDB"), ("Todos los archivos", "*.*")]
        )
        
        if nueva_ruta:
            # Guardar en configuración
            try:
                db_local.guardar_config('fdb_path', nueva_ruta)
                self.ruta_fdb = nueva_ruta
                messagebox.showinfo(
                    "Configuración Guardada", 
                    f"Nueva ruta de BD configurada:\n{nueva_ruta}\n\nLa aplicación usará esta ruta en próximas consultas."
                )
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la configuración:\n{e}")
    
    def _abrir_carpeta_datos(self):
        """Abre la carpeta donde está el archivo de datos SQLite."""
        import subprocess
        if sys.platform == 'win32':
            carpeta = os.path.dirname(db_local.DB_PATH)
            subprocess.Popen(f'explorer "{carpeta}"')
        else:
            carpeta = os.path.dirname(db_local.DB_PATH)
            subprocess.Popen(['xdg-open', carpeta])
    
    def _mostrar_acerca_de(self):
        """Muestra información de la aplicación."""
        info = (
            "LiquiVentas v3.0\n"
            "━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "Sistema de Liquidación de Repartidores\n\n"
            f"📁 Ruta BD Eleventa:\n{self.ruta_fdb}\n\n"
            f"💾 Datos locales:\n{db_local.DB_PATH}\n\n"
            "━━━━━━━━━━━━━━━━━━━━━━\n"
            "Desarrollado por K-Jaramillo\n"
            "Febrero 2026"
        )
        messagebox.showinfo("Acerca de LiquiVentas", info)
    
    def _cargar_datos_inicial(self):
        """Carga los datos de la fecha actual al iniciar la aplicación."""
        try:
            # Sincronizar fecha del DataStore con el selector
            if HAS_CALENDAR:
                fecha = self.fecha_global_entry.get_date().strftime('%Y-%m-%d')
            else:
                fecha = self.fecha_global_var.get().strip()
            
            self.ds.fecha = fecha
            if hasattr(self, 'fecha_asign_var'):
                self.fecha_asign_var.set(fecha)
            
            # Cargar facturas de Firebird
            self._cargar_facturas()
        except Exception as e:
            print(f"⚠️ Error cargando datos iniciales: {e}")

    # ------------------------------------------------------------------
    # CONFIGURAR ESTILOS MEJORADOS - MODO OSCURO/CLARO
    # ------------------------------------------------------------------
    def _configurar_estilos(self):
        """Configura estilos visuales profesionales para la interfaz."""
        estilo = ttk.Style()
        
        # Usar tema base 'clam' que permite mayor personalización
        try:
            estilo.theme_use('clam')
        except:
            pass
        
        # Variable para modo oscuro (por defecto True)
        if not hasattr(self, 'modo_oscuro'):
            self.modo_oscuro = True
        
        self._aplicar_tema()
    
    def _aplicar_tema(self):
        """Aplica el tema actual (oscuro o claro)."""
        estilo = ttk.Style()
        
        if self.modo_oscuro:
            # === COLORES DEL TEMA OSCURO ===
            PRIMARY = '#2196f3'          # Azul brillante
            PRIMARY_DARK = '#1976d2'     # Azul oscuro
            PRIMARY_LIGHT = '#64b5f6'    # Azul claro
            SUCCESS = '#4caf50'          # Verde
            SUCCESS_DARK = '#388e3c'
            WARNING = '#ff9800'          # Naranja
            ERROR = '#f44336'            # Rojo
            
            # Fondos oscuros
            BG_DARK = '#1e1e1e'          # Fondo principal
            BG_DARKER = '#121212'        # Fondo más oscuro
            BG_CARD = '#2d2d2d'          # Fondo de tarjetas/frames
            BG_HOVER = '#3d3d3d'         # Hover
            BG_SELECTED = '#0d47a1'      # Selección
            
            # Textos
            TEXT_PRIMARY = '#ffffff'      # Texto principal
            TEXT_SECONDARY = '#b0b0b0'    # Texto secundario
            TEXT_MUTED = '#757575'        # Texto apagado
            
            # Bordes
            BORDER_COLOR = '#404040'
            
            # Tags de Treeview
            TREE_ROW_PAR = '#2d2d2d'
            TREE_ROW_IMPAR = '#3d3d3d'
        else:
            # === COLORES DEL TEMA CLARO ===
            PRIMARY = '#1976d2'          # Azul
            PRIMARY_DARK = '#0d47a1'     # Azul oscuro
            PRIMARY_LIGHT = '#42a5f5'    # Azul claro
            SUCCESS = '#2e7d32'          # Verde
            SUCCESS_DARK = '#1b5e20'
            WARNING = '#f57c00'          # Naranja
            ERROR = '#c62828'            # Rojo
            
            # Fondos claros
            BG_DARK = '#f5f5f5'          # Fondo principal
            BG_DARKER = '#e0e0e0'        # Fondo más oscuro
            BG_CARD = '#ffffff'          # Fondo de tarjetas/frames
            BG_HOVER = '#eeeeee'         # Hover
            BG_SELECTED = '#bbdefb'      # Selección
            
            # Textos
            TEXT_PRIMARY = '#212121'      # Texto principal
            TEXT_SECONDARY = '#757575'    # Texto secundario
            TEXT_MUTED = '#9e9e9e'        # Texto apagado
            
            # Bordes
            BORDER_COLOR = '#bdbdbd'
            
            # Tags de Treeview
            TREE_ROW_PAR = '#ffffff'
            TREE_ROW_IMPAR = '#f5f5f5'
        
        # Guardar colores para uso en otras partes
        self.COLORS = {
            'bg_dark': BG_DARK,
            'bg_card': BG_CARD,
            'bg_hover': BG_HOVER,
            'primary': PRIMARY,
            'primary_dark': PRIMARY_DARK,
            'primary_light': PRIMARY_LIGHT,
            'success': SUCCESS,
            'success_dark': SUCCESS_DARK,
            'warning': WARNING,
            'error': ERROR,
            'text': TEXT_PRIMARY,
            'text_secondary': TEXT_SECONDARY,
            'border': BORDER_COLOR,
            'row_par': TREE_ROW_PAR,
            'row_impar': TREE_ROW_IMPAR
        }
        
        # Configurar fondo de ventana
        self.ventana.configure(bg=BG_DARK)
        
        # === ESTILOS DE TREEVIEW ===
        estilo.configure("Treeview",
                        background=BG_CARD,
                        foreground=TEXT_PRIMARY,
                        rowheight=28,
                        fieldbackground=BG_CARD,
                        font=("Segoe UI", 9),
                        borderwidth=0)
        estilo.configure("Treeview.Heading",
                        background=PRIMARY_DARK,
                        foreground="white",
                        font=("Segoe UI", 9, "bold"),
                        padding=(8, 6),
                        relief="flat")
        estilo.map("Treeview",
                   background=[("selected", PRIMARY)],
                   foreground=[("selected", "white")])
        estilo.map("Treeview.Heading",
                   background=[("active", PRIMARY), ("pressed", PRIMARY)])
        
        # === ESTILOS DE BOTONES ===
        estilo.configure("TButton",
                        padding=(12, 6),
                        font=("Segoe UI", 9),
                        background=BG_HOVER,
                        foreground=TEXT_PRIMARY)
        estilo.map("TButton",
                   background=[("active", PRIMARY), ("pressed", PRIMARY_DARK)],
                   foreground=[("active", "white"), ("pressed", "white")])
        
        estilo.configure("Primary.TButton",
                        padding=(16, 8),
                        font=("Segoe UI", 10, "bold"),
                        background=PRIMARY,
                        foreground="white")
        estilo.map("Primary.TButton",
                   background=[("active", PRIMARY_LIGHT), ("pressed", PRIMARY_DARK)])
        
        estilo.configure("Success.TButton",
                        background=SUCCESS,
                        foreground="white")
        estilo.map("Success.TButton",
                   background=[("active", SUCCESS_DARK)])
        
        # === ESTILOS DE LABELFRAME ===
        estilo.configure("TLabelframe",
                        background=BG_CARD,
                        borderwidth=1,
                        relief="solid",
                        bordercolor=BORDER_COLOR)
        estilo.configure("TLabelframe.Label",
                        background=BG_CARD,
                        foreground=PRIMARY_LIGHT,
                        font=("Segoe UI", 10, "bold"))
        
        # === ESTILOS DE NOTEBOOK (PESTAÑAS) ===
        estilo.configure("TNotebook",
                        background=BG_DARK,
                        borderwidth=0,
                        tabmargins=(4, 4, 4, 0))
        estilo.configure("TNotebook.Tab",
                        padding=(16, 8),
                        font=("Segoe UI", 9, "bold"),
                        background=BG_CARD,
                        foreground=TEXT_SECONDARY)
        estilo.map("TNotebook.Tab",
                   background=[("selected", PRIMARY_DARK), ("active", BG_HOVER)],
                   foreground=[("selected", "white"), ("active", TEXT_PRIMARY)],
                   expand=[("selected", (1, 1, 1, 0))])
        
        # === ESTILOS DE FRAME ===
        # Por defecto los frames usan BG_CARD para coincidir con LabelFrames
        estilo.configure("TFrame",
                        background=BG_CARD)
        estilo.configure("Card.TFrame",
                        background=BG_CARD)
        estilo.configure("Dark.TFrame",
                        background=BG_DARK)
        estilo.configure("Toolbar.TFrame",
                        background=BG_DARK)
        
        # === ESTILOS DE LABEL ===
        # Los labels por defecto usan BG_CARD para coincidir con LabelFrames
        estilo.configure("TLabel",
                        background=BG_CARD,
                        foreground=TEXT_PRIMARY,
                        font=("Segoe UI", 9))
        estilo.configure("Card.TLabel",
                        background=BG_CARD,
                        foreground=TEXT_PRIMARY)
        estilo.configure("Dark.TLabel",
                        background=BG_DARK,
                        foreground=TEXT_PRIMARY)
        estilo.configure("Toolbar.TLabel",
                        background=BG_DARK,
                        foreground=TEXT_PRIMARY)
        estilo.configure("Title.TLabel",
                        font=("Segoe UI", 14, "bold"),
                        foreground=PRIMARY_LIGHT,
                        background=BG_DARK)
        estilo.configure("Heading.TLabel",
                        font=("Segoe UI", 10, "bold"),
                        foreground=PRIMARY)
        estilo.configure("Success.TLabel",
                        foreground=SUCCESS)
        estilo.configure("Warning.TLabel",
                        foreground=WARNING)
        estilo.configure("Error.TLabel",
                        foreground=ERROR)
        
        # === ESTILOS DE ENTRY Y COMBOBOX ===
        estilo.configure("TEntry",
                        padding=(8, 4),
                        font=("Segoe UI", 9),
                        fieldbackground=BG_HOVER,
                        foreground=TEXT_PRIMARY,
                        insertcolor=TEXT_PRIMARY)
        estilo.map("TEntry",
                   fieldbackground=[("focus", BG_CARD)],
                   bordercolor=[("focus", PRIMARY)])
        
        estilo.configure("TCombobox",
                        padding=(8, 4),
                        font=("Segoe UI", 9),
                        fieldbackground=BG_HOVER,
                        foreground=TEXT_PRIMARY,
                        background=BG_HOVER,
                        arrowcolor=TEXT_PRIMARY)
        estilo.map("TCombobox",
                   fieldbackground=[("focus", BG_CARD), ("readonly", BG_HOVER)],
                   foreground=[("readonly", TEXT_PRIMARY)],
                   background=[("active", BG_CARD)])
        
        # === ESTILOS DE SCROLLBAR ===
        estilo.configure("TScrollbar",
                        background=BG_HOVER,
                        troughcolor=BG_DARK,
                        borderwidth=0,
                        arrowsize=14,
                        arrowcolor=TEXT_SECONDARY)
        estilo.map("TScrollbar",
                   background=[("active", PRIMARY), ("pressed", PRIMARY_DARK)])
        
        # === ESTILOS DE CHECKBUTTON Y RADIOBUTTON ===
        estilo.configure("TCheckbutton",
                        background=BG_CARD,
                        foreground=TEXT_PRIMARY)
        estilo.configure("TRadiobutton",
                        background=BG_CARD,
                        foreground=TEXT_PRIMARY)
        
        # === ESTILOS DE SEPARATOR ===
        estilo.configure("TSeparator",
                        background=BORDER_COLOR)
        
        # === ESTILOS DE PROGRESSBAR ===
        estilo.configure("TProgressbar",
                        background=PRIMARY,
                        troughcolor=BG_HOVER)
        
        # === ESTILOS DE SCALE ===
        estilo.configure("TScale",
                        background=BG_DARK,
                        troughcolor=BG_HOVER)

    # ------------------------------------------------------------------
    # FUNCIONES PARA COPIAR TEXTO DE TABLAS
    # ------------------------------------------------------------------
    def _copiar_seleccion_tree(self, tree, event=None):
        """Copia el contenido de la fila seleccionada al portapapeles."""
        seleccion = tree.selection()
        if not seleccion:
            return
        
        # Obtener valores de todas las filas seleccionadas
        lineas = []
        for item in seleccion:
            valores = tree.item(item, 'values')
            linea = '\t'.join(str(v) for v in valores)
            lineas.append(linea)
        
        texto = '\n'.join(lineas)
        
        # Copiar al portapapeles
        self.ventana.clipboard_clear()
        self.ventana.clipboard_append(texto)
        self.ventana.update()  # Necesario para que el portapapeles funcione
        
        # Mostrar notificación breve
        messagebox.showinfo("Copiado", f"Se copió al portapapeles:\n{texto[:100]}..." if len(texto) > 100 else f"Se copió al portapapeles:\n{texto}", parent=self.ventana)
    
    def _copiar_toda_tabla(self, tree):
        """Copia todo el contenido de la tabla al portapapeles."""
        items = tree.get_children()
        if not items:
            messagebox.showwarning("Sin datos", "No hay datos para copiar.")
            return
        
        # Obtener encabezados
        columnas = tree['columns']
        encabezados = [tree.heading(col, 'text') for col in columnas]
        lineas = ['\t'.join(encabezados)]
        
        # Obtener todas las filas
        for item in items:
            valores = tree.item(item, 'values')
            linea = '\t'.join(str(v) for v in valores)
            lineas.append(linea)
        
        texto = '\n'.join(lineas)
        
        # Copiar al portapapeles
        self.ventana.clipboard_clear()
        self.ventana.clipboard_append(texto)
        self.ventana.update()
        
        messagebox.showinfo("Copiado", f"Se copiaron {len(items)} filas al portapapeles.")
    
    def _mostrar_menu_copiar(self, tree, event):
        """Muestra menú contextual para copiar."""
        menu = tk.Menu(self.ventana, tearoff=0)
        menu.add_command(label="📋 Copiar fila seleccionada", 
                        command=lambda: self._copiar_seleccion_tree(tree))
        menu.add_command(label="📋 Copiar toda la tabla", 
                        command=lambda: self._copiar_toda_tabla(tree))
        menu.add_separator()
        menu.add_command(label="Cerrar", command=menu.destroy)
        
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
    
    def _crear_treeview_con_scroll(self, parent, columns, height=20, show="headings", selectmode="browse"):
        """Crea un Treeview con scrollbars horizontal y vertical."""
        # Frame contenedor
        container = ttk.Frame(parent)
        container.pack(fill=tk.BOTH, expand=True)
        
        # Crear Treeview
        tree = ttk.Treeview(container, columns=columns, height=height, 
                           show=show, selectmode=selectmode)
        
        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Scrollbar horizontal
        scroll_x = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=tree.xview)
        scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Configurar scrollbars
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Bindings para copiar (Ctrl+C y menú contextual)
        tree.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(tree))
        tree.bind("<Control-C>", lambda e: self._copiar_seleccion_tree(tree))
        
        return tree, container

    # ------------------------------------------------------------------
    # INTERFAZ PRINCIPAL
    # ------------------------------------------------------------------
    def _crear_interfaz(self):
        # ===== BARRA DE HERRAMIENTAS =====
        frame_config = ttk.Frame(self.ventana, padding=(10, 5))
        frame_config.pack(fill=tk.X, padx=8, pady=5)
        
        # ═══════════════════════════════════════════════════════════════
        # FILA 1: Conexión a BD y Tema
        # ═══════════════════════════════════════════════════════════════
        fila1 = ttk.Frame(frame_config)
        fila1.pack(fill=tk.X, pady=(0, 5))
        
        # Ruta FDB
        ttk.Label(fila1, text="📁 BD:").pack(side=tk.LEFT, padx=(0, 5))
        self.ruta_fdb_var = tk.StringVar(value=self.ruta_fdb)
        entry_fdb = ttk.Entry(fila1, textvariable=self.ruta_fdb_var, width=50)
        entry_fdb.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        
        ttk.Button(fila1, text="📂 Examinar", command=self._seleccionar_archivo_fdb, width=12).pack(side=tk.LEFT, padx=4)
        ttk.Button(fila1, text="🔗 Verificar", command=self._verificar_conexion_bd, width=10).pack(side=tk.LEFT, padx=4)
        
        # Indicador de estado
        self.lbl_estado_bd = ttk.Label(fila1, text="● Desconectado", foreground="red", width=15)
        self.lbl_estado_bd.pack(side=tk.LEFT, padx=(12, 0))
        
        # Separador
        ttk.Separator(fila1, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        # Switch de tema claro/oscuro con toggle visual
        self.tema_var = tk.BooleanVar(value=True)  # True = oscuro
        ttk.Label(fila1, text="☀️").pack(side=tk.LEFT, padx=(0, 5))
        # frame_bg inicial: modo oscuro = '#1e1e1e', modo claro = '#f5f5f5'
        self.switch_tema = ToggleSwitch(
            fila1, 
            variable=self.tema_var,
            command=self._toggle_tema,
            width=36, height=18,
            bg_off='#bdbdbd', bg_on='#4caf50',
            frame_bg='#1e1e1e'  # Fondo inicial modo oscuro
        )
        self.switch_tema.pack(side=tk.LEFT, padx=4)
        ttk.Label(fila1, text="🌙").pack(side=tk.LEFT, padx=(5, 0))
        
        # ═══════════════════════════════════════════════════════════════
        # FILA 2: Filtros globales (Fecha, Repartidor, Estado)
        # ═══════════════════════════════════════════════════════════════
        fila2 = ttk.Frame(frame_config)
        fila2.pack(fill=tk.X, pady=(5, 0))
        
        # Selector de fecha global
        ttk.Label(fila2, text="📅 Fecha:").pack(side=tk.LEFT, padx=(0, 5))
        if HAS_CALENDAR:
            self.fecha_global_entry = DateEntry(
                fila2, width=12,
                date_pattern='yyyy-mm-dd',
                background='#1e88e5', foreground='white', borderwidth=2
            )
            self.fecha_global_entry.set_date(datetime.now())
            self.fecha_global_entry.pack(side=tk.LEFT, padx=(0, 8))
            self.fecha_global_entry.bind("<<DateEntrySelected>>", self._on_fecha_global_cambio)
        else:
            self.fecha_global_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
            self.fecha_global_entry = ttk.Entry(fila2, textvariable=self.fecha_global_var, width=12)
            self.fecha_global_entry.pack(side=tk.LEFT, padx=(0, 5))
            self.fecha_global_entry.bind("<Return>", self._on_fecha_global_cambio)
            self.fecha_global_entry.bind("<FocusOut>", self._on_fecha_global_cambio)
        
        # Botón para cargar datos de la fecha seleccionada
        btn_cargar = ttk.Button(fila2, text="📥", width=3,
                   command=self._on_fecha_global_cambio)
        btn_cargar.pack(side=tk.LEFT, padx=(0, 8))
        self._crear_tooltip(btn_cargar, "Cargar")
        
        # Botones de navegación de fecha
        ttk.Button(fila2, text="◀", width=3,
                   command=lambda: self._cambiar_fecha_global(-1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(fila2, text="Hoy", width=5,
                   command=self._ir_a_fecha_hoy).pack(side=tk.LEFT, padx=2)
        ttk.Button(fila2, text="▶", width=3,
                   command=lambda: self._cambiar_fecha_global(1)).pack(side=tk.LEFT, padx=(2, 15))
        
        # Separador
        ttk.Separator(fila2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)
        
        # Filtro de repartidor global
        ttk.Label(fila2, text="🚚 Repartidor:").pack(side=tk.LEFT, padx=(0, 5))
        self.filtro_rep_global_var = tk.StringVar(value="(Todos)")
        self.combo_filtro_rep_global = ttk.Combobox(fila2, textvariable=self.filtro_rep_global_var,
                                    values=["(Todos)"],
                                    state="readonly", width=16)
        self.combo_filtro_rep_global.pack(side=tk.LEFT, padx=(0, 15))
        self.combo_filtro_rep_global.bind("<<ComboboxSelected>>", self._on_filtro_rep_global_cambio)
        
        # Separador
        ttk.Separator(fila2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)
        
        # Filtro general por estado (visible desde todas las pestañas)
        ttk.Label(fila2, text="▼ Estado:").pack(side=tk.LEFT, padx=(0, 5))
        self.filtro_estado_var = tk.StringVar(value="Todos")
        self.combo_filtro_estado = ttk.Combobox(fila2, textvariable=self.filtro_estado_var,
                                    values=["Todos", "Sin Repartidor", "Canceladas", "Crédito"],
                                    state="readonly", width=14)
        self.combo_filtro_estado.pack(side=tk.LEFT, padx=(0, 8))
        self.combo_filtro_estado.bind("<<ComboboxSelected>>", self._on_filtro_general_cambio)
        
        # Separador
        ttk.Separator(fila2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)
        
        # Buscador global (cliente/folio)
        ttk.Label(fila2, text="⌕ Cliente (F10):").pack(side=tk.LEFT, padx=(0, 5))
        self.buscar_global_var = tk.StringVar()
        self.entry_buscar_global = ttk.Entry(fila2, textvariable=self.buscar_global_var, width=20)
        self.entry_buscar_global.pack(side=tk.LEFT, padx=(0, 5))
        self.buscar_global_var.trace_add("write", lambda *a: self._on_buscar_global())
        # Enter en buscador -> saltar al listado
        self.entry_buscar_global.bind("<Return>", self._saltar_al_listado)
        self.entry_buscar_global.bind("<KP_Enter>", self._saltar_al_listado)
        ttk.Button(fila2, text="✕", width=2,
                   command=self._limpiar_buscar_global).pack(side=tk.LEFT, padx=(0, 8))
        
        # Separador antes de botones de acción
        ttk.Separator(fila2, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=12)
        
        # Botones de acción (movidos desde abajo)
        ttk.Button(fila2, text="💾 Liquidación",
                   command=self._guardar_liquidacion, style="Success.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(fila2, text="📄 Gen. Reporte",
                   command=self._generar_reporte).pack(side=tk.LEFT, padx=5)
        
        # ===== NOTEBOOK (PESTAÑAS) =====
        self.notebook = ttk.Notebook(self.ventana)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Pestaña 0 – Asignar Repartidor
        self.tab_asignacion = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_asignacion, text="  🚚 Asignar Repartidor  ")
        self._crear_tab_asignacion()

        # Pestaña 1 – Liquidación
        self.tab_liquidacion = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_liquidacion, text="  📊 Liquidación  ")
        self._crear_tab_liquidacion()

        # Pestaña 2 – Descuentos
        self.tab_descuentos = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_descuentos, text="  🏷 Descuentos  ")
        self._crear_tab_descuentos()

        # Pestaña 3 – Gastos Adicionales
        self.tab_gastos = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_gastos, text="  💸 Gastos  ")
        self._crear_tab_gastos()

        # Pestaña 4 – Conteo de Dinero
        self.tab_dinero = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_dinero, text="  💰 Conteo Dinero  ")
        self._crear_tab_dinero()

        # Pestaña 5 – Anotaciones (Sticky Notes) - antes era Auditoría
        if HAS_ANOTACIONES:
            self.tab_anotaciones = ttk.Frame(self.notebook)
            self.notebook.add(self.tab_anotaciones, text="  📝 Anotaciones  ")
            self._crear_tab_anotaciones()

        # Pestaña 7 – Créditos
        self.tab_creditos_punteados = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_creditos_punteados, text="  💳 Créditos  ")
        self._crear_tab_creditos_punteados()
        
        # Pestaña 8 – No Entregados
        self.tab_no_entregados = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_no_entregados, text="  📦 No Entregados  ")
        self._crear_tab_no_entregados()
        
        # Pestaña 9 – Préstamos
        self.tab_prestamos = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_prestamos, text="  💸 Préstamos  ")
        self._crear_tab_prestamos()
        
        # Pestaña 10 – Devoluciones Parciales
        self.tab_dev_parciales = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_dev_parciales, text="  🔄 Dev. Parciales  ")
        self._crear_tab_dev_parciales()
        
        # Pestaña 11 – Canceladas y Devoluciones
        self.tab_canceladas = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_canceladas, text="  ❌ Canceladas  ")
        self._crear_tab_canceladas()

    # ------------------------------------------------------------------
    # CREAR PESTAÑA DE CRÉDITOS
    # ------------------------------------------------------------------
    def _crear_tab_creditos_punteados(self):
        """Crea la pestaña de créditos unificada (Punteados + Eleventa)."""
        tab = self.tab_creditos_punteados
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE HERRAMIENTAS
        # ═══════════════════════════════════════════════════════════════
        toolbar = ttk.Frame(tab)
        toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        
        # Filtro de Estado interno
        ttk.Label(toolbar, text="Estado:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(5, 2))
        self.filtro_estado_creditos_var = tk.StringVar(value="PENDIENTE")
        self.combo_filtro_estado_creditos = ttk.Combobox(
            toolbar, 
            textvariable=self.filtro_estado_creditos_var,
            values=["Todos", "PENDIENTE", "PAGADO", "CANCELADA"],
            width=12,
            state="readonly"
        )
        self.combo_filtro_estado_creditos.pack(side=tk.LEFT, padx=5)
        self.combo_filtro_estado_creditos.bind("<<ComboboxSelected>>", lambda e: self._refrescar_creditos_tab())
        
        # Filtro de Origen
        ttk.Label(toolbar, text="Origen:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(10, 2))
        self.filtro_origen_creditos_var = tk.StringVar(value="Todos")
        self.combo_filtro_origen_creditos = ttk.Combobox(
            toolbar, 
            textvariable=self.filtro_origen_creditos_var,
            values=["Todos", "ELEVENTA", "PUNTEADO"],
            width=12,
            state="readonly"
        )
        self.combo_filtro_origen_creditos.pack(side=tk.LEFT, padx=5)
        self.combo_filtro_origen_creditos.bind("<<ComboboxSelected>>", lambda e: self._refrescar_creditos_tab())
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        # Botón para cargar TODOS los créditos de Firebird
        ttk.Button(toolbar, text="📥 Cargar Créditos Eleventa", 
                   command=self._cargar_todos_creditos_eleventa).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(toolbar, text="🔄 Actualizar", 
                   command=self._refrescar_creditos_tab).pack(side=tk.LEFT, padx=5)
        
        # Botón para saldar créditos viejos
        ttk.Button(toolbar, text="🗑️ Saldar Anteriores 2026", 
                   command=self._saldar_creditos_anteriores).pack(side=tk.LEFT, padx=5)
        
        # Total general
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)
        ttk.Label(toolbar, text="Total Pendiente:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=5)
        self.lbl_total_creditos_general = ttk.Label(toolbar, text="$0.00", 
                                                      font=("Segoe UI", 12, "bold"), 
                                                      foreground="#c62828")
        self.lbl_total_creditos_general.pack(side=tk.LEFT, padx=5)
        
        # Cantidad
        ttk.Label(toolbar, text="   Cantidad:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_cantidad_creditos = ttk.Label(toolbar, text="0", font=("Segoe UI", 10, "bold"))
        self.lbl_cantidad_creditos.pack(side=tk.LEFT, padx=5)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE FILTROS POR FECHA
        # ═══════════════════════════════════════════════════════════════
        toolbar2 = ttk.Frame(tab)
        toolbar2.grid(row=1, column=0, sticky="ew", padx=10, pady=2)
        
        # Filtro por Fecha Venta
        ttk.Label(toolbar2, text="📅 Fecha Venta:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_venta_desde = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                       background='#1976d2', foreground='white',
                                                       headersbackground='#1565c0')
            self.filtro_fecha_venta_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_venta_desde.delete(0, tk.END)
            self.filtro_fecha_venta_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_creditos_tab())
        else:
            self.filtro_fecha_venta_desde = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_venta_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_venta_desde.bind("<KeyRelease>", lambda e: self._refrescar_creditos_tab())
        
        ttk.Label(toolbar2, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_venta_hasta = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                       background='#1976d2', foreground='white',
                                                       headersbackground='#1565c0')
            self.filtro_fecha_venta_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_venta_hasta.delete(0, tk.END)
            self.filtro_fecha_venta_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_creditos_tab())
        else:
            self.filtro_fecha_venta_hasta = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_venta_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_venta_hasta.bind("<KeyRelease>", lambda e: self._refrescar_creditos_tab())
        
        # Separador visual
        ttk.Separator(toolbar2, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        # Filtro por Fecha Pagado
        ttk.Label(toolbar2, text="💰 Fecha Pagado:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_desde = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                        background='#2e7d32', foreground='white',
                                                        headersbackground='#1b5e20')
            self.filtro_fecha_pagado_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_pagado_desde.delete(0, tk.END)
            self.filtro_fecha_pagado_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_creditos_tab())
        else:
            self.filtro_fecha_pagado_desde = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_pagado_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_pagado_desde.bind("<KeyRelease>", lambda e: self._refrescar_creditos_tab())
        
        ttk.Label(toolbar2, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_hasta = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                        background='#2e7d32', foreground='white',
                                                        headersbackground='#1b5e20')
            self.filtro_fecha_pagado_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_pagado_hasta.delete(0, tk.END)
            self.filtro_fecha_pagado_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_creditos_tab())
        else:
            self.filtro_fecha_pagado_hasta = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_pagado_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_pagado_hasta.bind("<KeyRelease>", lambda e: self._refrescar_creditos_tab())
        
        ttk.Button(toolbar2, text="Hoy", width=5, 
                   command=lambda: self._set_fecha_hoy_creditos()).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar2, text="Limpiar", width=7, 
                   command=lambda: self._limpiar_filtro_fecha_creditos()).pack(side=tk.LEFT, padx=2)

        # ═══════════════════════════════════════════════════════════════
        # LISTADO UNIFICADO DE CRÉDITOS
        # ═══════════════════════════════════════════════════════════════
        frame_lista = ttk.Frame(tab)
        frame_lista.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        frame_lista.columnconfigure(0, weight=1)
        frame_lista.rowconfigure(0, weight=1)
        
        # Ajustar el peso de las filas
        tab.rowconfigure(2, weight=1)
        
        # Treeview unificado (orden: fecha, folio, cliente, valores, estado, fecha_pagado, repartidor, origen)
        columnas = ("fecha", "folio", "cliente", "valor_factura", "valor_credito", "abono", "saldo", "estado", "fecha_pagado", "repartidor", "origen")
        self.tree_creditos = ttk.Treeview(frame_lista, columns=columnas, show="headings", height=15)
        
        self.tree_creditos.heading("fecha", text="Fecha Venta", anchor=tk.CENTER)
        self.tree_creditos.heading("folio", text="Folio", anchor=tk.CENTER)
        self.tree_creditos.heading("cliente", text="Cliente", anchor=tk.W)
        self.tree_creditos.heading("valor_factura", text="Valor Factura", anchor=tk.E)
        self.tree_creditos.heading("valor_credito", text="Valor Crédito", anchor=tk.E)
        self.tree_creditos.heading("abono", text="Abono", anchor=tk.E)
        self.tree_creditos.heading("saldo", text="Saldo", anchor=tk.E)
        self.tree_creditos.heading("estado", text="Estado", anchor=tk.CENTER)
        self.tree_creditos.heading("fecha_pagado", text="Fecha Pagado", anchor=tk.CENTER)
        self.tree_creditos.heading("repartidor", text="Repartidor", anchor=tk.W)
        self.tree_creditos.heading("origen", text="Origen", anchor=tk.CENTER)
        
        self.tree_creditos.column("fecha", width=95, anchor=tk.CENTER)
        self.tree_creditos.column("folio", width=70, anchor=tk.CENTER)
        self.tree_creditos.column("cliente", width=180, anchor=tk.W)
        self.tree_creditos.column("valor_factura", width=100, anchor=tk.E)
        self.tree_creditos.column("valor_credito", width=100, anchor=tk.E)
        self.tree_creditos.column("abono", width=90, anchor=tk.E)
        self.tree_creditos.column("saldo", width=100, anchor=tk.E)
        self.tree_creditos.column("estado", width=90, anchor=tk.CENTER)
        self.tree_creditos.column("fecha_pagado", width=100, anchor=tk.CENTER)
        self.tree_creditos.column("repartidor", width=110, anchor=tk.W)
        self.tree_creditos.column("origen", width=90, anchor=tk.CENTER)
        
        scrolly = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=self.tree_creditos.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient=tk.HORIZONTAL, command=self.tree_creditos.xview)
        self.tree_creditos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_creditos.grid(row=0, column=0, sticky="nsew")
        scrolly.grid(row=0, column=1, sticky="ns")
        scrollx.grid(row=1, column=0, sticky="ew")
        
        # Clic simple para editar estado/abono in-place
        self.tree_creditos.bind("<Button-1>", self._on_clic_credito)
        # Doble clic para ver/editar observaciones
        self.tree_creditos.bind("<Double-1>", self._on_doble_clic_credito)
        
        # Tags para estados - colores más suaves y profesionales
        self.tree_creditos.tag_configure("pagado", background="#1b5e20", foreground="#a5d6a7")    # Verde
        self.tree_creditos.tag_configure("pendiente", background="#e65100", foreground="#ffe0b2") # Naranja
        self.tree_creditos.tag_configure("cancelada", background="#880e4f", foreground="#f8bbd0") # Rosa
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA INFERIOR DE TOTALES FILTRADOS
        # ═══════════════════════════════════════════════════════════════
        frame_totales_bottom = ttk.Frame(tab)
        frame_totales_bottom.grid(row=3, column=0, sticky="ew", padx=10, pady=5)
        
        ttk.Label(frame_totales_bottom, text="📊 TOTALES FILTRADOS:", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(frame_totales_bottom, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Label(frame_totales_bottom, text="Total Pagado:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_total_pagado_filtrado = ttk.Label(frame_totales_bottom, text="$0", 
                                                     font=("Segoe UI", 11, "bold"), foreground="#81c784")
        self.lbl_total_pagado_filtrado.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(frame_totales_bottom, text="(", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.lbl_cant_pagado_filtrado = ttk.Label(frame_totales_bottom, text="0", font=("Segoe UI", 9, "bold"))
        self.lbl_cant_pagado_filtrado.pack(side=tk.LEFT)
        ttk.Label(frame_totales_bottom, text=" registros)", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Separator(frame_totales_bottom, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Label(frame_totales_bottom, text="Total Pendiente:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_total_pendiente_filtrado = ttk.Label(frame_totales_bottom, text="$0", 
                                                        font=("Segoe UI", 11, "bold"), foreground="#ffb74d")
        self.lbl_total_pendiente_filtrado.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(frame_totales_bottom, text="(", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.lbl_cant_pendiente_filtrado = ttk.Label(frame_totales_bottom, text="0", font=("Segoe UI", 9, "bold"))
        self.lbl_cant_pendiente_filtrado.pack(side=tk.LEFT)
        ttk.Label(frame_totales_bottom, text=" registros)", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        
        # Widget flotante para edición in-place
        self.credito_edit_widget = None
        
        # Variables para compatibilidad (ya no usamos dos trees)
        self.tree_creditos_punt = self.tree_creditos
        self.tree_creditos_elev = self.tree_creditos
        self.lbl_total_creditos_punt_tab = self.lbl_total_creditos_general
        self.lbl_total_creditos_elev_tab = self.lbl_total_creditos_general
        self.lbl_cantidad_creditos_punt = self.lbl_cantidad_creditos
        self.lbl_cantidad_creditos_elev = self.lbl_cantidad_creditos
        
        # Cargar datos iniciales desde SQLite
        self._refrescar_creditos_tab()
    
    def _set_fecha_hoy_creditos(self):
        """Establece la fecha de hoy en los filtros de fecha venta de créditos."""
        from datetime import date
        hoy = date.today()
        # Solo setea fecha venta
        if HAS_CALENDAR:
            self.filtro_fecha_venta_desde.set_date(hoy)
            self.filtro_fecha_venta_hasta.set_date(hoy)
        else:
            self.filtro_fecha_venta_desde.delete(0, tk.END)
            self.filtro_fecha_venta_desde.insert(0, hoy.isoformat())
            self.filtro_fecha_venta_hasta.delete(0, tk.END)
            self.filtro_fecha_venta_hasta.insert(0, hoy.isoformat())
        self._refrescar_creditos_tab()
    
    def _limpiar_filtro_fecha_creditos(self):
        """Limpia los filtros de fecha de créditos."""
        # Limpiar fecha venta
        self.filtro_fecha_venta_desde.delete(0, tk.END)
        self.filtro_fecha_venta_hasta.delete(0, tk.END)
        # Limpiar fecha pagado
        self.filtro_fecha_pagado_desde.delete(0, tk.END)
        self.filtro_fecha_pagado_hasta.delete(0, tk.END)
        self._refrescar_creditos_tab()
    
    def _set_fecha_hoy_ne(self):
        """Establece la fecha de hoy en los filtros de No Entregados."""
        from datetime import date
        hoy = date.today()
        # Fecha Venta
        if HAS_CALENDAR:
            self.filtro_fecha_venta_ne_desde.set_date(hoy)
            self.filtro_fecha_venta_ne_hasta.set_date(hoy)
        else:
            self.filtro_fecha_venta_ne_desde.delete(0, tk.END)
            self.filtro_fecha_venta_ne_desde.insert(0, hoy.isoformat())
            self.filtro_fecha_venta_ne_hasta.delete(0, tk.END)
            self.filtro_fecha_venta_ne_hasta.insert(0, hoy.isoformat())
        self._refrescar_no_entregados_tab()
    
    def _set_fecha_hoy_pagado_ne(self):
        """Establece la fecha de hoy en filtro de Fecha Pagado de No Entregados."""
        from datetime import date
        hoy = date.today()
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_ne_desde.set_date(hoy)
            self.filtro_fecha_pagado_ne_hasta.set_date(hoy)
        else:
            self.filtro_fecha_pagado_ne_desde.delete(0, tk.END)
            self.filtro_fecha_pagado_ne_desde.insert(0, hoy.isoformat())
            self.filtro_fecha_pagado_ne_hasta.delete(0, tk.END)
            self.filtro_fecha_pagado_ne_hasta.insert(0, hoy.isoformat())
        self._refrescar_no_entregados_tab()
    
    def _limpiar_filtro_fecha_ne(self):
        """Limpia los filtros de fecha de No Entregados."""
        # Limpiar fecha venta
        self.filtro_fecha_venta_ne_desde.delete(0, tk.END)
        self.filtro_fecha_venta_ne_hasta.delete(0, tk.END)
        # Limpiar fecha pagado
        self.filtro_fecha_pagado_ne_desde.delete(0, tk.END)
        self.filtro_fecha_pagado_ne_hasta.delete(0, tk.END)
        self._refrescar_no_entregados_tab()
    
    def _on_clic_credito(self, event):
        """Maneja clic en créditos para editar abono o estado in-place."""
        # Cerrar widget de edición previo
        self._cerrar_edicion_credito()
        
        # Identificar fila y columna
        item_id = self.tree_creditos.identify_row(event.y)
        column = self.tree_creditos.identify_column(event.x)
        
        if not item_id or not column:
            return
        
        col_idx = int(column.replace('#', '')) - 1
        # Orden: fecha, folio, cliente, valor_factura, valor_credito, abono, saldo, estado, fecha_pagado, repartidor, origen
        # Índices: 0      1      2        3              4              5      6      7       8             9           10
        columnas = ("fecha", "folio", "cliente", "valor_factura", "valor_credito", "abono", "saldo", "estado", "fecha_pagado", "repartidor", "origen")
        
        if col_idx < 0 or col_idx >= len(columnas):
            return
        
        col_name = columnas[col_idx]
        
        # Solo editar columnas abono, estado y repartidor
        if col_name not in ("abono", "estado", "repartidor"):
            # Seleccionar la fila normalmente
            self.tree_creditos.selection_set(item_id)
            return
        
        values = self.tree_creditos.item(item_id, 'values')
        # Índices: fecha=0, folio=1, cliente=2, valor_factura=3, valor_credito=4, abono=5, saldo=6, estado=7, fecha_pagado=8, repartidor=9, origen=10
        fecha = values[0]
        folio = int(values[1])
        origen = values[10]  # ELEVENTA o PUNTEADO
        tipo = 'eleventa' if origen == 'ELEVENTA' else 'punteado'
        
        # Obtener coordenadas de la celda
        bbox = self.tree_creditos.bbox(item_id, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        if col_name == "abono":
            self._crear_entry_abono(item_id, tipo, fecha, folio, x, y, width, height, values)
        elif col_name == "estado":
            self._crear_combo_estado(item_id, tipo, fecha, folio, x, y, width, height, values)
        elif col_name == "repartidor":
            self._crear_combo_repartidor_credito(item_id, tipo, fecha, folio, x, y, width, height, values)
    
    def _cerrar_edicion_credito(self, event=None):
        """Cierra el widget de edición in-place."""
        if hasattr(self, 'credito_edit_widget') and self.credito_edit_widget:
            try:
                self.credito_edit_widget.destroy()
            except:
                pass
            self.credito_edit_widget = None
        if hasattr(self, 'credito_edit_frame') and self.credito_edit_frame:
            try:
                self.credito_edit_frame.destroy()
            except:
                pass
            self.credito_edit_frame = None
    
    def _crear_entry_abono(self, item_id, tipo, fecha, folio, x, y, width, height, values):
        """Crea Entry in-place con botón para editar abono."""
        # Índices: fecha=0, folio=1, cliente=2, valor_factura=3, valor_credito=4, abono=5, saldo=6, estado=7, fecha_pagado=8, repartidor=9, origen=10
        abono_actual = values[5].replace('$', '').replace(',', '') if len(values) > 5 and values[5] else '0'
        valor_credito_str = values[4].replace('$', '').replace(',', '') if len(values) > 4 and values[4] else '0'
        cliente = values[2] if len(values) > 2 else ''
        
        try:
            valor_credito = float(valor_credito_str)
        except:
            valor_credito = 0
        
        # Frame contenedor para Entry + Botón
        frame = tk.Frame(self.tree_creditos, bg='white', highlightbackground='#1976d2', highlightthickness=2)
        frame.place(x=x-5, y=y, width=width+80, height=height+4)
        
        self.credito_edit_frame = frame
        
        entry = tk.Entry(frame, font=("Segoe UI", 10), justify='right', bd=0)
        entry.insert(0, abono_actual)
        entry.select_range(0, tk.END)
        entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        entry.focus_set()
        
        self.credito_edit_widget = entry
        
        def guardar(event=None):
            try:
                nuevo_abono = round(float(entry.get().replace(',', '').replace('$', '')))
                # Validar que el abono no sea mayor al valor del crédito
                if nuevo_abono > valor_credito:
                    entry.config(background='#ffcccc')
                    messagebox.showwarning("Advertencia", 
                        f"El abono (${nuevo_abono:,.0f}) no puede ser mayor al valor del crédito (${valor_credito:,.0f})",
                        parent=self.ventana)
                    return
                
                # Guardar en SQLite (las funciones ahora retornan dict con info completa)
                if tipo == 'punteado':
                    resultado = db_local.actualizar_abono_credito_punteado(fecha, folio, nuevo_abono)
                else:
                    resultado = db_local.actualizar_abono_credito_eleventa(fecha, folio, nuevo_abono)
                
                if isinstance(resultado, dict) and resultado.get('success'):
                    nuevo_saldo = resultado.get('nuevo_saldo', 0)
                    nuevo_estado = resultado.get('nuevo_estado', '')
                    cambio_estado = resultado.get('cambio_estado', False)
                    
                    msg = f"✅ Folio {folio} | Abono: ${nuevo_abono:,.0f} | Saldo: ${nuevo_saldo:,.0f}"
                    if cambio_estado:
                        msg += f" | Estado: {nuevo_estado}"
                        if nuevo_estado == 'PAGADO':
                            messagebox.showinfo("Crédito Pagado", 
                                f"¡El crédito del folio {folio} ha sido pagado completamente!\n\n"
                                f"Cliente: {cliente}\n"
                                f"Valor Crédito: ${valor_credito:,.0f}\n"
                                f"Total Abonado: ${nuevo_abono:,.0f}\n"
                                f"Saldo: ${nuevo_saldo:,.0f}",
                                parent=self.ventana)
                    print(msg)
                elif isinstance(resultado, dict):
                    messagebox.showerror("Error", resultado.get('error', 'Error desconocido'), parent=self.ventana)
                
                self._cerrar_edicion_credito()
                self._refrescar_creditos_tab()
                # Actualizar también la liquidación para reflejar los créditos cobrados
                self._actualizar_datos_liquidacion()
            except ValueError:
                entry.config(background='#ffcccc')
        
        def cancelar(event=None):
            self._cerrar_edicion_credito()
        
        # Botón Guardar
        btn_guardar = tk.Button(frame, text="✓", font=("Segoe UI", 9, "bold"), 
                                bg='#4caf50', fg='white', bd=0, width=3,
                                command=guardar, cursor='hand2')
        btn_guardar.pack(side=tk.LEFT, padx=1)
        
        # Botón Cancelar
        btn_cancelar = tk.Button(frame, text="✗", font=("Segoe UI", 9, "bold"), 
                                 bg='#f44336', fg='white', bd=0, width=3,
                                 command=cancelar, cursor='hand2')
        btn_cancelar.pack(side=tk.LEFT, padx=1)
        
        entry.bind("<Return>", guardar)
        entry.bind("<Escape>", cancelar)
    
    def _crear_combo_estado(self, item_id, tipo, fecha, folio, x, y, width, height, values):
        """Crea Combobox in-place para seleccionar estado."""
        # Índice 7 = estado (fecha=0, folio=1, cliente=2, valor_factura=3, valor_credito=4, abono=5, saldo=6, estado=7)
        estado_actual = values[7] if len(values) > 7 else 'PENDIENTE'
        
        # Frame contenedor
        frame = tk.Frame(self.tree_creditos, bg='white', highlightbackground='#1976d2', highlightthickness=2)
        frame.place(x=x-5, y=y, width=width+50, height=height+4)
        
        self.credito_edit_frame = frame
        
        combo = ttk.Combobox(frame, values=["PENDIENTE", "PAGADO", "CANCELADA"],
                             state="readonly", font=("Segoe UI", 9), width=12)
        combo.set(estado_actual)
        combo.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        combo.focus_set()
        
        self.credito_edit_widget = combo
        
        def guardar(event=None):
            nuevo_estado = combo.get()
            # Guardar en SQLite
            if tipo == 'punteado':
                resultado = db_local.actualizar_estado_credito_punteado(fecha, folio, nuevo_estado)
            else:
                resultado = db_local.actualizar_estado_credito_eleventa(fecha, folio, nuevo_estado)
            
            if resultado:
                print(f"✅ Estado actualizado: Folio {folio} | Estado: {nuevo_estado}")
            
            self._cerrar_edicion_credito()
            self._refrescar_creditos_tab()
        
        def cancelar(event=None):
            self._cerrar_edicion_credito()
        
        # Botón Cancelar
        btn_cancelar = tk.Button(frame, text="✗", font=("Segoe UI", 9, "bold"), 
                                 bg='#f44336', fg='white', bd=0, width=3,
                                 command=cancelar, cursor='hand2')
        btn_cancelar.pack(side=tk.LEFT, padx=1)
        
        combo.bind("<<ComboboxSelected>>", guardar)
        combo.bind("<Escape>", cancelar)
        combo.bind("<FocusOut>", cancelar)
    
    def _crear_combo_repartidor_credito(self, item_id, tipo, fecha, folio, x, y, width, height, values):
        """Crea Combobox in-place para seleccionar repartidor del crédito."""
        # Índice 9 = repartidor (fecha=0, folio=1, cliente=2, valor_factura=3, valor_credito=4, abono=5, saldo=6, estado=7, fecha_pagado=8, repartidor=9, origen=10)
        repartidor_actual = values[9] if len(values) > 9 else ''
        
        # Obtener lista de repartidores
        repartidores = [""] + list(self.ds.repartidores.keys())
        
        # Frame contenedor
        frame = tk.Frame(self.tree_creditos, bg='white', highlightbackground='#1976d2', highlightthickness=2)
        frame.place(x=x-5, y=y, width=width+50, height=height+4)
        
        self.credito_edit_frame = frame
        
        combo = ttk.Combobox(frame, values=repartidores, font=("Segoe UI", 9), width=15)
        combo.set(repartidor_actual)
        combo.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        combo.focus_set()
        
        self.credito_edit_widget = combo
        
        def guardar(event=None):
            nuevo_repartidor = combo.get()
            # Guardar en SQLite
            if tipo == 'punteado':
                resultado = db_local.actualizar_repartidor_credito_punteado(fecha, folio, nuevo_repartidor)
            else:
                resultado = db_local.actualizar_repartidor_credito_eleventa(fecha, folio, nuevo_repartidor)
            
            if resultado:
                print(f"✅ Repartidor actualizado: Folio {folio} | Repartidor: {nuevo_repartidor}")
            
            self._cerrar_edicion_credito()
            self._refrescar_creditos_tab()
        
        def cancelar(event=None):
            self._cerrar_edicion_credito()
        
        # Botón Guardar
        btn_guardar = tk.Button(frame, text="✓", font=("Segoe UI", 9, "bold"), 
                                bg='#4caf50', fg='white', bd=0, width=3,
                                command=guardar, cursor='hand2')
        btn_guardar.pack(side=tk.LEFT, padx=1)
        
        combo.bind("<Return>", guardar)
        combo.bind("<Escape>", cancelar)
    
    def _on_doble_clic_credito(self, event):
        """Maneja doble clic para ver/editar observaciones del crédito."""
        item_id = self.tree_creditos.identify_row(event.y)
        if not item_id:
            return
        
        values = self.tree_creditos.item(item_id, 'values')
        # Índices: 0=fecha, 1=folio, 2=cliente, 3=valor_factura, 4=valor_credito, 5=abono, 6=saldo, 7=estado, 8=fecha_pagado, 9=repartidor, 10=origen
        fecha = values[0]
        folio = int(values[1])
        cliente = values[2]
        origen = values[10]  # ELEVENTA o PUNTEADO
        tipo = 'eleventa' if origen == 'ELEVENTA' else 'punteado'
        
        # Obtener observaciones actuales
        if tipo == 'punteado':
            credito = db_local.obtener_credito_punteado(fecha, folio)
        else:
            credito = db_local.obtener_credito_eleventa(fecha, folio)
        
        obs_actual = credito.get('observaciones', '') if credito else ''
        
        # Crear diálogo para editar observaciones
        dialog = tk.Toplevel(self.ventana)
        dialog.title(f"Observaciones - Folio {folio}")
        dialog.geometry("500x320")
        dialog.transient(self.ventana)
        dialog.grab_set()
        dialog.configure(bg='#f5f5f5')
        dialog.resizable(False, False)
        
        # Centrar
        dialog.update_idletasks()
        x = self.ventana.winfo_x() + (self.ventana.winfo_width() // 2) - 250
        y = self.ventana.winfo_y() + (self.ventana.winfo_height() // 2) - 160
        dialog.geometry(f"+{x}+{y}")
        
        # Header con info del crédito
        frame_header = tk.Frame(dialog, bg='#1976d2', height=60)
        frame_header.pack(fill=tk.X)
        frame_header.pack_propagate(False)
        
        tk.Label(frame_header, text=f"Folio: {folio}", font=("Segoe UI", 14, "bold"),
                 bg='#1976d2', fg='white').pack(anchor='w', padx=15, pady=(10, 0))
        tk.Label(frame_header, text=f"{cliente} | {fecha} | {origen}",
                 font=("Segoe UI", 9), bg='#1976d2', fg='#bbdefb').pack(anchor='w', padx=15)
        
        # Contenido
        frame_content = tk.Frame(dialog, bg='#f5f5f5')
        frame_content.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        tk.Label(frame_content, text="Observaciones / Novedades:", font=("Segoe UI", 10, "bold"),
                 bg='#f5f5f5', fg='#333333', anchor='w').pack(fill=tk.X, pady=(0, 5))
        
        # Frame para el texto con borde visual
        frame_text = tk.Frame(frame_content, bg='#cccccc', bd=1, relief=tk.SOLID)
        frame_text.pack(fill=tk.BOTH, expand=True)
        
        text_obs = tk.Text(frame_text, height=8, font=("Segoe UI", 10), wrap=tk.WORD,
                           bg='white', fg='#333333', insertbackground='#333333',
                           relief=tk.FLAT, padx=10, pady=10)
        text_obs.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        text_obs.insert('1.0', obs_actual)
        text_obs.focus_set()
        
        # Botones al fondo
        frame_btns = tk.Frame(dialog, bg='#f5f5f5')
        frame_btns.pack(fill=tk.X, padx=20, pady=(0, 20))
        
        def guardar():
            nueva_obs = text_obs.get('1.0', tk.END).strip()
            if tipo == 'punteado':
                db_local.actualizar_observaciones_credito_punteado(fecha, folio, nueva_obs)
            else:
                db_local.actualizar_observaciones_credito_eleventa(fecha, folio, nueva_obs)
            print(f"Observaciones actualizadas: Folio {folio}")
            dialog.destroy()
        
        btn_guardar = tk.Button(frame_btns, text="Guardar", font=("Segoe UI", 10, "bold"),
                  bg='#4caf50', fg='white', width=12, cursor='hand2', relief=tk.FLAT,
                  activebackground='#388e3c', activeforeground='white',
                  command=guardar)
        btn_guardar.pack(side=tk.RIGHT, padx=(10, 0))
        
        btn_cancelar = tk.Button(frame_btns, text="Cancelar", font=("Segoe UI", 10),
                  bg='#9e9e9e', fg='white', width=12, cursor='hand2', relief=tk.FLAT,
                  activebackground='#757575', activeforeground='white',
                  command=dialog.destroy)
        btn_cancelar.pack(side=tk.RIGHT)
        
        dialog.bind("<Escape>", lambda e: dialog.destroy())
    
    def _saldar_creditos_anteriores(self):
        """Salda automáticamente todos los créditos anteriores al 01 de enero de 2026."""
        respuesta = messagebox.askyesno(
            "Confirmar Saldado Masivo",
            "¿Está seguro de saldar TODOS los créditos anteriores al 01 de Enero de 2026?\n\n"
            "Esto pondrá el ABONO igual al VALOR CRÉDITO para todos esos registros,\n"
            "cambiando su estado a PAGADO.\n\n"
            "Esta acción quedará registrada en el historial y NO se puede deshacer.",
            parent=self.ventana
        )
        
        if not respuesta:
            return
        
        try:
            resultado = db_local.saldar_creditos_anteriores_a_fecha('2026-01-01')
            
            total_creditos = resultado['eleventa_count'] + resultado['punteados_count']
            total_monto = resultado['total_saldado_eleventa'] + resultado['total_saldado_punteados']
            
            if total_creditos > 0:
                messagebox.showinfo(
                    "Saldado Completado",
                    f"Se saldaron {total_creditos} créditos anteriores al 01/01/2026:\n\n"
                    f"• Eleventa: {resultado['eleventa_count']} créditos (${resultado['total_saldado_eleventa']:,.2f})\n"
                    f"• Punteados: {resultado['punteados_count']} créditos (${resultado['total_saldado_punteados']:,.2f})\n\n"
                    f"Total saldado: ${total_monto:,.2f}\n\n"
                    f"Todos los cambios quedaron registrados en el historial.",
                    parent=self.ventana
                )
            else:
                messagebox.showinfo(
                    "Sin Cambios",
                    "No se encontraron créditos pendientes anteriores al 01/01/2026.",
                    parent=self.ventana
                )
            
            # Refrescar la tabla
            self._refrescar_creditos_tab()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al saldar créditos: {e}", parent=self.ventana)
    
    def _cargar_todos_creditos_eleventa(self):
        """Consulta TODOS los créditos de Firebird y los guarda en SQLite."""
        if not self.ruta_fdb or not os.path.exists(self.ruta_fdb):
            messagebox.showerror("Error", "No se ha configurado la ruta de la base de datos Firebird.")
            return
        
        # Consulta SQL para obtener TODOS los créditos (sin filtro de fecha)
        sql = (
            "SET HEADING ON;\n"
            "SELECT V.ID, V.FOLIO, V.NOMBRE, V.SUBTOTAL, V.TOTAL_CREDITO, "
            "CAST(V.CREADO_EN AS DATE) AS FECHA_CREACION\n"
            "FROM VENTATICKETS V\n"
            "WHERE V.TOTAL_CREDITO > 0\n"
            "ORDER BY V.CREADO_EN DESC, V.FOLIO;\n"
        )
        
        ok, stdout, stderr = self._ejecutar_sql(sql)
        
        if not ok or not stdout:
            error_msg = stderr or "No se recibieron datos de la BD"
            messagebox.showerror("Error BD", f"No se pudo consultar créditos:\n{error_msg}")
            return
        
        creditos = []
        header_visto = False
        
        try:
            for linea in stdout.split('\n'):
                linea = linea.strip()
                if not linea or linea.startswith('='):
                    continue
                # Detectar header
                if 'ID' in linea and 'FOLIO' in linea:
                    header_visto = True
                    continue
                if not header_visto:
                    continue
                
                partes = linea.split()
                if len(partes) < 5:
                    continue
                
                try:
                    id_v = int(partes[0])
                    folio_s = partes[1]
                    if folio_s == '<null>':
                        continue
                    folio = int(folio_s)
                    
                    # Fecha está al final
                    fecha = partes[-1] if partes[-1] != '<null>' else ''
                    # Total crédito está antes de la fecha
                    total_credito = float(partes[-2]) if partes[-2] != '<null>' else 0.0
                    # Subtotal está antes del total crédito
                    subtotal = float(partes[-3]) if partes[-3] != '<null>' else 0.0
                    # Nombre está entre FOLIO y SUBTOTAL
                    nombre = ' '.join(partes[2:-3]).replace('<null>', '').strip()
                    if not nombre:
                        nombre = 'MOSTRADOR'
                    
                    if folio <= 0 or total_credito <= 0:
                        continue
                    
                    creditos.append({
                        'fecha': fecha,
                        'folio': folio,
                        'id': id_v,
                        'nombre': nombre,
                        'subtotal': subtotal,
                        'total_credito': total_credito,
                        'repartidor': ''
                    })
                except (ValueError, IndexError):
                    continue
            
            if not creditos:
                messagebox.showinfo("Info", "No se encontraron créditos en el sistema Eleventa.")
                return
            
            # Guardar en SQLite por fecha
            creditos_por_fecha = {}
            for c in creditos:
                fecha = c['fecha']
                if fecha not in creditos_por_fecha:
                    creditos_por_fecha[fecha] = []
                creditos_por_fecha[fecha].append(c)
            
            total_guardados = 0
            for fecha, lista in creditos_por_fecha.items():
                count = db_local.guardar_creditos_eleventa_bulk(fecha, lista)
                total_guardados += count
            
            messagebox.showinfo("Carga Completa", 
                f"Se cargaron {total_guardados} créditos de {len(creditos_por_fecha)} fechas diferentes.")
            
            # Refrescar vista
            self._refrescar_creditos_tab()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error procesando créditos: {e}")
    
    def _refrescar_creditos_tab(self):
        """Refresca la lista unificada de créditos (Punteados + Eleventa)."""
        # Limpiar treeview unificado
        self.tree_creditos.delete(*self.tree_creditos.get_children())
        
        if not USE_SQLITE:
            return
        
        # Obtener filtros
        filtro_cliente = self.buscar_global_var.get().strip().lower() if hasattr(self, 'buscar_global_var') else ""
        filtro_estado = self.filtro_estado_creditos_var.get() if hasattr(self, 'filtro_estado_creditos_var') else "Todos"
        filtro_origen = self.filtro_origen_creditos_var.get() if hasattr(self, 'filtro_origen_creditos_var') else "Todos"
        
        # Filtros de fecha venta
        filtro_venta_desde = self.filtro_fecha_venta_desde.get().strip() if hasattr(self, 'filtro_fecha_venta_desde') else ""
        filtro_venta_hasta = self.filtro_fecha_venta_hasta.get().strip() if hasattr(self, 'filtro_fecha_venta_hasta') else ""
        
        # Filtros de fecha pagado
        filtro_pagado_desde = self.filtro_fecha_pagado_desde.get().strip() if hasattr(self, 'filtro_fecha_pagado_desde') else ""
        filtro_pagado_hasta = self.filtro_fecha_pagado_hasta.get().strip() if hasattr(self, 'filtro_fecha_pagado_hasta') else ""
        
        # Cargar asignaciones para obtener repartidores
        asignaciones = cargar_asignaciones()
        
        # Configurar tags para colores según estado - colores suaves y profesionales
        self.tree_creditos.tag_configure("pagado", background="#1b5e20", foreground="#a5d6a7")    # Verde
        self.tree_creditos.tag_configure("pendiente", background="#e65100", foreground="#ffe0b2") # Naranja
        self.tree_creditos.tag_configure("cancelada", background="#880e4f", foreground="#f8bbd0") # Rosa
        
        # Lista unificada de todos los créditos
        creditos_unificados = []
        
        # ═══════════════════════════════════════════════════════════════
        # CRÉDITOS PUNTEADOS (Manuales)
        # ═══════════════════════════════════════════════════════════════
        if filtro_origen in ("Todos", "PUNTEADO"):
            creditos_punt = db_local.obtener_todos_creditos_punteados()
            for cp in creditos_punt:
                fecha = cp.get('fecha', '')
                folio = cp.get('folio', '')
                # Buscar repartidor en asignaciones
                key = f"{fecha}_{folio}"
                repartidor = asignaciones.get(key, '') or cp.get('repartidor', '') or ''
                
                # Obtener cliente (mostrar tal cual viene de la BD)
                cliente = cp.get('cliente', '') or 'MOSTRADOR'
                
                creditos_unificados.append({
                    'fecha': fecha,
                    'folio': folio,
                    'cliente': cliente,
                    'repartidor': repartidor,
                    'valor_factura': cp.get('subtotal', 0) or 0,
                    'valor_credito': cp.get('valor_credito', 0) or cp.get('subtotal', 0) or 0,
                    'abono': cp.get('abono', 0) or 0,
                    'estado': cp.get('estado', 'PENDIENTE') or 'PENDIENTE',
                    'fecha_pagado': cp.get('fecha_pagado', '') or '',
                    'origen': 'PUNTEADO',
                    'observaciones': cp.get('observaciones', '') or ''
                })
        
        # ═══════════════════════════════════════════════════════════════
        # CRÉDITOS ELEVENTA (Sistema)
        # ═══════════════════════════════════════════════════════════════
        if filtro_origen in ("Todos", "ELEVENTA"):
            creditos_elev = db_local.obtener_todos_creditos_eleventa()
            for ce in creditos_elev:
                fecha = ce.get('fecha', '')
                folio = ce.get('folio', '')
                valor_factura = ce.get('subtotal', 0) or 0
                valor_credito = ce.get('total_credito', 0) or 0
                estado_guardado = ce.get('estado', 'PENDIENTE') or 'PENDIENTE'
                
                # Buscar repartidor en asignaciones
                key = f"{fecha}_{folio}"
                repartidor = asignaciones.get(key, '') or ce.get('repartidor', '') or ''
                
                # Detectar facturas CANCELADAS: valor_factura=0 pero valor_credito>0
                if valor_factura == 0 and valor_credito > 0 and estado_guardado not in ('PAGADO',):
                    estado = 'CANCELADA'
                else:
                    estado = estado_guardado
                
                # Obtener cliente (mostrar tal cual viene de la BD)
                cliente = ce.get('cliente', '') or 'MOSTRADOR'
                
                creditos_unificados.append({
                    'fecha': fecha,
                    'folio': folio,
                    'cliente': cliente,
                    'repartidor': repartidor,
                    'valor_factura': valor_factura,
                    'valor_credito': valor_credito,
                    'abono': ce.get('abono', 0) or 0,
                    'estado': estado,
                    'fecha_pagado': ce.get('fecha_pagado', '') or '',
                    'origen': 'ELEVENTA',
                    'observaciones': ce.get('observaciones', '') or ''
                })
        
        # Ordenar por fecha descendente
        creditos_unificados.sort(key=lambda x: x['fecha'], reverse=True)
        
        # Poblar treeview unificado
        total_credito = 0
        total_pendiente = 0
        count_total = 0
        
        # Totales para barra inferior
        total_pagado_filtrado = 0
        count_pagado_filtrado = 0
        total_pendiente_filtrado = 0
        count_pendiente_filtrado = 0
        
        for c in creditos_unificados:
            fecha = c['fecha']
            folio = c['folio']
            cliente = c['cliente']
            repartidor = c['repartidor']
            valor_factura = c['valor_factura']
            valor_credito = c['valor_credito']
            abono = c['abono']
            estado = c['estado']
            fecha_pagado = c['fecha_pagado']
            origen = c['origen']
            
            # Aplicar filtro de cliente
            if filtro_cliente:
                if filtro_cliente not in cliente.lower() and filtro_cliente not in str(folio):
                    continue
            
            # Aplicar filtro de estado
            if filtro_estado != "Todos" and estado != filtro_estado:
                continue
            
            # Aplicar filtro de fecha venta
            if filtro_venta_desde or filtro_venta_hasta:
                if filtro_venta_desde and fecha < filtro_venta_desde:
                    continue
                if filtro_venta_hasta and fecha > filtro_venta_hasta:
                    continue
            
            # Aplicar filtro de fecha pagado (solo para PAGADO)
            if filtro_pagado_desde or filtro_pagado_hasta:
                if estado == 'PAGADO' and fecha_pagado:
                    if filtro_pagado_desde and fecha_pagado < filtro_pagado_desde:
                        continue
                    if filtro_pagado_hasta and fecha_pagado > filtro_pagado_hasta:
                        continue
                elif estado != 'PENDIENTE':
                    continue  # Si tiene filtro de fecha pagado y no es PAGADO ni PENDIENTE, omitir
            
            total_credito += valor_credito
            saldo = valor_credito - abono
            if estado == 'PENDIENTE':
                total_pendiente += saldo
                total_pendiente_filtrado += saldo
                count_pendiente_filtrado += 1
            elif estado == 'PAGADO':
                total_pagado_filtrado += valor_credito
                count_pagado_filtrado += 1
            count_total += 1
            
            # Asignar tag según estado
            if estado == "PAGADO":
                tag = "pagado"
            elif estado == "CANCELADA":
                tag = "cancelada"
            else:
                tag = "pendiente"
            # Orden: fecha, folio, cliente, valor_factura, valor_credito, abono, saldo, estado, fecha_pagado, repartidor, origen
            self.tree_creditos.insert("", tk.END, values=(
                fecha,
                folio,
                cliente,
                f"${valor_factura:,.0f}",
                f"${valor_credito:,.0f}",
                f"${abono:,.0f}",
                f"${saldo:,.0f}",
                estado,
                fecha_pagado,
                repartidor,
                origen
            ), tags=(tag,))
        
        # Actualizar etiquetas de totales (barra superior)
        self.lbl_cantidad_creditos.config(text=str(count_total))
        self.lbl_total_creditos_general.config(text=f"${total_pendiente:,.0f}")
        
        # Actualizar etiquetas de totales filtrados (barra inferior)
        if hasattr(self, 'lbl_total_pagado_filtrado'):
            self.lbl_total_pagado_filtrado.config(text=f"${total_pagado_filtrado:,.0f}")
            self.lbl_cant_pagado_filtrado.config(text=str(count_pagado_filtrado))
            self.lbl_total_pendiente_filtrado.config(text=f"${total_pendiente_filtrado:,.0f}")
            self.lbl_cant_pendiente_filtrado.config(text=str(count_pendiente_filtrado))

    # ------------------------------------------------------------------
    # CREAR PESTAÑA DE NO ENTREGADOS
    # ------------------------------------------------------------------
    def _crear_tab_no_entregados(self):
        """Crea la pestaña de No Entregados."""
        tab = self.tab_no_entregados
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)  # Lista en row=2
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE HERRAMIENTAS
        # ═══════════════════════════════════════════════════════════════
        frame_toolbar = ttk.Frame(tab)
        frame_toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
        
        # Buscar
        ttk.Label(frame_toolbar, text="🔍 Buscar:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.buscar_no_entregados_var = tk.StringVar()
        entry_buscar = ttk.Entry(frame_toolbar, textvariable=self.buscar_no_entregados_var, width=20)
        entry_buscar.pack(side=tk.LEFT, padx=(0, 15))
        entry_buscar.bind("<KeyRelease>", lambda e: self._refrescar_no_entregados_tab())
        
        # Filtro Estado
        ttk.Label(frame_toolbar, text="Estado:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.filtro_estado_ne_var = tk.StringVar(value="Todos")
        combo_estado = ttk.Combobox(frame_toolbar, textvariable=self.filtro_estado_ne_var,
                                    values=["Todos", "PENDIENTE", "PAGADO", "CANCELADA"],
                                    state="readonly", width=12)
        combo_estado.pack(side=tk.LEFT, padx=(0, 15))
        combo_estado.bind("<<ComboboxSelected>>", lambda e: self._refrescar_no_entregados_tab())
        
        # Botón Refrescar
        btn_refrescar = ttk.Button(frame_toolbar, text="🔄 Refrescar", 
                                   command=self._refrescar_no_entregados_tab)
        btn_refrescar.pack(side=tk.LEFT, padx=5)
        
        # Totales a la derecha
        frame_totales = ttk.Frame(frame_toolbar)
        frame_totales.pack(side=tk.RIGHT, padx=10)
        
        ttk.Label(frame_totales, text="Total:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        self.lbl_cantidad_ne = ttk.Label(frame_totales, text="0", font=("Segoe UI", 10, "bold"),
                                         foreground="#1976d2")
        self.lbl_cantidad_ne.pack(side=tk.LEFT, padx=(5, 15))
        
        ttk.Label(frame_totales, text="Pendiente:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        self.lbl_total_ne_pendiente = ttk.Label(frame_totales, text="$0", 
                                                 font=("Segoe UI", 11, "bold"),
                                                 foreground="#f44336")
        self.lbl_total_ne_pendiente.pack(side=tk.LEFT, padx=5)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE FILTROS POR FECHA
        # ═══════════════════════════════════════════════════════════════
        toolbar2_ne = ttk.Frame(tab)
        toolbar2_ne.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 5))
        
        # Filtro por Fecha Venta
        ttk.Label(toolbar2_ne, text="📅 Fecha Venta:", 
                  font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2_ne, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_venta_ne_desde = DateEntry(toolbar2_ne, width=10, date_pattern='yyyy-mm-dd',
                                                          background='#1976d2', foreground='white',
                                                          headersbackground='#1565c0')
            self.filtro_fecha_venta_ne_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_venta_ne_desde.delete(0, tk.END)
            self.filtro_fecha_venta_ne_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_no_entregados_tab())
        else:
            self.filtro_fecha_venta_ne_desde = ttk.Entry(toolbar2_ne, width=10)
            self.filtro_fecha_venta_ne_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_venta_ne_desde.bind("<KeyRelease>", lambda e: self._refrescar_no_entregados_tab())
        
        ttk.Label(toolbar2_ne, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_venta_ne_hasta = DateEntry(toolbar2_ne, width=10, date_pattern='yyyy-mm-dd',
                                                          background='#1976d2', foreground='white',
                                                          headersbackground='#1565c0')
            self.filtro_fecha_venta_ne_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_venta_ne_hasta.delete(0, tk.END)
            self.filtro_fecha_venta_ne_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_no_entregados_tab())
        else:
            self.filtro_fecha_venta_ne_hasta = ttk.Entry(toolbar2_ne, width=10)
            self.filtro_fecha_venta_ne_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_venta_ne_hasta.bind("<KeyRelease>", lambda e: self._refrescar_no_entregados_tab())
        
        # Separador visual
        ttk.Separator(toolbar2_ne, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        # Filtro por Fecha Pagado
        ttk.Label(toolbar2_ne, text="💰 Fecha Pagado:", 
                  font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2_ne, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_ne_desde = DateEntry(toolbar2_ne, width=10, date_pattern='yyyy-mm-dd',
                                                           background='#2e7d32', foreground='white',
                                                           headersbackground='#1b5e20')
            self.filtro_fecha_pagado_ne_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_pagado_ne_desde.delete(0, tk.END)
            self.filtro_fecha_pagado_ne_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_no_entregados_tab())
        else:
            self.filtro_fecha_pagado_ne_desde = ttk.Entry(toolbar2_ne, width=10)
            self.filtro_fecha_pagado_ne_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_pagado_ne_desde.bind("<KeyRelease>", lambda e: self._refrescar_no_entregados_tab())
        
        ttk.Label(toolbar2_ne, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_ne_hasta = DateEntry(toolbar2_ne, width=10, date_pattern='yyyy-mm-dd',
                                                           background='#2e7d32', foreground='white',
                                                           headersbackground='#1b5e20')
            self.filtro_fecha_pagado_ne_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_pagado_ne_hasta.delete(0, tk.END)
            self.filtro_fecha_pagado_ne_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_no_entregados_tab())
        else:
            self.filtro_fecha_pagado_ne_hasta = ttk.Entry(toolbar2_ne, width=10)
            self.filtro_fecha_pagado_ne_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_pagado_ne_hasta.bind("<KeyRelease>", lambda e: self._refrescar_no_entregados_tab())
        
        btn_hoy_ne = ttk.Button(toolbar2_ne, text="Hoy", width=5, 
                                 command=self._set_fecha_hoy_ne)
        btn_hoy_ne.pack(side=tk.LEFT, padx=2)
        
        btn_limpiar_ne = ttk.Button(toolbar2_ne, text="Limpiar", width=7,
                                     command=self._limpiar_filtro_fecha_ne)
        btn_limpiar_ne.pack(side=tk.LEFT, padx=2)
        
        # ═══════════════════════════════════════════════════════════════
        # LISTA DE NO ENTREGADOS
        # ═══════════════════════════════════════════════════════════════
        frame_lista = ttk.LabelFrame(tab, text=" 📦 No Entregados ", padding=5)
        frame_lista.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        frame_lista.columnconfigure(0, weight=1)
        frame_lista.rowconfigure(0, weight=1)
        
        # Treeview
        columnas = ("fecha", "folio", "cliente", "valor", "abono", "saldo", "estado", "fecha_devuelto", "repartidor")
        self.tree_no_entregados = ttk.Treeview(frame_lista, columns=columnas, show="headings", height=15)
        
        self.tree_no_entregados.heading("fecha", text="Fecha Venta", anchor=tk.CENTER)
        self.tree_no_entregados.heading("folio", text="Folio", anchor=tk.CENTER)
        self.tree_no_entregados.heading("cliente", text="Cliente", anchor=tk.W)
        self.tree_no_entregados.heading("valor", text="Valor", anchor=tk.E)
        self.tree_no_entregados.heading("abono", text="Abono", anchor=tk.E)
        self.tree_no_entregados.heading("saldo", text="Saldo", anchor=tk.E)
        self.tree_no_entregados.heading("estado", text="Estado", anchor=tk.CENTER)
        self.tree_no_entregados.heading("fecha_devuelto", text="Fecha Pagado", anchor=tk.CENTER)
        self.tree_no_entregados.heading("repartidor", text="Repartidor", anchor=tk.W)
        
        self.tree_no_entregados.column("fecha", width=95, anchor=tk.CENTER)
        self.tree_no_entregados.column("folio", width=70, anchor=tk.CENTER)
        self.tree_no_entregados.column("cliente", width=200, anchor=tk.W)
        self.tree_no_entregados.column("valor", width=100, anchor=tk.E)
        self.tree_no_entregados.column("abono", width=90, anchor=tk.E)
        self.tree_no_entregados.column("saldo", width=100, anchor=tk.E)
        self.tree_no_entregados.column("estado", width=90, anchor=tk.CENTER)
        self.tree_no_entregados.column("fecha_devuelto", width=110, anchor=tk.CENTER)
        self.tree_no_entregados.column("repartidor", width=120, anchor=tk.W)
        
        scrolly = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=self.tree_no_entregados.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient=tk.HORIZONTAL, command=self.tree_no_entregados.xview)
        self.tree_no_entregados.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_no_entregados.grid(row=0, column=0, sticky="nsew")
        scrolly.grid(row=0, column=1, sticky="ns")
        scrollx.grid(row=1, column=0, sticky="ew")
        
        # Bindings
        self.tree_no_entregados.bind("<Button-1>", self._on_clic_no_entregado)
        self.tree_no_entregados.bind("<Double-1>", self._on_doble_clic_no_entregado)
        
        # Tags
        self.tree_no_entregados.tag_configure("pagado", background="#1b5e20", foreground="#a5d6a7")
        self.tree_no_entregados.tag_configure("pendiente", background="#e65100", foreground="#ffe0b2")
        self.tree_no_entregados.tag_configure("cancelada", background="#880e4f", foreground="#f8bbd0")
        
        # Widget flotante para edición
        self.ne_edit_widget = None
        self.ne_edit_frame = None
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA INFERIOR DE TOTALES FILTRADOS
        # ═══════════════════════════════════════════════════════════════
        frame_totales_ne = ttk.Frame(tab)
        frame_totales_ne.grid(row=3, column=0, sticky="ew", padx=10, pady=5)
        
        # Total Pagado (filtrado por fecha)
        ttk.Label(frame_totales_ne, text="📊 Total Pagado (filtrado):", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_total_pagado_ne_filtrado = ttk.Label(frame_totales_ne, text="$0",
                                                       font=("Segoe UI", 11, "bold"),
                                                       foreground="#81c784")
        self.lbl_total_pagado_ne_filtrado.pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_cant_pagado_ne_filtrado = ttk.Label(frame_totales_ne, text="(0)",
                                                      font=("Segoe UI", 9),
                                                      foreground="gray")
        self.lbl_cant_pagado_ne_filtrado.pack(side=tk.LEFT, padx=(0, 30))
        
        # Total Pendiente (filtrado)
        ttk.Label(frame_totales_ne, text="Total Pendiente (filtrado):", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_total_pendiente_ne_filtrado = ttk.Label(frame_totales_ne, text="$0",
                                                          font=("Segoe UI", 11, "bold"),
                                                          foreground="#ffb74d")
        self.lbl_total_pendiente_ne_filtrado.pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_cant_pendiente_ne_filtrado = ttk.Label(frame_totales_ne, text="(0)",
                                                          font=("Segoe UI", 9),
                                                          foreground="gray")
        self.lbl_cant_pendiente_ne_filtrado.pack(side=tk.LEFT)
        
        # Cargar datos
        self._refrescar_no_entregados_tab()
    
    def _on_clic_no_entregado(self, event):
        """Maneja clic en no entregados para editar."""
        self._cerrar_edicion_ne()
        
        item_id = self.tree_no_entregados.identify_row(event.y)
        column = self.tree_no_entregados.identify_column(event.x)
        
        if not item_id or not column:
            return
        
        col_idx = int(column.replace('#', '')) - 1
        columnas = ("fecha", "folio", "cliente", "valor", "abono", "saldo", "estado", "fecha_devuelto", "repartidor")
        
        if col_idx < 0 or col_idx >= len(columnas):
            return
        
        col_name = columnas[col_idx]
        
        if col_name not in ("abono", "estado", "repartidor"):
            self.tree_no_entregados.selection_set(item_id)
            return
        
        values = self.tree_no_entregados.item(item_id, 'values')
        fecha = values[0]
        folio = int(values[1])
        
        bbox = self.tree_no_entregados.bbox(item_id, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        if col_name == "abono":
            self._crear_entry_abono_ne(item_id, fecha, folio, x, y, width, height, values)
        elif col_name == "estado":
            self._crear_combo_estado_ne(item_id, fecha, folio, x, y, width, height, values)
        elif col_name == "repartidor":
            self._crear_combo_repartidor_ne(item_id, fecha, folio, x, y, width, height, values)
    
    def _cerrar_edicion_ne(self, event=None):
        """Cierra el widget de edición in-place de no entregados."""
        if hasattr(self, 'ne_edit_widget') and self.ne_edit_widget:
            try:
                self.ne_edit_widget.destroy()
            except:
                pass
            self.ne_edit_widget = None
        if hasattr(self, 'ne_edit_frame') and self.ne_edit_frame:
            try:
                self.ne_edit_frame.destroy()
            except:
                pass
            self.ne_edit_frame = None
    
    def _crear_entry_abono_ne(self, item_id, fecha, folio, x, y, width, height, values):
        """Crea Entry para editar abono de no entregado."""
        abono_actual = values[4].replace('$', '').replace(',', '') if len(values) > 4 else '0'
        valor_str = values[3].replace('$', '').replace(',', '') if len(values) > 3 else '0'
        cliente = values[2] if len(values) > 2 else ''
        
        try:
            valor = float(valor_str)
        except:
            valor = 0
        
        frame = tk.Frame(self.tree_no_entregados, bg='white', highlightbackground='#1976d2', highlightthickness=2)
        frame.place(x=x-5, y=y, width=width+80, height=height+4)
        
        self.ne_edit_frame = frame
        
        entry = tk.Entry(frame, font=("Segoe UI", 10), justify='right', bd=0)
        entry.insert(0, abono_actual)
        entry.select_range(0, tk.END)
        entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        entry.focus_set()
        
        self.ne_edit_widget = entry
        
        def guardar(event=None):
            try:
                nuevo_abono = round(float(entry.get().replace(',', '').replace('$', '')))
                if nuevo_abono > valor:
                    entry.config(background='#ffcccc')
                    messagebox.showwarning("Advertencia", 
                        f"El abono (${nuevo_abono:,.0f}) no puede ser mayor al valor (${valor:,.0f})",
                        parent=self.ventana)
                    return
                
                resultado = db_local.actualizar_abono_no_entregado(fecha, folio, nuevo_abono)
                
                if isinstance(resultado, dict) and resultado.get('success'):
                    nuevo_saldo = resultado.get('nuevo_saldo', 0)
                    nuevo_estado = resultado.get('nuevo_estado', '')
                    cambio_estado = resultado.get('cambio_estado', False)
                    
                    if cambio_estado and nuevo_estado == 'PAGADO':
                        messagebox.showinfo("Pagado", 
                            f"¡El no entregado del folio {folio} ha sido pagado completamente!\n\n"
                            f"Cliente: {cliente}\n"
                            f"Valor: ${valor:,.0f}\n"
                            f"Total Abonado: ${nuevo_abono:,.0f}",
                            parent=self.ventana)
                elif isinstance(resultado, dict):
                    messagebox.showerror("Error", resultado.get('error', 'Error desconocido'), parent=self.ventana)
                
                self._cerrar_edicion_ne()
                self._refrescar_no_entregados_tab()
            except ValueError:
                entry.config(background='#ffcccc')
        
        def cancelar(event=None):
            self._cerrar_edicion_ne()
        
        btn_guardar = tk.Button(frame, text="✓", font=("Segoe UI", 9, "bold"), 
                                bg='#4caf50', fg='white', bd=0, width=3,
                                command=guardar, cursor='hand2')
        btn_guardar.pack(side=tk.LEFT, padx=1)
        
        btn_cancelar = tk.Button(frame, text="✗", font=("Segoe UI", 9, "bold"), 
                                 bg='#f44336', fg='white', bd=0, width=3,
                                 command=cancelar, cursor='hand2')
        btn_cancelar.pack(side=tk.LEFT, padx=1)
        
        entry.bind("<Return>", guardar)
        entry.bind("<Escape>", cancelar)
    
    def _crear_combo_estado_ne(self, item_id, fecha, folio, x, y, width, height, values):
        """Crea Combobox para seleccionar estado de no entregado."""
        estado_actual = values[6] if len(values) > 6 else 'PENDIENTE'
        
        frame = tk.Frame(self.tree_no_entregados, bg='white', highlightbackground='#1976d2', highlightthickness=2)
        frame.place(x=x-5, y=y, width=width+50, height=height+4)
        
        self.ne_edit_frame = frame
        
        combo = ttk.Combobox(frame, values=["PENDIENTE", "PAGADO", "CANCELADA"],
                             state="readonly", font=("Segoe UI", 9), width=12)
        combo.set(estado_actual)
        combo.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        combo.focus_set()
        
        self.ne_edit_widget = combo
        
        def guardar(event=None):
            nuevo_estado = combo.get()
            resultado = db_local.actualizar_estado_no_entregado(fecha, folio, nuevo_estado)
            if resultado:
                print(f"✅ Estado NE actualizado: Folio {folio} | Estado: {nuevo_estado}")
            self._cerrar_edicion_ne()
            self._refrescar_no_entregados_tab()
        
        def cancelar(event=None):
            self._cerrar_edicion_ne()
        
        btn_cancelar = tk.Button(frame, text="✗", font=("Segoe UI", 9, "bold"), 
                                 bg='#f44336', fg='white', bd=0, width=3,
                                 command=cancelar, cursor='hand2')
        btn_cancelar.pack(side=tk.LEFT, padx=1)
        
        combo.bind("<<ComboboxSelected>>", guardar)
        combo.bind("<Escape>", cancelar)
    
    def _crear_combo_repartidor_ne(self, item_id, fecha, folio, x, y, width, height, values):
        """Crea Combobox para seleccionar repartidor de no entregado."""
        repartidor_actual = values[8] if len(values) > 8 else ''
        
        repartidores = [""] + list(self.ds.repartidores.keys())
        
        frame = tk.Frame(self.tree_no_entregados, bg='white', highlightbackground='#1976d2', highlightthickness=2)
        frame.place(x=x-5, y=y, width=width+50, height=height+4)
        
        self.ne_edit_frame = frame
        
        combo = ttk.Combobox(frame, values=repartidores, font=("Segoe UI", 9), width=15)
        combo.set(repartidor_actual)
        combo.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2)
        combo.focus_set()
        
        self.ne_edit_widget = combo
        
        def guardar(event=None):
            nuevo_repartidor = combo.get()
            resultado = db_local.actualizar_repartidor_no_entregado(fecha, folio, nuevo_repartidor)
            if resultado:
                print(f"✅ Repartidor NE actualizado: Folio {folio} | Repartidor: {nuevo_repartidor}")
            self._cerrar_edicion_ne()
            self._refrescar_no_entregados_tab()
        
        def cancelar(event=None):
            self._cerrar_edicion_ne()
        
        btn_guardar = tk.Button(frame, text="✓", font=("Segoe UI", 9, "bold"), 
                                bg='#4caf50', fg='white', bd=0, width=3,
                                command=guardar, cursor='hand2')
        btn_guardar.pack(side=tk.LEFT, padx=1)
        
        combo.bind("<Return>", guardar)
        combo.bind("<Escape>", cancelar)
    
    def _on_doble_clic_no_entregado(self, event):
        """Maneja doble clic para ver/editar observaciones."""
        item_id = self.tree_no_entregados.identify_row(event.y)
        if not item_id:
            return
        
        values = self.tree_no_entregados.item(item_id, 'values')
        fecha = values[0]
        folio = int(values[1])
        cliente = values[2]
        
        ne = db_local.obtener_no_entregado(fecha, folio)
        observaciones = ne.get('observaciones', '') if ne else ''
        
        # Ventana de observaciones
        dialog = tk.Toplevel(self.ventana)
        dialog.title(f"Observaciones - Folio {folio}")
        dialog.geometry("500x300")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        ttk.Label(dialog, text=f"Cliente: {cliente}", font=("Segoe UI", 10, "bold")).pack(pady=10)
        
        text = tk.Text(dialog, font=("Segoe UI", 10), wrap=tk.WORD, height=10)
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        text.insert("1.0", observaciones)
        
        def guardar():
            nuevas_obs = text.get("1.0", tk.END).strip()
            db_local.actualizar_observaciones_no_entregado(fecha, folio, nuevas_obs)
            dialog.destroy()
            self._refrescar_no_entregados_tab()
        
        frame_btns = ttk.Frame(dialog)
        frame_btns.pack(pady=10)
        ttk.Button(frame_btns, text="Guardar", command=guardar).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_btns, text="Cancelar", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    # ------------------------------------------------------------------
    # CREAR PESTAÑA DE PRÉSTAMOS
    # ------------------------------------------------------------------
    def _crear_tab_prestamos(self):
        """Crea la pestaña de Préstamos."""
        tab = self.tab_prestamos
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE HERRAMIENTAS
        # ═══════════════════════════════════════════════════════════════
        frame_toolbar = ttk.Frame(tab)
        frame_toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
        
        # Buscar
        ttk.Label(frame_toolbar, text="🔍 Buscar:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.buscar_prestamos_var = tk.StringVar()
        entry_buscar = ttk.Entry(frame_toolbar, textvariable=self.buscar_prestamos_var, width=20)
        entry_buscar.pack(side=tk.LEFT, padx=(0, 15))
        entry_buscar.bind("<KeyRelease>", lambda e: self._refrescar_prestamos_tab())
        
        # Filtro Estado
        ttk.Label(frame_toolbar, text="Estado:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.filtro_estado_prestamos_var = tk.StringVar(value="PENDIENTE")
        combo_estado = ttk.Combobox(frame_toolbar, textvariable=self.filtro_estado_prestamos_var,
                                    values=["Todos", "PENDIENTE", "PAGADO", "CANCELADA"],
                                    state="readonly", width=12)
        combo_estado.pack(side=tk.LEFT, padx=(0, 15))
        combo_estado.bind("<<ComboboxSelected>>", lambda e: self._refrescar_prestamos_tab())
        
        # Botón Nuevo Préstamo
        btn_nuevo = ttk.Button(frame_toolbar, text="➕ Nuevo Préstamo", 
                               command=self._nuevo_prestamo)
        btn_nuevo.pack(side=tk.LEFT, padx=5)
        
        # Botón Refrescar
        btn_refrescar = ttk.Button(frame_toolbar, text="🔄 Refrescar", 
                                   command=self._refrescar_prestamos_tab)
        btn_refrescar.pack(side=tk.LEFT, padx=5)
        
        # Totales a la derecha
        frame_totales = ttk.Frame(frame_toolbar)
        frame_totales.pack(side=tk.RIGHT, padx=10)
        
        ttk.Label(frame_totales, text="Total:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        self.lbl_cantidad_prestamos = ttk.Label(frame_totales, text="0", font=("Segoe UI", 10, "bold"),
                                                 foreground="#1976d2")
        self.lbl_cantidad_prestamos.pack(side=tk.LEFT, padx=(5, 15))
        
        ttk.Label(frame_totales, text="Pendiente:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        self.lbl_total_prestamos_pendiente = ttk.Label(frame_totales, text="$0", 
                                                        font=("Segoe UI", 11, "bold"),
                                                        foreground="#f44336")
        self.lbl_total_prestamos_pendiente.pack(side=tk.LEFT, padx=5)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE FILTROS POR FECHA
        # ═══════════════════════════════════════════════════════════════
        toolbar2 = ttk.Frame(tab)
        toolbar2.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 5))
        
        # Filtro por Fecha Préstamo
        ttk.Label(toolbar2, text="📅 Fecha Préstamo:", 
                  font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_prestamo_desde = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                          background='#1976d2', foreground='white',
                                                          headersbackground='#1565c0')
            self.filtro_fecha_prestamo_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_prestamo_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_prestamos_tab())
        else:
            self.filtro_fecha_prestamo_desde = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_prestamo_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_prestamo_desde.bind("<KeyRelease>", lambda e: self._refrescar_prestamos_tab())
        
        ttk.Label(toolbar2, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_prestamo_hasta = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                          background='#1976d2', foreground='white',
                                                          headersbackground='#1565c0')
            self.filtro_fecha_prestamo_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_prestamo_hasta.delete(0, tk.END)
            self.filtro_fecha_prestamo_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_prestamos_tab())
        else:
            self.filtro_fecha_prestamo_hasta = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_prestamo_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_prestamo_hasta.bind("<KeyRelease>", lambda e: self._refrescar_prestamos_tab())
        
        # Separador visual
        ttk.Separator(toolbar2, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        # Filtro por Fecha Pagado
        ttk.Label(toolbar2, text="💰 Fecha Pagado:", 
                  font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_prestamo_desde = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                                 background='#2e7d32', foreground='white',
                                                                 headersbackground='#1b5e20')
            self.filtro_fecha_pagado_prestamo_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_pagado_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_pagado_prestamo_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_prestamos_tab())
        else:
            self.filtro_fecha_pagado_prestamo_desde = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_pagado_prestamo_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_pagado_prestamo_desde.bind("<KeyRelease>", lambda e: self._refrescar_prestamos_tab())
        
        ttk.Label(toolbar2, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_pagado_prestamo_hasta = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                                 background='#2e7d32', foreground='white',
                                                                 headersbackground='#1b5e20')
            self.filtro_fecha_pagado_prestamo_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_pagado_prestamo_hasta.delete(0, tk.END)
            self.filtro_fecha_pagado_prestamo_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_prestamos_tab())
        else:
            self.filtro_fecha_pagado_prestamo_hasta = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_pagado_prestamo_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_pagado_prestamo_hasta.bind("<KeyRelease>", lambda e: self._refrescar_prestamos_tab())
        
        btn_hoy = ttk.Button(toolbar2, text="Hoy", width=5, 
                              command=self._set_fecha_hoy_prestamos)
        btn_hoy.pack(side=tk.LEFT, padx=2)
        
        btn_limpiar = ttk.Button(toolbar2, text="Limpiar", width=7,
                                  command=self._limpiar_filtro_fecha_prestamos)
        btn_limpiar.pack(side=tk.LEFT, padx=2)
        
        # ═══════════════════════════════════════════════════════════════
        # LISTA DE PRÉSTAMOS
        # ═══════════════════════════════════════════════════════════════
        frame_lista = ttk.LabelFrame(tab, text=" 💸 Préstamos ", padding=5)
        frame_lista.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        frame_lista.columnconfigure(0, weight=1)
        frame_lista.rowconfigure(0, weight=1)
        
        # Treeview
        columnas = ("id", "fecha", "beneficiario", "concepto", "monto", "abono", "saldo", "estado", "fecha_pagado", "responsable")
        self.tree_prestamos = ttk.Treeview(frame_lista, columns=columnas, show="headings", height=15)
        
        self.tree_prestamos.heading("id", text="ID", anchor=tk.CENTER)
        self.tree_prestamos.heading("fecha", text="Fecha", anchor=tk.CENTER)
        self.tree_prestamos.heading("beneficiario", text="Beneficiario", anchor=tk.W)
        self.tree_prestamos.heading("concepto", text="Concepto", anchor=tk.W)
        self.tree_prestamos.heading("monto", text="Monto", anchor=tk.E)
        self.tree_prestamos.heading("abono", text="Abono", anchor=tk.E)
        self.tree_prestamos.heading("saldo", text="Saldo", anchor=tk.E)
        self.tree_prestamos.heading("estado", text="Estado", anchor=tk.CENTER)
        self.tree_prestamos.heading("fecha_pagado", text="Fecha Pagado", anchor=tk.CENTER)
        self.tree_prestamos.heading("responsable", text="Responsable", anchor=tk.W)
        
        self.tree_prestamos.column("id", width=50, anchor=tk.CENTER)
        self.tree_prestamos.column("fecha", width=95, anchor=tk.CENTER)
        self.tree_prestamos.column("beneficiario", width=180, anchor=tk.W)
        self.tree_prestamos.column("concepto", width=200, anchor=tk.W)
        self.tree_prestamos.column("monto", width=100, anchor=tk.E)
        self.tree_prestamos.column("abono", width=90, anchor=tk.E)
        self.tree_prestamos.column("saldo", width=100, anchor=tk.E)
        self.tree_prestamos.column("estado", width=90, anchor=tk.CENTER)
        self.tree_prestamos.column("fecha_pagado", width=110, anchor=tk.CENTER)
        self.tree_prestamos.column("responsable", width=120, anchor=tk.W)
        
        scrolly = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=self.tree_prestamos.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient=tk.HORIZONTAL, command=self.tree_prestamos.xview)
        self.tree_prestamos.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_prestamos.grid(row=0, column=0, sticky="nsew")
        scrolly.grid(row=0, column=1, sticky="ns")
        scrollx.grid(row=1, column=0, sticky="ew")
        
        # Bindings
        self.tree_prestamos.bind("<Button-1>", self._on_clic_prestamo)
        self.tree_prestamos.bind("<Double-1>", self._on_doble_clic_prestamo)
        
        # Tags
        self.tree_prestamos.tag_configure("pagado", background="#1b5e20", foreground="#a5d6a7")
        self.tree_prestamos.tag_configure("pendiente", background="#e65100", foreground="#ffe0b2")
        self.tree_prestamos.tag_configure("cancelada", background="#880e4f", foreground="#f8bbd0")
        
        # Widget flotante para edición
        self.prestamo_edit_widget = None
        self.prestamo_edit_frame = None
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA INFERIOR DE TOTALES FILTRADOS
        # ═══════════════════════════════════════════════════════════════
        frame_totales_bottom = ttk.Frame(tab)
        frame_totales_bottom.grid(row=3, column=0, sticky="ew", padx=10, pady=5)
        
        # Total Pagado (filtrado)
        ttk.Label(frame_totales_bottom, text="📊 Total Pagado (filtrado):", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_total_pagado_prestamos_filtrado = ttk.Label(frame_totales_bottom, text="$0",
                                                              font=("Segoe UI", 11, "bold"),
                                                              foreground="#81c784")
        self.lbl_total_pagado_prestamos_filtrado.pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_cant_pagado_prestamos_filtrado = ttk.Label(frame_totales_bottom, text="(0)",
                                                             font=("Segoe UI", 9),
                                                             foreground="gray")
        self.lbl_cant_pagado_prestamos_filtrado.pack(side=tk.LEFT, padx=(0, 30))
        
        # Total Pendiente (filtrado)
        ttk.Label(frame_totales_bottom, text="Total Pendiente (filtrado):", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_total_pendiente_prestamos_filtrado = ttk.Label(frame_totales_bottom, text="$0",
                                                                 font=("Segoe UI", 11, "bold"),
                                                                 foreground="#ffb74d")
        self.lbl_total_pendiente_prestamos_filtrado.pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_cant_pendiente_prestamos_filtrado = ttk.Label(frame_totales_bottom, text="(0)",
                                                                font=("Segoe UI", 9),
                                                                foreground="gray")
        self.lbl_cant_pendiente_prestamos_filtrado.pack(side=tk.LEFT)
        
        # Cargar datos
        self._refrescar_prestamos_tab()
    
    def _set_fecha_hoy_prestamos(self):
        """Establece la fecha de hoy en los filtros de préstamos."""
        from datetime import date
        hoy = date.today()
        if HAS_CALENDAR:
            self.filtro_fecha_prestamo_desde.set_date(hoy)
            self.filtro_fecha_prestamo_hasta.set_date(hoy)
        else:
            self.filtro_fecha_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_prestamo_desde.insert(0, hoy.isoformat())
            self.filtro_fecha_prestamo_hasta.delete(0, tk.END)
            self.filtro_fecha_prestamo_hasta.insert(0, hoy.isoformat())
        self._refrescar_prestamos_tab()
    
    def _limpiar_filtro_fecha_prestamos(self):
        """Limpia los filtros de fecha de préstamos."""
        if HAS_CALENDAR:
            self.filtro_fecha_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_prestamo_hasta.delete(0, tk.END)
            self.filtro_fecha_pagado_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_pagado_prestamo_hasta.delete(0, tk.END)
        else:
            self.filtro_fecha_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_prestamo_hasta.delete(0, tk.END)
            self.filtro_fecha_pagado_prestamo_desde.delete(0, tk.END)
            self.filtro_fecha_pagado_prestamo_hasta.delete(0, tk.END)
        self._refrescar_prestamos_tab()
    
    def _refrescar_prestamos_tab(self):
        """Refresca la lista de préstamos aplicando filtros."""
        # Limpiar
        for item in self.tree_prestamos.get_children():
            self.tree_prestamos.delete(item)
        
        # Obtener todos los préstamos
        prestamos = db_local.obtener_todos_prestamos()
        
        # Obtener filtros
        buscar = self.buscar_prestamos_var.get().lower().strip()
        estado_filtro = self.filtro_estado_prestamos_var.get()
        
        # Filtros de fecha
        try:
            fecha_desde = self.filtro_fecha_prestamo_desde.get().strip()
            fecha_hasta = self.filtro_fecha_prestamo_hasta.get().strip()
            fecha_pagado_desde = self.filtro_fecha_pagado_prestamo_desde.get().strip()
            fecha_pagado_hasta = self.filtro_fecha_pagado_prestamo_hasta.get().strip()
        except:
            fecha_desde = fecha_hasta = fecha_pagado_desde = fecha_pagado_hasta = ""
        
        total_pendiente = 0
        total_pagado = 0
        cant_pendiente = 0
        cant_pagado = 0
        total_general = 0
        count = 0
        
        for p in prestamos:
            # Filtrar por estado
            estado = p.get('estado', 'PENDIENTE')
            if estado_filtro != "Todos" and estado != estado_filtro:
                continue
            
            # Filtrar por búsqueda
            beneficiario = p.get('beneficiario', '') or ''
            concepto = p.get('concepto', '') or ''
            if buscar and buscar not in beneficiario.lower() and buscar not in concepto.lower():
                continue
            
            # Filtrar por fecha préstamo
            fecha = p.get('fecha', '') or ''
            if fecha_desde and fecha < fecha_desde:
                continue
            if fecha_hasta and fecha > fecha_hasta:
                continue
            
            # Filtrar por fecha pagado
            fecha_pagado = p.get('fecha_pagado', '') or ''
            if fecha_pagado_desde and (not fecha_pagado or fecha_pagado < fecha_pagado_desde):
                continue
            if fecha_pagado_hasta and (not fecha_pagado or fecha_pagado > fecha_pagado_hasta):
                continue
            
            # Calcular valores
            monto = round(p.get('monto', 0) or 0)
            abono = round(p.get('abono', 0) or 0)
            saldo = monto - abono
            
            # Preparar valores para mostrar
            id_prestamo = p.get('id', 0)
            responsable = p.get('responsable', '') or ''
            
            # Tag según estado
            tag = estado.lower() if estado else "pendiente"
            
            self.tree_prestamos.insert("", tk.END, values=(
                id_prestamo,
                fecha,
                beneficiario,
                concepto,
                f"${monto:,.0f}",
                f"${abono:,.0f}",
                f"${saldo:,.0f}",
                estado,
                fecha_pagado,
                responsable
            ), tags=(tag,))
            
            count += 1
            total_general += saldo
            
            if estado == 'PAGADO':
                total_pagado += monto
                cant_pagado += 1
            elif estado == 'PENDIENTE':
                total_pendiente += saldo
                cant_pendiente += 1
        
        # Actualizar labels
        self.lbl_cantidad_prestamos.config(text=str(count))
        self.lbl_total_prestamos_pendiente.config(text=f"${total_general:,.0f}")
        self.lbl_total_pagado_prestamos_filtrado.config(text=f"${total_pagado:,.0f}")
        self.lbl_cant_pagado_prestamos_filtrado.config(text=f"({cant_pagado})")
        self.lbl_total_pendiente_prestamos_filtrado.config(text=f"${total_pendiente:,.0f}")
        self.lbl_cant_pendiente_prestamos_filtrado.config(text=f"({cant_pendiente})")
    
    def _nuevo_prestamo(self):
        """Abre diálogo para crear un nuevo préstamo."""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("Nuevo Préstamo")
        dialog.geometry("450x350")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        # Formulario
        frame_form = ttk.Frame(dialog, padding=20)
        frame_form.pack(fill=tk.BOTH, expand=True)
        
        # Fecha
        ttk.Label(frame_form, text="Fecha:", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=5)
        from datetime import date
        fecha_var = tk.StringVar(value=date.today().isoformat())
        if HAS_CALENDAR:
            entry_fecha = DateEntry(frame_form, textvariable=fecha_var, width=15, date_pattern='yyyy-mm-dd')
        else:
            entry_fecha = ttk.Entry(frame_form, textvariable=fecha_var, width=15)
        entry_fecha.grid(row=0, column=1, sticky="w", pady=5, padx=5)
        
        # Beneficiario
        ttk.Label(frame_form, text="Beneficiario:", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=5)
        beneficiario_var = tk.StringVar()
        entry_benef = ttk.Entry(frame_form, textvariable=beneficiario_var, width=30)
        entry_benef.grid(row=1, column=1, sticky="w", pady=5, padx=5)
        
        # Concepto
        ttk.Label(frame_form, text="Concepto:", font=("Segoe UI", 10)).grid(row=2, column=0, sticky="w", pady=5)
        concepto_var = tk.StringVar()
        entry_concepto = ttk.Entry(frame_form, textvariable=concepto_var, width=30)
        entry_concepto.grid(row=2, column=1, sticky="w", pady=5, padx=5)
        
        # Monto
        ttk.Label(frame_form, text="Monto:", font=("Segoe UI", 10)).grid(row=3, column=0, sticky="w", pady=5)
        monto_var = tk.StringVar()
        entry_monto = ttk.Entry(frame_form, textvariable=monto_var, width=15)
        entry_monto.grid(row=3, column=1, sticky="w", pady=5, padx=5)
        
        # Responsable
        ttk.Label(frame_form, text="Responsable:", font=("Segoe UI", 10)).grid(row=4, column=0, sticky="w", pady=5)
        responsable_var = tk.StringVar()
        combo_responsable = ttk.Combobox(frame_form, textvariable=responsable_var, width=20, state="readonly")
        # Cargar repartidores
        repartidores = db_local.obtener_todos_repartidores()
        nombres = [""] + [r.get('nombre', '') for r in repartidores]
        combo_responsable['values'] = nombres
        combo_responsable.grid(row=4, column=1, sticky="w", pady=5, padx=5)
        
        # Observaciones
        ttk.Label(frame_form, text="Observaciones:", font=("Segoe UI", 10)).grid(row=5, column=0, sticky="nw", pady=5)
        text_obs = tk.Text(frame_form, width=30, height=4, font=("Segoe UI", 9))
        text_obs.grid(row=5, column=1, sticky="w", pady=5, padx=5)
        
        def guardar():
            fecha = fecha_var.get().strip()
            beneficiario = beneficiario_var.get().strip()
            concepto = concepto_var.get().strip()
            try:
                monto = float(monto_var.get().replace(',', '').replace('$', '').strip())
            except:
                messagebox.showerror("Error", "Monto inválido")
                return
            responsable = responsable_var.get().strip()
            observaciones = text_obs.get("1.0", tk.END).strip()
            
            if not beneficiario:
                messagebox.showerror("Error", "El beneficiario es requerido")
                return
            if monto <= 0:
                messagebox.showerror("Error", "El monto debe ser mayor a 0")
                return
            
            if db_local.guardar_prestamo(fecha, beneficiario, concepto, monto, responsable, observaciones):
                messagebox.showinfo("Éxito", "Préstamo guardado correctamente")
                dialog.destroy()
                self._refrescar_prestamos_tab()
            else:
                messagebox.showerror("Error", "No se pudo guardar el préstamo")
        
        # Botones
        frame_btns = ttk.Frame(frame_form)
        frame_btns.grid(row=6, column=0, columnspan=2, pady=20)
        ttk.Button(frame_btns, text="💾 Guardar", command=guardar).pack(side=tk.LEFT, padx=10)
        ttk.Button(frame_btns, text="❌ Cancelar", command=dialog.destroy).pack(side=tk.LEFT, padx=10)
        
        entry_benef.focus_set()
    
    def _on_clic_prestamo(self, event):
        """Maneja clic en préstamos para editar."""
        self._cerrar_edicion_prestamo()
        
        item_id = self.tree_prestamos.identify_row(event.y)
        column = self.tree_prestamos.identify_column(event.x)
        
        if not item_id or not column:
            return
        
        col_idx = int(column.replace('#', '')) - 1
        columnas = ("id", "fecha", "beneficiario", "concepto", "monto", "abono", "saldo", "estado", "fecha_pagado", "responsable")
        
        if col_idx < 0 or col_idx >= len(columnas):
            return
        
        col_name = columnas[col_idx]
        
        if col_name not in ("abono", "estado", "responsable"):
            self.tree_prestamos.selection_set(item_id)
            return
        
        values = self.tree_prestamos.item(item_id, 'values')
        id_prestamo = int(values[0])
        
        bbox = self.tree_prestamos.bbox(item_id, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        if col_name == "abono":
            self._crear_entry_abono_prestamo(item_id, id_prestamo, x, y, width, height, values)
        elif col_name == "estado":
            self._crear_combo_estado_prestamo(item_id, id_prestamo, x, y, width, height, values)
        elif col_name == "responsable":
            self._crear_combo_responsable_prestamo(item_id, id_prestamo, x, y, width, height, values)
    
    def _cerrar_edicion_prestamo(self, event=None):
        """Cierra el widget de edición in-place de préstamos."""
        if hasattr(self, 'prestamo_edit_widget') and self.prestamo_edit_widget:
            try:
                self.prestamo_edit_widget.destroy()
            except:
                pass
            self.prestamo_edit_widget = None
        if hasattr(self, 'prestamo_edit_frame') and self.prestamo_edit_frame:
            try:
                self.prestamo_edit_frame.destroy()
            except:
                pass
            self.prestamo_edit_frame = None
    
    def _crear_entry_abono_prestamo(self, item_id, id_prestamo, x, y, width, height, values):
        """Crea entry para editar abono de préstamo."""
        frame = ttk.Frame(self.tree_prestamos)
        frame.place(x=x, y=y, width=width+60, height=height)
        
        abono_actual = values[5].replace('$', '').replace(',', '').strip()
        
        entry = ttk.Entry(frame, width=10)
        entry.insert(0, abono_actual)
        entry.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        entry.focus_set()
        entry.select_range(0, tk.END)
        
        def guardar(event=None):
            try:
                nuevo_abono = float(entry.get().replace(',', '').replace('$', '').strip())
            except:
                messagebox.showerror("Error", "Valor inválido")
                self._cerrar_edicion_prestamo()
                return
            
            resultado = db_local.actualizar_abono_prestamo(id_prestamo, nuevo_abono)
            if resultado.get('success'):
                if resultado.get('cambio_estado'):
                    messagebox.showinfo("Estado Actualizado", 
                                        f"El préstamo pasó a estado {resultado.get('nuevo_estado')}")
            self._cerrar_edicion_prestamo()
            self._refrescar_prestamos_tab()
        
        entry.bind("<Return>", guardar)
        entry.bind("<Escape>", lambda e: self._cerrar_edicion_prestamo())
        
        btn = ttk.Button(frame, text="✓", width=3, command=guardar)
        btn.pack(side=tk.LEFT, padx=2)
        
        self.prestamo_edit_frame = frame
        self.prestamo_edit_widget = entry
    
    def _crear_combo_estado_prestamo(self, item_id, id_prestamo, x, y, width, height, values):
        """Crea combo para editar estado de préstamo."""
        estado_actual = values[7]
        
        combo = ttk.Combobox(self.tree_prestamos, values=["PENDIENTE", "PAGADO", "CANCELADA"],
                             state="readonly", width=12)
        combo.set(estado_actual)
        combo.place(x=x, y=y, width=width, height=height)
        combo.focus_set()
        
        def cambiar(event=None):
            nuevo_estado = combo.get()
            db_local.actualizar_estado_prestamo(id_prestamo, nuevo_estado)
            self._cerrar_edicion_prestamo()
            self._refrescar_prestamos_tab()
        
        combo.bind("<<ComboboxSelected>>", cambiar)
        combo.bind("<Escape>", lambda e: self._cerrar_edicion_prestamo())
        
        self.prestamo_edit_widget = combo
    
    def _crear_combo_responsable_prestamo(self, item_id, id_prestamo, x, y, width, height, values):
        """Crea combo para editar responsable de préstamo."""
        responsable_actual = values[9]
        
        repartidores = db_local.obtener_todos_repartidores()
        nombres = [""] + [r.get('nombre', '') for r in repartidores]
        
        combo = ttk.Combobox(self.tree_prestamos, values=nombres, state="readonly", width=15)
        combo.set(responsable_actual)
        combo.place(x=x, y=y, width=max(width, 120), height=height)
        combo.focus_set()
        
        def cambiar(event=None):
            nuevo_responsable = combo.get()
            db_local.actualizar_responsable_prestamo(id_prestamo, nuevo_responsable)
            self._cerrar_edicion_prestamo()
            self._refrescar_prestamos_tab()
        
        combo.bind("<<ComboboxSelected>>", cambiar)
        combo.bind("<Escape>", lambda e: self._cerrar_edicion_prestamo())
        
        self.prestamo_edit_widget = combo
    
    def _on_doble_clic_prestamo(self, event):
        """Maneja doble clic en préstamos para ver/editar observaciones."""
        item_id = self.tree_prestamos.identify_row(event.y)
        if not item_id:
            return
        
        values = self.tree_prestamos.item(item_id, 'values')
        id_prestamo = int(values[0])
        beneficiario = values[2]
        
        # Obtener datos completos
        prestamo = db_local.obtener_prestamo(id_prestamo)
        if not prestamo:
            return
        
        observaciones = prestamo.get('observaciones', '') or ''
        
        dialog = tk.Toplevel(self.ventana)
        dialog.title(f"Observaciones - {beneficiario}")
        dialog.geometry("500x300")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        ttk.Label(dialog, text=f"Beneficiario: {beneficiario}", font=("Segoe UI", 10, "bold")).pack(pady=10)
        
        text = tk.Text(dialog, font=("Segoe UI", 10), wrap=tk.WORD, height=10)
        text.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        text.insert("1.0", observaciones)
        
        def guardar():
            nuevas_obs = text.get("1.0", tk.END).strip()
            db_local.actualizar_observaciones_prestamo(id_prestamo, nuevas_obs)
            dialog.destroy()
            self._refrescar_prestamos_tab()
        
        frame_btns = ttk.Frame(dialog)
        frame_btns.pack(pady=10)
        ttk.Button(frame_btns, text="Guardar", command=guardar).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_btns, text="Cancelar", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    # ------------------------------------------------------------------
    # CREAR PESTAÑA DE DEVOLUCIONES PARCIALES
    # ------------------------------------------------------------------
    def _crear_tab_dev_parciales(self):
        """Crea la pestaña de Devoluciones Parciales."""
        tab = self.tab_dev_parciales
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE HERRAMIENTAS
        # ═══════════════════════════════════════════════════════════════
        toolbar = ttk.Frame(tab)
        toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        
        # Botón cargar desde Firebird
        ttk.Button(toolbar, text="📥 Cargar desde Eleventa", 
                   command=self._cargar_dev_parciales_firebird).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(toolbar, text="🔄 Refrescar", 
                   command=self._refrescar_dev_parciales_tab).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        # Total general en toolbar
        ttk.Label(toolbar, text="Total:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=5)
        self.lbl_total_dev_parciales = ttk.Label(toolbar, text="$0.00", 
                                                   font=("Segoe UI", 12, "bold"), 
                                                   foreground="#7b1fa2")
        self.lbl_total_dev_parciales.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(toolbar, text="   Cantidad:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_cantidad_dev_parciales = ttk.Label(toolbar, text="0", font=("Segoe UI", 10, "bold"))
        self.lbl_cantidad_dev_parciales.pack(side=tk.LEFT, padx=5)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE FILTROS POR FECHA
        # ═══════════════════════════════════════════════════════════════
        toolbar2 = ttk.Frame(tab)
        toolbar2.grid(row=1, column=0, sticky="ew", padx=10, pady=2)
        
        # Filtro por Fecha Venta
        ttk.Label(toolbar2, text="📅 Fecha Venta:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_venta_dp_desde = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                          background='#1976d2', foreground='white',
                                                          headersbackground='#1565c0')
            self.filtro_fecha_venta_dp_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_venta_dp_desde.delete(0, tk.END)
            self.filtro_fecha_venta_dp_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_dev_parciales_tab())
        else:
            self.filtro_fecha_venta_dp_desde = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_venta_dp_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_venta_dp_desde.bind("<KeyRelease>", lambda e: self._refrescar_dev_parciales_tab())
        
        ttk.Label(toolbar2, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_venta_dp_hasta = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                          background='#1976d2', foreground='white',
                                                          headersbackground='#1565c0')
            self.filtro_fecha_venta_dp_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_venta_dp_hasta.delete(0, tk.END)
            self.filtro_fecha_venta_dp_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_dev_parciales_tab())
        else:
            self.filtro_fecha_venta_dp_hasta = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_venta_dp_hasta.pack(side=tk.LEFT, padx=(3, 15))
            self.filtro_fecha_venta_dp_hasta.bind("<KeyRelease>", lambda e: self._refrescar_dev_parciales_tab())
        
        # Separador visual
        ttk.Separator(toolbar2, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        # Filtro por Fecha Devolución
        ttk.Label(toolbar2, text="🔄 Fecha Devolución:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Label(toolbar2, text="Desde:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_dev_dp_desde = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                        background='#7b1fa2', foreground='white',
                                                        headersbackground='#6a1b9a')
            self.filtro_fecha_dev_dp_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_dev_dp_desde.delete(0, tk.END)
            self.filtro_fecha_dev_dp_desde.bind("<<DateEntrySelected>>", lambda e: self._refrescar_dev_parciales_tab())
        else:
            self.filtro_fecha_dev_dp_desde = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_dev_dp_desde.pack(side=tk.LEFT, padx=(3, 8))
            self.filtro_fecha_dev_dp_desde.bind("<KeyRelease>", lambda e: self._refrescar_dev_parciales_tab())
        
        ttk.Label(toolbar2, text="Hasta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        if HAS_CALENDAR:
            self.filtro_fecha_dev_dp_hasta = DateEntry(toolbar2, width=10, date_pattern='yyyy-mm-dd',
                                                        background='#7b1fa2', foreground='white',
                                                        headersbackground='#6a1b9a')
            self.filtro_fecha_dev_dp_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_dev_dp_hasta.delete(0, tk.END)
            self.filtro_fecha_dev_dp_hasta.bind("<<DateEntrySelected>>", lambda e: self._refrescar_dev_parciales_tab())
        else:
            self.filtro_fecha_dev_dp_hasta = ttk.Entry(toolbar2, width=10)
            self.filtro_fecha_dev_dp_hasta.pack(side=tk.LEFT, padx=(3, 10))
            self.filtro_fecha_dev_dp_hasta.bind("<KeyRelease>", lambda e: self._refrescar_dev_parciales_tab())
        
        ttk.Button(toolbar2, text="Hoy", width=5, 
                   command=lambda: self._set_fecha_hoy_dp()).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar2, text="Limpiar", width=7, 
                   command=lambda: self._limpiar_filtro_fecha_dp()).pack(side=tk.LEFT, padx=2)

        # ═══════════════════════════════════════════════════════════════
        # LISTADO DE DEVOLUCIONES PARCIALES
        # ═══════════════════════════════════════════════════════════════
        frame_lista = ttk.Frame(tab)
        frame_lista.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        frame_lista.columnconfigure(0, weight=1)
        frame_lista.rowconfigure(0, weight=1)
        
        # Treeview (orden: fecha_venta, folio, cliente, total_devolucion, fecha_devolucion)
        columnas = ("fecha_venta", "folio", "cliente", "total_devolucion", "fecha_devolucion")
        self.tree_dev_parciales = ttk.Treeview(frame_lista, columns=columnas, show="headings", height=15)
        
        self.tree_dev_parciales.heading("fecha_venta", text="Fecha Venta", anchor=tk.CENTER)
        self.tree_dev_parciales.heading("folio", text="Folio", anchor=tk.CENTER)
        self.tree_dev_parciales.heading("cliente", text="Cliente", anchor=tk.W)
        self.tree_dev_parciales.heading("total_devolucion", text="Total Devolución", anchor=tk.E)
        self.tree_dev_parciales.heading("fecha_devolucion", text="Fecha Devolución", anchor=tk.CENTER)
        
        self.tree_dev_parciales.column("fecha_venta", width=100, anchor=tk.CENTER)
        self.tree_dev_parciales.column("folio", width=80, anchor=tk.CENTER)
        self.tree_dev_parciales.column("cliente", width=250, anchor=tk.W)
        self.tree_dev_parciales.column("total_devolucion", width=130, anchor=tk.E)
        self.tree_dev_parciales.column("fecha_devolucion", width=120, anchor=tk.CENTER)
        
        scrolly = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=self.tree_dev_parciales.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient=tk.HORIZONTAL, command=self.tree_dev_parciales.xview)
        self.tree_dev_parciales.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_dev_parciales.grid(row=0, column=0, sticky="nsew")
        scrolly.grid(row=0, column=1, sticky="ns")
        scrollx.grid(row=1, column=0, sticky="ew")
        
        # Tags para colores según si es del mismo día o de otro día
        self.tree_dev_parciales.tag_configure("mismo_dia", background="#1b5e20", foreground="#a5d6a7")
        self.tree_dev_parciales.tag_configure("otro_dia", background="#bf360c", foreground="#ffcc80")
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA INFERIOR DE TOTALES FILTRADOS
        # ═══════════════════════════════════════════════════════════════
        frame_totales = ttk.Frame(tab)
        frame_totales.grid(row=3, column=0, sticky="ew", padx=10, pady=5)
        
        ttk.Label(frame_totales, text="📊 TOTALES FILTRADOS:", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(frame_totales, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Label(frame_totales, text="Mismo Día:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_total_dp_mismo_dia = ttk.Label(frame_totales, text="$0", 
                                                  font=("Segoe UI", 11, "bold"), foreground="#81c784")
        self.lbl_total_dp_mismo_dia.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(frame_totales, text="(", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.lbl_cant_dp_mismo_dia = ttk.Label(frame_totales, text="0", font=("Segoe UI", 9, "bold"))
        self.lbl_cant_dp_mismo_dia.pack(side=tk.LEFT)
        ttk.Label(frame_totales, text=" registros)", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Separator(frame_totales, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Label(frame_totales, text="Otro Día:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_total_dp_otro_dia = ttk.Label(frame_totales, text="$0", 
                                                 font=("Segoe UI", 11, "bold"), foreground="#ffcc80")
        self.lbl_total_dp_otro_dia.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(frame_totales, text="(", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.lbl_cant_dp_otro_dia = ttk.Label(frame_totales, text="0", font=("Segoe UI", 9, "bold"))
        self.lbl_cant_dp_otro_dia.pack(side=tk.LEFT)
        ttk.Label(frame_totales, text=" registros)", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        
        # Cargar datos iniciales
        self._refrescar_dev_parciales_tab()
    
    def _set_fecha_hoy_dp(self):
        """Establece la fecha de hoy en los filtros de fecha de Dev. Parciales."""
        from datetime import date
        hoy = date.today()
        # Setea fecha venta
        if HAS_CALENDAR:
            self.filtro_fecha_venta_dp_desde.set_date(hoy)
            self.filtro_fecha_venta_dp_hasta.set_date(hoy)
        else:
            self.filtro_fecha_venta_dp_desde.delete(0, tk.END)
            self.filtro_fecha_venta_dp_desde.insert(0, hoy.isoformat())
            self.filtro_fecha_venta_dp_hasta.delete(0, tk.END)
            self.filtro_fecha_venta_dp_hasta.insert(0, hoy.isoformat())
        self._refrescar_dev_parciales_tab()
    
    def _limpiar_filtro_fecha_dp(self):
        """Limpia los filtros de fecha de Devoluciones Parciales."""
        # Limpiar fecha venta
        self.filtro_fecha_venta_dp_desde.delete(0, tk.END)
        self.filtro_fecha_venta_dp_hasta.delete(0, tk.END)
        # Limpiar fecha devolución
        self.filtro_fecha_dev_dp_desde.delete(0, tk.END)
        self.filtro_fecha_dev_dp_hasta.delete(0, tk.END)
        self._refrescar_dev_parciales_tab()
    
    def _cargar_dev_parciales_firebird(self):
        """Carga las devoluciones parciales desde Firebird/Eleventa."""
        if not self.db_manager:
            messagebox.showwarning("Sin conexión", "No hay conexión a la base de datos Firebird.")
            return
        
        # Obtener fechas de filtro o usar fechas por defecto (último mes)
        fecha_desde = self.filtro_fecha_venta_dp_desde.get().strip() if hasattr(self, 'filtro_fecha_venta_dp_desde') else ""
        fecha_hasta = self.filtro_fecha_venta_dp_hasta.get().strip() if hasattr(self, 'filtro_fecha_venta_dp_hasta') else ""
        
        from datetime import date, timedelta, datetime
        if not fecha_desde:
            fecha_desde = (date.today() - timedelta(days=30)).isoformat()
        if not fecha_hasta:
            fecha_hasta = date.today().isoformat()
        
        # Calcular fecha siguiente para el filtro (exclusivo)
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_siguiente = (fecha_hasta_dt + timedelta(days=1)).strftime('%Y-%m-%d')
        except:
            fecha_siguiente = fecha_hasta
        
        # Consulta SQL para devoluciones parciales (TIPO_DEVOLUCION = 'P')
        sql = f"""
SET HEADING ON;
SELECT D.ID, D.TURNO_ID, V.FOLIO, V.NOMBRE, 
       CAST(V.CREADO_EN AS DATE) AS FECHA_VENTA,
       CAST(D.DEVUELTO_EN AS DATE) AS FECHA_DEVOLUCION,
       D.TOTAL_DEVUELTO
FROM DEVOLUCIONES D
JOIN VENTATICKETS V ON D.TICKET_ID = V.ID
WHERE D.TIPO_DEVOLUCION = 'P'
  AND V.CREADO_EN >= '{fecha_desde}'
  AND V.CREADO_EN < '{fecha_siguiente}'
ORDER BY D.DEVUELTO_EN DESC, V.FOLIO;
"""
        
        resultado, error = self.db_manager.ejecutar_sql(sql)
        
        if error:
            messagebox.showerror("Error", f"Error al consultar Firebird:\n{error}")
            return
        
        # Parsear resultados
        count = 0
        for linea in resultado.split('\n'):
            linea = linea.strip()
            if not linea or linea.startswith('=') or 'ID' in linea or 'FOLIO' in linea or 'Database' in linea or 'SQL>' in linea:
                continue
            
            partes = linea.split()
            if len(partes) >= 7:
                try:
                    dev_id = int(partes[0])
                    turno_id = int(partes[1])
                    folio = int(partes[2])
                    # Nombre puede tener espacios, tomar hasta encontrar la fecha
                    # Buscar el índice donde empieza la fecha (formato YYYY-MM-DD)
                    nombre_parts = []
                    fecha_venta_idx = -1
                    for i, p in enumerate(partes[3:], start=3):
                        if len(p) == 10 and p[4] == '-' and p[7] == '-':
                            fecha_venta_idx = i
                            break
                        nombre_parts.append(p)
                    
                    if fecha_venta_idx == -1:
                        continue
                    
                    nombre = ' '.join(nombre_parts)
                    fecha_venta = partes[fecha_venta_idx]
                    fecha_devolucion = partes[fecha_venta_idx + 1]
                    total_devuelto = float(partes[fecha_venta_idx + 2])
                    
                    # Guardar en SQLite
                    db_local.guardar_devolucion_parcial(
                        fecha=fecha_venta,
                        folio=folio,
                        devolucion_id=dev_id,
                        codigo='',
                        descripcion=nombre,
                        cantidad=1,
                        valor_unitario=total_devuelto,
                        dinero=total_devuelto,
                        fecha_devolucion=fecha_devolucion
                    )
                    count += 1
                except Exception as e:
                    print(f"Error parseando línea: {linea} - {e}")
                    continue
        
        messagebox.showinfo("Carga completa", f"Se cargaron {count} devoluciones parciales desde Eleventa.")
        self._refrescar_dev_parciales_tab()
    
    def _refrescar_dev_parciales_tab(self):
        """Refresca la lista de devoluciones parciales."""
        if not hasattr(self, 'tree_dev_parciales'):
            return
            
        self.tree_dev_parciales.delete(*self.tree_dev_parciales.get_children())
        
        if not USE_SQLITE:
            return
        
        # Obtener filtros de fecha venta
        filtro_venta_desde = ""
        filtro_venta_hasta = ""
        if hasattr(self, 'filtro_fecha_venta_dp_desde'):
            filtro_venta_desde = self.filtro_fecha_venta_dp_desde.get().strip()
        if hasattr(self, 'filtro_fecha_venta_dp_hasta'):
            filtro_venta_hasta = self.filtro_fecha_venta_dp_hasta.get().strip()
        
        # Obtener filtros de fecha devolución
        filtro_dev_desde = ""
        filtro_dev_hasta = ""
        if hasattr(self, 'filtro_fecha_dev_dp_desde'):
            filtro_dev_desde = self.filtro_fecha_dev_dp_desde.get().strip()
        if hasattr(self, 'filtro_fecha_dev_dp_hasta'):
            filtro_dev_hasta = self.filtro_fecha_dev_dp_hasta.get().strip()
        
        # Obtener todas las devoluciones parciales de SQLite
        devoluciones = db_local.obtener_todas_dev_parciales()
        
        total_general = 0
        count_total = 0
        total_mismo_dia = 0
        count_mismo_dia = 0
        total_otro_dia = 0
        count_otro_dia = 0
        
        for dev in devoluciones:
            fecha_venta = dev.get('fecha', '')
            fecha_devolucion = dev.get('fecha_devolucion', '')
            folio = dev.get('folio', 0)
            cliente = dev.get('descripcion_producto', '')  # Usamos descripción como nombre
            total = dev.get('dinero_devuelto', 0)
            
            # Aplicar filtros de fecha venta
            if filtro_venta_desde and fecha_venta < filtro_venta_desde:
                continue
            if filtro_venta_hasta and fecha_venta > filtro_venta_hasta:
                continue
            
            # Aplicar filtros de fecha devolución
            if filtro_dev_desde and fecha_devolucion and fecha_devolucion < filtro_dev_desde:
                continue
            if filtro_dev_hasta and fecha_devolucion and fecha_devolucion > filtro_dev_hasta:
                continue
            
            # Determinar si es mismo día u otro día
            es_mismo_dia = fecha_venta == fecha_devolucion
            tag = "mismo_dia" if es_mismo_dia else "otro_dia"
            
            self.tree_dev_parciales.insert("", tk.END, values=(
                fecha_venta,
                folio,
                cliente,
                f"${total:,.2f}",
                fecha_devolucion or ""
            ), tags=(tag,))
            
            total_general += total
            count_total += 1
            
            if es_mismo_dia:
                total_mismo_dia += total
                count_mismo_dia += 1
            else:
                total_otro_dia += total
                count_otro_dia += 1
        
        # Actualizar totales
        self.lbl_total_dev_parciales.config(text=f"${total_general:,.2f}")
        self.lbl_cantidad_dev_parciales.config(text=str(count_total))
        self.lbl_total_dp_mismo_dia.config(text=f"${total_mismo_dia:,.2f}")
        self.lbl_cant_dp_mismo_dia.config(text=str(count_mismo_dia))
        self.lbl_total_dp_otro_dia.config(text=f"${total_otro_dia:,.2f}")
        self.lbl_cant_dp_otro_dia.config(text=str(count_otro_dia))
    
    # ------------------------------------------------------------------
    # PESTAÑA CANCELADAS Y DEVOLUCIONES
    # ------------------------------------------------------------------
    def _crear_tab_canceladas(self):
        """Crea la pestaña de Canceladas y Devoluciones con detalle de bugs."""
        tab = self.tab_canceladas
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)
        
        # ═══════════════════════════════════════════════════════════════
        # BARRA DE HERRAMIENTAS
        # ═══════════════════════════════════════════════════════════════
        toolbar = ttk.Frame(tab)
        toolbar.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))
        
        ttk.Button(toolbar, text="📥 Cargar Canceladas", 
                   command=self._cargar_canceladas_firebird).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(toolbar, text="🔄 Refrescar", 
                   command=self._refrescar_canceladas_tab).pack(side=tk.LEFT, padx=5)
        
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        # Totales en toolbar
        ttk.Label(toolbar, text="Total Cancelaciones:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=5)
        self.lbl_total_canceladas = ttk.Label(toolbar, text="$0.00", 
                                               font=("Segoe UI", 11, "bold"), 
                                               foreground="#e53935")
        self.lbl_total_canceladas.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(toolbar, text="   Total Dev. Parciales:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=5)
        self.lbl_total_dev_parc_cancel = ttk.Label(toolbar, text="$0.00", 
                                                    font=("Segoe UI", 11, "bold"), 
                                                    foreground="#ff9800")
        self.lbl_total_dev_parc_cancel.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(toolbar, text="   Bugs Duplicados:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=5)
        self.lbl_total_bugs_cancel = ttk.Label(toolbar, text="$0.00", 
                                                font=("Segoe UI", 11, "bold"), 
                                                foreground="#ff5722")
        self.lbl_total_bugs_cancel.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(toolbar, text="   Cantidad:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=5)
        self.lbl_cantidad_canceladas = ttk.Label(toolbar, text="0", font=("Segoe UI", 10, "bold"))
        self.lbl_cantidad_canceladas.pack(side=tk.LEFT, padx=5)

        # ═══════════════════════════════════════════════════════════════
        # LISTADO DE CANCELADAS/DEVOLUCIONES
        # ═══════════════════════════════════════════════════════════════
        frame_lista = ttk.Frame(tab)
        frame_lista.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        frame_lista.columnconfigure(0, weight=1)
        frame_lista.rowconfigure(0, weight=1)
        
        # Columnas: Tipo, Folio, Cliente, Fecha Venta, Fecha Cancel/Dev, Total Original, Después Dev, Cancelación, Bug
        columnas = ("tipo", "folio", "cliente", "fecha_venta", "fecha_cancel", 
                    "total_original", "despues_dev", "cancelacion", "bug")
        self.tree_canceladas = ttk.Treeview(frame_lista, columns=columnas, show="headings", height=20)
        
        self.tree_canceladas.heading("tipo", text="Tipo", anchor=tk.CENTER)
        self.tree_canceladas.heading("folio", text="Folio", anchor=tk.CENTER)
        self.tree_canceladas.heading("cliente", text="Cliente", anchor=tk.W)
        self.tree_canceladas.heading("fecha_venta", text="Fecha Venta", anchor=tk.CENTER)
        self.tree_canceladas.heading("fecha_cancel", text="Fecha Cancel", anchor=tk.CENTER)
        self.tree_canceladas.heading("total_original", text="Total Original", anchor=tk.E)
        self.tree_canceladas.heading("despues_dev", text="Después Dev.", anchor=tk.E)
        self.tree_canceladas.heading("cancelacion", text="Cancelación", anchor=tk.E)
        self.tree_canceladas.heading("bug", text="Bug Dup.", anchor=tk.E)
        
        self.tree_canceladas.column("tipo", width=100, anchor=tk.CENTER)
        self.tree_canceladas.column("folio", width=70, anchor=tk.CENTER)
        self.tree_canceladas.column("cliente", width=200, anchor=tk.W)
        self.tree_canceladas.column("fecha_venta", width=100, anchor=tk.CENTER)
        self.tree_canceladas.column("fecha_cancel", width=100, anchor=tk.CENTER)
        self.tree_canceladas.column("total_original", width=110, anchor=tk.E)
        self.tree_canceladas.column("despues_dev", width=110, anchor=tk.E)
        self.tree_canceladas.column("cancelacion", width=110, anchor=tk.E)
        self.tree_canceladas.column("bug", width=100, anchor=tk.E)
        
        scrolly = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL, command=self.tree_canceladas.yview)
        scrollx = ttk.Scrollbar(frame_lista, orient=tk.HORIZONTAL, command=self.tree_canceladas.xview)
        self.tree_canceladas.configure(yscrollcommand=scrolly.set, xscrollcommand=scrollx.set)
        
        self.tree_canceladas.grid(row=0, column=0, sticky="nsew")
        scrolly.grid(row=0, column=1, sticky="ns")
        scrollx.grid(row=1, column=0, sticky="ew")
        
        # Tags para colores según tipo
        self.tree_canceladas.tag_configure("cancelacion", background="#b71c1c", foreground="#ffcdd2")
        self.tree_canceladas.tag_configure("dev_parcial", background="#e65100", foreground="#ffe0b2")
        self.tree_canceladas.tag_configure("bug_dup", background="#4a148c", foreground="#e1bee7")
        self.tree_canceladas.tag_configure("normal", background="", foreground="")
        
        # Cargar datos iniciales
        self._refrescar_canceladas_tab()
    
    def _cargar_canceladas_firebird(self):
        """Carga las cancelaciones y devoluciones desde Firebird."""
        # Usar la fecha actual del selector si está disponible
        fecha = self.ds.fecha if hasattr(self, 'ds') and self.ds and self.ds.fecha else None
        if not fecha:
            from datetime import date
            fecha = date.today().isoformat()
        
        try:
            import fdb
            import re
            from corte_cajero import DB_PATH_DEFAULT
            
            conn = fdb.connect(
                dsn=DB_PATH_DEFAULT,
                user='SYSDBA',
                password='masterkey',
                charset='UTF8'
            )
            cur = conn.cursor()
            
            # Obtener turnos de la fecha
            cur.execute(f"""
                SELECT ID FROM TURNOS 
                WHERE CAST(INICIO_EN AS DATE) = '{fecha}'
            """)
            turnos = [row[0] for row in cur.fetchall()]
            
            if not turnos:
                messagebox.showinfo("Sin datos", f"No hay turnos para la fecha {fecha}")
                conn.close()
                return
            
            turnos_str = ','.join(map(str, turnos))
            
            # Limpiar tabla SQLite para esta fecha
            conn_local = db_local.get_connection()
            cursor_local = conn_local.cursor()
            cursor_local.execute('DELETE FROM cancelaciones_detalle WHERE fecha = ?', (fecha,))
            
            count = 0
            
            # 1. Obtener tickets cancelados
            cur.execute(f'''
                SELECT V.FOLIO, V.NOMBRE, V.TOTAL, V.TURNO_ID,
                       CAST(V.CREADO_EN AS DATE) AS FECHA_VENTA
                FROM VENTATICKETS V
                WHERE V.TURNO_ID IN ({turnos_str})
                AND V.ESTA_CANCELADO = 't'
            ''')
            
            tickets_cancelados = {}
            for row in cur.fetchall():
                folio = row[0]
                tickets_cancelados[folio] = {
                    'folio': folio,
                    'cliente': row[1] or 'MOSTRADOR',
                    'total_original': float(row[2]) if row[2] else 0.0,
                    'turno_id': row[3],
                    'fecha_venta': str(row[4]) if row[4] else fecha
                }
            
            # 2. Obtener devoluciones por folio de CORTE_MOVIMIENTOS
            cur.execute(f'''
                SELECT ID_TURNO, DESCRIPCION, MONTO,
                       CAST(CUANDO_FUE AS DATE) AS FECHA_CANCEL
                FROM CORTE_MOVIMIENTOS
                WHERE ID_TURNO IN ({turnos_str})
                AND TIPO CONTAINING 'Devol'
            ''')
            
            devoluciones_cm = {}
            for row in cur.fetchall():
                turno_id = row[0]
                desc = row[1] or ''
                monto = float(row[2]) if row[2] else 0.0
                fecha_cancel = str(row[3]) if row[3] else fecha
                
                # Extraer folio de descripción
                match = re.search(r'#(\d+)', desc)
                if match:
                    folio = int(match.group(1))
                    if folio not in devoluciones_cm:
                        devoluciones_cm[folio] = {
                            'monto_total': 0.0,
                            'fecha_cancel': fecha_cancel,
                            'turnos': [],
                            'detalle': []
                        }
                    devoluciones_cm[folio]['monto_total'] += monto
                    if turno_id not in devoluciones_cm[folio]['turnos']:
                        devoluciones_cm[folio]['turnos'].append(turno_id)
                    devoluciones_cm[folio]['detalle'].append({
                        'turno': turno_id,
                        'desc': desc,
                        'monto': monto
                    })
            
            # 3. Guardar en SQLite combinando información
            for folio, info in tickets_cancelados.items():
                cm_info = devoluciones_cm.get(folio, {})
                monto_cm = cm_info.get('monto_total', 0.0)
                fecha_cancel = cm_info.get('fecha_cancel', fecha)
                
                # Calcular bug de duplicidad (si CM > ticket total)
                bug_dup = max(0, monto_cm - info['total_original'])
                despues_dev = info['total_original'] - (monto_cm - bug_dup) if monto_cm > 0 else info['total_original']
                
                tipo = 'CANCELACIÓN' if monto_cm >= info['total_original'] * 0.95 else 'DEV. PARCIAL'
                if bug_dup > 0:
                    tipo = 'BUG DUP.'
                
                cursor_local.execute('''
                    INSERT OR REPLACE INTO cancelaciones_detalle 
                    (fecha, folio, cliente, fecha_venta, fecha_cancel, 
                     total_original, despues_dev, cancelacion, bug_dup, tipo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (fecha, folio, info['cliente'], info['fecha_venta'], fecha_cancel,
                      info['total_original'], despues_dev, monto_cm, bug_dup, tipo))
                count += 1
            
            # 4. Agregar devoluciones que no son cancelaciones (folios no cancelados pero con devol)
            for folio, cm_info in devoluciones_cm.items():
                if folio not in tickets_cancelados:
                    # Buscar info del ticket
                    cur.execute(f'''
                        SELECT NOMBRE, TOTAL, CAST(CREADO_EN AS DATE) 
                        FROM VENTATICKETS WHERE FOLIO = {folio}
                    ''')
                    row = cur.fetchone()
                    if row:
                        cliente = row[0] or 'MOSTRADOR'
                        total_orig = float(row[1]) if row[1] else 0.0
                        fecha_venta = str(row[2]) if row[2] else ''
                        
                        monto_cm = cm_info['monto_total']
                        despues_dev = total_orig - monto_cm
                        
                        cursor_local.execute('''
                            INSERT OR REPLACE INTO cancelaciones_detalle 
                            (fecha, folio, cliente, fecha_venta, fecha_cancel, 
                             total_original, despues_dev, cancelacion, bug_dup, tipo)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (fecha, folio, cliente, fecha_venta, cm_info['fecha_cancel'],
                              total_orig, despues_dev, monto_cm, 0, 'DEV. PARCIAL'))
                        count += 1
            
            conn_local.commit()
            conn_local.close()
            conn.close()
            
            messagebox.showinfo("Carga completa", f"Se cargaron {count} registros de cancelaciones/devoluciones.")
            self._refrescar_canceladas_tab()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Error al cargar cancelaciones:\n{str(e)}")
    
    def _refrescar_canceladas_tab(self):
        """Refresca la lista de canceladas y devoluciones."""
        if not hasattr(self, 'tree_canceladas'):
            return
            
        self.tree_canceladas.delete(*self.tree_canceladas.get_children())
        
        if not USE_SQLITE:
            return
        
        # Usar la fecha actual del selector
        fecha = self.ds.fecha if hasattr(self, 'ds') and self.ds and self.ds.fecha else None
        if not fecha:
            return
        
        try:
            conn = db_local.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT tipo, folio, cliente, fecha_venta, fecha_cancel,
                       total_original, despues_dev, cancelacion, bug_dup
                FROM cancelaciones_detalle
                WHERE fecha = ?
                ORDER BY folio
            ''', (fecha,))
            
            total_cancelaciones = 0
            total_dev_parciales = 0
            total_bugs = 0
            count = 0
            
            for row in cursor.fetchall():
                tipo = row[0] or ''
                folio = row[1]
                cliente = row[2] or 'MOSTRADOR'
                fecha_venta = row[3] or ''
                fecha_cancel = row[4] or ''
                total_original = row[5] or 0
                despues_dev = row[6] or 0
                cancelacion = row[7] or 0
                bug_dup = row[8] or 0
                
                # Determinar tag según tipo
                if 'BUG' in tipo.upper():
                    tag = 'bug_dup'
                    total_bugs += bug_dup
                elif 'CANCEL' in tipo.upper():
                    tag = 'cancelacion'
                    total_cancelaciones += cancelacion
                else:
                    tag = 'dev_parcial'
                    total_dev_parciales += cancelacion
                
                self.tree_canceladas.insert("", tk.END, values=(
                    tipo,
                    folio,
                    cliente[:40],
                    fecha_venta,
                    fecha_cancel,
                    f"${total_original:,.2f}",
                    f"${despues_dev:,.2f}" if despues_dev else "",
                    f"${cancelacion:,.2f}",
                    f"${bug_dup:,.2f}" if bug_dup > 0 else ""
                ), tags=(tag,))
                
                count += 1
            
            conn.close()
            
            # Actualizar totales
            self.lbl_total_canceladas.config(text=f"${total_cancelaciones:,.2f}")
            self.lbl_total_dev_parc_cancel.config(text=f"${total_dev_parciales:,.2f}")
            self.lbl_total_bugs_cancel.config(text=f"${total_bugs:,.2f}")
            self.lbl_cantidad_canceladas.config(text=str(count))
            
        except Exception as e:
            print(f"Error refrescando canceladas: {e}")
    
    def _refrescar_no_entregados_tab(self):
        """Refresca la lista de no entregados."""
        self.tree_no_entregados.delete(*self.tree_no_entregados.get_children())
        
        if not USE_SQLITE:
            return
        
        filtro_cliente = self.buscar_no_entregados_var.get().strip().lower() if hasattr(self, 'buscar_no_entregados_var') else ""
        filtro_estado = self.filtro_estado_ne_var.get() if hasattr(self, 'filtro_estado_ne_var') else "Todos"
        
        # Obtener filtros de fecha venta
        filtro_venta_desde = ""
        filtro_venta_hasta = ""
        if hasattr(self, 'filtro_fecha_venta_ne_desde'):
            filtro_venta_desde = self.filtro_fecha_venta_ne_desde.get().strip()
        if hasattr(self, 'filtro_fecha_venta_ne_hasta'):
            filtro_venta_hasta = self.filtro_fecha_venta_ne_hasta.get().strip()
        
        # Obtener filtros de fecha pagado
        filtro_pagado_desde = ""
        filtro_pagado_hasta = ""
        if hasattr(self, 'filtro_fecha_pagado_ne_desde'):
            filtro_pagado_desde = self.filtro_fecha_pagado_ne_desde.get().strip()
        if hasattr(self, 'filtro_fecha_pagado_ne_hasta'):
            filtro_pagado_hasta = self.filtro_fecha_pagado_ne_hasta.get().strip()
        
        no_entregados = db_local.obtener_todos_no_entregados()
        
        total_valor = 0
        total_pendiente = 0
        count_total = 0
        
        # Totales para barra inferior
        total_pagado_filtrado = 0
        count_pagado_filtrado = 0
        total_pendiente_filtrado = 0
        count_pendiente_filtrado = 0
        
        for ne in no_entregados:
            fecha = ne.get('fecha', '')
            folio = ne.get('folio', '')
            cliente = ne.get('cliente', '') or 'MOSTRADOR'
            valor = ne.get('subtotal', 0) or 0
            abono = ne.get('abono', 0) or 0
            estado = ne.get('estado', 'PENDIENTE') or 'PENDIENTE'
            fecha_pagado = ne.get('fecha_devuelto', '') or ''
            repartidor = ne.get('repartidor', '') or ''
            
            # Filtros de texto
            if filtro_cliente:
                if filtro_cliente not in cliente.lower() and filtro_cliente not in str(folio):
                    continue
            
            if filtro_estado != "Todos" and estado != filtro_estado:
                continue
            
            # Aplicar filtro de fecha venta
            if filtro_venta_desde or filtro_venta_hasta:
                if filtro_venta_desde and fecha < filtro_venta_desde:
                    continue
                if filtro_venta_hasta and fecha > filtro_venta_hasta:
                    continue
            
            # Aplicar filtro de fecha pagado (solo para PAGADO)
            if filtro_pagado_desde or filtro_pagado_hasta:
                if estado == 'PAGADO' and fecha_pagado:
                    if filtro_pagado_desde and fecha_pagado < filtro_pagado_desde:
                        continue
                    if filtro_pagado_hasta and fecha_pagado > filtro_pagado_hasta:
                        continue
                elif estado != 'PENDIENTE':
                    continue  # Si tiene filtro de fecha pagado y no es PAGADO ni PENDIENTE, omitir
            
            total_valor += valor
            saldo = valor - abono
            if estado == 'PENDIENTE':
                total_pendiente += saldo
                total_pendiente_filtrado += saldo
                count_pendiente_filtrado += 1
            elif estado == 'PAGADO':
                total_pagado_filtrado += valor
                count_pagado_filtrado += 1
            count_total += 1
            
            # Tag
            if estado == "PAGADO":
                tag = "pagado"
            elif estado == "CANCELADA":
                tag = "cancelada"
            else:
                tag = "pendiente"
            
            self.tree_no_entregados.insert("", tk.END, values=(
                fecha,
                folio,
                cliente,
                f"${valor:,.0f}",
                f"${abono:,.0f}",
                f"${saldo:,.0f}",
                estado,
                fecha_pagado,
                repartidor
            ), tags=(tag,))
        
        # Actualizar etiquetas de totales (barra superior)
        self.lbl_cantidad_ne.config(text=str(count_total))
        self.lbl_total_ne_pendiente.config(text=f"${total_pendiente:,.0f}")
        
        # Actualizar etiquetas de totales filtrados (barra inferior)
        if hasattr(self, 'lbl_total_pagado_ne_filtrado'):
            self.lbl_total_pagado_ne_filtrado.config(text=f"${total_pagado_filtrado:,.0f}")
            self.lbl_cant_pagado_ne_filtrado.config(text=f"({count_pagado_filtrado})")
            self.lbl_total_pendiente_ne_filtrado.config(text=f"${total_pendiente_filtrado:,.0f}")
            self.lbl_cant_pendiente_ne_filtrado.config(text=f"({count_pendiente_filtrado})")

    # ------------------------------------------------------------------
    # CREAR PESTAÑA DE ANOTACIONES
    # ------------------------------------------------------------------
    def _crear_tab_anotaciones(self):
        """Crea la pestaña de anotaciones (sticky notes)."""
        try:
            self.anotaciones_widget = TabAnotaciones(self.tab_anotaciones, self, self.ds)
        except Exception as e:
            print(f"⚠️ Error creando tab anotaciones: {e}")
            ttk.Label(self.tab_anotaciones, 
                     text=f"Error al cargar anotaciones: {e}",
                     foreground="red").pack(padx=20, pady=20)

    # ------------------------------------------------------------------
    # FILTRO GENERAL: se ejecuta cuando cambia el filtro de estado
    # ------------------------------------------------------------------
    def _on_filtro_general_cambio(self, event=None):
        """Actualiza todas las vistas cuando cambia el filtro de estado."""
        self._filtrar_facturas_asign()
        self._refrescar_liquidacion()

    # ------------------------------------------------------------------
    # BUSCADOR GLOBAL: se ejecuta al escribir en el buscador
    # ------------------------------------------------------------------
    def _on_buscar_global(self):
        """Aplica el filtro de búsqueda en todos los módulos."""
        # Filtrar en asignación
        self._filtrar_facturas_asign()
        # Filtrar en liquidación
        self._refrescar_liquidacion()
        # Filtrar en créditos (usa el buscador global)
        if hasattr(self, 'tree_creditos_punt'):
            self._refrescar_creditos_tab()
    
    def _limpiar_buscar_global(self):
        """Limpia el buscador global."""
        self.buscar_global_var.set("")
        self.entry_buscar_global.focus_set()
    
    def _enfocar_buscador_seleccionar(self, event=None):
        """Enfoca el buscador y selecciona todo el texto (F10)."""
        self.entry_buscar_global.focus_set()
        self.entry_buscar_global.select_range(0, tk.END)
        self.entry_buscar_global.icursor(tk.END)
        return "break"
    
    def _saltar_al_listado(self, event=None):
        """Salta al listado de facturas para asignar repartidor (Enter en buscador)."""
        # Obtener la pestaña actual
        tab_actual = self.notebook.index(self.notebook.select())
        
        if tab_actual == 0:  # Pestaña de Asignación
            # Seleccionar el primer item si hay resultados
            items = self.tree_asign.get_children()
            if items:
                self.tree_asign.focus_set()
                self.tree_asign.selection_set(items[0])
                self.tree_asign.focus(items[0])
                self.tree_asign.see(items[0])
        elif tab_actual == 1:  # Pestaña de Liquidación
            if hasattr(self, 'tree_liq'):
                items = self.tree_liq.get_children()
                if items:
                    self.tree_liq.focus_set()
                    self.tree_liq.selection_set(items[0])
                    self.tree_liq.focus(items[0])
                    self.tree_liq.see(items[0])
        elif tab_actual == 2:  # Pestaña de Descuentos
            if hasattr(self, 'tree_descuentos'):
                items = self.tree_descuentos.get_children()
                if items:
                    self.tree_descuentos.focus_set()
                    self.tree_descuentos.selection_set(items[0])
                    self.tree_descuentos.focus(items[0])
                    self.tree_descuentos.see(items[0])
        return "break"

    # ------------------------------------------------------------------
    # FILTRO REPARTIDOR GLOBAL: se ejecuta cuando cambia el repartidor
    # ------------------------------------------------------------------
    def _on_filtro_rep_global_cambio(self, event=None):
        """Actualiza TODAS las vistas cuando cambia el repartidor global."""
        rep_seleccionado = self.filtro_rep_global_var.get()
        
        # Sincronizar con el combo de liquidación
        if hasattr(self, 'repartidor_filtro_var'):
            self.repartidor_filtro_var.set(rep_seleccionado)
        
        # Sincronizar con el combo de gastos
        if hasattr(self, 'gasto_rep_var') and rep_seleccionado not in ("(Todos)", "(Sin Asignar)"):
            reps_gastos = self.gasto_rep_combo.cget('values') or []
            if rep_seleccionado in reps_gastos:
                self.gasto_rep_var.set(rep_seleccionado)
        
        # Sincronizar con el combo de conteo de dinero
        if hasattr(self, 'dinero_rep_var') and rep_seleccionado not in ("(Todos)", "(Sin Asignar)"):
            reps_dinero = self.dinero_rep_combo.cget('values') or []
            if rep_seleccionado in reps_dinero:
                # Guardar conteo anterior primero
                if hasattr(self, '_rep_dinero_anterior') and self._rep_dinero_anterior:
                    self._guardar_dinero_rep_especifico(self._rep_dinero_anterior)
                self.dinero_rep_var.set(rep_seleccionado)
                self._cargar_dinero_rep(rep_seleccionado)
                self._rep_dinero_anterior = rep_seleccionado
        
        # Refrescar todas las vistas
        self._filtrar_facturas_asign()
        self._refrescar_liquidacion()
        self._refrescar_tab_gastos()
        self._refrescar_tab_dinero()

    # ------------------------------------------------------------------
    # FECHA GLOBAL: Navegación de fecha (anterior/hoy/siguiente)
    # ------------------------------------------------------------------
    def _cambiar_fecha_global(self, dias: int):
        """Cambia la fecha global por N días (positivo o negativo)."""
        from datetime import timedelta
        try:
            if HAS_CALENDAR:
                fecha_actual = self.fecha_global_entry.get_date()
            else:
                fecha_actual = datetime.strptime(self.fecha_global_var.get(), '%Y-%m-%d')
            
            nueva_fecha = fecha_actual + timedelta(days=dias)
            
            if HAS_CALENDAR:
                self.fecha_global_entry.set_date(nueva_fecha)
            else:
                self.fecha_global_var.set(nueva_fecha.strftime('%Y-%m-%d'))
            
            self._on_fecha_global_cambio()
        except (ValueError, AttributeError):
            pass
    
    def _ir_a_fecha_hoy(self):
        """Establece la fecha global a hoy."""
        hoy = datetime.now()
        if HAS_CALENDAR:
            self.fecha_global_entry.set_date(hoy)
        else:
            self.fecha_global_var.set(hoy.strftime('%Y-%m-%d'))
        self._on_fecha_global_cambio()

    # ------------------------------------------------------------------
    # FECHA GLOBAL: se ejecuta cuando cambia la fecha en la barra
    # ------------------------------------------------------------------
    def _on_fecha_global_cambio(self, event=None):
        """Cambia la fecha del DataStore y recarga los datos de TODOS los módulos."""
        # Obtener nueva fecha
        if HAS_CALENDAR:
            nueva_fecha = self.fecha_global_entry.get_date().strftime('%Y-%m-%d')
        else:
            nueva_fecha = self.fecha_global_var.get().strip()
        
        # Validar formato
        try:
            datetime.strptime(nueva_fecha, '%Y-%m-%d')
        except ValueError:
            messagebox.showwarning("Fecha Inválida", "El formato debe ser YYYY-MM-DD")
            return
        
        # Actualizar fecha en DataStore y variable de sincronización
        self.ds.fecha = nueva_fecha
        if hasattr(self, 'fecha_asign_var'):
            self.fecha_asign_var.set(nueva_fecha)
        
        # Cargar facturas de Firebird (pestaña Asignación)
        self._cargar_facturas()
        
        # Refrescar todos los módulos SQLite (descuentos, gastos, conteo, etc.)
        self._on_data_changed()
        
        # Actualizar corte cajero con la nueva fecha
        self._actualizar_corte_cajero_async()

    # ------------------------------------------------------------------
    # ACTUALIZAR COMBO DE REPARTIDORES GLOBAL
    # ------------------------------------------------------------------
    def _actualizar_combo_rep_global(self):
        """Actualiza el combo de repartidores global con los disponibles."""
        reps = self.ds.get_repartidores()
        valores = ["(Todos)"] + reps
        self.combo_filtro_rep_global['values'] = valores
        # Si el valor actual no está en la lista, resetear a (Todos)
        if self.filtro_rep_global_var.get() not in valores:
            self.filtro_rep_global_var.set("(Todos)")

    # ------------------------------------------------------------------
    # CALLBACK GLOBAL: se ejecuta cada vez que el DataStore cambia
    # ------------------------------------------------------------------
    def _on_data_changed(self):
        self._refrescar_tree_asignacion()
        self._filtrar_facturas_asign()  # Actualiza TOTALES (Monto Efectivo, etc.)
        self._refrescar_liquidacion()
        self._refrescar_folio_combo_descuentos()
        self._refrescar_lista_descuentos()
        self._refrescar_tab_gastos()
        self._refrescar_tab_dinero()
        self._actualizar_combo_rep_global()
        # NOTA: _refrescar_creditos NO se llama aquí, tiene su propio filtro de fecha independiente

    # ------------------------------------------------------------------
    # CONFIGURACION: Seleccionar ruta FDB y verificar conexion
    # ------------------------------------------------------------------
    def _seleccionar_archivo_fdb(self):
        """Abre diálogo para seleccionar archivo FDB"""
        from tkinter import filedialog
        archivo = filedialog.askopenfilename(
            title="Seleccionar PDVDATA.FDB",
            filetypes=[("Firebird Database", "*.FDB"), ("Todos", "*.*")]
        )
        if archivo:
            self.ruta_fdb_var.set(archivo)
            self.ruta_fdb = archivo
            self._verificar_conexion_bd()
    
    def _verificar_conexion_bd(self):
        """Verifica conexión a la BD y actualiza indicador"""
        self.ruta_fdb = self.ruta_fdb_var.get()
        
        if not self.ruta_fdb.strip():
            self.lbl_estado_bd.config(text="● Error: Ruta vacía", foreground="red")
            return
        
        if not os.path.exists(self.ruta_fdb):
            self.lbl_estado_bd.config(text="● Error: Archivo no encontrado", foreground="red")
            messagebox.showerror("Error", f"Archivo no encontrado:\n{self.ruta_fdb}")
            return
        
        # Intentar conectar
        sql = "SELECT COUNT(*) as TEST FROM RDB$RELATIONS;"
        ok, stdout, stderr = self._ejecutar_sql(sql)
        
        if ok and ('TEST' in stdout or 'COUNT' in stdout or stdout.strip()):
            self.lbl_estado_bd.config(text="● Conectado ✓", foreground="green")
            messagebox.showinfo("Conexión", "✓ Conexión a BD establecida correctamente")
        else:
            self.lbl_estado_bd.config(text="● Error de conexión ✗", foreground="red")
            # Mostrar error más detallado
            if "firebird" in stderr.lower() or "no se encontró" in stderr.lower():
                error_msg = (
                    "No se encontró Firebird/isql.\n\n"
                    "Instala Firebird desde: https://www.firebirdsql.org/\n"
                    "O agrega la carpeta 'bin' de Firebird al PATH.\n\n"
                    "Error técnico:\n" + stderr[:200]
                )
            else:
                error_msg = f"No se pudo conectar:\n{stderr[:300]}"
            
            messagebox.showerror("Error de Conexión", error_msg)

    # ==================================================================
    # PESTAÑA 0 – ASIGNAR REPARTIDORES
    # ==================================================================
    def _crear_tab_asignacion(self):
        # --- barra superior: Solo guardar y buscar (fecha está en barra global) ---
        frame_top = ttk.Frame(self.tab_asignacion)
        frame_top.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

        # Variable para sincronizar con fecha global
        self.fecha_asign_var = tk.StringVar(value=self.ds.fecha)
        
        # Diccionario para rastrear cambios pendientes {folio: nuevo_repartidor}
        self._cambios_pendientes = {}
        
        # Variable para compatibilidad con buscador (usa el global)
        self.buscar_asign_var = self.buscar_global_var
        
        # Binding global F10 para enfocar buscador y seleccionar todo el texto
        self.ventana.bind("<F10>", self._enfocar_buscador_seleccionar)
        # Binding global Ctrl+S para guardar cambios pendientes
        self.ventana.bind("<Control-s>", lambda e: self._guardar_cambios_pendientes())
        self.ventana.bind("<Control-S>", lambda e: self._guardar_cambios_pendientes())

        # ============================================================
        # RESUMEN INFERIOR - Se empaqueta PRIMERO con side=BOTTOM
        # para que siempre sea visible
        # ============================================================
        self.resumen_var = tk.StringVar(value="Carga las facturas para ver el resumen.")
        frame_res = ttk.LabelFrame(self.tab_asignacion, text="📊 RESUMEN DE ASIGNACIÓN", padding=(12, 8))
        frame_res.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(4, 10))
        
        # Frame interno para organizar el resumen en columnas
        resumen_content = ttk.Frame(frame_res)
        resumen_content.pack(fill=tk.X, expand=True)
        
        # Columna 1: Totales
        col1 = ttk.Frame(resumen_content)
        col1.pack(side=tk.LEFT, padx=(0, 30))
        
        ttk.Label(col1, text="TOTALES", font=("Segoe UI", 9, "bold"), 
                 foreground="#64b5f6").grid(row=0, column=0, columnspan=2, sticky=tk.W)
        ttk.Label(col1, text="Total Facturas:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_total_facturas_asign = ttk.Label(col1, text="0", font=("Segoe UI", 9, "bold"))
        self.lbl_total_facturas_asign.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col1, text="Monto Efectivo:").grid(row=2, column=0, sticky=tk.W)
        self.lbl_monto_efectivo_asign = ttk.Label(col1, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#81c784")
        self.lbl_monto_efectivo_asign.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col1, text="Canceladas:", foreground="#ef5350").grid(row=3, column=0, sticky=tk.W)
        self.lbl_canceladas_monto_asign = ttk.Label(col1, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#ef5350")
        self.lbl_canceladas_monto_asign.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col1, text="Dev. Parciales:", foreground="#ffb74d").grid(row=4, column=0, sticky=tk.W)
        self.lbl_dev_parciales_asign = ttk.Label(col1, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#ffb74d")
        self.lbl_dev_parciales_asign.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col1, text="Monto Total:", font=("Segoe UI", 9, "bold")).grid(row=5, column=0, sticky=tk.W)
        self.lbl_monto_total_asign = ttk.Label(col1, text="$0.00", font=("Segoe UI", 10, "bold"), foreground="#64b5f6")
        self.lbl_monto_total_asign.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # Columna 2: Asignación
        col2 = ttk.Frame(resumen_content)
        col2.pack(side=tk.LEFT, padx=(0, 30))
        
        ttk.Label(col2, text="ASIGNACIÓN", font=("Segoe UI", 9, "bold"), 
                 foreground="#81c784").grid(row=0, column=0, columnspan=2, sticky=tk.W)
        ttk.Label(col2, text="Asignadas:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_asignadas = ttk.Label(col2, text="0", font=("Segoe UI", 9, "bold"), foreground="#81c784")
        self.lbl_asignadas.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col2, text="Sin Asignar:", foreground="#ef5350").grid(row=2, column=0, sticky=tk.W)
        self.lbl_sin_asignar = ttk.Label(col2, text="0", font=("Segoe UI", 9, "bold"), foreground="#ef5350")
        self.lbl_sin_asignar.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Columna 3: Estados especiales
        col3 = ttk.Frame(resumen_content)
        col3.pack(side=tk.LEFT, padx=(0, 30))
        
        ttk.Label(col3, text="ESTADOS", font=("Segoe UI", 9, "bold"), 
                 foreground="#ffb74d").grid(row=0, column=0, columnspan=2, sticky=tk.W)
        ttk.Label(col3, text="Canceladas:", foreground="#ef5350").grid(row=1, column=0, sticky=tk.W)
        self.lbl_canceladas_asign = ttk.Label(col3, text="0", font=("Segoe UI", 9, "bold"), foreground="#ef5350")
        self.lbl_canceladas_asign.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col3, text="Crédito:", foreground="#ffb74d").grid(row=2, column=0, sticky=tk.W)
        self.lbl_credito_asign = ttk.Label(col3, text="0", font=("Segoe UI", 9, "bold"), foreground="#ffb74d")
        self.lbl_credito_asign.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Columna 4: Repartidores
        col4 = ttk.Frame(resumen_content)
        col4.pack(side=tk.LEFT, padx=(0, 30))
        
        ttk.Label(col4, text="REPARTIDORES", font=("Segoe UI", 9, "bold"), 
                 foreground="#ce93d8").grid(row=0, column=0, columnspan=2, sticky=tk.W)
        ttk.Label(col4, text="Activos:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_repartidores_activos = ttk.Label(col4, text="0", font=("Segoe UI", 9, "bold"), foreground="#ce93d8")
        self.lbl_repartidores_activos.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col4, text="% Completado:").grid(row=2, column=0, sticky=tk.W)
        self.lbl_porcentaje_asign = ttk.Label(col4, text="0%", font=("Segoe UI", 9, "bold"), foreground="#64b5f6")
        self.lbl_porcentaje_asign.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Columna 5: Filtro activo
        col5 = ttk.Frame(resumen_content)
        col5.pack(side=tk.LEFT, padx=(0, 0))
        
        ttk.Label(col5, text="FILTRO", font=("Segoe UI", 9, "bold"), 
                 foreground="#4db6ac").grid(row=0, column=0, columnspan=2, sticky=tk.W)
        ttk.Label(col5, text="Mostrando:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_filtro_activo = ttk.Label(col5, text="Todos", font=("Segoe UI", 9, "bold"), foreground="#4db6ac")
        self.lbl_filtro_activo.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        ttk.Label(col5, text="Facturas:").grid(row=2, column=0, sticky=tk.W)
        self.lbl_facturas_filtradas = ttk.Label(col5, text="0", font=("Segoe UI", 9, "bold"), foreground="#4db6ac")
        self.lbl_facturas_filtradas.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))

        # ============================================================
        # RESUMEN CORTE CAJERO ELEVENTA - Compacto y atractivo
        # ============================================================
        frame_corte_asign = ttk.LabelFrame(self.tab_asignacion, text="💰 CORTE CAJERO ELEVENTA", padding=(12, 6))
        frame_corte_asign.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 4))
        
        # Contenedor horizontal para los indicadores
        corte_content = ttk.Frame(frame_corte_asign)
        corte_content.pack(fill=tk.X, expand=True)
        
        # Indicador 1: Dinero en Caja (con icono grande)
        ind1 = ttk.Frame(corte_content)
        ind1.pack(side=tk.LEFT, padx=(0, 40))
        ttk.Label(ind1, text="🏦", font=("Segoe UI", 18)).pack(side=tk.LEFT, padx=(0, 8))
        ind1_text = ttk.Frame(ind1)
        ind1_text.pack(side=tk.LEFT)
        ttk.Label(ind1_text, text="DINERO EN CAJA", font=("Segoe UI", 8), 
                 foreground="#90caf9").pack(anchor=tk.W)
        self.lbl_corte_asign_dinero = ttk.Label(ind1_text, text="$0.00", 
                 font=("Segoe UI", 14, "bold"), foreground="#4fc3f7")
        self.lbl_corte_asign_dinero.pack(anchor=tk.W)
        
        # Indicador 2: Total Ventas
        ind2 = ttk.Frame(corte_content)
        ind2.pack(side=tk.LEFT, padx=(0, 40))
        ttk.Label(ind2, text="📈", font=("Segoe UI", 18)).pack(side=tk.LEFT, padx=(0, 8))
        ind2_text = ttk.Frame(ind2)
        ind2_text.pack(side=tk.LEFT)
        ttk.Label(ind2_text, text="TOTAL VENTAS", font=("Segoe UI", 8), 
                 foreground="#a5d6a7").pack(anchor=tk.W)
        self.lbl_corte_asign_ventas = ttk.Label(ind2_text, text="$0.00", 
                 font=("Segoe UI", 14, "bold"), foreground="#81c784")
        self.lbl_corte_asign_ventas.pack(anchor=tk.W)
        
        # Indicador 3: Ganancia
        ind3 = ttk.Frame(corte_content)
        ind3.pack(side=tk.LEFT, padx=(0, 40))
        ttk.Label(ind3, text="💎", font=("Segoe UI", 18)).pack(side=tk.LEFT, padx=(0, 8))
        ind3_text = ttk.Frame(ind3)
        ind3_text.pack(side=tk.LEFT)
        ttk.Label(ind3_text, text="GANANCIA", font=("Segoe UI", 8), 
                 foreground="#ce93d8").pack(anchor=tk.W)
        self.lbl_corte_asign_ganancia = ttk.Label(ind3_text, text="$0.00", 
                 font=("Segoe UI", 14, "bold"), foreground="#ba68c8")
        self.lbl_corte_asign_ganancia.pack(anchor=tk.W)
        
        # Indicador 4: Devoluciones
        ind4 = ttk.Frame(corte_content)
        ind4.pack(side=tk.LEFT, padx=(0, 40))
        ttk.Label(ind4, text="🔄", font=("Segoe UI", 18)).pack(side=tk.LEFT, padx=(0, 8))
        ind4_text = ttk.Frame(ind4)
        ind4_text.pack(side=tk.LEFT)
        ttk.Label(ind4_text, text="DEVOLUCIONES", font=("Segoe UI", 8), 
                 foreground="#ef9a9a").pack(anchor=tk.W)
        self.lbl_corte_asign_devs = ttk.Label(ind4_text, text="$0.00", 
                 font=("Segoe UI", 14, "bold"), foreground="#e57373")
        self.lbl_corte_asign_devs.pack(anchor=tk.W)
        
        # Indicador 5: Turno actual
        ind5 = ttk.Frame(corte_content)
        ind5.pack(side=tk.LEFT)
        ttk.Label(ind5, text="#", font=("Segoe UI", 22, "bold"), foreground="#fdd835").pack(side=tk.LEFT, padx=(0, 8))
        ind5_text = ttk.Frame(ind5)
        ind5_text.pack(side=tk.LEFT)
        ttk.Label(ind5_text, text="TURNO", font=("Segoe UI", 8), 
                 foreground="#fff59d").pack(anchor=tk.W)
        self.lbl_corte_asign_turno = ttk.Label(ind5_text, text="#---", 
                 font=("Segoe UI", 14, "bold"), foreground="#fdd835")
        self.lbl_corte_asign_turno.pack(anchor=tk.W)
        
        # Botón GUARDAR y indicador de cambios (lado derecho, centrado verticalmente)
        frame_guardar = ttk.Frame(corte_content)
        frame_guardar.pack(side=tk.RIGHT, padx=(20, 0), fill=tk.Y)
        
        # Frame interno para centrar verticalmente
        frame_guardar_inner = ttk.Frame(frame_guardar)
        frame_guardar_inner.pack(expand=True)
        
        self.btn_guardar_asign = ttk.Button(
            frame_guardar_inner, 
            text="💾 GUARDAR",
            command=self._guardar_cambios_repartidores,
            state="disabled",
            style="Success.TButton"
        )
        self.btn_guardar_asign.pack()
        
        # Label indicador de cambios pendientes
        self.lbl_cambios_pendientes = ttk.Label(
            frame_guardar_inner, 
            text="",
            foreground="#ff9800",
            font=("Segoe UI", 8)
        )
        self.lbl_cambios_pendientes.pack(pady=(2, 0))

        # ============================================================
        # TABLA DE FACTURAS - Ocupa el espacio restante
        # ============================================================
        frame_tree = ttk.LabelFrame(self.tab_asignacion,
                                    text="📋 FACTURAS DEL DÍA",
                                    padding=(5, 5))
        frame_tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        # Contenedor para Treeview con scrollbars
        tree_container = ttk.Frame(frame_tree)
        tree_container.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        self.tree_asign = ttk.Treeview(
            tree_container,
            columns=("folio", "cliente", "subtotal", "total", "repartidor", "usuario", "estado", "f_venta", "f_cancel"),
            selectmode="extended",  # Permite selección múltiple
            height=20
        )
        self.tree_asign.column("#0", width=0, stretch=tk.NO)
        self.tree_asign.column("folio",      anchor=tk.CENTER, width=70,  minwidth=60)
        self.tree_asign.column("cliente",    anchor=tk.W,      width=220, minwidth=150)
        self.tree_asign.column("subtotal",   anchor=tk.E,      width=90, minwidth=70)
        self.tree_asign.column("total",      anchor=tk.E,      width=90, minwidth=70)
        self.tree_asign.column("repartidor", anchor=tk.CENTER, width=110, minwidth=90)
        self.tree_asign.column("usuario",    anchor=tk.CENTER, width=100, minwidth=80)
        self.tree_asign.column("estado",     anchor=tk.CENTER, width=90, minwidth=70)
        self.tree_asign.column("f_venta",    anchor=tk.CENTER, width=90,  minwidth=80)
        self.tree_asign.column("f_cancel",   anchor=tk.CENTER, width=90,  minwidth=80)

        self.tree_asign.heading("folio",      text="📋 Folio")
        self.tree_asign.heading("cliente",    text="👤 Cliente")
        self.tree_asign.heading("subtotal",   text="💵 Subtotal")
        self.tree_asign.heading("total",      text="💰 Total")
        self.tree_asign.heading("repartidor", text="🚚 Repartidor ✎")
        self.tree_asign.heading("usuario",    text="👨‍💼 Usuario")
        self.tree_asign.heading("estado",     text="📊 Estado")
        self.tree_asign.heading("f_venta",    text="📅 F.Venta")
        self.tree_asign.heading("f_cancel",   text="❌ F.Cancel")

        # Tags con colores para modo oscuro
        # Colores por repartidor específico (tonos oscuros suaves)
        self.tree_asign.tag_configure("rep_cristian",  background="#37474f", foreground="#80cbc4", font=("Segoe UI", 9))  # Gris azulado + cyan
        self.tree_asign.tag_configure("rep_cajero",    background="#33691e", foreground="#c5e1a5", font=("Segoe UI", 9))  # Verde oliva
        self.tree_asign.tag_configure("rep_david",     background="#4a148c", foreground="#ce93d8", font=("Segoe UI", 9))  # Morado profundo
        self.tree_asign.tag_configure("rep_otro",      background="#004d40", foreground="#80cbc4", font=("Segoe UI", 9))  # Teal oscuro
        
        # Estados especiales (tienen prioridad)
        self.tree_asign.tag_configure("sin_asignar",  background="#4a4a4a", foreground="#ff8a80", font=("Segoe UI", 9))
        self.tree_asign.tag_configure("cancelada",    background="#b71c1c", foreground="#ffffff", font=("Segoe UI", 9))
        self.tree_asign.tag_configure("cancelada_otro_dia", background="#880e4f", foreground="#ffffff", font=("Segoe UI", 9, "bold"))
        self.tree_asign.tag_configure("credito",      background="#e65100", foreground="#ffffff", font=("Segoe UI", 9))
        self.tree_asign.tag_configure("pendiente",    background="#ffc107", foreground="#000000", font=("Segoe UI", 9, "italic"))

        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.tree_asign.yview)
        # Scrollbar horizontal
        scroll_x = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.tree_asign.xview)
        
        self.tree_asign.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        # Grid layout para scrollbars
        self.tree_asign.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Binding para copiar con Ctrl+C
        self.tree_asign.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(self.tree_asign))
        self.tree_asign.bind("<Control-C>", lambda e: self._copiar_seleccion_tree(self.tree_asign))

        # Bindings
        self.tree_asign.bind("<Double-1>",  self._on_tree_double_click)
        self.tree_asign.bind("<Button-1>",  self._on_tree_single_click)  # Clic simple selecciona
        self.tree_asign.bind("<Button-3>",  self._on_tree_right_click_asign)
        # Enter abre editor en la fila seleccionada
        self.tree_asign.bind("<Return>",    self._abrir_editor_fila_seleccionada)
        # Flechas navegan normalmente (cierra editor si está abierto)
        self.tree_asign.bind("<Up>",        self._on_tree_navigate)
        self.tree_asign.bind("<Down>",      self._on_tree_navigate)
        # Cierra editor si el usuario hace scroll
        self.tree_asign.bind("<MouseWheel>", self._cerrar_editor)
        self.tree_asign.bind("<Escape>",    self._cerrar_editor)

    # --- cargar facturas desde BD ---
    def _cargar_facturas(self):
        fecha = self.fecha_asign_var.get().strip()
        if not fecha:
            messagebox.showwarning("Fecha", "Ingresa una fecha válida (YYYY-MM-DD)")
            return

        self.ds.fecha = fecha  # sincronizar fecha global
        
        # Limpiar cambios pendientes al cargar nuevas facturas
        if hasattr(self, '_cambios_pendientes'):
            self._cambios_pendientes.clear()
            self._actualizar_estado_boton_guardar()

        # Consulta principal usando VENTATICKETS con campo TOTAL para coincidir con corte de caja
        # Incluye TURNO_ID para identificar el turno de cada venta
        sql = (
            "SET HEADING ON;\n"
            "SELECT V.ID, V.FOLIO, V.NOMBRE, V.SUBTOTAL, V.TOTAL, V.ESTA_CANCELADO, V.TOTAL_CREDITO, "
            "CAST(V.CREADO_EN AS DATE) AS FECHA_CREACION, "
            "CAST(D.DEVUELTO_EN AS DATE) AS FECHA_CANCELACION, "
            "V.TURNO_ID\n"
            "FROM VENTATICKETS V\n"
            "LEFT JOIN DEVOLUCIONES D ON D.TICKET_ID = V.ID AND D.TIPO_DEVOLUCION = 'C'\n"
            f"WHERE CAST(V.CREADO_EN AS DATE) = '{fecha}'\n"
            "ORDER BY V.FOLIO;\n"
        )
        ok, stdout, stderr = self._ejecutar_sql(sql)

        if not ok or not stdout:
            error_msg = stderr or "No se recibieron datos de la BD"
            messagebox.showerror("Error BD", f"No se pudo consultar:\n{error_msg}")
            return

        ventas = []
        header_visto = False
        try:
            for linea in stdout.split('\n'):
                linea = linea.strip()
                if not linea or linea.startswith('='):
                    continue
                # Detectar header de VENTATICKETS
                if 'ID' in linea and 'FOLIO' in linea:
                    header_visto = True
                    continue
                if not header_visto:
                    continue
                partes = linea.split()
                if len(partes) < 9:  # Campos con TURNO_ID
                    continue
                try:
                    # Formato: ID, FOLIO, NOMBRE..., SUBTOTAL, TOTAL, ESTA_CANCELADO, TOTAL_CREDITO, FECHA_CREACION, FECHA_CANCELACION, TURNO_ID
                    id_v = int(partes[0])
                    folio_s = partes[1]
                    if folio_s == '<null>':
                        continue
                    folio = int(folio_s)
                    
                    # Obtener campos desde el final (más confiable)
                    # Los últimos campos son: ..., FECHA_CANCELACION, TURNO_ID
                    turno_id_venta = partes[-1] if partes[-1] != '<null>' else ''
                    fecha_cancelacion = partes[-2] if partes[-2] != '<null>' else ''
                    fecha_creacion = partes[-3] if partes[-3] != '<null>' else fecha
                    total_credito = float(partes[-4]) if partes[-4] != '<null>' else 0.0
                    # ESTA_CANCELADO puede ser 't'/'f' o 1/0
                    cancelado_val = partes[-5].lower()
                    esta_cancelado = cancelado_val == 't' or cancelado_val == '1'
                    total_original = float(partes[-6]) if partes[-6] != '<null>' else 0.0
                    subtotal = float(partes[-7]) if partes[-7] != '<null>' else 0.0
                    # El nombre está entre FOLIO (índice 1) y SUBTOTAL (índice -7)
                    nombre = ' '.join(partes[2:-7]).replace('<null>', '').strip()
                    if not nombre:
                        nombre = 'MOSTRADOR'
                    
                    # Usar turno_id como identificador de usuario (Turno X)
                    usuario = f"Turno {turno_id_venta}" if turno_id_venta else ''
                    
                    if folio <= 0:
                        continue

                    # Obtener repartidor asignado
                    rep = obtener_repartidor_factura(folio, fecha) or ''
                    
                    # Si el nombre es "Ticket X", "MOSTRADOR" o similar, asignar a CAJERO
                    nombre_lower = nombre.lower()
                    if not rep and (nombre_lower.startswith('ticket ') or nombre_lower == 'ticket' or nombre_lower == 'mostrador'):
                        rep = 'CAJERO'
                        # Guardar automáticamente esta asignación
                        asignar_repartidor(folio, fecha, 'CAJERO')
                    
                    es_credito = total_credito > 0
                    
                    # Para facturas canceladas del MISMO DÍA: subtotal = total
                    # Esto hace que sumen al total de facturas
                    subtotal_final = subtotal
                    if esta_cancelado:
                        subtotal_final = total_original  # Subtotal = Total para canceladas del día
                    
                    # Si la factura está cancelada y es del mismo día = cancelada normal
                    # (Las canceladas de otro día vendrán de la segunda consulta)
                    ventas.append({
                        'id': id_v,
                        'folio': folio,
                        'nombre': nombre, 
                        'subtotal': subtotal_final,  # Usar subtotal ajustado
                        'total_original': total_original,
                        'repartidor': rep, 
                        'cancelada': esta_cancelado,
                        'cancelada_otro_dia': False,  # Las del mismo día no son de otro día
                        'total_credito': total_credito,
                        'es_credito': es_credito,
                        'fecha_creacion': fecha_creacion,
                        'fecha_cancelacion': fecha_cancelacion,
                        'turno_id': turno_id_venta,
                        'usuario': usuario
                    })
                except (ValueError, IndexError):
                    continue

            # --- SEGUNDA CONSULTA: Facturas canceladas de otros días ---
            # Buscar facturas canceladas cuya fecha de creación NO es la fecha consultada
            # pero que el repartidor reporta como canceladas ese día (usando asignaciones previas)
            canceladas_otro_dia = self._cargar_canceladas_otro_dia(fecha)
            if canceladas_otro_dia:
                ventas.extend(canceladas_otro_dia)

            self.ds.set_ventas(ventas)
            
            # Cargar devoluciones del día
            self._cargar_devoluciones(fecha)
            
            # Asignar cajero que canceló como repartidor en las canceladas
            self._asignar_cajero_cancelaciones()
            
            # Cargar devoluciones parciales (artículos devueltos sin cancelar factura)
            self._cargar_devoluciones_parciales(fecha)
            
            # Cargar movimientos (ingresos y salidas) del día
            self._cargar_movimientos(fecha)

            if ventas:
                # Usar total_original para coincidir con Firebird
                total_facturas = sum(v.get('total_original', v['subtotal']) for v in ventas if not v['cancelada'])
                total_canceladas = sum(v['total_original'] for v in ventas if v['cancelada'] and not v.get('cancelada_otro_dia', False))
                total_canceladas_otro_dia = sum(v['total_original'] for v in ventas if v.get('cancelada_otro_dia', False))
                total_credito = sum(v['total_credito'] for v in ventas if v['es_credito'])
                
                msg = f"Se cargaron {len(ventas)} facturas.\n"
                msg += f"Total Facturas: ${total_facturas:,.2f}\n"
                msg += f"Total Canceladas: ${total_canceladas:,.2f}\n"
                if total_canceladas_otro_dia > 0:
                    msg += f"Total Canceladas (otro día): ${total_canceladas_otro_dia:,.2f}\n"
                msg += f"Total a Crédito: ${total_credito:,.2f}"
                
                messagebox.showinfo("Carga exitosa", msg)
            else:
                messagebox.showwarning("Sin datos", f"No hay ventas para {fecha}.")
        except Exception as e:
            messagebox.showerror("Error", f"Error procesando facturas:\n{str(e)}")

    def _cargar_canceladas_otro_dia(self, fecha: str) -> list:
        """
        Carga facturas de días anteriores que fueron CANCELADAS el día consultado.
        Estas son solo informativas y NO suman al total de facturas.
        El movimiento de cancelación se hizo HOY pero la factura es de otro día.
        """
        # Buscar facturas que fueron creadas ANTES de hoy pero CANCELADAS hoy
        # Usando VENTATICKETS con DEVOLUCIONES
        sql = (
            "SET HEADING ON;\n"
            "SELECT V.ID, V.FOLIO, V.NOMBRE, V.SUBTOTAL, V.TOTAL, V.ESTA_CANCELADO, V.TOTAL_CREDITO, "
            "CAST(V.CREADO_EN AS DATE) AS FECHA_CREACION, "
            "CAST(D.DEVUELTO_EN AS DATE) AS FECHA_CANCELACION\n"
            "FROM VENTATICKETS V\n"
            "INNER JOIN DEVOLUCIONES D ON D.TICKET_ID = V.ID AND D.TIPO_DEVOLUCION = 'C'\n"
            f"WHERE V.ESTA_CANCELADO = 't'\n"
            f"AND CAST(V.CREADO_EN AS DATE) < '{fecha}'\n"
            f"AND CAST(D.DEVUELTO_EN AS DATE) = '{fecha}'\n"  # Canceladas HOY
            "ORDER BY V.FOLIO;\n"
        )
        ok, stdout, stderr = self._ejecutar_sql(sql)
        
        canceladas = []
        if ok and stdout:
            header_visto = False
            for linea in stdout.split('\n'):
                linea = linea.strip()
                if not linea or linea.startswith('='):
                    continue
                if 'ID' in linea and 'FOLIO' in linea:
                    header_visto = True
                    continue
                if not header_visto:
                    continue
                partes = linea.split()
                if len(partes) < 8:
                    continue
                try:
                    id_v = int(partes[0])
                    folio_s = partes[1]
                    if folio_s == '<null>':
                        continue
                    folio = int(folio_s)
                    
                    # Parsear campos (de derecha a izquierda)
                    fecha_cancelacion = partes[-1] if partes[-1] != '<null>' else fecha
                    fecha_creacion = partes[-2] if partes[-2] != '<null>' else ''
                    total_credito = float(partes[-3]) if partes[-3] != '<null>' else 0.0
                    total_original = float(partes[-5]) if partes[-5] != '<null>' else 0.0
                    subtotal = float(partes[-6]) if partes[-6] != '<null>' else 0.0
                    nombre = ' '.join(partes[2:-6]).replace('<null>', '').strip()
                    if not nombre:
                        nombre = 'MOSTRADOR'
                    
                    if folio <= 0:
                        continue

                    # Para canceladas de otro día, buscar si tiene repartidor asignado
                    rep = obtener_repartidor_factura(folio, fecha_creacion) or ''
                    
                    canceladas.append({
                        'id': id_v, 
                        'folio': folio,
                        'nombre': f"⚠️ {nombre}", 
                        'subtotal': 0,  # NO suma al total (informativa)
                        'total_original': total_original,
                        'repartidor': rep, 
                        'cancelada': True,
                        'cancelada_otro_dia': True,  # Indica que es de otro día
                        'total_credito': total_credito,
                        'es_credito': total_credito > 0,
                        'fecha_creacion': fecha_creacion,
                        'fecha_cancelacion': fecha_cancelacion
                    })
                except (ValueError, IndexError):
                    continue
        
        return canceladas

    def _cargar_devoluciones(self, fecha: str):
        """Carga las devoluciones del día desde la BD."""
        sql = (
            "SET HEADING ON;\n"
            "SELECT ID, TICKET_ID, TOTAL_DEVUELTO, CAJERO, TIPO_DEVOLUCION\n"
            "FROM DEVOLUCIONES\n"
            f"WHERE CAST(DEVUELTO_EN AS DATE) = '{fecha}'\n"
            "ORDER BY ID;\n"
        )
        ok, stdout, stderr = self._ejecutar_sql(sql)
        
        devoluciones = []
        if ok and stdout:
            header_visto = False
            for linea in stdout.split('\n'):
                linea = linea.strip()
                if not linea or linea.startswith('='):
                    continue
                if 'ID' in linea and 'TICKET_ID' in linea:
                    header_visto = True
                    continue
                if not header_visto:
                    continue
                partes = linea.split()
                if len(partes) >= 5:
                    try:
                        dev_id = int(partes[0])
                        ticket_id = int(partes[1]) if partes[1] != '<null>' else 0
                        monto = float(partes[2]) if partes[2] != '<null>' else 0.0
                        cajero = partes[3] if partes[3] != '<null>' else ''
                        tipo = partes[4] if partes[4] != '<null>' else ''
                        
                        devoluciones.append({
                            'id': dev_id,
                            'ticket_id': ticket_id,
                            'monto': monto,
                            'cajero': cajero,
                            'tipo': tipo
                        })
                    except (ValueError, IndexError):
                        continue
        
        self.ds.set_devoluciones(devoluciones)

    def _asignar_cajero_cancelaciones(self):
        """
        Asigna el cajero que canceló como 'repartidor' en las facturas canceladas.
        Usa las devoluciones cargadas (tipo='C') para encontrar quién canceló cada factura.
        También guarda el detalle en SQLite para persistencia.
        Calcula y guarda totales de cancelaciones en efectivo por CAJERO y ADMIN.
        """
        fecha = self.ds.fecha
        
        # Crear mapa de ticket_id -> {cajero, monto} para devoluciones completas (tipo C)
        cancelaciones_por_ticket = {}
        for dev in self.ds.devoluciones:
            if dev.get('tipo') == 'C':  # Solo cancelaciones completas
                ticket_id = dev.get('ticket_id', 0)
                cajero = dev.get('cajero', '')
                monto = dev.get('monto', 0)
                if ticket_id and cajero:
                    cancelaciones_por_ticket[ticket_id] = {
                        'cajero': cajero.upper(),
                        'monto': monto
                    }
        
        # Lista para guardar en SQLite
        cancelaciones_para_sqlite = []
        
        # Totales por cajero (solo cancelaciones en efectivo del día actual)
        totales_efectivo_por_cajero = {}  # {'CAJERO': total, 'ADMIN': total}
        
        # Asignar el cajero como repartidor en las facturas canceladas
        for venta in self.ds.ventas:
            if venta.get('cancelada', False):
                ticket_id = venta.get('id', 0)
                folio = venta.get('folio', 0)
                info_canc = cancelaciones_por_ticket.get(ticket_id, {})
                cajero = info_canc.get('cajero', '')
                monto = info_canc.get('monto', 0)
                
                # Verificar si es cancelada del día (no de otro día)
                es_del_dia = not venta.get('cancelada_otro_dia', False)
                # Verificar si es en efectivo (no es crédito)
                es_efectivo = venta.get('total_credito', 0) == 0
                
                if cajero:
                    venta['repartidor'] = cajero  # ADMIN o CAJERO
                    
                    # Preparar para guardar en SQLite
                    cancelaciones_para_sqlite.append({
                        'folio': folio,
                        'ticket_id': ticket_id,
                        'cajero_cancelo': cajero,
                        'monto': monto,
                        'fecha_cancelacion': venta.get('fecha_cancelacion'),
                        'es_efectivo': es_efectivo,
                        'es_del_dia': es_del_dia
                    })
                    
                    # Sumar al total por cajero (solo efectivo del día)
                    if es_del_dia and es_efectivo:
                        total_factura = venta.get('total_original', 0)
                        totales_efectivo_por_cajero[cajero] = totales_efectivo_por_cajero.get(cajero, 0) + total_factura
        
        # Guardar en SQLite si hay cancelaciones
        if USE_SQLITE:
            if cancelaciones_para_sqlite:
                db_local.guardar_cancelaciones_detalle_lote(fecha, cancelaciones_para_sqlite)
                print(f"✅ Guardadas {len(cancelaciones_para_sqlite)} cancelaciones con detalle de cajero")
            
            # Guardar totales de cancelaciones en efectivo por cajero
            if totales_efectivo_por_cajero:
                db_local.guardar_totales_cancelaciones_efectivo(fecha, totales_efectivo_por_cajero)
                for cajero, total in totales_efectivo_por_cajero.items():
                    print(f"   💰 {cajero}: ${total:,.2f} en cancelaciones efectivo")

    def _cargar_devoluciones_parciales(self, fecha: str):
        """Carga las devoluciones parciales de artículos desde Firebird y las guarda en SQLite.
        
        Extrae: código, descripción, cantidad devuelta, precio de venta y total devuelto.
        El Precio de Venta = DINERO_DEVUELTO / CANTIDAD_DEVUELTA (en centavos).
        Las devoluciones se asocian a la FECHA DE LA VENTA original.
        """
        if not USE_SQLITE:
            return
        
        # Limpiar devoluciones previas de esta fecha
        db_local.limpiar_devoluciones_parciales_fecha(fecha)
        
        # Consultar devoluciones parciales (TIPO_DEVOLUCION = 'P')
        # El Precio de Venta = DINERO_DEVUELTO / CANTIDAD_DEVUELTA
        sql = f"""
SELECT 
    DA.DEVOLUCION_ID,
    V.FOLIO,
    DA.CODIGO_PRODUCTO,
    DA.DESCRIPCION_PRODUCTO,
    DA.CANTIDAD_DEVUELTA,
    DA.DINERO_DEVUELTO,
    D.DEVUELTO_EN
FROM DEVOLUCIONES_ARTICULOS DA
INNER JOIN DEVOLUCIONES D ON DA.DEVOLUCION_ID = D.ID
INNER JOIN VENTATICKETS V ON DA.TICKET_ID = V.ID
WHERE CAST(V.VENDIDO_EN AS DATE) = '{fecha}'
AND D.TIPO_DEVOLUCION = 'P'
ORDER BY V.FOLIO, DA.ID;
"""
        ok, stdout, stderr = self._ejecutar_sql(sql)
        
        if not ok or not stdout:
            print(f"⚠️ No se pudieron cargar devoluciones parciales: {stderr}")
            return
        
        # Parsear resultado
        lineas = stdout.strip().split('\n')
        datos_inicio = False
        
        for linea in lineas:
            linea = linea.strip()
            
            # Saltar líneas vacías y separadores
            if not linea or linea.startswith('=') or linea.startswith('-'):
                continue
            
            # Detectar inicio de datos (después de la línea de encabezados)
            if 'DEVOLUCION_ID' in linea or 'FOLIO' in linea:
                datos_inicio = True
                continue
            
            if not datos_inicio:
                continue
            
            # Parsear la línea de datos
            # Formato: DEVOLUCION_ID FOLIO CODIGO DESCRIPCION... CANTIDAD DINERO FECHA
            partes = linea.split()
            if len(partes) < 6:
                continue
            
            try:
                devolucion_id = int(partes[0])
                folio = int(partes[1])
                codigo = partes[2] if partes[2] else ""
                
                # Buscar la fecha al final (formato YYYY-MM-DD HH:MM:SS o similar)
                fecha_dev = None
                idx_fecha = -1
                for i in range(len(partes) - 1, 2, -1):
                    if '-' in partes[i] and len(partes[i]) >= 10:
                        fecha_dev = partes[i][:10]  # Tomar solo la fecha
                        idx_fecha = i
                        break
                
                if idx_fecha == -1:
                    continue
                
                # El dinero devuelto está antes de la fecha (en PESOS COLOMBIANOS)
                dinero_raw = partes[idx_fecha - 1]
                dinero = float(dinero_raw.replace(',', ''))
                
                # La cantidad está antes del dinero
                cantidad_raw = partes[idx_fecha - 2]
                cantidad = float(cantidad_raw.replace(',', ''))
                
                # La descripción es todo entre el código y la cantidad
                descripcion = ' '.join(partes[3:idx_fecha - 2])
                
                # PRECIO DE VENTA = DINERO_DEVUELTO / CANTIDAD (ya en pesos colombianos)
                # Ej: 48000 pesos / 10 unidades = $4,800 pesos por unidad
                precio_venta = dinero / cantidad if cantidad > 0 else 0
                
                # Guardar en SQLite
                db_local.guardar_devolucion_parcial(
                    fecha=fecha,
                    folio=folio,
                    devolucion_id=devolucion_id,
                    codigo=codigo,
                    descripcion=descripcion.strip(),
                    cantidad=cantidad,
                    valor_unitario=precio_venta,
                    dinero=dinero,
                    fecha_devolucion=fecha_dev
                )
                print(f"✅ Dev: Folio {folio}, {descripcion.strip()[:25]}, Cant: {int(cantidad)}, Precio: ${precio_venta:,.0f}, Total: ${dinero:,.0f}")
                
            except (ValueError, IndexError) as e:
                print(f"⚠️ Error parseando línea: {linea} - {e}")
                continue

    def _cargar_movimientos(self, fecha: str):
        """Carga los movimientos (ingresos y salidas) del día desde la BD."""
        sql = (
            "SET HEADING ON;\n"
            "SELECT ID, TIPO, MONTO, COMENTARIOS\n"
            "FROM MOVIMIENTOS\n"
            f"WHERE CAST(CUANDO_FUE AS DATE) = '{fecha}'\n"
            "ORDER BY ID;\n"
        )
        ok, stdout, stderr = self._ejecutar_sql(sql)
        
        entradas = []
        salidas = []
        if ok and stdout:
            header_visto = False
            for linea in stdout.split('\n'):
                linea = linea.strip()
                if not linea or linea.startswith('='):
                    continue
                if 'ID' in linea and 'TIPO' in linea:
                    header_visto = True
                    continue
                if not header_visto:
                    continue
                partes = linea.split()
                if len(partes) >= 3:
                    try:
                        mov_id = int(partes[0])
                        tipo = partes[1].upper()
                        monto = float(partes[2]) if partes[2] != '<null>' else 0.0
                        comentario = ' '.join(partes[3:]).replace('<null>', '').strip() if len(partes) > 3 else ''
                        
                        mov = {
                            'id': mov_id,
                            'tipo': tipo,
                            'monto': monto,
                            'comentario': comentario
                        }
                        
                        if tipo == 'E':  # Entrada/Ingreso
                            entradas.append(mov)
                        elif tipo == 'S':  # Salida
                            salidas.append(mov)
                    except (ValueError, IndexError):
                        continue
        
        self.ds.set_movimientos(entradas, salidas)

    def _get_repartidor_tag(self, repartidor):
        """Devuelve el tag de color según el nombre del repartidor."""
        if not repartidor:
            return "sin_asignar"
        
        nombre = repartidor.upper().strip()
        if "CRISTIAN" in nombre:
            return "rep_cristian"
        elif "CAJERO" in nombre:
            return "rep_cajero"
        elif "DAVID" in nombre:
            return "rep_david"
        else:
            return "rep_otro"

    # --- refrescar tree de asignación (desde DataStore) ---
    def _refrescar_tree_asignacion(self):
        self._cerrar_editor()
        self.tree_asign.delete(*self.tree_asign.get_children())

        for v in self.ds.get_ventas():
            folio = v['folio']
            cancelada = v.get('cancelada', False)
            es_credito = v.get('es_credito', False)
            cancelada_otro_dia = v.get('cancelada_otro_dia', False)
            
            # Usar repartidor de cambios pendientes si existe, sino del DataStore
            repartidor = self._cambios_pendientes.get(folio, v['repartidor'])
            tiene_cambio_pendiente = folio in self._cambios_pendientes
            
            if cancelada_otro_dia:
                tag = "cancelada_otro_dia"
            elif cancelada:
                tag = "cancelada"
            elif es_credito:
                tag = "credito"
            elif tiene_cambio_pendiente:
                tag = "pendiente"  # Color especial para cambios sin guardar
            else:
                tag = self._get_repartidor_tag(repartidor)
            
            # Obtener subtotal y total
            subtotal = v.get('subtotal', 0)
            total = v.get('total_original', subtotal)
            
            # Determinar estado
            if cancelada_otro_dia:
                estado = "CANC. OTRO DÍA"
            elif cancelada:
                estado = "CANCELADA"
            elif es_credito:
                estado = "CRÉDITO"
            else:
                estado = "—"
            
            # Obtener fechas de venta y cancelación
            fecha_venta = v.get('fecha_creacion', '')
            fecha_cancel = v.get('fecha_cancelacion', '')
            
            # Obtener usuario de Eleventa
            usuario = v.get('usuario', '')
            
            # Mostrar indicador de cambio pendiente
            repartidor_display = repartidor or '— Sin asignar'
            if tiene_cambio_pendiente:
                repartidor_display = f"✏️ {repartidor}" if repartidor else '— Sin asignar'
            
            # Para canceladas, mostrar repartidor + estado
            if cancelada or cancelada_otro_dia:
                if repartidor:
                    repartidor_display = f"{repartidor}, {estado}"
                else:
                    repartidor_display = estado
            
            self.tree_asign.insert("", tk.END, iid=str(folio),
                                   values=(folio, v['nombre'],
                                           f"${subtotal:,.2f}",
                                           f"${total:,.2f}",
                                           repartidor_display,
                                           usuario,
                                           estado,
                                           fecha_venta,
                                           fecha_cancel),
                                   tags=(tag,))

        # resumen
        total = len(self.ds.ventas)
        asign = sum(1 for v in self.ds.ventas if v['repartidor'] and not v.get('cancelada', False) and not v.get('cancelada_otro_dia', False))
        canceladas = sum(1 for v in self.ds.ventas if v.get('cancelada', False) and not v.get('cancelada_otro_dia', False))
        canceladas_otro_dia = sum(1 for v in self.ds.ventas if v.get('cancelada_otro_dia', False))
        credito = sum(1 for v in self.ds.ventas if v.get('es_credito', False))
        # Facturas sin asignar = total - asignadas - canceladas (las canceladas no cuentan como sin asignar)
        total_no_canceladas = total - canceladas - canceladas_otro_dia
        sin_asignar = total_no_canceladas - asign
        
        # Montos de cada categoría
        total_canceladas = self.ds.get_total_canceladas()  # Canceladas del mismo día
        total_canceladas_otro_dia_val = self.ds.get_total_canceladas_otro_dia()  # Canceladas de otro día
        total_credito = self.ds.get_total_credito()
        
        # Devoluciones parciales
        dev_parciales = 0
        dev_parciales_otro_dia = 0
        self._facturas_dev_parciales_otro_dia = []  # Para el modal
        self._total_dev_parciales_otro_dia = 0  # Para usar en corte cajero
        if USE_SQLITE and self.ds.fecha:
            dev_parciales = db_local.obtener_total_devoluciones_parciales_fecha(self.ds.fecha)
            # Obtener dev. parciales de facturas de OTROS días procesadas HOY
            dev_parciales_otro_dia, self._facturas_dev_parciales_otro_dia = db_local.obtener_dev_parciales_otro_dia(self.ds.fecha)
            self._total_dev_parciales_otro_dia = dev_parciales_otro_dia
        
        # --- MONTO FACTURAS: Solo facturas efectivo NO canceladas del día ---
        # NO se restan las canceladas de otro día ni dev. parciales de otro día
        # porque eso afecta el CUADRE DEL DÍA DE LA VENTA, no el día en que se procesa.
        #
        # El comportamiento correcto es:
        # - DÍA DE LA VENTA: La factura cancelada ya NO suma (es excluida por cancelada=True)
        # - DÍA DE LA CANCELACIÓN: Solo mostrar info de canceladas/dev otro día (sin restar)
        monto_facturas_efectivo = self.ds.get_monto_facturas_efectivo()
        monto_efectivo = monto_facturas_efectivo  # NO se restan canceladas/dev de otro día
        
        # Monto Total = Facturas del día + Canceladas + Crédito (para referencia)
        total_monto = monto_facturas_efectivo + total_canceladas + total_credito
        
        reps = self.ds.get_repartidores()
        
        # Actualizar combo de filtros (sin repartidores, solo estados)
        filtros_base = ["Todos", "Sin Repartidor", "Canceladas", "Crédito"]
        self.combo_filtro_estado['values'] = filtros_base
        
        # Actualizar labels del resumen
        self.lbl_total_facturas_asign.config(text=str(total))
        self.lbl_monto_efectivo_asign.config(text=f"${monto_efectivo:,.2f}")
        self.lbl_canceladas_monto_asign.config(text=f"${total_canceladas:,.2f}")
        self.lbl_dev_parciales_asign.config(text=f"${dev_parciales:,.2f}")
        self.lbl_monto_total_asign.config(text=f"${total_monto:,.2f}")
        self.lbl_asignadas.config(text=str(asign))
        self.lbl_sin_asignar.config(text=str(sin_asignar))
        self.lbl_canceladas_asign.config(text=f"{canceladas + canceladas_otro_dia}")
        self.lbl_credito_asign.config(text=str(credito))
        self.lbl_repartidores_activos.config(text=str(len(reps)))
        
        # Actualizar dev parciales otro día
        if hasattr(self, 'lbl_dev_parciales_otro_dia'):
            self.lbl_dev_parciales_otro_dia.config(text=f"${dev_parciales_otro_dia:,.2f}")
        
        # Filtro activo
        self.lbl_filtro_activo.config(text="Todos")
        self.lbl_facturas_filtradas.config(text=str(total))
        
        porcentaje = (asign / total_no_canceladas * 100) if total_no_canceladas > 0 else 0
        self.lbl_porcentaje_asign.config(text=f"{porcentaje:.1f}%")
        
        # Cambiar color del porcentaje según el valor
        if porcentaje >= 90:
            self.lbl_porcentaje_asign.config(foreground="#81c784")  # Verde
        elif porcentaje >= 50:
            self.lbl_porcentaje_asign.config(foreground="#ffb74d")  # Naranja
        else:
            self.lbl_porcentaje_asign.config(foreground="#ef5350")  # Rojo
        
        # También mantener el resumen_var para compatibilidad
        resumen_text = (
            f"Total facturas: {total}  |  Asignadas: {asign}  |  Sin asignar: {total - asign}\n"
            f"Canceladas: {canceladas} (${total_canceladas:,.2f})"
        )
        if canceladas_otro_dia > 0:
            resumen_text += f"  |  Canceladas otro día: {canceladas_otro_dia} (${total_canceladas_otro_dia_val:,.2f})"
        resumen_text += f"\nA Crédito: {credito} (${total_credito:,.2f})  |  Repartidores: {', '.join(reps) if reps else 'Ninguno'}"
        
        self.resumen_var.set(resumen_text)

    # --- Filtrar facturas en el buscador ---
    def _filtrar_facturas_asign(self):
        """Filtra las facturas según el texto de búsqueda, estado y repartidor seleccionado."""
        texto = self.buscar_asign_var.get().strip().lower()
        estado_filtro = self.filtro_estado_var.get()
        rep_filtro = self.filtro_rep_global_var.get() if hasattr(self, 'filtro_rep_global_var') else "(Todos)"
        
        self.tree_asign.delete(*self.tree_asign.get_children())
        
        # Obtener devoluciones parciales por folio para esta fecha
        dev_parciales_por_folio = {}
        creditos_punteados_folios = set()
        if USE_SQLITE and self.ds.fecha:
            dev_parciales_por_folio = db_local.obtener_devoluciones_parciales_por_folio_fecha(self.ds.fecha)
            # Obtener créditos punteados
            creditos_punteados = db_local.obtener_creditos_punteados_fecha(self.ds.fecha)
            creditos_punteados_folios = {c['folio'] for c in creditos_punteados}
        
        # Variables para resumen del filtro
        facturas_mostradas = 0
        monto_efectivo_mostrado = 0
        monto_canceladas_mostrado = 0
        dev_parciales_mostradas = 0
        asignadas_mostradas = 0
        sin_asignar_mostradas = 0
        canceladas_mostradas = 0
        credito_mostradas = 0
        
        for v in self.ds.get_ventas():
            folio = v['folio']
            # Aplicar filtro de estado
            cancelada = v.get('cancelada', False)
            es_credito = v.get('es_credito', False)
            es_credito_punteado = folio in creditos_punteados_folios
            cancelada_otro_dia = v.get('cancelada_otro_dia', False)
            
            # Usar repartidor de cambios pendientes si existe
            repartidor = self._cambios_pendientes.get(folio, v['repartidor'])
            tiene_cambio_pendiente = folio in self._cambios_pendientes
            tiene_repartidor = bool(repartidor)
            
            # Aplicar filtro de repartidor global
            if rep_filtro and rep_filtro != "(Todos)":
                if repartidor != rep_filtro:
                    continue
            
            # Aplicar filtros de estado
            if estado_filtro == "─────────":
                # Es el separador, ignorar y mostrar todos
                pass
            elif estado_filtro == "Sin Repartidor":
                if tiene_repartidor:
                    continue
            elif estado_filtro == "Canceladas":
                if not (cancelada or cancelada_otro_dia):
                    continue
            elif estado_filtro == "Crédito":
                # Incluir créditos originales Y créditos punteados
                if not (es_credito or es_credito_punteado):
                    continue
            elif estado_filtro != "Todos":
                # Es un nombre de repartidor específico (compatibilidad)
                if repartidor != estado_filtro:
                    continue
            
            # Aplicar filtro de texto (también buscar en repartidor del caché)
            if texto:
                nombre_match = texto in v['nombre'].lower()
                folio_match = texto in str(folio)
                rep_match = texto in (repartidor or '').lower()
                if not (nombre_match or folio_match or rep_match):
                    continue
            
            # Determinar tag de color
            if cancelada_otro_dia:
                tag = "cancelada_otro_dia"
            elif cancelada:
                tag = "cancelada"
            elif es_credito:
                tag = "credito"
            elif tiene_cambio_pendiente:
                tag = "pendiente"  # Color especial para cambios sin guardar
            else:
                tag = self._get_repartidor_tag(repartidor)
            
            subtotal = v.get('subtotal', 0)
            total = v.get('total_original', subtotal)
            
            if cancelada_otro_dia:
                estado = "CANC. OTRO DÍA"
            elif cancelada:
                estado = "CANCELADA"
            elif es_credito:
                estado = "CRÉDITO"
            else:
                estado = "—"
            
            fecha_venta = v.get('fecha_creacion', '')
            fecha_cancel = v.get('fecha_cancelacion', '')
            
            # Obtener usuario de Eleventa
            usuario = v.get('usuario', '')
            
            # Mostrar indicador de cambio pendiente
            repartidor_display = repartidor or '— Sin asignar'
            if tiene_cambio_pendiente:
                repartidor_display = f"✏️ {repartidor}" if repartidor else '— Sin asignar'
            
            # Para canceladas, mostrar repartidor + estado
            if cancelada or cancelada_otro_dia:
                if repartidor:
                    repartidor_display = f"{repartidor}, {estado}"
                else:
                    repartidor_display = estado
            
            self.tree_asign.insert("", tk.END, iid=str(folio),
                                   values=(folio, v['nombre'],
                                           f"${subtotal:,.2f}",
                                           f"${total:,.2f}",
                                           repartidor_display,
                                           usuario,
                                           estado,
                                           fecha_venta,
                                           fecha_cancel),
                                   tags=(tag,))
            
            # Acumular para resumen del filtro
            facturas_mostradas += 1
            # Separar: Efectivo (no crédito, no cancelado), Canceladas y Crédito
            if cancelada or cancelada_otro_dia:
                monto_canceladas_mostrado += subtotal
            elif not es_credito:
                # Solo facturas en efectivo (no crédito, no canceladas)
                monto_efectivo_mostrado += subtotal
            # Sumar devoluciones parciales de este folio si existen
            dev_parciales_mostradas += dev_parciales_por_folio.get(folio, 0)
            if not (cancelada or cancelada_otro_dia):
                if repartidor:
                    asignadas_mostradas += 1
                else:
                    sin_asignar_mostradas += 1
            if cancelada or cancelada_otro_dia:
                canceladas_mostradas += 1
            if es_credito:
                credito_mostradas += 1
        
        # Actualizar resumen según filtro
        self._actualizar_resumen_filtrado(estado_filtro, facturas_mostradas, monto_efectivo_mostrado,
                                          monto_canceladas_mostrado, dev_parciales_mostradas, 
                                          asignadas_mostradas, sin_asignar_mostradas,
                                          canceladas_mostradas, credito_mostradas)
    
    def _actualizar_resumen_filtrado(self, filtro, facturas, monto_efectivo, monto_canceladas, dev_parciales, asignadas, sin_asignar, canceladas, credito):
        """Actualiza el resumen basado en el filtro aplicado."""
        # Actualizar label de filtro activo
        self.lbl_filtro_activo.config(text=filtro if filtro else "Todos")
        self.lbl_facturas_filtradas.config(text=str(facturas))
        
        # Calcular monto total = monto efectivo + canceladas + devoluciones parciales
        monto_total = monto_efectivo + monto_canceladas + dev_parciales
        
        # Actualizar labels con valores del filtro
        self.lbl_monto_efectivo_asign.config(text=f"${monto_efectivo:,.2f}")
        self.lbl_canceladas_monto_asign.config(text=f"${monto_canceladas:,.2f}")
        self.lbl_dev_parciales_asign.config(text=f"${dev_parciales:,.2f}")
        self.lbl_monto_total_asign.config(text=f"${monto_total:,.2f}")
        self.lbl_asignadas.config(text=str(asignadas))
        self.lbl_sin_asignar.config(text=str(sin_asignar))
        self.lbl_canceladas_asign.config(text=str(canceladas))
        self.lbl_credito_asign.config(text=str(credito))
        
        # Calcular porcentaje
        total_asignables = asignadas + sin_asignar
        porcentaje = (asignadas / total_asignables * 100) if total_asignables > 0 else 0
        self.lbl_porcentaje_asign.config(text=f"{porcentaje:.1f}%")
        
        # Cambiar color del porcentaje según el valor
        if porcentaje >= 90:
            self.lbl_porcentaje_asign.config(foreground="#81c784")
        elif porcentaje >= 50:
            self.lbl_porcentaje_asign.config(foreground="#ffb74d")
        else:
            self.lbl_porcentaje_asign.config(foreground="#ef5350")
    
    def _enfocar_resultados_busqueda(self, event=None):
        """Enfoca el treeview de resultados y selecciona la primera fila."""
        children = self.tree_asign.get_children()
        if children:
            self.tree_asign.focus_set()
            self.tree_asign.selection_set(children[0])
            self.tree_asign.focus(children[0])
            self.tree_asign.see(children[0])
        return "break"  # Evitar comportamiento por defecto del Enter

    # --- limpiar todas las asignaciones ---
    def _limpiar_asignaciones(self):
        if not self.ds.ventas:
            messagebox.showinfo("Sin datos", "No hay facturas cargadas.")
            return
        if messagebox.askyesno("Confirmar",
                               "¿Limpiar TODAS las asignaciones de repartidores para este día?"):
            self.ds.clear_all_asignaciones()
            messagebox.showinfo("Listo", "Asignaciones eliminadas.")

    # ================================================================
    # EDICIÓN INLINE TIPO EXCEL
    # ================================================================
    def _cerrar_editor(self, event=None):
        """Destruye cualquier editor flotante activo sin guardar."""
        if self._editor_activo is not None:
            try:
                self._editor_activo.destroy()
            except tk.TclError:
                pass
            self._editor_activo = None

    def _on_tree_single_click(self, event):
        """Al hacer clic en una fila, solo selecciona (no abre editor)."""
        self._cerrar_editor()  # Cerrar editor si está abierto
        # La selección la maneja el Treeview automáticamente
    
    def _on_tree_navigate(self, event):
        """Al presionar flechas arriba/abajo, navega y cierra el editor."""
        self._cerrar_editor()
        # Permitir que el Treeview maneje la navegación normalmente
        # No retornar "break" para que funcione la navegación por defecto
    
    def _abrir_editor_fila_seleccionada(self, event=None):
        """Abre el editor en la fila actualmente seleccionada (con Enter)."""
        seleccion = self.tree_asign.selection()
        if seleccion:
            self._abrir_editor_en_fila(seleccion[0])
    
    def _abrir_editor_en_fila(self, row):
        """Abre el editor de repartidor en una fila específica."""
        self._cerrar_editor()
        
        valores = self.tree_asign.item(row, 'values')
        if not valores or len(valores) < 5:
            return
        
        folio = int(valores[0])
        valor_actual = str(valores[4]) if valores[4] and valores[4] != '— Sin asignar' else ''
        reps_conocidos = self.ds.get_repartidores()
        
        # bbox de la columna repartidor
        bbox = self.tree_asign.bbox(row, "repartidor")
        if not bbox:
            return
        
        self._crear_editor_repartidor(folio, valor_actual, reps_conocidos, bbox)

    def _on_tree_double_click(self, event):
        """Doble-clic también abre el editor (por compatibilidad)."""
        self._cerrar_editor()
        row = self.tree_asign.identify_row(event.y)
        if row:
            self._abrir_editor_en_fila(row)
    
    def _crear_editor_repartidor(self, folio: int, valor_actual: str, reps_conocidos: list, bbox: tuple):
        """Crea el editor de repartidor con autocompletado."""
        x, y, w, h = bbox

        # Obtener lista de repartidores ya usados en esta sesión (prioridad) + conocidos
        reps_en_uso = set()
        for item in self.tree_asign.get_children():
            vals = self.tree_asign.item(item, 'values')
            if vals and len(vals) > 4 and vals[4] and vals[4] != '— Sin asignar':
                reps_en_uso.add(vals[4])
        
        # Combinar: primero los usados en esta lista, luego los conocidos
        reps_lista = sorted(reps_en_uso) + [r for r in sorted(reps_conocidos) if r not in reps_en_uso]

        # Crear Combobox flotante encima de la celda
        combo = ttk.Combobox(
            self.tree_asign,
            values=reps_lista,
            state="normal",
            width=max(w // 7, 16)
        )
        combo.set(valor_actual)
        combo.place(x=x, y=y, width=w, height=h)
        combo.focus_set()
        combo.selection_range(0, tk.END)
        self._editor_activo = combo
        
        # Variable para sugerencia actual
        sugerencia_actual = [None]
        
        # Autocompletado inteligente con sugerencia
        def _autocompletar(event):
            # Ignorar teclas de navegación
            if event.keysym in ('Return', 'Escape', 'Shift_L', 'Shift_R',
                               'Control_L', 'Control_R', 'Alt_L', 'Alt_R',
                               'Up', 'Down'):
                return
            
            # Si es Tab, completar con la sugerencia
            if event.keysym == 'Tab':
                return  # Se maneja en el binding de Tab
            
            texto = combo.get()
            cursor_pos = combo.index(tk.INSERT)
            
            if not texto:
                combo['values'] = reps_lista
                sugerencia_actual[0] = None
                return
            
            texto_lower = texto.lower()
            
            # Buscar coincidencias que empiecen con el texto
            coincidencias = [r for r in reps_lista if r.lower().startswith(texto_lower)]
            
            if coincidencias:
                sugerencia_actual[0] = coincidencias[0]
                # Mostrar sugerencia en el campo (texto gris después del cursor)
                if len(coincidencias[0]) > len(texto):
                    # Mostrar sugerencia completando el texto
                    combo.set(coincidencias[0])
                    combo.selection_range(len(texto), tk.END)
                    combo.icursor(len(texto))
            else:
                sugerencia_actual[0] = None
                # Filtrar lista por coincidencia parcial
                filtrados = [r for r in reps_lista if texto_lower in r.lower()]
                combo['values'] = filtrados if filtrados else reps_lista
        
        combo.bind("<KeyRelease>", _autocompletar)

        # --- funciones de guardar / cancelar ---
        _guardado = [False]
        _tab_pressed = [False]  # Flag para saber si se presionó Tab

        def _guardar(evt=None):
            if _guardado[0]:
                return
            _guardado[0] = True

            nuevo = combo.get().strip() if combo.winfo_exists() else valor_actual

            self._editor_activo = None
            try:
                combo.destroy()
            except tk.TclError:
                pass

            if nuevo != valor_actual:
                self._registrar_cambio_pendiente(folio, nuevo, valor_actual)
            
            # Devolver el foco al Treeview para poder navegar con flechas
            self.ventana.after(50, lambda: self.tree_asign.focus_set())
        
        def _guardar_y_siguiente(evt=None):
            """Guarda y mueve al siguiente ticket."""
            if _guardado[0]:
                return
            _guardado[0] = True
            _tab_pressed[0] = True
            
            nuevo = combo.get().strip() if combo.winfo_exists() else valor_actual
            
            self._editor_activo = None
            try:
                combo.destroy()
            except tk.TclError:
                pass
            
            if nuevo != valor_actual:
                self._registrar_cambio_pendiente(folio, nuevo, valor_actual)
            
            # Mover a la siguiente fila y abrir editor
            self.ventana.after(50, self._mover_siguiente_fila)
            return "break"  # Evitar comportamiento por defecto del Tab

        def _cancelar(evt=None):
            if _guardado[0]:
                return
            _guardado[0] = True
            self._editor_activo = None
            try:
                combo.destroy()
            except tk.TclError:
                pass
        
        def _on_focus_out(evt=None):
            # No guardar si se presionó Tab (ya se manejó)
            if _tab_pressed[0]:
                return
            _guardar(evt)

        combo.bind("<Return>",   _guardar)
        combo.bind("<Tab>",      _guardar_y_siguiente)
        combo.bind("<Escape>",   _cancelar)
        combo.bind("<FocusOut>", _on_focus_out)
    
    def _mover_siguiente_fila(self):
        """Mueve la selección a la siguiente fila y abre el editor."""
        items = self.tree_asign.get_children()
        seleccion = self.tree_asign.selection()
        
        if not items or not seleccion:
            return
        
        current = seleccion[0]
        try:
            idx = items.index(current)
            if idx < len(items) - 1:
                next_item = items[idx + 1]
                self.tree_asign.selection_set(next_item)
                self.tree_asign.focus(next_item)
                self.tree_asign.see(next_item)
                # Abrir editor en la siguiente fila con delay suficiente
                self.ventana.after(100, lambda: self._abrir_editor_en_fila(next_item))
            else:
                # Si es la última fila, mantener el foco en el tree
                self.tree_asign.focus_set()
        except (ValueError, IndexError):
            pass

    # ================================================================
    # GESTIÓN DE CAMBIOS PENDIENTES EN REPARTIDORES
    # ================================================================
    def _registrar_cambio_pendiente(self, folio: int, nuevo_valor: str, valor_original: str):
        """Registra un cambio pendiente de repartidor y actualiza la UI."""
        # Si el nuevo valor es igual al original guardado, eliminar de pendientes
        # Obtener el valor real guardado en BD
        valor_bd = ''
        for v in self.ds.get_ventas():
            if v['folio'] == folio:
                valor_bd = v.get('repartidor', '') or ''
                break
        
        if nuevo_valor == valor_bd:
            # Eliminar de cambios pendientes si existe
            if folio in self._cambios_pendientes:
                del self._cambios_pendientes[folio]
        else:
            # Agregar/actualizar cambio pendiente
            self._cambios_pendientes[folio] = nuevo_valor
        
        # Actualizar la celda en el Treeview para mostrar el cambio
        self._actualizar_celda_repartidor(folio, nuevo_valor)
        
        # Actualizar estado del botón guardar
        self._actualizar_estado_boton_guardar()
    
    def _actualizar_celda_repartidor(self, folio: int, nuevo_valor: str):
        """Actualiza visualmente la celda de repartidor en el Treeview."""
        for item in self.tree_asign.get_children():
            valores = self.tree_asign.item(item, 'values')
            if int(valores[0]) == folio:
                # Actualizar solo la columna de repartidor
                nuevos_valores = list(valores)
                # Mostrar emoji de lápiz si hay cambio pendiente
                if folio in self._cambios_pendientes:
                    nuevos_valores[4] = f"✏️ {nuevo_valor}" if nuevo_valor else '— Sin asignar'
                else:
                    nuevos_valores[4] = nuevo_valor if nuevo_valor else '— Sin asignar'
                self.tree_asign.item(item, values=tuple(nuevos_valores))
                
                # Cambiar el tag para indicar cambio pendiente
                if folio in self._cambios_pendientes:
                    self.tree_asign.item(item, tags=("pendiente",))
                break
    
    def _actualizar_estado_boton_guardar(self):
        """Habilita/deshabilita el botón guardar según haya cambios pendientes."""
        if hasattr(self, '_cambios_pendientes') and self._cambios_pendientes:
            self.btn_guardar_asign.config(state="normal")
            n = len(self._cambios_pendientes)
            self.lbl_cambios_pendientes.config(
                text=f"⚠️ {n} cambio{'s' if n > 1 else ''} sin guardar"
            )
        else:
            self.btn_guardar_asign.config(state="disabled")
            self.lbl_cambios_pendientes.config(text="")
    
    def _guardar_cambios_repartidores(self):
        """Guarda todos los cambios pendientes de repartidores."""
        if not self._cambios_pendientes:
            return
        
        guardados = 0
        for folio, nuevo_rep in self._cambios_pendientes.items():
            if nuevo_rep:
                self.ds.set_repartidor_factura(folio, nuevo_rep)
            else:
                self.ds.clear_repartidor_factura(folio)
            guardados += 1
        
        # Limpiar cambios pendientes
        self._cambios_pendientes.clear()
        
        # Actualizar UI
        self._actualizar_estado_boton_guardar()
        self._refrescar_tree_asignacion()
        
        messagebox.showinfo("Guardado", f"Se guardaron {guardados} cambios correctamente.")
    
    def _guardar_cambios_pendientes(self, event=None):
        """Atajo Ctrl+S para guardar cambios pendientes."""
        if self._cambios_pendientes:
            self._guardar_cambios_repartidores()
        return "break"  # Evitar comportamiento por defecto
    
    # ================================================================
    # TOGGLE DE TEMA CLARO/OSCURO
    # ================================================================
    def _toggle_tema(self):
        """Cambia entre tema claro y oscuro."""
        self.modo_oscuro = self.tema_var.get()
        self._aplicar_tema()
        self._actualizar_tags_treeviews()
        # Actualizar fondo del switch de tema
        if hasattr(self, 'switch_tema'):
            self.switch_tema.set_frame_bg(self.COLORS['bg_dark'])
    
    def _actualizar_tags_treeviews(self):
        """Actualiza los tags de colores en todos los Treeviews."""
        C = self.COLORS
        
        # Tags para tree_asign
        if hasattr(self, 'tree_asign'):
            # Tags por repartidor específico
            self.tree_asign.tag_configure("rep_cristian",
                background="#37474f" if self.modo_oscuro else "#cfd8dc",
                foreground="#80cbc4" if self.modo_oscuro else "#37474f")
            self.tree_asign.tag_configure("rep_cajero",
                background="#33691e" if self.modo_oscuro else "#dcedc8",
                foreground="#c5e1a5" if self.modo_oscuro else "#33691e")
            self.tree_asign.tag_configure("rep_david",
                background="#4a148c" if self.modo_oscuro else "#e1bee7",
                foreground="#ce93d8" if self.modo_oscuro else "#4a148c")
            self.tree_asign.tag_configure("rep_otro",
                background="#004d40" if self.modo_oscuro else "#b2dfdb",
                foreground="#80cbc4" if self.modo_oscuro else "#004d40")
            
            # Estados especiales
            self.tree_asign.tag_configure("sin_asignar", 
                background="#4a4a4a" if self.modo_oscuro else "#ffcdd2", 
                foreground="#ff8a80" if self.modo_oscuro else "#c62828")
            self.tree_asign.tag_configure("cancelada", 
                background="#b71c1c" if self.modo_oscuro else "#ef9a9a", 
                foreground="#ffffff" if self.modo_oscuro else "#b71c1c")
            self.tree_asign.tag_configure("cancelada_otro_dia", 
                background="#880e4f" if self.modo_oscuro else "#f8bbd9", 
                foreground="#ffffff" if self.modo_oscuro else "#880e4f")
            self.tree_asign.tag_configure("credito", 
                background="#e65100" if self.modo_oscuro else "#ffe0b2", 
                foreground="#ffffff" if self.modo_oscuro else "#e65100")
            self.tree_asign.tag_configure("pendiente", 
                background="#ffc107" if self.modo_oscuro else "#fff9c4", 
                foreground="#000000" if self.modo_oscuro else "#f57f17")
        
        # Tags para tree_liq (Liquidaciones)
        if hasattr(self, 'tree_liq'):
            # Tags por repartidor (tonos oscuros suaves)
            self.tree_liq.tag_configure("rep_cristian",
                background="#37474f" if self.modo_oscuro else "#cfd8dc",
                foreground="#80cbc4" if self.modo_oscuro else "#37474f")
            self.tree_liq.tag_configure("rep_cajero",
                background="#33691e" if self.modo_oscuro else "#dcedc8",
                foreground="#c5e1a5" if self.modo_oscuro else "#33691e")
            self.tree_liq.tag_configure("rep_david",
                background="#4a148c" if self.modo_oscuro else "#e1bee7",
                foreground="#ce93d8" if self.modo_oscuro else "#4a148c")
            self.tree_liq.tag_configure("rep_otro",
                background="#004d40" if self.modo_oscuro else "#b2dfdb",
                foreground="#80cbc4" if self.modo_oscuro else "#004d40")
            # Estados especiales
            self.tree_liq.tag_configure("sin_asignar", 
                background="#4a4a4a" if self.modo_oscuro else "#ffcdd2", 
                foreground="#ff8a80" if self.modo_oscuro else "#c62828")
            self.tree_liq.tag_configure("cancelada", 
                background="#b71c1c" if self.modo_oscuro else "#ef9a9a", 
                foreground="#ffffff" if self.modo_oscuro else "#b71c1c")
            self.tree_liq.tag_configure("cancelada_otro_dia", 
                background="#880e4f" if self.modo_oscuro else "#f8bbd9", 
                foreground="#ffffff" if self.modo_oscuro else "#880e4f")
            self.tree_liq.tag_configure("credito", 
                background="#e65100" if self.modo_oscuro else "#ffe0b2", 
                foreground="#ffffff" if self.modo_oscuro else "#e65100")
        
        # Tags para tree_dev_parciales (Detalle de devoluciones parciales)
        if hasattr(self, 'tree_dev_parciales'):
            self.tree_dev_parciales.tag_configure("devolucion", 
                background="#4a1f1f" if self.modo_oscuro else "#ffebee", 
                foreground="#ff8a80" if self.modo_oscuro else "#b71c1c",
                font=("Segoe UI", 9))
        
        # Tags comunes para otros Treeviews
        for tree_name in ['tree_buscar', 'tree_desc', 'tree_productos', 'tree_devparciales']:
            if hasattr(self, tree_name):
                tree = getattr(self, tree_name)
                tree.tag_configure("par", 
                    background=C['row_par'], 
                    foreground=C['text'])
                tree.tag_configure("impar", 
                    background=C['row_impar'], 
                    foreground=C['text'])
                tree.tag_configure("selec", 
                    background=C['primary_dark'], 
                    foreground="#ffffff")
        
        # Refrescar las vistas para aplicar cambios
        if hasattr(self, 'ds') and self.ds.get_ventas():
            self._refrescar_tree_asignacion()
            if hasattr(self, 'tree_liq'):
                self._refrescar_liquidacion()

    # --- clic derecho → menú contextual para asignación ---
    def _on_tree_right_click_asign(self, event):
        row = self.tree_asign.identify_row(event.y)
        if not row:
            # Si no hay fila, mostrar solo opción de copiar toda la tabla
            self._mostrar_menu_copiar(self.tree_asign, event)
            return
        self.tree_asign.selection_set(row)
        valores = self.tree_asign.item(row, 'values')
        folio = int(valores[0])

        menu = tk.Menu(self.ventana, tearoff=0)
        menu.add_command(
            label="📋 Copiar fila seleccionada",
            command=lambda: self._copiar_seleccion_tree(self.tree_asign)
        )
        menu.add_command(
            label="📋 Copiar toda la tabla",
            command=lambda: self._copiar_toda_tabla(self.tree_asign)
        )
        menu.add_separator()
        menu.add_command(
            label="🗑️ Limpiar asignación de repartidor",
            command=lambda: self.ds.clear_repartidor_factura(folio)
        )
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    # ==================================================================
    # PESTAÑA 1 – LIQUIDACIÓN  (datos en tiempo real desde DataStore)
    # ==================================================================
    def _crear_tab_liquidacion(self):
        # ============================================================
        # FRAMES INFERIORES - Se empaquetan PRIMERO con side=BOTTOM
        # (El primero queda más abajo)
        # ============================================================
        
        # --- Labels ocultos para mantener lógica (no se muestran pero se actualizan) ---
        # Estos labels son necesarios porque el código los actualiza en otras partes
        frame_inf = ttk.Frame(self.tab_liquidacion)
        # No empaquetamos frame_inf - solo creamos los labels para que existan
        self.lbl_dinero_contado = ttk.Label(frame_inf, text="$0.00", font=("Segoe UI", 10, "bold"))
        self.lbl_diferencia = ttk.Label(frame_inf, text="$0.00",
                                        font=("Segoe UI", 10, "bold"), foreground="red")

        # --- CONTENEDOR PARA CUADRE (sin scroll) ---
        self.frame_cuadre_content = ttk.Frame(self.tab_liquidacion)
        self.frame_cuadre_content.pack(side=tk.BOTTOM, fill=tk.X, padx=0, pady=0)

        # --- CUADRE ALMACEN (CAJA) ---
        frame_fin = ttk.Frame(self.frame_cuadre_content, padding=(5, 2))
        frame_fin.pack(side=tk.TOP, fill=tk.X, padx=5, pady=0)
        
        # Configurar grid para que las columnas se expandan uniformemente
        for i in range(3):
            frame_fin.columnconfigure(i, weight=1)

        # ═══════════════════════════════════════════════════════════════════════
        # COLUMNA 1: DESCUENTOS Y AJUSTES
        # ═══════════════════════════════════════════════════════════════════════
        col2 = ttk.Frame(frame_fin)
        col2.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        ttk.Label(col2, text="📉 DESCUENTOS Y AJUSTES", font=("Segoe UI", 9, "bold"), 
                  foreground="#ff5252").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        
        # 1. Ajustes de Precios
        ttk.Label(col2, text="(-) Ajustes de Precios:", foreground="#ffb74d").grid(row=1, column=0, sticky=tk.W)
        self.lbl_total_ajustes = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#ffb74d")
        self.lbl_total_ajustes.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # 2. Gastos Repartidores
        ttk.Label(col2, text="(-) Gastos Repartidores:", foreground="#ce93d8").grid(row=2, column=0, sticky=tk.W)
        self.lbl_total_gastos_liq = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#ce93d8")
        self.lbl_total_gastos_liq.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # 3. Gastos Cajero
        ttk.Label(col2, text="(-) Gastos Cajero:", foreground="#b39ddb").grid(row=3, column=0, sticky=tk.W)
        self.lbl_total_gastos_cajero = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#b39ddb")
        self.lbl_total_gastos_cajero.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # 4. Pago Proveedores
        ttk.Label(col2, text="(-) Pago Proveedores:", foreground="#64b5f6").grid(row=4, column=0, sticky=tk.W)
        self.lbl_total_desc = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#64b5f6")
        self.lbl_total_desc.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        
        # 5. Préstamos
        ttk.Label(col2, text="(-) Préstamos:", foreground="#4db6ac").grid(row=5, column=0, sticky=tk.W)
        self.lbl_total_prestamos = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#4db6ac")
        self.lbl_total_prestamos.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # 6. Nómina
        ttk.Label(col2, text="(-) Nómina:", foreground="#bcaaa4").grid(row=6, column=0, sticky=tk.W)
        self.lbl_total_nomina_desc = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#bcaaa4")
        self.lbl_total_nomina_desc.grid(row=6, column=1, sticky=tk.E, padx=(10, 0))
        
        # 7. Socios
        ttk.Label(col2, text="(-) Socios:", foreground="#ffab91").grid(row=7, column=0, sticky=tk.W)
        self.lbl_total_socios_desc = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#ffab91")
        self.lbl_total_socios_desc.grid(row=7, column=1, sticky=tk.E, padx=(10, 0))
        
        # 8. Transferencias
        ttk.Label(col2, text="(-) Transferencias:", foreground="#81d4fa").grid(row=8, column=0, sticky=tk.W)
        self.lbl_total_transferencias_desc = ttk.Label(col2, text="$0", font=("Segoe UI", 9, "bold"), foreground="#81d4fa")
        self.lbl_total_transferencias_desc.grid(row=8, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador
        ttk.Separator(col2, orient="horizontal").grid(row=9, column=0, columnspan=2, sticky="ew", pady=3)
        
        # 9. Total Descuentos (suma de los anteriores)
        ttk.Label(col2, text="= Total Descuentos:", font=("Segoe UI", 9, "bold"), foreground="#ff8a80").grid(row=10, column=0, sticky=tk.W)
        self.lbl_total_devoluciones = ttk.Label(col2, text="$0", font=("Segoe UI", 10, "bold"), foreground="#ff8a80")
        self.lbl_total_devoluciones.grid(row=10, column=1, sticky=tk.E, padx=(10, 0))

        # ═══════════════════════════════════════════════════════════════════════
        # COLUMNA 2: CUADRE GENERAL
        # ═══════════════════════════════════════════════════════════════════════
        col3 = ttk.Frame(frame_fin)
        col3.grid(row=0, column=1, sticky="nsew", padx=(0, 15))
        
        ttk.Label(col3, text="📊 CUADRE GENERAL", font=("Segoe UI", 9, "bold"), 
                  foreground="#1565c0").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        
        # Total Dinero Caja (copia del valor de DINERO EN CAJA del corte cajero)
        ttk.Label(col3, text="Total Dinero Caja:", font=("Segoe UI", 9, "bold")).grid(row=1, column=0, sticky=tk.W)
        self.lbl_total_dinero_cuadre = ttk.Label(col3, text="$0", font=("Segoe UI", 10, "bold"), foreground="#2e7d32")
        self.lbl_total_dinero_cuadre.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # Total Descuentos (copiado de columna 1)
        ttk.Label(col3, text="(-) Total Descuentos:", foreground="#c62828").grid(row=2, column=0, sticky=tk.W)
        self.lbl_total_desc_cuadre = ttk.Label(col3, text="$0", font=("Segoe UI", 9, "bold"), foreground="#c62828")
        self.lbl_total_desc_cuadre.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Total Créditos Punteados (NUEVO)
        ttk.Label(col3, text="(-) Créditos Punteados:", foreground="#e65100").grid(row=3, column=0, sticky=tk.W)
        self.lbl_total_creditos_punteados = ttk.Label(col3, text="$0", font=("Segoe UI", 9, "bold"), foreground="#e65100")
        self.lbl_total_creditos_punteados.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # Total No Entregados
        ttk.Label(col3, text="(-) No Entregados:", foreground="#00bcd4").grid(row=4, column=0, sticky=tk.W)
        self.lbl_total_no_entregados = ttk.Label(col3, text="$0", font=("Segoe UI", 9, "bold"), foreground="#00bcd4")
        self.lbl_total_no_entregados.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador
        ttk.Separator(col3, orient="horizontal").grid(row=5, column=0, columnspan=2, sticky="ew", pady=3)
        
        # TOTAL EFECTIVO CAJA (Total Vendido - Total Descuentos - Créditos Punteados - No Entregados)
        ttk.Label(col3, text="= TOTAL EFECTIVO CAJA:", font=("Segoe UI", 9, "bold")).grid(row=6, column=0, sticky=tk.W)
        self.lbl_total_efectivo_caja = ttk.Label(col3, text="$0", font=("Segoe UI", 11, "bold"), foreground="#1565c0")
        self.lbl_total_efectivo_caja.grid(row=6, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador
        ttk.Separator(col3, orient="horizontal").grid(row=7, column=0, columnspan=2, sticky="ew", pady=3)
        
        # Conteo de Dinero (copia del valor del módulo conteo de dinero) - DEBAJO DE TOTAL EFECTIVO CAJA
        ttk.Label(col3, text="💵 Conteo de Dinero:", font=("Segoe UI", 9, "bold")).grid(row=8, column=0, sticky=tk.W)
        self.lbl_conteo_dinero_cuadre = ttk.Label(col3, text="$0", font=("Segoe UI", 10, "bold"), foreground="#1565c0")
        self.lbl_conteo_dinero_cuadre.grid(row=8, column=1, sticky=tk.E, padx=(10, 0))
        
        # Diferencia Final (Conteo - Efectivo)
        ttk.Label(col3, text="📊 Diferencia Final:", font=("Segoe UI", 9, "bold")).grid(row=9, column=0, sticky=tk.W)
        self.lbl_diferencia_cuadre = ttk.Label(col3, text="$0", font=("Segoe UI", 10, "bold"), foreground="#9e9e9e")
        self.lbl_diferencia_cuadre.grid(row=9, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador
        ttk.Separator(col3, orient="horizontal").grid(row=10, column=0, columnspan=2, sticky="ew", pady=3)
        
        # Total Facturadas a Crédito (desde Firebird)
        ttk.Label(col3, text="Total a Crédito (FB):").grid(row=11, column=0, sticky=tk.W)
        self.lbl_total_credito = ttk.Label(col3, text="$0", font=("Segoe UI", 9, "bold"), foreground="#f57c00")
        self.lbl_total_credito.grid(row=11, column=1, sticky=tk.E, padx=(10, 0))

        # ═══════════════════════════════════════════════════════════════════════
        # COLUMNA 3: CUADRE REPARTIDOR
        # ═══════════════════════════════════════════════════════════════════════
        col4 = ttk.Frame(frame_fin)
        col4.grid(row=0, column=2, sticky="nsew")
        
        ttk.Label(col4, text="💰 CUADRE REPARTIDOR", font=("Segoe UI", 9, "bold"), 
                  foreground="#2e7d32").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        
        # 1. Conteo de Dinero (del módulo conteo de dinero)
        ttk.Label(col4, text="💵 Conteo de Dinero:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_conteo_dinero_resultado = ttk.Label(col4, text="$0.00", font=("Segoe UI", 10, "bold"), foreground="#1565c0")
        self.lbl_conteo_dinero_resultado.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # 2. Monto Facturas (valor del módulo Asignar Repartidores - TOTALES)
        ttk.Label(col4, text="📊 Monto Facturas:").grid(row=2, column=0, sticky=tk.W)
        self.lbl_monto_facturas_resultado = ttk.Label(col4, text="$0.00", font=("Segoe UI", 10, "bold"), foreground="#2e7d32")
        self.lbl_monto_facturas_resultado.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # 3. Total Descuentos
        ttk.Label(col4, text="(-) Total Descuentos:", foreground="#ff9800").grid(row=3, column=0, sticky=tk.W)
        self.lbl_total_desc_resultado = ttk.Label(col4, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#ff9800")
        self.lbl_total_desc_resultado.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # 4. Créditos Punteados
        ttk.Label(col4, text="(-) Créditos Punteados:", foreground="#e65100").grid(row=4, column=0, sticky=tk.W)
        self.lbl_creditos_punt_resultado = ttk.Label(col4, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#e65100")
        self.lbl_creditos_punt_resultado.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        
        # 5. No Entregados
        ttk.Label(col4, text="(-) No Entregados:", foreground="#00bcd4").grid(row=5, column=0, sticky=tk.W)
        self.lbl_no_entreg_resultado = ttk.Label(col4, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#00bcd4")
        self.lbl_no_entreg_resultado.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador visual
        ttk.Separator(col4, orient="horizontal").grid(row=6, column=0, columnspan=2, sticky="ew", pady=5)
        
        # TOTAL DINERO A ENTREGAR (grande y destacado)
        ttk.Label(col4, text="💵 TOTAL DINERO A ENTREGAR:", font=("Segoe UI", 10, "bold")).grid(row=7, column=0, sticky=tk.W)
        self.lbl_neto = ttk.Label(col4, text="$0.00", font=("Segoe UI", 12, "bold"), foreground="#2e7d32")
        self.lbl_neto.grid(row=7, column=1, sticky=tk.E, padx=(10, 0))
        
        # Diferencia con dinero entregado
        ttk.Label(col4, text="Diferencia:", foreground="#9e9e9e").grid(row=8, column=0, sticky=tk.W)
        self.lbl_diferencia_global = ttk.Label(col4, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#9e9e9e")
        self.lbl_diferencia_global.grid(row=8, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador
        ttk.Separator(col4, orient="horizontal").grid(row=9, column=0, columnspan=2, sticky="ew", pady=5)
        
        # Total Créditos Cobrados (abonos + pagos completos del día)
        ttk.Label(col4, text="💳 Total Créditos Cobrados:", font=("Segoe UI", 9, "bold"), foreground="#7b1fa2").grid(row=10, column=0, sticky=tk.W)
        self.lbl_total_creditos_cobrados = ttk.Label(col4, text="$0.00", font=("Segoe UI", 10, "bold"), foreground="#7b1fa2")
        self.lbl_total_creditos_cobrados.grid(row=10, column=1, sticky=tk.E, padx=(10, 0))

        # ═══════════════════════════════════════════════════════════════════════
        # FILA 2: CORTE CAJERO (DATOS DE ELEVENTA) - SCROLLEABLE
        # ═══════════════════════════════════════════════════════════════════════
        # Contenedor con scroll para CORTE CAJERO
        frame_corte_container = ttk.LabelFrame(self.frame_cuadre_content, text="📊 CUADRE CAJA (ELEVENTA)", padding=(0, 0))
        frame_corte_container.pack(side=tk.TOP, fill=tk.X, padx=5, pady=(5, 0))
        
        # Canvas scrolleable
        self.canvas_corte = tk.Canvas(frame_corte_container, highlightthickness=0, height=130)
        scrollbar_corte = ttk.Scrollbar(frame_corte_container, orient="vertical", command=self.canvas_corte.yview)
        self.canvas_corte.configure(yscrollcommand=scrollbar_corte.set)
        
        scrollbar_corte.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas_corte.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Frame interno
        frame_corte = ttk.Frame(self.canvas_corte, padding=(5, 1))
        self.canvas_corte.create_window((0, 0), window=frame_corte, anchor="nw")
        
        # Actualizar scroll region
        def _on_corte_configure(event):
            self.canvas_corte.configure(scrollregion=self.canvas_corte.bbox("all"))
        frame_corte.bind("<Configure>", _on_corte_configure)
        
        # Ajustar ancho
        def _on_canvas_corte_configure(event):
            self.canvas_corte.itemconfig(self.canvas_corte.find_withtag("all")[0], width=event.width)
        self.canvas_corte.bind("<Configure>", _on_canvas_corte_configure)
        
        # Scroll con rueda del mouse
        def _on_mousewheel_corte(event):
            self.canvas_corte.yview_scroll(int(-1*(event.delta/120)), "units")
        self.canvas_corte.bind("<MouseWheel>", _on_mousewheel_corte)
        frame_corte.bind("<MouseWheel>", _on_mousewheel_corte)
        
        # Propagar scroll a hijos
        def _bind_corte_mousewheel(widget):
            widget.bind("<MouseWheel>", _on_mousewheel_corte)
            for child in widget.winfo_children():
                _bind_corte_mousewheel(child)
        
        # Configurar grid para 4 columnas
        for i in range(4):
            frame_corte.columnconfigure(i, weight=1)
        
        # --- COLUMNA 1: DINERO EN CAJA ---
        col_dinero = ttk.Frame(frame_corte)
        col_dinero.grid(row=0, column=0, sticky="nsew", padx=(0, 20))
        
        ttk.Label(col_dinero, text="💵 DINERO EN CAJA", font=("Segoe UI", 8, "bold"), 
                  foreground="#1565c0").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 2))
        
        # Fondo de Caja
        ttk.Label(col_dinero, text="Fondo de Caja:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_corte_fondo_caja = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_fondo_caja.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # Ventas en Efectivo
        ttk.Label(col_dinero, text="(+) Ventas en Efectivo:", foreground="#2e7d32").grid(row=2, column=0, sticky=tk.W)
        self.lbl_corte_ventas_efectivo = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9, "bold"), foreground="#2e7d32")
        self.lbl_corte_ventas_efectivo.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Abonos en Efectivo
        ttk.Label(col_dinero, text="(+) Abonos en Efectivo:", foreground="#00695c").grid(row=3, column=0, sticky=tk.W)
        self.lbl_corte_abonos_efectivo = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_abonos_efectivo.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # Entradas
        ttk.Label(col_dinero, text="(+) Entradas:", foreground="#1565c0").grid(row=4, column=0, sticky=tk.W)
        self.lbl_corte_entradas = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_entradas.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador y Total Efectivo (Ventas Efectivo + Entradas)
        ttk.Separator(col_dinero, orient="horizontal").grid(row=5, column=0, columnspan=2, sticky="ew", pady=2)
        ttk.Label(col_dinero, text="= Total Efectivo:", font=("Segoe UI", 9, "bold"), foreground="#2e7d32").grid(row=6, column=0, sticky=tk.W)
        self.lbl_corte_total_efectivo = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9, "bold"), foreground="#2e7d32")
        self.lbl_corte_total_efectivo.grid(row=6, column=1, sticky=tk.E, padx=(10, 0))
        
        # Canceladas día (del resumen de asignación)
        ttk.Label(col_dinero, text="(-) Canceladas:", foreground="#d32f2f").grid(row=7, column=0, sticky=tk.W)
        self.lbl_corte_canceladas_dia = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9, "bold"), foreground="#d32f2f")
        self.lbl_corte_canceladas_dia.grid(row=7, column=1, sticky=tk.E, padx=(10, 0))
        
        # Dev. Parciales (del resumen de asignación)
        ttk.Label(col_dinero, text="(-) Dev. Parciales:", foreground="#ff8a65").grid(row=8, column=0, sticky=tk.W)
        self.lbl_corte_dev_parciales_dia = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 9, "bold"), foreground="#ff8a65")
        self.lbl_corte_dev_parciales_dia.grid(row=8, column=1, sticky=tk.E, padx=(10, 0))
        
        # Labels ocultos para compatibilidad
        self.lbl_corte_salidas = ttk.Label(col_dinero, text="$0")
        self.lbl_corte_dev_efectivo = ttk.Label(col_dinero, text="$0")
        
        # Separador y Total
        ttk.Separator(col_dinero, orient="horizontal").grid(row=9, column=0, columnspan=2, sticky="ew", pady=2)
        ttk.Label(col_dinero, text="= Total Dinero Caja:", font=("Segoe UI", 9, "bold")).grid(row=10, column=0, sticky=tk.W)
        self.lbl_corte_total_dinero = ttk.Label(col_dinero, text="$0", font=("Segoe UI", 10, "bold"), foreground="#1565c0")
        self.lbl_corte_total_dinero.grid(row=10, column=1, sticky=tk.E, padx=(10, 0))
        
        # --- COLUMNA 2: VENTAS (dividida en 2 sub-columnas) ---
        col_ventas = ttk.Frame(frame_corte)
        col_ventas.grid(row=0, column=1, sticky="nsew", padx=(0, 20))
        
        # Sub-columna izquierda: Ventas principales
        col_ventas_izq = ttk.Frame(col_ventas)
        col_ventas_izq.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        ttk.Label(col_ventas_izq, text="🛒 VENTAS", font=("Segoe UI", 8, "bold"), 
                  foreground="#81c784").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 2))
        
        # En Efectivo
        ttk.Label(col_ventas_izq, text="En Efectivo:").grid(row=1, column=0, sticky=tk.W)
        self.lbl_corte_v_efectivo = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9, "bold"), foreground="#81c784")
        self.lbl_corte_v_efectivo.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # Con Tarjeta (oculto pero necesario para el código)
        self.lbl_corte_v_tarjeta = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9))
        
        # A Crédito
        ttk.Label(col_ventas_izq, text="A Crédito:").grid(row=2, column=0, sticky=tk.W)
        self.lbl_corte_v_credito = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_v_credito.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Con Vales (oculto pero necesario para el código)
        self.lbl_corte_v_vales = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9))
        
        # Separador y Total Vendido (En Efectivo + A Crédito)
        ttk.Separator(col_ventas_izq, orient="horizontal").grid(row=3, column=0, columnspan=2, sticky="ew", pady=2)
        ttk.Label(col_ventas_izq, text="= Total Vendido:", font=("Segoe UI", 9, "bold")).grid(row=4, column=0, sticky=tk.W)
        self.lbl_corte_total_ventas = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 10, "bold"), foreground="#81c784")
        self.lbl_corte_total_ventas.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        
        # Devoluciones de Ventas (TODAS)
        ttk.Label(col_ventas_izq, text="(-) Devoluciones Ventas:", foreground="#ef5350").grid(row=5, column=0, sticky=tk.W)
        self.lbl_corte_dev_ventas = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9, "bold"), foreground="#ef5350")
        self.lbl_corte_dev_ventas.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # Canceladas (facturas completas canceladas - informativo)
        ttk.Label(col_ventas_izq, text="    └─ Canceladas:", foreground="#ff8a80").grid(row=6, column=0, sticky=tk.W)
        self.lbl_corte_canceladas = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9), foreground="#ff8a80")
        self.lbl_corte_canceladas.grid(row=6, column=1, sticky=tk.E, padx=(10, 0))
        
        # Devoluciones Parciales (informativo)
        ttk.Label(col_ventas_izq, text="    └─ Dev. Parciales:", foreground="#ffb74d").grid(row=7, column=0, sticky=tk.W)
        self.lbl_corte_dev_parciales = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9), foreground="#ffb74d")
        self.lbl_corte_dev_parciales.grid(row=7, column=1, sticky=tk.E, padx=(10, 0))
        
        # Total Eleventa (Total Vendido - Devoluciones Ventas)
        ttk.Separator(col_ventas_izq, orient="horizontal").grid(row=8, column=0, columnspan=2, sticky="ew", pady=2)
        ttk.Label(col_ventas_izq, text="= Total Eleventa:", font=("Segoe UI", 9, "bold")).grid(row=9, column=0, sticky=tk.W)
        self.lbl_corte_total_eleventa = ttk.Label(col_ventas_izq, text="$0", font=("Segoe UI", 9, "bold"), foreground="#81c784")
        self.lbl_corte_total_eleventa.grid(row=9, column=1, sticky=tk.E, padx=(10, 0))
        
        # Sub-columna derecha: Ajustes y bugs
        col_ventas_der = ttk.Frame(col_ventas)
        col_ventas_der.grid(row=0, column=1, sticky="nsew")
        
        ttk.Label(col_ventas_der, text="🔧 AJUSTES", font=("Segoe UI", 8, "bold"), 
                  foreground="#ff5722").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 2))
        
        # Bug Dev. Parc. (Bug TURNOS > CORTE_MOV - informativo)
        ttk.Label(col_ventas_der, text="(+) Bug Dev. Parc:", foreground="#ff5722").grid(row=1, column=0, sticky=tk.W)
        self.lbl_corte_bug_dev_parc = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ff5722")
        self.lbl_corte_bug_dev_parc.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # Bug Duplicados (Bug duplicados en CORTE_MOVIMIENTOS - informativo)
        ttk.Label(col_ventas_der, text="(+) Bug Duplicados:", foreground="#ff5722").grid(row=2, column=0, sticky=tk.W)
        self.lbl_corte_bug_duplicados = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ff5722")
        self.lbl_corte_bug_duplicados.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Cancelaciones No Formalizadas (en CORTE_MOV pero no en tabla DEVOLUCIONES)
        ttk.Label(col_ventas_der, text="(+) Cancel. No Form.:", foreground="#ff5722").grid(row=3, column=0, sticky=tk.W)
        self.lbl_corte_cancel_no_form = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ff5722")
        self.lbl_corte_cancel_no_form.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # Bug Dev. Parc. OT (Devoluciones Parciales de Otro Turno duplicadas)
        ttk.Label(col_ventas_der, text="(+) Bug Dup.Dev.Parc.OT:", foreground="#ff5722").grid(row=4, column=0, sticky=tk.W)
        self.lbl_corte_bug_dev_parc_ot = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ff5722")
        self.lbl_corte_bug_dev_parc_ot.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        
        # Canceladas otro día (cancelaciones de ventas de otros días procesadas hoy - SUMA)
        ttk.Label(col_ventas_der, text="(+) Canceladas otro día:", foreground="#ce93d8").grid(row=5, column=0, sticky=tk.W)
        self.lbl_corte_cancel_otro_dia = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ce93d8")
        self.lbl_corte_cancel_otro_dia.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # Dev. Parciales NO registradas (ventas de este día devueltas en otro día - RESTA)
        ttk.Label(col_ventas_der, text="(-) Dev. Parc. No Regis.:", foreground="#ff80ab").grid(row=6, column=0, sticky=tk.W)
        self.lbl_corte_dev_parc_no_reg = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ff80ab")
        self.lbl_corte_dev_parc_no_reg.grid(row=6, column=1, sticky=tk.E, padx=(10, 0))
        
        # Dev. Parciales de otro día (afectan el corte del día que se procesan - SUMA)
        ttk.Label(col_ventas_der, text="(+) Dev. Parc. otro día:", foreground="#ce93d8").grid(row=7, column=0, sticky=tk.W)
        self.lbl_corte_dev_parc_otro_dia = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 9), foreground="#ce93d8")
        self.lbl_corte_dev_parc_otro_dia.grid(row=7, column=1, sticky=tk.E, padx=(10, 0))
        
        # Total Ventas después de descuentos
        ttk.Separator(col_ventas_der, orient="horizontal").grid(row=8, column=0, columnspan=2, sticky="ew", pady=2)
        ttk.Label(col_ventas_der, text="= Total Ventas Netas:", font=("Segoe UI", 9, "bold")).grid(row=9, column=0, sticky=tk.W)
        self.lbl_corte_ventas_netas = ttk.Label(col_ventas_der, text="$0", font=("Segoe UI", 10, "bold"), foreground="#64b5f6")
        self.lbl_corte_ventas_netas.grid(row=9, column=1, sticky=tk.E, padx=(10, 0))
        
        # --- COLUMNA 3: EXPLICACIÓN DIFERENCIA ---
        col_info = ttk.Frame(frame_corte)
        col_info.grid(row=0, column=2, sticky="nsew")
        
        ttk.Label(col_info, text="💡 DIFERENCIA EN CANCELACIONES", font=("Segoe UI", 8, "bold"), 
                  foreground="#ce93d8").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 2))
        
        # Cancelaciones en Efectivo
        ttk.Label(col_info, text="Cancel. en Efectivo:", foreground="#ff8a80").grid(row=1, column=0, sticky=tk.W)
        self.lbl_corte_exp_dev_ef = ttk.Label(col_info, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_exp_dev_ef.grid(row=1, column=1, sticky=tk.E, padx=(10, 0))
        
        # Cancelaciones a Crédito
        ttk.Label(col_info, text="Cancel. a Crédito:", foreground="#ffcc80").grid(row=2, column=0, sticky=tk.W)
        self.lbl_corte_exp_dev_cr = ttk.Label(col_info, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_exp_dev_cr.grid(row=2, column=1, sticky=tk.E, padx=(10, 0))
        
        # Cancelaciones con Tarjeta
        ttk.Label(col_info, text="Cancel. con Tarjeta:", foreground="#90caf9").grid(row=3, column=0, sticky=tk.W)
        self.lbl_corte_exp_dev_tar = ttk.Label(col_info, text="$0", font=("Segoe UI", 9))
        self.lbl_corte_exp_dev_tar.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # Separador
        ttk.Separator(col_info, orient="horizontal").grid(row=4, column=0, columnspan=2, sticky="ew", pady=2)
        
        # Diferencia
        ttk.Label(col_info, text="= Total Cancel. Ventas:", font=("Segoe UI", 9, "bold")).grid(row=5, column=0, sticky=tk.W)
        self.lbl_corte_exp_total_dev = ttk.Label(col_info, text="$0", font=("Segoe UI", 9, "bold"), foreground="#ef5350")
        self.lbl_corte_exp_total_dev.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # Botón para actualizar corte (usa versión async)
        ttk.Button(col_info, text="🔄 Actualizar Corte", 
                   command=self._actualizar_corte_cajero_async).grid(row=6, column=0, columnspan=2, sticky="ew", pady=(3, 0))
        
        # Ganancia (debajo del botón)
        ttk.Label(col_info, text="💰 Ganancia:", font=("Segoe UI", 9, "bold"), foreground="#ff6f00").grid(row=7, column=0, sticky=tk.W, pady=(5, 0))
        self.lbl_corte_ganancia = ttk.Label(col_info, text="$0", font=("Segoe UI", 10, "bold"), foreground="#ff6f00")
        self.lbl_corte_ganancia.grid(row=7, column=1, sticky=tk.E, padx=(10, 0), pady=(5, 0))

        # --- COLUMNA 4: CANCELACIONES POR USUARIO ---
        col_cancel = ttk.Frame(frame_corte)
        col_cancel.grid(row=0, column=3, sticky="nsew", padx=(20, 0))
        
        ttk.Label(col_cancel, text="❌ CANCELACIONES POR USUARIO", font=("Segoe UI", 8, "bold"), 
                  foreground="#ef5350").grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 2))
        
        # Labels dinámicos para usuarios (se llenarán al cargar datos)
        # Usaremos un diccionario para los labels de cancelaciones por usuario
        self.frame_cancel_usuarios = ttk.Frame(col_cancel)
        self.frame_cancel_usuarios.grid(row=1, column=0, columnspan=2, sticky="nsew")
        
        # Separador antes de Canceladas otro día
        ttk.Separator(col_cancel, orient="horizontal").grid(row=2, column=0, columnspan=2, sticky="ew", pady=3)
        
        # (-) Canceladas otro día (al final de cancelaciones por usuario)
        ttk.Label(col_cancel, text="📅 Canceladas otro día:", foreground="#b71c1c").grid(row=3, column=0, sticky=tk.W)
        self.lbl_total_canceladas_otro_dia = ttk.Label(col_cancel, text="$0.00", font=("Segoe UI", 9, "bold"), foreground="#b71c1c")
        self.lbl_total_canceladas_otro_dia.grid(row=3, column=1, sticky=tk.E, padx=(10, 0))
        
        # (-) Dev. Parciales otro día (devoluciones de facturas de otros días procesadas hoy)
        lbl_dev_parciales_otro_dia_titulo = ttk.Label(col_cancel, text="📦 Dev. Parciales otro día:", 
                                                       foreground="#6a1b9a", cursor="hand2")
        lbl_dev_parciales_otro_dia_titulo.grid(row=4, column=0, sticky=tk.W)
        lbl_dev_parciales_otro_dia_titulo.bind("<Button-1>", lambda e: self._mostrar_modal_dev_parciales_otro_dia())
        
        self.lbl_dev_parciales_otro_dia = ttk.Label(col_cancel, text="$0.00", 
                                                     font=("Segoe UI", 9, "bold"), foreground="#6a1b9a", cursor="hand2")
        self.lbl_dev_parciales_otro_dia.grid(row=4, column=1, sticky=tk.E, padx=(10, 0))
        self.lbl_dev_parciales_otro_dia.bind("<Button-1>", lambda e: self._mostrar_modal_dev_parciales_otro_dia())
        
        # Cantidad de turnos del día
        ttk.Label(col_cancel, text="🔄 Cantidad de turnos:", foreground="#1565c0").grid(row=5, column=0, sticky=tk.W)
        self.lbl_cantidad_turnos = ttk.Label(col_cancel, text="0", font=("Segoe UI", 9, "bold"), foreground="#1565c0")
        self.lbl_cantidad_turnos.grid(row=5, column=1, sticky=tk.E, padx=(10, 0))
        
        # Aplicar scroll de rueda a todos los widgets del CORTE CAJERO
        _bind_corte_mousewheel(frame_corte)

        # ============================================================
        # PANEL DE ASIGNACIÓN DE REPARTIDOR (se empaqueta primero con side=BOTTOM)
        # ============================================================
        frame_asignar_rep = ttk.LabelFrame(self.tab_liquidacion, text="📝 ASIGNAR REPARTIDOR", padding=(5, 3))
        frame_asignar_rep.pack(side=tk.BOTTOM, fill=tk.X, padx=5, pady=(0, 5))
        
        # Folio seleccionado
        ttk.Label(frame_asignar_rep, text="Folio:").pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_folio_asignar = ttk.Label(frame_asignar_rep, text="—", font=("Segoe UI", 9, "bold"), foreground="#1565c0")
        self.lbl_folio_asignar.pack(side=tk.LEFT, padx=(0, 15))
        
        # Cliente
        ttk.Label(frame_asignar_rep, text="Cliente:").pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_cliente_asignar = ttk.Label(frame_asignar_rep, text="—", font=("Segoe UI", 9))
        self.lbl_cliente_asignar.pack(side=tk.LEFT, padx=(0, 15))
        
        # Repartidor actual
        ttk.Label(frame_asignar_rep, text="Rep. Actual:").pack(side=tk.LEFT, padx=(0, 5))
        self.lbl_rep_actual = ttk.Label(frame_asignar_rep, text="—", font=("Segoe UI", 9), foreground="#e65100")
        self.lbl_rep_actual.pack(side=tk.LEFT, padx=(0, 15))
        
        # Combo para nuevo repartidor
        ttk.Label(frame_asignar_rep, text="Nuevo Rep.:").pack(side=tk.LEFT, padx=(0, 5))
        self.combo_nuevo_rep_liq = ttk.Combobox(frame_asignar_rep, width=15, state="readonly")
        self.combo_nuevo_rep_liq.pack(side=tk.LEFT, padx=(0, 10))
        
        # Botón guardar
        ttk.Button(frame_asignar_rep, text="💾 Guardar Cambio", 
                   command=self._guardar_cambio_repartidor_liq).pack(side=tk.LEFT, padx=5)

        # ============================================================
        # TABLA DE VENTAS - Se empaqueta AL FINAL para ocupar el espacio restante
        # ============================================================
        frame_tabla = ttk.LabelFrame(self.tab_liquidacion, text="📊 VENTAS DEL DÍA", padding=(5, 2))
        frame_tabla.pack(fill=tk.BOTH, expand=True, padx=5, pady=0)

        # Contenedor para Treeview con scrollbars
        tree_liq_container = ttk.Frame(frame_tabla)
        tree_liq_container.pack(fill=tk.BOTH, expand=True)
        
        self.tree_liq = ttk.Treeview(
            tree_liq_container,
            columns=("credito", "no_entreg", "folio", "cliente", "subtotal", "art_dev", "precio_dev", "cant_dev", "total_dev", "ajuste", "total_desp_aj", "repartidor", "estado"),
            selectmode="extended", height=8
        )
        self.tree_liq.column("#0",           width=0, stretch=tk.NO)
        self.tree_liq.column("credito",      anchor=tk.CENTER, width=50)
        self.tree_liq.column("no_entreg",    anchor=tk.CENTER, width=55)
        self.tree_liq.column("folio",        anchor=tk.CENTER, width=60)
        self.tree_liq.column("cliente",      anchor=tk.W,      width=160)
        self.tree_liq.column("subtotal",     anchor=tk.E,      width=85)
        self.tree_liq.column("art_dev",      anchor=tk.W,      width=110)
        self.tree_liq.column("precio_dev",   anchor=tk.E,      width=65)
        self.tree_liq.column("cant_dev",     anchor=tk.CENTER, width=45)
        self.tree_liq.column("total_dev",    anchor=tk.E,      width=70)
        self.tree_liq.column("ajuste",       anchor=tk.E,      width=70)
        self.tree_liq.column("total_desp_aj",anchor=tk.E,      width=95)
        self.tree_liq.column("repartidor",   anchor=tk.CENTER, width=80)
        self.tree_liq.column("estado",       anchor=tk.CENTER, width=75)

        self.tree_liq.heading("credito",      text="Crédito")
        self.tree_liq.heading("no_entreg",    text="No Entreg")
        self.tree_liq.heading("folio",        text="📋 Folio")
        self.tree_liq.heading("cliente",      text="👤 Cliente")
        self.tree_liq.heading("subtotal",     text="💵 Subtotal")
        self.tree_liq.heading("art_dev",      text="↩️ Art.Dev")
        self.tree_liq.heading("precio_dev",   text="💲 Precio")
        self.tree_liq.heading("cant_dev",     text="📦 Cant")
        self.tree_liq.heading("total_dev",    text="💸 TotalDev")
        self.tree_liq.heading("ajuste",       text="📉 Ajuste")
        self.tree_liq.heading("total_desp_aj",text="✅ TotalDespAj")
        self.tree_liq.heading("repartidor",   text="🚚 Rep.")
        self.tree_liq.heading("estado",       text="📊 Estado")

        # Tags con colores para modo oscuro
        # Colores por repartidor (tonos oscuros suaves)
        self.tree_liq.tag_configure("rep_cristian",  background="#37474f", foreground="#80cbc4", font=("Segoe UI", 9))  # Gris azulado
        self.tree_liq.tag_configure("rep_cajero",    background="#33691e", foreground="#c5e1a5", font=("Segoe UI", 9))  # Verde oliva
        self.tree_liq.tag_configure("rep_david",     background="#4a148c", foreground="#ce93d8", font=("Segoe UI", 9))  # Morado profundo
        self.tree_liq.tag_configure("rep_otro",      background="#004d40", foreground="#80cbc4", font=("Segoe UI", 9))  # Teal oscuro
        # Estados especiales
        self.tree_liq.tag_configure("sin_asignar", background="#4a4a4a", foreground="#ff8a80", font=("Segoe UI", 9))
        self.tree_liq.tag_configure("cancelada",   background="#b71c1c", foreground="#ffffff", font=("Segoe UI", 9))
        self.tree_liq.tag_configure("cancelada_otro_dia", background="#880e4f", foreground="#ffffff", font=("Segoe UI", 9, "bold"))
        self.tree_liq.tag_configure("credito",     background="#e65100", foreground="#ffffff", font=("Segoe UI", 9))
        self.tree_liq.tag_configure("credito_punteado", background="#0d47a1", foreground="#ffffff", font=("Segoe UI", 9, "bold"))
        self.tree_liq.tag_configure("no_entregado", background="#6a1b9a", foreground="#ffffff", font=("Segoe UI", 9, "bold"))

        # Scrollbars
        scroll_liq_y = ttk.Scrollbar(tree_liq_container, orient=tk.VERTICAL, command=self.tree_liq.yview)
        scroll_liq_x = ttk.Scrollbar(tree_liq_container, orient=tk.HORIZONTAL, command=self.tree_liq.xview)
        self.tree_liq.configure(yscrollcommand=scroll_liq_y.set, xscrollcommand=scroll_liq_x.set)
        
        # Grid layout
        self.tree_liq.grid(row=0, column=0, sticky="nsew")
        scroll_liq_y.grid(row=0, column=1, sticky="ns")
        scroll_liq_x.grid(row=1, column=0, sticky="ew")
        
        tree_liq_container.grid_rowconfigure(0, weight=1)
        tree_liq_container.grid_columnconfigure(0, weight=1)
        
        # Bindings para copiar
        self.tree_liq.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(self.tree_liq))
        self.tree_liq.bind("<Control-C>", lambda e: self._copiar_seleccion_tree(self.tree_liq))
        self.tree_liq.bind("<Button-3>", lambda e: self._mostrar_menu_copiar(self.tree_liq, e))
        # Binding para mostrar detalle de devoluciones al seleccionar
        self.tree_liq.bind("<<TreeviewSelect>>", self._on_select_venta_liq)
        # Binding para toggle de crédito punteado al hacer click en columna crédito
        self.tree_liq.bind("<Button-1>", self._on_click_tree_liq)
        
        # Diccionario para guardar info de devoluciones por folio (usado internamente)
        self.devoluciones_detalle = {}

    def _on_click_tree_liq(self, event):
        """Maneja el click en la tabla de liquidación para toggle de crédito punteado o no entregado."""
        # Identificar la región del click
        region = self.tree_liq.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        # Identificar la columna
        column = self.tree_liq.identify_column(event.x)
        # #1 = credito, #2 = no_entregado
        if column not in ("#1", "#2"):
            return
        
        # Obtener la fila
        row_id = self.tree_liq.identify_row(event.y)
        if not row_id:
            return
        
        # Obtener valores de la fila
        values = self.tree_liq.item(row_id, "values")
        if not values or len(values) < 5:
            return
        
        # Verificar que tenga folio (no es fila de continuación)
        folio_str = values[2]  # Columna folio (índice 2: credito=0, no_entreg=1, folio=2)
        if not folio_str or folio_str == "":
            return
        
        # Verificar si es cancelada - NO permitir
        estado = values[12] if len(values) > 12 else ""
        if "CANCEL" in estado.upper():
            messagebox.showwarning("No permitido", "No se puede marcar facturas canceladas.")
            return
        
        try:
            folio = int(folio_str)
        except ValueError:
            return
        
        # Obtener cliente y subtotal
        cliente = values[3] if len(values) > 3 else ""
        subtotal_str = values[4] if len(values) > 4 else "$0"
        try:
            subtotal = float(subtotal_str.replace("$", "").replace(",", ""))
        except:
            subtotal = 0
        
        # Obtener repartidor
        repartidor = values[11] if len(values) > 11 else ""
        
        if column == "#1":
            # Toggle crédito punteado en la BD
            if USE_SQLITE and self.ds.fecha:
                if db_local.es_credito_punteado(self.ds.fecha, folio):
                    # Eliminar de créditos punteados
                    db_local.eliminar_credito_punteado(self.ds.fecha, folio)
                else:
                    # Agregar a créditos punteados
                    db_local.agregar_credito_punteado(
                        self.ds.fecha, folio, cliente, subtotal, repartidor
                    )
        elif column == "#2":
            # Toggle no entregado en la BD
            if USE_SQLITE and self.ds.fecha:
                if db_local.es_no_entregado(self.ds.fecha, folio):
                    # Eliminar de no entregados
                    db_local.eliminar_no_entregado(self.ds.fecha, folio)
                else:
                    # Agregar a no entregados
                    db_local.agregar_no_entregado(
                        self.ds.fecha, folio, cliente, subtotal, repartidor
                    )
        
        # Refrescar la tabla para mostrar el cambio
        self._refrescar_liquidacion()

    def _on_select_venta_liq(self, event=None):
        """Maneja la selección de una venta en liquidación para actualizar panel de asignación."""
        # Actualizar combo de repartidores
        reps = self.ds.get_repartidores()
        self.combo_nuevo_rep_liq['values'] = reps
        
        # Obtener selección
        sel = self.tree_liq.selection()
        if not sel:
            self.lbl_folio_asignar.config(text="—")
            self.lbl_cliente_asignar.config(text="—")
            self.lbl_rep_actual.config(text="—")
            self.combo_nuevo_rep_liq.set("")
            # También actualizar panel de devoluciones
            self._mostrar_detalle_devoluciones(event)
            return
        
        # Obtener valores de la fila
        values = self.tree_liq.item(sel[0], "values")
        if len(values) < 12:
            self._mostrar_detalle_devoluciones(event)
            return
        
        folio = values[2]  # Columna folio (índice 2: credito=0, no_entreg=1, folio=2)
        if not folio or folio == "":  # Fila de continuación
            self._mostrar_detalle_devoluciones(event)
            return
        
        cliente = values[3]  # Cliente (índice 3)
        repartidor = values[11]  # Repartidor actual (índice 11)
        
        # Actualizar panel de asignación
        self.lbl_folio_asignar.config(text=f"#{folio}")
        self.lbl_cliente_asignar.config(text=cliente if cliente else "—")
        
        if repartidor and repartidor != "—":
            self.lbl_rep_actual.config(text=repartidor, foreground="#2e7d32")
        else:
            self.lbl_rep_actual.config(text="SIN ASIGNAR", foreground="#c62828")
        
        # Pre-seleccionar repartidor actual en el combo
        if repartidor and repartidor in reps:
            self.combo_nuevo_rep_liq.set(repartidor)
        else:
            self.combo_nuevo_rep_liq.set("")
        
        # Guardar folio seleccionado para usarlo después
        self._folio_seleccionado_liq = folio
        
        # También actualizar panel de devoluciones
        self._mostrar_detalle_devoluciones(event)

    def _guardar_cambio_repartidor_liq(self):
        """Guarda el cambio de repartidor para la factura seleccionada."""
        # Verificar que haya folio seleccionado
        if not hasattr(self, '_folio_seleccionado_liq') or not self._folio_seleccionado_liq:
            messagebox.showwarning("Selección", "Selecciona una factura primero.")
            return
        
        # Verificar que se haya seleccionado un repartidor
        nuevo_rep = self.combo_nuevo_rep_liq.get()
        if not nuevo_rep:
            messagebox.showwarning("Repartidor", "Selecciona un repartidor.")
            return
        
        folio = self._folio_seleccionado_liq
        try:
            folio_int = int(folio)
        except:
            messagebox.showerror("Error", "Folio inválido.")
            return
        
        # Guardar asignación en SQLite
        if USE_SQLITE and self.ds.fecha:
            db_local.guardar_asignacion(self.ds.fecha, folio_int, nuevo_rep)
        
        # También actualizar en DataStore (memoria)
        for v in self.ds.get_ventas():
            if v['folio'] == folio_int:
                v['repartidor'] = nuevo_rep
                break
        
        # Notificar cambio
        self.ds._notificar()
        
        messagebox.showinfo("Guardado", f"Repartidor '{nuevo_rep}' asignado a factura #{folio}")
        
        # Refrescar
        self._refrescar_liquidacion()

    def _mostrar_detalle_devoluciones(self, event=None):
        """Muestra el detalle de las devoluciones parciales de la factura seleccionada en el panel inferior"""
        # Verificar que exista el diccionario
        if not hasattr(self, 'devoluciones_detalle'):
            self.devoluciones_detalle = {}
            
        # Limpiar tabla de devoluciones
        for item in self.tree_dev_parciales.get_children():
            self.tree_dev_parciales.delete(item)
        
        # Obtener selección
        sel = self.tree_liq.selection()
        if not sel:
            self.lbl_dev_folio.config(text="—")
            self.lbl_dev_total_orig.config(text="$0.00")
            self.lbl_dev_total_dev.config(text="$0.00")
            self.lbl_dev_total_final.config(text="$0.00")
            return
        
        # Obtener datos de la fila seleccionada (columnas: credito=0, no_entreg=1, folio=2, ...)
        values = self.tree_liq.item(sel[0], "values")
        if len(values) < 12:
            return
        
        folio = values[2]  # Columna folio (índice 2)
        if not folio:  # Fila de continuación de devolución
            return
            
        subtotal = values[4]  # Columna subtotal (índice 4)
        total_venta = values[10]  # Columna total después ajustes (índice 10)
        
        # Actualizar etiquetas
        self.lbl_dev_folio.config(text=f"#{folio}")
        
        # Buscar el total de devoluciones del folio
        try:
            folio_int = int(folio)
        except:
            folio_int = folio
        
        total_dev = 0
        if folio_int in self.devoluciones_detalle:
            for detalle in self.devoluciones_detalle[folio_int]:
                total_dev += detalle.get("dinero", 0)
                
                # Agregar al tree de detalle
                self.tree_dev_parciales.insert(
                    "", tk.END,
                    values=(
                        detalle.get("codigo", "—"),
                        detalle.get("articulo", "—"),
                        f"{int(detalle.get('cantidad', 0))}",
                        f"${detalle.get('valor_unitario', 0):,.2f}",
                        f"${detalle.get('dinero', 0):,.2f}"
                    ),
                    tags=("devolucion",)
                )
        
        # Calcular totales
        try:
            total_venta_num = float(total_venta.replace("$", "").replace(",", "")) if total_venta and total_venta != "—" else 0
            total_original = total_venta_num + total_dev
        except:
            total_venta_num = 0
            total_original = 0
        
        self.lbl_dev_total_orig.config(text=f"${total_original:,.2f}")
        self.lbl_dev_total_dev.config(text=f"${total_dev:,.2f}")
        self.lbl_dev_total_final.config(text=f"${total_venta_num:,.2f}")

    # --- refrescar tabla y resumen de liquidación ---
    def _refrescar_liquidacion(self):
        # actualizar lista de repartidores
        reps = self.ds.get_repartidores()
        opciones_filtro = ["(Todos)", "(Sin Asignar)"] + reps
        
        # Sincronizar con filtro global de repartidor si existe
        if hasattr(self, 'filtro_rep_global_var'):
            rep_global = self.filtro_rep_global_var.get()
            if rep_global and rep_global != "(Todos)" and rep_global in reps:
                self.repartidor_filtro_var.set(rep_global)
        
        filtro = self.repartidor_filtro_var.get()
        if filtro not in opciones_filtro:
            self.repartidor_filtro_var.set("(Todos)")

        filtro = self.repartidor_filtro_var.get()

        # filtrar ventas (excluyendo canceladas del total de ventas efectivas)
        ventas = self.ds.get_ventas()
        if filtro == "(Sin Asignar)":
            # Mostrar solo ventas sin repartidor asignado
            ventas = [v for v in ventas if not v['repartidor'] or v['repartidor'].strip() == '']
        elif filtro and filtro != "(Todos)":
            ventas = [v for v in ventas if v['repartidor'] == filtro]
        
        # Obtener créditos punteados para el filtro de estado
        creditos_punteados_folios_filtro = set()
        if USE_SQLITE and self.ds.fecha:
            creditos_punteados = db_local.obtener_creditos_punteados_fecha(self.ds.fecha)
            creditos_punteados_folios_filtro = {c['folio'] for c in creditos_punteados}
        
        # Aplicar filtro de estado (Crédito, Canceladas, Sin Repartidor, Todos)
        estado_filtro = self.filtro_estado_var.get() if hasattr(self, 'filtro_estado_var') else "Todos"
        if estado_filtro == "Sin Repartidor":
            ventas = [v for v in ventas if not v['repartidor'] or v['repartidor'].strip() == '']
        elif estado_filtro == "Canceladas":
            ventas = [v for v in ventas if v.get('cancelada', False) or v.get('cancelada_otro_dia', False)]
        elif estado_filtro == "Crédito":
            # Incluir créditos originales Y créditos punteados
            ventas = [v for v in ventas if v.get('es_credito', False) or v['folio'] in creditos_punteados_folios_filtro]
        
        # Aplicar filtro de búsqueda global
        texto_buscar = self.buscar_global_var.get().strip().lower() if hasattr(self, 'buscar_global_var') else ""
        if texto_buscar:
            ventas_filtradas = []
            for v in ventas:
                # Buscar en folio, nombre del cliente y repartidor
                folio_str = str(v.get('folio', ''))
                nombre = v.get('nombre', '').lower()
                repartidor = (v.get('repartidor') or '').lower()
                if (texto_buscar in folio_str or 
                    texto_buscar in nombre or 
                    texto_buscar in repartidor):
                    ventas_filtradas.append(v)
            ventas = ventas_filtradas

        # Obtener devoluciones parciales por folio
        dev_parciales_por_folio = {}
        self.devoluciones_detalle = {}  # Reiniciar detalle de devoluciones
        ajustes_por_folio = {}  # Ajustes de precio por folio
        creditos_punteados_folios = set()  # Folios marcados como crédito punteado
        no_entregados_folios = set()  # Folios marcados como no entregado
        if USE_SQLITE and self.ds.fecha:
            dev_parciales_por_folio = db_local.obtener_devoluciones_parciales_por_folio_fecha(self.ds.fecha)
            # Obtener detalle de artículos devueltos
            self.devoluciones_detalle = db_local.obtener_detalle_devoluciones_por_fecha(self.ds.fecha)
            # Obtener ajustes de precio por folio
            desc_lista = db_local.obtener_descuentos_fecha(self.ds.fecha)
            for d in desc_lista:
                if d.get('tipo') == 'ajuste':
                    fol = d.get('folio', 0)
                    ajustes_por_folio[fol] = ajustes_por_folio.get(fol, 0) + d.get('monto', 0)
            # Obtener créditos punteados
            creditos_punteados = db_local.obtener_creditos_punteados_fecha(self.ds.fecha)
            creditos_punteados_folios = {c['folio'] for c in creditos_punteados}
            # Obtener no entregados
            no_entregados = db_local.obtener_no_entregados_fecha(self.ds.fecha)
            no_entregados_folios = {n['folio'] for n in no_entregados}

        # poblar tree
        self.tree_liq.delete(*self.tree_liq.get_children())
        for v in ventas:
            cancelada = v.get('cancelada', False)
            es_credito = v.get('es_credito', False)
            cancelada_otro_dia = v.get('cancelada_otro_dia', False)
            folio = v['folio']
            es_credito_punteado = folio in creditos_punteados_folios
            es_no_entregado = folio in no_entregados_folios
            
            # Determinar tag
            if cancelada_otro_dia:
                tag = "cancelada_otro_dia"
            elif cancelada:
                tag = "cancelada"
            elif es_no_entregado:
                tag = "no_entregado"
            elif es_credito_punteado:
                tag = "credito_punteado"
            elif es_credito:
                tag = "credito"
            else:
                tag = self._get_repartidor_tag(v['repartidor'])
            
            # Checkbox de crédito punteado
            checkbox_credito = "☑" if es_credito_punteado else "☐"
            # Checkbox de no entregado
            checkbox_no_entreg = "☑" if es_no_entregado else "☐"
            
            # Obtener subtotal y total
            subtotal = v.get('subtotal', 0)
            total = v.get('total_original', subtotal)
            
            # Obtener el folio y calcular total devoluciones de esta factura
            total_dev_factura = dev_parciales_por_folio.get(folio, 0)
            total_ajuste_factura = ajustes_por_folio.get(folio, 0)
            
            # Calcular Nuevo Total = Subtotal - Devoluciones - Ajustes
            nuevo_total = subtotal - total_dev_factura - total_ajuste_factura
            
            # Determinar estado
            if cancelada_otro_dia:
                estado = "CANC. OTRO DÍA"
            elif cancelada:
                estado = "CANCELADA"
            elif es_no_entregado:
                estado = "NO ENTREG. ✓"
            elif es_credito_punteado:
                estado = "CRÉDITO ✓"
            elif es_credito:
                estado = "CRÉDITO"
            else:
                estado = "—"
            
            # Obtener detalle de devoluciones parciales de este folio
            dev_detalle = self.devoluciones_detalle.get(folio, [])
            
            if dev_detalle:
                # Insertar una fila por cada artículo devuelto
                for i, detalle in enumerate(dev_detalle):
                    art_dev = detalle.get("articulo", "—")
                    precio_dev = detalle.get("valor_unitario", 0)
                    cant_dev = detalle.get("cantidad", 0)
                    total_dev = detalle.get("dinero", 0)
                    
                    if i == 0:
                        # Primera fila: mostrar todos los datos de la venta
                        self.tree_liq.insert("", tk.END,
                                             values=(checkbox_credito, checkbox_no_entreg, folio, v['nombre'],
                                                     f"${subtotal:,.0f}",
                                                     art_dev,
                                                     f"${precio_dev:,.0f}",
                                                     f"{int(cant_dev)}",
                                                     f"${total_dev:,.0f}",
                                                     f"${total_ajuste_factura:,.0f}" if total_ajuste_factura > 0 else "—",
                                                     f"${nuevo_total:,.0f}",
                                                     v['repartidor'],
                                                     estado),
                                             tags=(tag,))
                    else:
                        # Filas adicionales: solo mostrar datos de devolución
                        self.tree_liq.insert("", tk.END,
                                             values=("", "", "", "",
                                                     "",
                                                     art_dev,
                                                     f"${precio_dev:,.0f}",
                                                     f"{int(cant_dev)}",
                                                     f"${total_dev:,.0f}",
                                                     "",
                                                     "",
                                                     "",
                                                     ""),
                                             tags=(tag,))
            else:
                # Sin devoluciones: fila normal
                self.tree_liq.insert("", tk.END,
                                     values=(checkbox_credito, checkbox_no_entreg, folio, v['nombre'],
                                             f"${subtotal:,.0f}",
                                             "—",
                                             "—",
                                             "—",
                                             "—",
                                             f"${total_ajuste_factura:,.0f}" if total_ajuste_factura > 0 else "—",
                                             f"${nuevo_total:,.0f}",
                                             v['repartidor'],
                                             estado),
                                     tags=(tag,))

        # ═══════════════════════════════════════════════════════════════════════
        # CALCULAR TODOS LOS TOTALES
        # ═══════════════════════════════════════════════════════════════════════
        
        # 1. Total de TODAS las facturas del día (por fecha de venta) - canceladas + no canceladas
        # Esto debe coincidir con el corte de caja de Firebird
        total_todas_facturas = sum(v.get('total_original', v['subtotal']) for v in ventas 
                                   if not v.get('cancelada_otro_dia', False))
        
        # 2. Total Canceladas del mismo día (valor original de facturas canceladas con fecha de hoy)
        total_canceladas = sum(v.get('total_original', 0) for v in ventas 
                               if v.get('cancelada', False) and not v.get('cancelada_otro_dia', False))
        
        # 3. Total Canceladas de otro día (facturas de otros días canceladas hoy)
        total_canceladas_otro_dia = sum(v.get('total_original', 0) for v in ventas 
                                        if v.get('cancelada_otro_dia', False))
        
        # 4. Total Canceladas General = Canceladas + Canceladas otro día
        total_canceladas_general = total_canceladas + total_canceladas_otro_dia
        
        # 5. Total Devoluciones Parciales (artículos devueltos sin cancelar factura)
        total_dev_parciales = 0
        if USE_SQLITE and self.ds.fecha:
            total_dev_parciales = db_local.obtener_total_devoluciones_parciales_fecha(self.ds.fecha)
        
        # 6. Total Vendido = Total Facturas del día - Total Canceladas General - Dev.Parciales
        total_vendido = total_todas_facturas - total_canceladas_general - total_dev_parciales
        
        # 7. Total Facturas + Canceladas (este es el valor que debe coincidir con Firebird)
        total_mas_cancel = total_todas_facturas
        
        # 8. Total a Crédito (facturas con crédito, de las válidas/no canceladas)
        total_credito = sum(v.get('total_credito', 0) for v in ventas if v.get('es_credito', False) and not v.get('cancelada', False))
        
        # 9. Total en Efectivo = Total Vendido - Total a Crédito
        total_efectivo = total_vendido - total_credito
        
        # 10. Total Créditos Punteados (desde SQLite) - dinámico según filtro
        total_creditos_punteados = 0
        if USE_SQLITE and self.ds.fecha:
            if filtro and filtro not in ("(Todos)", "(Sin Asignar)"):
                # Filtro específico: solo créditos punteados de los folios del repartidor
                folios_repartidor = [v['folio'] for v in ventas]
                total_creditos_punteados = db_local.obtener_total_creditos_punteados_por_folios(self.ds.fecha, folios_repartidor)
            else:
                # Sin filtro: todos los créditos punteados
                total_creditos_punteados = db_local.obtener_total_creditos_punteados(self.ds.fecha)
        
        # 10b. Total No Entregados (desde SQLite) - dinámico según filtro
        total_no_entregados = 0
        if USE_SQLITE and self.ds.fecha:
            if filtro and filtro not in ("(Todos)", "(Sin Asignar)"):
                # Filtro específico: solo no entregados de los folios del repartidor
                folios_repartidor = [v['folio'] for v in ventas]
                total_no_entregados = db_local.obtener_total_no_entregados_por_folios(self.ds.fecha, folios_repartidor)
            else:
                # Sin filtro: todos los no entregados
                total_no_entregados = db_local.obtener_total_no_entregados(self.ds.fecha)
        
        # Filtro para gastos: solo aplica si es un repartidor específico
        filtro_gastos = filtro if filtro and filtro not in ("(Todos)", "(Sin Asignar)") else ''
        
        # 11. Total Ajustes de Precios (tipo 'ajuste' en descuentos)
        total_ajustes = self.ds.get_total_ajustes(filtro_gastos)
        
        # 12. Total Gastos de repartidores (excluyendo cajero)
        total_gastos = self.ds.get_total_gastos_repartidores(filtro_gastos)
        
        # 12b. Total Gastos de Cajero
        total_gastos_cajero = self.ds.get_total_gastos_cajero(filtro_gastos)
        
        # 13. Pago a Proveedores (desde SQLite)
        total_pago_proveedores = self.ds.get_total_pagos_proveedores(filtro_gastos)
        
        # 14. Préstamos (desde SQLite)
        total_prestamos = self.ds.get_total_prestamos(filtro_gastos)
        
        # 15. Ingresos extras y Salidas (DataStore - movimientos generales)
        total_ingresos = self.ds.get_total_ingresos_extras()
        total_salidas = self.ds.get_total_salidas()
        
        # 16. Pagos de Nómina (desde SQLite)
        total_pago_nomina = self.ds.get_total_pagos_nomina(filtro_gastos)
        
        # 17. Pagos a Socios (desde SQLite)
        total_pago_socios = self.ds.get_total_pagos_socios(filtro_gastos)
        
        # 18. Transferencias (desde SQLite)
        total_transferencias = self.ds.get_total_transferencias(filtro_gastos)
        
        # ═══════════════════════════════════════════════════════════════════════
        # CÁLCULOS FINALES
        # ═══════════════════════════════════════════════════════════════════════
        
        # Total Después de Ajustes = Efectivo - Ajustes de Precios  
        total_despues_ajustes = total_efectivo - total_ajustes
        
        # NETO A ENTREGAR = Total Después Ajustes + Ingresos - Gastos - Gastos Cajero - Pago Proveedores - Préstamos - Nómina - Socios - Transferencias - Salidas
        neto = total_despues_ajustes + total_ingresos - total_gastos - total_gastos_cajero - total_pago_proveedores - total_prestamos - total_pago_nomina - total_pago_socios - total_transferencias - total_salidas

        # ═══════════════════════════════════════════════════════════════════════
        # ACTUALIZAR LABELS
        # ═══════════════════════════════════════════════════════════════════════
        
        # COLUMNA 1: DESCUENTOS Y AJUSTES
        self.lbl_total_ajustes.config(text=f"${total_ajustes:,.2f}")
        self.lbl_total_gastos_liq.config(text=f"${total_gastos:,.2f}")
        self.lbl_total_gastos_cajero.config(text=f"${total_gastos_cajero:,.2f}")  # Gastos Cajero
        self.lbl_total_desc.config(text=f"${total_pago_proveedores:,.2f}")  # Pago Proveedores
        self.lbl_total_prestamos.config(text=f"${total_prestamos:,.2f}")  # Préstamos
        self.lbl_total_nomina_desc.config(text=f"${total_pago_nomina:,.2f}")  # Nómina
        self.lbl_total_socios_desc.config(text=f"${total_pago_socios:,.2f}")  # Socios
        self.lbl_total_transferencias_desc.config(text=f"${total_transferencias:,.2f}")  # Transferencias
        # Total Descuentos = Ajustes + Gastos + Gastos Cajero + Proveedores + Préstamos + Nómina + Socios + Transferencias
        total_descuentos_col2 = total_ajustes + total_gastos + total_gastos_cajero + total_pago_proveedores + total_prestamos + total_pago_nomina + total_pago_socios + total_transferencias
        self.lbl_total_devoluciones.config(text=f"${total_descuentos_col2:,.2f}")
        
        # COLUMNA 2: CUADRE GENERAL
        # Obtener Total Dinero Caja del corte cajero (si está disponible)
        total_dinero_caja = 0
        if hasattr(self, 'lbl_corte_total_dinero'):
            try:
                texto_dinero = self.lbl_corte_total_dinero.cget("text")
                total_dinero_caja = float(texto_dinero.replace("$", "").replace(",", ""))
            except (ValueError, AttributeError):
                total_dinero_caja = 0
        self.lbl_total_dinero_cuadre.config(text=f"${total_dinero_caja:,.2f}")
        
        # Para CUADRE GENERAL: calcular Total Descuentos de TODOS (sin filtro de repartidor)
        total_ajustes_general = self.ds.get_total_ajustes('')
        total_gastos_general = self.ds.get_total_gastos_repartidores('')
        total_gastos_cajero_general = self.ds.get_total_gastos_cajero('')
        total_pago_proveedores_general = self.ds.get_total_pagos_proveedores('')
        total_prestamos_general = self.ds.get_total_prestamos('')
        total_pago_nomina_general = self.ds.get_total_pagos_nomina()
        total_pago_socios_general = self.ds.get_total_pagos_socios()
        total_transferencias_general = self.ds.get_total_transferencias()
        
        total_descuentos_cuadre_general = (total_ajustes_general + total_gastos_general + 
                                           total_gastos_cajero_general + total_pago_proveedores_general + 
                                           total_prestamos_general + total_pago_nomina_general + 
                                           total_pago_socios_general + total_transferencias_general)
        
        # Créditos punteados general (sin filtro)
        total_creditos_punteados_general = 0
        if USE_SQLITE and self.ds.fecha:
            total_creditos_punteados_general = db_local.obtener_total_creditos_punteados(self.ds.fecha)
        
        # No entregados general (sin filtro)
        total_no_entregados_general = 0
        if USE_SQLITE and self.ds.fecha:
            total_no_entregados_general = db_local.obtener_total_no_entregados(self.ds.fecha)
        
        self.lbl_total_desc_cuadre.config(text=f"${total_descuentos_cuadre_general:,.2f}")
        self.lbl_total_creditos_punteados.config(text=f"${total_creditos_punteados_general:,.2f}")
        self.lbl_total_no_entregados.config(text=f"${total_no_entregados_general:,.2f}")
        # Total Efectivo Caja = Total Dinero Caja - Total Descuentos - Créditos Punteados - No Entregados
        total_efectivo_caja = total_dinero_caja - total_descuentos_cuadre_general - total_creditos_punteados_general - total_no_entregados_general
        self.lbl_total_efectivo_caja.config(text=f"${total_efectivo_caja:,.2f}",
                                             foreground="#2e7d32" if total_efectivo_caja >= 0 else "#c62828")
        self.lbl_total_credito.config(text=f"${total_credito:,.2f}")
        
        # CONTEO DE DINERO Y DIFERENCIA EN CUADRE GENERAL
        # Obtener total de conteo de dinero para el cuadre (SIN filtro, todos los repartidores)
        total_conteo_cuadre = self.ds.get_total_dinero('')  # Sin filtro para obtener el total general
        self.lbl_conteo_dinero_cuadre.config(text=f"${total_conteo_cuadre:,.2f}")
        
        # La diferencia se calculará después cuando se cargue el corte cajero
        # porque en este punto el TOTAL EFECTIVO CAJA aún no tiene el valor correcto
        
        # COLUMNA 3: CUADRE REPARTIDOR
        # Obtener total de conteo de dinero
        filtro_dinero = filtro if filtro and filtro not in ("(Todos)", "(Sin Asignar)") else ''
        total_conteo_dinero = self.ds.get_total_dinero(filtro_dinero)
        self.lbl_conteo_dinero_resultado.config(text=f"${total_conteo_dinero:,.2f}")
        
        # Monto Facturas: pintar el valor de la etiqueta "Monto Efectivo" de TOTALES
        # Forzar actualización de UI para asegurar que el valor esté sincronizado
        self.ventana.update_idletasks()
        monto_facturas_resultado = 0
        if hasattr(self, 'lbl_monto_efectivo_asign'):
            try:
                texto_monto = self.lbl_monto_efectivo_asign.cget("text")
                monto_facturas_resultado = float(texto_monto.replace("$", "").replace(",", ""))
            except (ValueError, AttributeError):
                monto_facturas_resultado = 0
        
        self.lbl_monto_facturas_resultado.config(text=f"${monto_facturas_resultado:,.2f}",
                                                  foreground="#2e7d32" if monto_facturas_resultado >= 0 else "#c62828")
        
        # Total Descuentos
        self.lbl_total_desc_resultado.config(text=f"${total_descuentos_col2:,.2f}")
        
        # Créditos Punteados
        self.lbl_creditos_punt_resultado.config(text=f"${total_creditos_punteados:,.2f}")
        
        # No Entregados
        self.lbl_no_entreg_resultado.config(text=f"${total_no_entregados:,.2f}")
        
        # TOTAL DINERO A ENTREGAR = Monto Facturas - Total Descuentos - Créditos Punteados - No Entregados
        total_dinero_entregar = monto_facturas_resultado - total_descuentos_col2 - total_creditos_punteados - total_no_entregados
        
        self.lbl_neto.config(text=f"${total_dinero_entregar:,.2f}",
                             foreground="#2e7d32" if total_dinero_entregar >= 0 else "#c62828")

        # Diferencia con dinero contado
        diferencia = total_conteo_dinero - total_dinero_entregar
        if abs(diferencia) < 0.01:
            self.lbl_diferencia_global.config(text="$0.00 ✓", foreground="#2e7d32")
        elif diferencia > 0:
            self.lbl_diferencia_global.config(text=f"+${diferencia:,.2f}", foreground="#1565c0")
        else:
            self.lbl_diferencia_global.config(text=f"${diferencia:,.2f}", foreground="#c62828")
        
        # Total Créditos Cobrados (abonos + pagos completos del día)
        if USE_SQLITE and self.ds.fecha:
            resultado_cobros = db_local.obtener_total_creditos_cobrados_fecha(self.ds.fecha)
            total_cobrado = resultado_cobros.get('total_cobrado', 0)
            self.lbl_total_creditos_cobrados.config(text=f"${total_cobrado:,.2f}")
        else:
            self.lbl_total_creditos_cobrados.config(text="$0.00")
        
        # Actualizar tablas de pagos y préstamos
        self._actualizar_tabla_pagos_proveedores()
        self._actualizar_tabla_prestamos()
        self._actualizar_combo_repartidores_prestamos()
        
        # Actualizar sección de Corte Cajero (Eleventa) - diferido para no bloquear GUI
        self.ventana.after(100, self._actualizar_corte_cajero_async)

    def _actualizar_tabla_pagos_proveedores(self):
        """Actualiza la tabla de pagos a proveedores."""
        if not hasattr(self, 'tree_pagos_prov'):
            return
        # Limpiar tabla
        for item in self.tree_pagos_prov.get_children():
            self.tree_pagos_prov.delete(item)
        
        # Cargar pagos del DataStore
        filtro = self.repartidor_filtro_var.get()
        pagos = self.ds.get_pagos_proveedores(filtro if filtro != "(Todos)" else '')
        
        for pago in pagos:
            self.tree_pagos_prov.insert(
                "", tk.END,
                values=(
                    pago.get('id', ''),
                    pago.get('proveedor', ''),
                    pago.get('concepto', ''),
                    f"${pago.get('monto', 0):,.2f}"
                ),
                tags=("pago",)
            )
    
    def _actualizar_tabla_prestamos(self):
        """Actualiza la tabla de préstamos (deshabilitada - movida a tab aparte)."""
        # La funcionalidad de préstamos ahora está en el tab de Préstamos
        pass
    
    def _actualizar_combo_repartidores_prestamos(self):
        """Actualiza el combobox de repartidores para préstamos (deshabilitada)."""
        # La funcionalidad ahora está en el tab de Préstamos
        pass
    
    def _actualizar_corte_cajero_async(self):
        """
        Versión asíncrona que ejecuta la consulta en un hilo separado
        para no bloquear la interfaz gráfica.
        Obtiene el corte COMBINADO de TODOS los turnos del día.
        """
        import threading
        
        def cargar_en_hilo():
            try:
                from corte_cajero import CorteCajeroManager
                import database_local as db
                
                # Usar la misma ruta FDB que el resto de la app (self.ruta_fdb)
                fdb_path = self.ruta_fdb
                print(f"[Corte Cajero] Usando FDB: {fdb_path}")
                
                if not fdb_path or not os.path.exists(fdb_path):
                    print(f"[Corte Cajero] ERROR: FDB no encontrado: {fdb_path}")
                    self.ventana.after(0, lambda: self._mostrar_error_corte_cajero("FDB no configurado"))
                    return
                
                manager = CorteCajeroManager(db_path=fdb_path)
                
                # Usar la fecha seleccionada o la actual
                fecha = self.ds.fecha if hasattr(self.ds, 'fecha') and self.ds.fecha else None
                print(f"[Corte Cajero] Fecha: {fecha}")
                
                if fecha:
                    # Primero intentar obtener turnos del día
                    print(f"[Corte Cajero] Buscando turnos para fecha: {fecha}")
                    turnos = manager.obtener_todos_turnos_por_fecha(fecha)
                    print(f"[Corte Cajero] Turnos encontrados: {turnos}")
                    
                    if turnos:
                        # Hay turnos para esta fecha, usar método por turno
                        print(f"[Corte Cajero] Obteniendo corte completo por turnos...")
                        corte = manager.obtener_corte_completo_por_fecha(fecha)
                        turno_id = turnos[-1]
                        num_turnos = len(turnos)
                    else:
                        # NO hay turnos para esta fecha - usar método por FECHA DE VENTAS
                        # Esto resuelve el caso donde el cajero dejó el turno abierto de un día anterior
                        print(f"[Corte Cajero] No hay turnos, usando método por fecha de ventas...")
                        corte = manager.obtener_corte_por_fecha_ventas(fecha)
                        turno_id = 0
                        num_turnos = 0
                    
                    if corte is None:
                        print(f"[Corte Cajero] No se encontraron datos para la fecha {fecha}")
                        self.ventana.after(0, self._limpiar_corte_cajero)
                        return
                else:
                    # Sin fecha, usar turno actual
                    turno_id = manager.obtener_turno_actual()
                    if turno_id is None:
                        turno_id = manager.obtener_ultimo_turno()
                    
                    if turno_id is None:
                        self.ventana.after(0, self._limpiar_corte_cajero)
                        return
                    
                    corte = manager.obtener_corte_por_turno(turno_id)
                    num_turnos = 1
                
                if corte is None:
                    self.ventana.after(0, self._limpiar_corte_cajero)
                    return
                
                # ═══════════════════════════════════════════════════════════
                # GUARDAR EN SQLite - Persistir los datos del corte cajero
                # ═══════════════════════════════════════════════════════════
                datos_guardar = {
                    'dinero_en_caja': {
                        'fondo_de_caja': corte.dinero_en_caja.fondo_de_caja,
                        'ventas_en_efectivo': corte.dinero_en_caja.ventas_en_efectivo,
                        'abonos_en_efectivo': corte.dinero_en_caja.abonos_en_efectivo,
                        'entradas': corte.dinero_en_caja.entradas,
                        'salidas': corte.dinero_en_caja.salidas,
                        'devoluciones_en_efectivo': corte.dinero_en_caja.devoluciones_en_efectivo,
                        'total': corte.dinero_en_caja.total
                    },
                    'ventas': {
                        'ventas_efectivo': corte.ventas.ventas_efectivo,
                        'ventas_tarjeta': corte.ventas.ventas_tarjeta,
                        'ventas_credito': corte.ventas.ventas_credito,
                        'ventas_vales': corte.ventas.ventas_vales,
                        'devoluciones_ventas': corte.ventas.devoluciones_ventas,
                        'total': corte.ventas.total
                    },
                    'devoluciones_por_forma_pago': corte.ventas.devoluciones_por_forma_pago,
                    'ganancia': corte.ganancia,
                    'num_turnos': num_turnos
                }

                db.guardar_corte_cajero(self.ds.fecha, turno_id, datos_guardar)

                # Guardar resumen de cancelaciones por usuario en SQLite
                from corte_cajero import obtener_cancelaciones_por_usuario
                resumen_cancel = obtener_cancelaciones_por_usuario(self.ds.fecha)
                db.guardar_cancelaciones_usuario(self.ds.fecha, resumen_cancel)

                # ═══════════════════════════════════════════════════════════
                # DETECTAR Y GUARDAR BUGS DE ELEVENTA
                # ═══════════════════════════════════════════════════════════
                try:
                    from utils_devoluciones import detectar_bugs_devoluciones
                    bugs_info = detectar_bugs_devoluciones(self.ds.fecha)
                    if bugs_info and bugs_info.get('total_bugs', 0) > 0:
                        db.guardar_bugs_eleventa_lote(self.ds.fecha, bugs_info.get('bugs', []))
                        print(f"[Bugs Eleventa] Detectados {len(bugs_info.get('bugs', []))} bugs, total: ${bugs_info.get('total_bugs', 0):,.2f}")
                    else:
                        # Limpiar bugs previos si ya no hay bugs
                        db.guardar_bugs_eleventa_lote(self.ds.fecha, [])
                        print(f"[Bugs Eleventa] Sin bugs detectados para {self.ds.fecha}")
                except Exception as bug_err:
                    print(f"⚠️ Error al detectar bugs Eleventa: {bug_err}")

                # Actualizar GUI en el hilo principal (pasar turno_id y num_turnos para mostrar)
                self.ventana.after(0, lambda t=turno_id, c=corte, n=num_turnos: self._aplicar_datos_corte(c, t, n))
                
            except Exception as e:
                import traceback
                print(f"⚠️ Error al cargar corte cajero: {e}")
                traceback.print_exc()
                # Mostrar el error en la interfaz
                self.ventana.after(0, lambda err=str(e): self._mostrar_error_corte_cajero(err))
                self.ventana.after(0, self._limpiar_corte_cajero)
        
        # Ejecutar en hilo separado
        hilo = threading.Thread(target=cargar_en_hilo, daemon=True)
        hilo.start()
    
    def _aplicar_datos_corte(self, corte, turno_id=None, num_turnos=1):
        """Aplica los datos del corte a los labels de la GUI."""
        try:
            # --- ACTUALIZAR LABELS DE DINERO EN CAJA (Liquidación) ---
            fondo_caja = corte.dinero_en_caja.fondo_de_caja
            self.lbl_corte_fondo_caja.config(text=f"${fondo_caja:,.2f}")
            self.lbl_corte_ventas_efectivo.config(text=f"${corte.dinero_en_caja.ventas_en_efectivo:,.2f}")
            self.lbl_corte_abonos_efectivo.config(text=f"${corte.dinero_en_caja.abonos_en_efectivo:,.2f}")
            self.lbl_corte_entradas.config(text=f"${corte.dinero_en_caja.entradas:,.2f}")
            
            # Total Efectivo = Ventas en Efectivo + Entradas
            total_efectivo = corte.dinero_en_caja.ventas_en_efectivo + corte.dinero_en_caja.entradas
            self.lbl_corte_total_efectivo.config(text=f"${total_efectivo:,.2f}")
            
            # Canceladas y Dev. Parciales del datastore (módulo asignación)
            total_canceladas_dia = self.ds.get_total_canceladas() if hasattr(self, 'ds') else 0
            dev_parciales_dia = 0
            if USE_SQLITE and self.ds.fecha:
                dev_parciales_dia = db_local.obtener_total_devoluciones_parciales_fecha(self.ds.fecha)
            
            self.lbl_corte_canceladas_dia.config(text=f"${total_canceladas_dia:,.2f}")
            self.lbl_corte_dev_parciales_dia.config(text=f"${dev_parciales_dia:,.2f}")
            
            # Labels ocultos para compatibilidad
            self.lbl_corte_salidas.config(text=f"${corte.dinero_en_caja.salidas:,.2f}")
            self.lbl_corte_dev_efectivo.config(text=f"${corte.dinero_en_caja.devoluciones_en_efectivo:,.2f}")
            
            # Total Dinero Caja = Fondo de Caja + Total Efectivo - Canceladas - Dev. Parciales
            total_dinero_caja = fondo_caja + total_efectivo - total_canceladas_dia - dev_parciales_dia
            self.lbl_corte_total_dinero.config(text=f"${total_dinero_caja:,.2f}")
            self.lbl_total_dinero_cuadre.config(text=f"${total_dinero_caja:,.2f}")
            
            # Recalcular TOTAL EFECTIVO CAJA en CUADRE GENERAL
            try:
                total_desc_texto = self.lbl_total_desc_cuadre.cget("text")
                total_descuentos = float(total_desc_texto.replace("$", "").replace(",", ""))
                total_cred_punt_texto = self.lbl_total_creditos_punteados.cget("text")
                total_creditos_punteados = float(total_cred_punt_texto.replace("$", "").replace(",", ""))
                total_efectivo_caja = total_dinero_caja - total_descuentos - total_creditos_punteados
                self.lbl_total_efectivo_caja.config(text=f"${total_efectivo_caja:,.2f}",
                                                     foreground="#2e7d32" if total_efectivo_caja >= 0 else "#c62828")
                
                # NUEVA LÓGICA DE DIFERENCIA FINAL
                # Leer conteo mostrado
                t_conteo = self.lbl_conteo_dinero_cuadre.cget("text")
                val_conteo = float(t_conteo.replace("$", "").replace(",", "").replace("✓", "").strip()) if t_conteo and t_conteo != "$0" else 0.0
                
                # Si el conteo está en 0, obtenerlo del datastore
                if val_conteo == 0:
                    val_conteo = self.ds.get_total_dinero('')
                
                # Diferencia = Conteo de Dinero - TOTAL EFECTIVO CAJA
                diferencia_final = val_conteo - total_efectivo_caja
                
                if abs(diferencia_final) < 0.01:
                    self.lbl_diferencia_cuadre.config(text="$0.00 ✓", foreground="#2e7d32")
                elif diferencia_final > 0:
                    self.lbl_diferencia_cuadre.config(text=f"+${diferencia_final:,.2f}", foreground="#1565c0")
                else:
                    self.lbl_diferencia_cuadre.config(text=f"-${abs(diferencia_final):,.2f}", foreground="#c62828")
            except (ValueError, AttributeError):
                pass
                pass
            
            # --- ACTUALIZAR LABELS DE VENTAS (Liquidación) ---
            self.lbl_corte_v_efectivo.config(text=f"${corte.ventas.ventas_efectivo:,.2f}")
            self.lbl_corte_v_tarjeta.config(text=f"${corte.ventas.ventas_tarjeta:,.2f}")
            
            # A Crédito = ventas a crédito + cancelaciones a crédito
            devs = corte.ventas.devoluciones_por_forma_pago
            total_credito = corte.ventas.ventas_credito + devs.get('credito', 0)
            self.lbl_corte_v_credito.config(text=f"${total_credito:,.2f}")
            
            self.lbl_corte_v_vales.config(text=f"${corte.ventas.ventas_vales:,.2f}")
            
            # Total Vendido = En Efectivo + A Crédito
            total_vendido = corte.ventas.ventas_efectivo + total_credito
            self.lbl_corte_total_ventas.config(text=f"${total_vendido:,.2f}")
            
            self.lbl_corte_dev_ventas.config(text=f"${corte.ventas.devoluciones_ventas:,.2f}")
            
            # Devoluciones Parciales y Canceladas (informativo - desde SQLite)
            dev_parc_otro_dia = 0
            dev_parc_no_reg = 0
            bug_dev_parc = 0  # Bug TURNOS > CORTE_MOV
            bug_duplicados = 0  # Bug duplicados
            cancel_no_form = 0  # Cancelaciones no formalizadas
            bug_dev_parc_ot = 0  # Bug Dev. Parc. de Otro Turno
            cancel_otro_dia = 0  # Canceladas de otro día
            total_eleventa = 0
            
            if USE_SQLITE and self.ds.fecha:
                total_dev_parciales = db_local.obtener_total_devoluciones_parciales_fecha(self.ds.fecha)
                self.lbl_corte_dev_parciales.config(text=f"${total_dev_parciales:,.2f}")
                # Canceladas = Devoluciones Totales - Parciales
                total_canceladas = corte.ventas.devoluciones_ventas - total_dev_parciales
                self.lbl_corte_canceladas.config(text=f"${total_canceladas:,.2f}")
                # Total Eleventa = Total Vendido - Devoluciones Ventas
                total_eleventa = total_vendido - corte.ventas.devoluciones_ventas
                self.lbl_corte_total_eleventa.config(text=f"${total_eleventa:,.2f}")
                
                # Obtener bugs de Eleventa ANTES de calcular ventas_netas
                bugs = db_local.obtener_bugs_eleventa_fecha(self.ds.fecha)
                for bug in bugs:
                    if bug['tipo_bug'] == 'turnos_mayor_corte':
                        bug_dev_parc += bug['monto_bug']
                    elif bug['tipo_bug'] == 'duplicado_corte':
                        bug_duplicados += bug['monto_bug']
                    elif bug['tipo_bug'] == 'cancelacion_no_formalizada':
                        cancel_no_form += bug['monto_bug']
                    elif bug['tipo_bug'] == 'dev_parc_otro_turno':
                        bug_dev_parc_ot += bug['monto_bug']
                
                # Actualizar labels de bugs
                if hasattr(self, 'lbl_corte_bug_dev_parc'):
                    self.lbl_corte_bug_dev_parc.config(text=f"${bug_dev_parc:,.2f}")
                if hasattr(self, 'lbl_corte_bug_duplicados'):
                    self.lbl_corte_bug_duplicados.config(text=f"${bug_duplicados:,.2f}")
                if hasattr(self, 'lbl_corte_cancel_no_form'):
                    self.lbl_corte_cancel_no_form.config(text=f"${cancel_no_form:,.2f}")
                if hasattr(self, 'lbl_corte_bug_dev_parc_ot'):
                    self.lbl_corte_bug_dev_parc_ot.config(text=f"${bug_dev_parc_ot:,.2f}")
                
                # Canceladas otro día (usar valor de DataSource - sección CANCELACIONES POR USUARIO)
                cancel_otro_dia = self.ds.get_total_canceladas_otro_dia()
                if hasattr(self, 'lbl_corte_cancel_otro_dia'):
                    self.lbl_corte_cancel_otro_dia.config(text=f"${cancel_otro_dia:,.2f}")
                
                # Dev. Parciales no registradas (ventas de hoy procesadas otro día - RESTA)
                dev_parc_no_reg, _ = db_local.obtener_dev_parciales_no_registradas(self.ds.fecha)
                self.lbl_corte_dev_parc_no_reg.config(text=f"${dev_parc_no_reg:,.2f}")
                # Dev. Parciales de otro día (procesadas hoy de ventas de otros días - SUMA)
                dev_parc_otro_dia = getattr(self, '_total_dev_parciales_otro_dia', 0) or 0
                self.lbl_corte_dev_parc_otro_dia.config(text=f"${dev_parc_otro_dia:,.2f}")
            
            # Total Ventas Netas = Total Eleventa + Bugs + Cancel Otro Día + Dev Parc Otro Día - Dev Parc No Reg
            ventas_netas = total_eleventa + bug_dev_parc + cancel_no_form + bug_dev_parc_ot + cancel_otro_dia + dev_parc_otro_dia - dev_parc_no_reg
            self.lbl_corte_ventas_netas.config(text=f"${ventas_netas:,.2f}")
            
            # NOTA: Ya no sincronizamos Monto Facturas con Total Ventas Netas de Eleventa
            # porque Monto Facturas = facturas NO canceladas del día (sin considerar cancelaciones de otro día)
            
            self.lbl_corte_ganancia.config(text=f"${corte.ganancia:,.2f}")
            
            # --- ACTUALIZAR EXPLICACIÓN DE DIFERENCIA EN CANCELACIONES ---
            self.lbl_corte_exp_dev_ef.config(text=f"${devs.get('efectivo', 0):,.2f}")
            self.lbl_corte_exp_dev_cr.config(text=f"${devs.get('credito', 0):,.2f}")
            self.lbl_corte_exp_dev_tar.config(text=f"${devs.get('tarjeta', 0):,.2f}")
            self.lbl_corte_exp_total_dev.config(text=f"${corte.ventas.devoluciones_ventas:,.2f}")
            
            # --- ACTUALIZAR RESUMEN EN MÓDULO DE ASIGNACIÓN ---
            if hasattr(self, 'lbl_corte_asign_dinero'):
                self.lbl_corte_asign_dinero.config(text=f"${corte.dinero_en_caja.total:,.2f}")
                self.lbl_corte_asign_ventas.config(text=f"${corte.ventas.total:,.2f}")
                self.lbl_corte_asign_ganancia.config(text=f"${corte.ganancia:,.2f}")
                self.lbl_corte_asign_devs.config(text=f"${corte.ventas.devoluciones_ventas:,.2f}")
                if turno_id and turno_id > 0:
                    # Mostrar número de turnos si hay más de 1
                    if num_turnos > 1:
                        self.lbl_corte_asign_turno.config(text=f"#{turno_id} ({num_turnos} turnos)")
                    else:
                        self.lbl_corte_asign_turno.config(text=f"#{turno_id}")
                else:
                    # Cuando no hay turno específico (corte por fecha de ventas)
                    self.lbl_corte_asign_turno.config(text="#POR FECHA")
            
            # --- ACTUALIZAR CANTIDAD DE TURNOS ---
            if hasattr(self, 'lbl_cantidad_turnos'):
                if num_turnos > 1:
                    self.lbl_cantidad_turnos.config(text=f"{num_turnos} turnos", foreground="#ff9800")  # Naranja si hay más de 1
                else:
                    self.lbl_cantidad_turnos.config(text=f"{num_turnos}", foreground="#1565c0")  # Azul normal
            
            # --- ACTUALIZAR CANCELACIONES POR USUARIO ---
            self._actualizar_cancelaciones_por_usuario(corte.dinero_en_caja.devoluciones_en_efectivo)
            
            # --- ACTUALIZAR TOTAL CRÉDITOS COBRADOS (abonos del día) ---
            if USE_SQLITE and self.ds.fecha:
                resultado_cobros = db_local.obtener_total_creditos_cobrados_fecha(self.ds.fecha)
                total_cobrado = resultado_cobros.get('total_cobrado', 0)
                self.lbl_total_creditos_cobrados.config(text=f"${total_cobrado:,.2f}")
                    
        except Exception as e:
            print(f"⚠️ Error al aplicar datos corte: {e}")
    
    def _actualizar_cancelaciones_por_usuario(self, devoluciones_efectivo=0.0):
        """Actualiza la sección de cancelaciones por usuario.
        
        CANCEL. ADMIN = viene de totales_cancelaciones_efectivo (cancelaciones hechas por ADMIN)
        CANCEL. CAJERO = devoluciones_efectivo - CANCEL. ADMIN (el resto lo hizo el cajero)
        
        Args:
            devoluciones_efectivo: Total de devoluciones en efectivo del corte de caja
        """
        try:
            import database_local as db
            # Limpiar labels anteriores
            for widget in self.frame_cancel_usuarios.winfo_children():
                widget.destroy()
            
            # Obtener fecha
            fecha = self.ds.fecha if hasattr(self.ds, 'fecha') and self.ds.fecha else None
            if not fecha:
                return
            
            # Obtener CANCEL. ADMIN desde SQLite (cancelaciones hechas por ADMIN)
            totales_efectivo = db.obtener_totales_cancelaciones_efectivo(fecha)
            
            # CANCEL. ADMIN
            datos_admin = totales_efectivo.get('ADMIN', {})
            cancel_admin = datos_admin.get('total', 0.0)
            num_admin = datos_admin.get('num', 0)
            
            # CANCEL. CAJERO = Devoluciones Efectivo - CANCEL. ADMIN
            cancel_cajero = devoluciones_efectivo - cancel_admin
            if cancel_cajero < 0:
                cancel_cajero = 0.0
            
            # Número de facturas CAJERO desde SQLite (si existe)
            datos_cajero = totales_efectivo.get('CAJERO', {})
            num_cajero = datos_cajero.get('num', 0)
            
            # Si no hay registro en SQLite para CAJERO, calcular desde devoluciones
            if num_cajero == 0 and cancel_cajero > 0:
                for dev in self.ds.devoluciones:
                    cajero = dev.get('cajero', '').upper()
                    if cajero and cajero != 'ADMIN':
                        num_cajero += 1
            
            row = 0
            
            # Mostrar CANCEL. ADMIN con conteo
            ttk.Label(self.frame_cancel_usuarios, text=f"CANCEL. ADMIN ({num_admin}):", 
                     foreground="#e57373", font=("Segoe UI", 9, "bold")).grid(row=row, column=0, sticky=tk.W)
            ttk.Label(self.frame_cancel_usuarios, text=f"${cancel_admin:,.2f}", 
                     font=("Segoe UI", 9, "bold"), foreground="#e57373").grid(row=row, column=1, sticky=tk.E, padx=(10, 0))
            row += 1
            
            # Mostrar CANCEL. CAJERO con conteo
            ttk.Label(self.frame_cancel_usuarios, text=f"CANCEL. CAJERO ({num_cajero}):", 
                     foreground="#ffb74d", font=("Segoe UI", 9, "bold")).grid(row=row, column=0, sticky=tk.W)
            ttk.Label(self.frame_cancel_usuarios, text=f"${cancel_cajero:,.2f}", 
                     font=("Segoe UI", 9, "bold"), foreground="#ffb74d").grid(row=row, column=1, sticky=tk.E, padx=(10, 0))
            row += 1
            
            # Separador
            ttk.Separator(self.frame_cancel_usuarios, orient="horizontal").grid(
                row=row, column=0, columnspan=2, sticky="ew", pady=5)
            row += 1
            
            # Total Cancelaciones (debe ser igual a Devoluciones Efectivo)
            total_facturas = num_admin + num_cajero
            ttk.Label(self.frame_cancel_usuarios, text=f"= Total Cancelaciones ({total_facturas}):", 
                     foreground="#ffffff", font=("Segoe UI", 9, "bold")).grid(row=row, column=0, sticky=tk.W)
            ttk.Label(self.frame_cancel_usuarios, text=f"${devoluciones_efectivo:,.2f}", 
                     font=("Segoe UI", 9, "bold"), foreground="#ff8a80").grid(row=row, column=1, sticky=tk.E, padx=(10, 0))
            
            # Actualizar Canceladas otro día (informativo, TOTAL no subtotal)
            total_canceladas_otro_dia = self.ds.get_total_canceladas_otro_dia()
            if hasattr(self, 'lbl_total_canceladas_otro_dia'):
                self.lbl_total_canceladas_otro_dia.config(text=f"${total_canceladas_otro_dia:,.2f}")
                
        except Exception as e:
            print(f"⚠️ Error al actualizar cancelaciones por usuario: {e}")
    
    def _actualizar_corte_cajero(self):
        """
        Actualiza los datos del Corte Cajero desde Eleventa (Firebird).
        Obtiene información en tiempo real del turno actual.
        """
        try:
            # Importar el módulo de corte cajero
            from corte_cajero import CorteCajeroManager
            
            manager = CorteCajeroManager()
            
            # Obtener turno actual o el último con datos
            turno_id = manager.obtener_turno_actual()
            if turno_id is None:
                turno_id = manager.obtener_ultimo_turno()
            
            if turno_id is None:
                # No hay datos - mostrar ceros
                self._limpiar_corte_cajero()
                return
            
            # Obtener datos del turno
            corte = manager.obtener_corte_por_turno(turno_id)
            
            # Si el turno actual no tiene ventas, intentar con el anterior
            if corte.ventas.ventas_efectivo == 0 and turno_id > 1:
                turno_id_anterior = turno_id - 1
                corte_anterior = manager.obtener_corte_por_turno(turno_id_anterior)
                if corte_anterior.ventas.ventas_efectivo > 0:
                    corte = corte_anterior
            
            # --- ACTUALIZAR LABELS DE DINERO EN CAJA ---
            fondo_caja = corte.dinero_en_caja.fondo_de_caja
            self.lbl_corte_fondo_caja.config(text=f"${fondo_caja:,.2f}")
            self.lbl_corte_ventas_efectivo.config(text=f"${corte.dinero_en_caja.ventas_en_efectivo:,.2f}")
            self.lbl_corte_abonos_efectivo.config(text=f"${corte.dinero_en_caja.abonos_en_efectivo:,.2f}")
            self.lbl_corte_entradas.config(text=f"${corte.dinero_en_caja.entradas:,.2f}")
            
            # Total Efectivo = Ventas en Efectivo + Entradas
            total_efectivo = corte.dinero_en_caja.ventas_en_efectivo + corte.dinero_en_caja.entradas
            self.lbl_corte_total_efectivo.config(text=f"${total_efectivo:,.2f}")
            
            # Canceladas y Dev. Parciales del datastore (módulo asignación)
            total_canceladas_dia = self.ds.get_total_canceladas() if hasattr(self, 'ds') else 0
            dev_parciales_dia = 0
            if USE_SQLITE and self.ds.fecha:
                dev_parciales_dia = db_local.obtener_total_devoluciones_parciales_fecha(self.ds.fecha)
            
            self.lbl_corte_canceladas_dia.config(text=f"${total_canceladas_dia:,.2f}")
            self.lbl_corte_dev_parciales_dia.config(text=f"${dev_parciales_dia:,.2f}")
            
            # Labels ocultos para compatibilidad
            self.lbl_corte_salidas.config(text=f"${corte.dinero_en_caja.salidas:,.2f}")
            self.lbl_corte_dev_efectivo.config(text=f"${corte.dinero_en_caja.devoluciones_en_efectivo:,.2f}")
            
            # Total Dinero Caja = Fondo de Caja + Total Efectivo - Canceladas - Dev. Parciales
            total_dinero_caja = fondo_caja + total_efectivo - total_canceladas_dia - dev_parciales_dia
            self.lbl_corte_total_dinero.config(text=f"${total_dinero_caja:,.2f}")
            self.lbl_total_dinero_cuadre.config(text=f"${total_dinero_caja:,.2f}")
            
            # Recalcular TOTAL EFECTIVO CAJA en CUADRE GENERAL
            try:
                total_desc_texto = self.lbl_total_desc_cuadre.cget("text")
                total_descuentos = float(total_desc_texto.replace("$", "").replace(",", ""))
                total_cred_punt_texto = self.lbl_total_creditos_punteados.cget("text")
                total_creditos_punteados = float(total_cred_punt_texto.replace("$", "").replace(",", ""))
                total_efectivo_caja = total_dinero_caja - total_descuentos - total_creditos_punteados
                self.lbl_total_efectivo_caja.config(text=f"${total_efectivo_caja:,.2f}",
                                                     foreground="#2e7d32" if total_efectivo_caja >= 0 else "#c62828")
                
                # NUEVA LÓGICA DE DIFERENCIA FINAL
                t_conteo = self.lbl_conteo_dinero_cuadre.cget("text")
                val_conteo = float(t_conteo.replace("$", "").replace(",", "").replace("✓", "").strip()) if t_conteo and t_conteo != "$0" else 0.0
                
                # Si el conteo está en 0, obtenerlo del datastore
                if val_conteo == 0:
                    val_conteo = self.ds.get_total_dinero('')
                
                # Diferencia = Conteo de Dinero - TOTAL EFECTIVO CAJA
                diferencia_final = val_conteo - total_efectivo_caja
                
                if abs(diferencia_final) < 0.01:
                    self.lbl_diferencia_cuadre.config(text="$0.00 ✓", foreground="#2e7d32")
                elif diferencia_final > 0:
                    self.lbl_diferencia_cuadre.config(text=f"+${diferencia_final:,.2f}", foreground="#1565c0")
                else:
                    self.lbl_diferencia_cuadre.config(text=f"-${abs(diferencia_final):,.2f}", foreground="#c62828")
            except (ValueError, AttributeError):
                pass
            
            # --- ACTUALIZAR LABELS DE VENTAS ---
            self.lbl_corte_v_efectivo.config(text=f"${corte.ventas.ventas_efectivo:,.2f}")
            self.lbl_corte_v_tarjeta.config(text=f"${corte.ventas.ventas_tarjeta:,.2f}")
            
            # A Crédito = ventas a crédito + cancelaciones a crédito
            devs = corte.ventas.devoluciones_por_forma_pago
            total_credito = corte.ventas.ventas_credito + devs.get('credito', 0)
            self.lbl_corte_v_credito.config(text=f"${total_credito:,.2f}")
            
            self.lbl_corte_v_vales.config(text=f"${corte.ventas.ventas_vales:,.2f}")
            
            # Total Vendido = En Efectivo + A Crédito
            total_vendido = corte.ventas.ventas_efectivo + total_credito
            self.lbl_corte_total_ventas.config(text=f"${total_vendido:,.2f}")
            
            self.lbl_corte_dev_ventas.config(text=f"${corte.ventas.devoluciones_ventas:,.2f}")
            
            # Devoluciones Parciales y Canceladas (informativo - desde SQLite)
            dev_parc_otro_dia = 0
            dev_parc_no_reg = 0
            bug_dev_parc = 0  # Bug TURNOS > CORTE_MOV
            bug_duplicados = 0  # Bug duplicados
            cancel_no_form = 0  # Cancelaciones no formalizadas
            bug_dev_parc_ot = 0  # Bug Dev. Parc. de Otro Turno
            cancel_otro_dia = 0  # Canceladas de otro día
            total_eleventa = 0
            
            if USE_SQLITE and self.ds.fecha:
                total_dev_parciales = db_local.obtener_total_devoluciones_parciales_fecha(self.ds.fecha)
                self.lbl_corte_dev_parciales.config(text=f"${total_dev_parciales:,.2f}")
                # Canceladas = Devoluciones Totales - Parciales
                total_canceladas = corte.ventas.devoluciones_ventas - total_dev_parciales
                self.lbl_corte_canceladas.config(text=f"${total_canceladas:,.2f}")
                # Total Eleventa = Total Vendido - Devoluciones Ventas
                total_eleventa = total_vendido - corte.ventas.devoluciones_ventas
                self.lbl_corte_total_eleventa.config(text=f"${total_eleventa:,.2f}")
                
                # Obtener bugs de Eleventa ANTES de calcular ventas_netas
                bugs = db_local.obtener_bugs_eleventa_fecha(self.ds.fecha)
                for bug in bugs:
                    if bug['tipo_bug'] == 'turnos_mayor_corte':
                        bug_dev_parc += bug['monto_bug']
                    elif bug['tipo_bug'] == 'duplicado_corte':
                        bug_duplicados += bug['monto_bug']
                    elif bug['tipo_bug'] == 'cancelacion_no_formalizada':
                        cancel_no_form += bug['monto_bug']
                    elif bug['tipo_bug'] == 'dev_parc_otro_turno':
                        bug_dev_parc_ot += bug['monto_bug']
                
                # Actualizar labels de bugs
                if hasattr(self, 'lbl_corte_bug_dev_parc'):
                    self.lbl_corte_bug_dev_parc.config(text=f"${bug_dev_parc:,.2f}")
                if hasattr(self, 'lbl_corte_bug_duplicados'):
                    self.lbl_corte_bug_duplicados.config(text=f"${bug_duplicados:,.2f}")
                if hasattr(self, 'lbl_corte_cancel_no_form'):
                    self.lbl_corte_cancel_no_form.config(text=f"${cancel_no_form:,.2f}")
                if hasattr(self, 'lbl_corte_bug_dev_parc_ot'):
                    self.lbl_corte_bug_dev_parc_ot.config(text=f"${bug_dev_parc_ot:,.2f}")
                
                # Canceladas otro día (usar valor de DataSource - sección CANCELACIONES POR USUARIO)
                cancel_otro_dia = self.ds.get_total_canceladas_otro_dia()
                if hasattr(self, 'lbl_corte_cancel_otro_dia'):
                    self.lbl_corte_cancel_otro_dia.config(text=f"${cancel_otro_dia:,.2f}")
                
                # Dev. Parciales no registradas (ventas de hoy procesadas otro día - RESTA)
                dev_parc_no_reg, _ = db_local.obtener_dev_parciales_no_registradas(self.ds.fecha)
                self.lbl_corte_dev_parc_no_reg.config(text=f"${dev_parc_no_reg:,.2f}")
                # Dev. Parciales de otro día (procesadas hoy de ventas de otros días - SUMA)
                dev_parc_otro_dia = getattr(self, '_total_dev_parciales_otro_dia', 0) or 0
                self.lbl_corte_dev_parc_otro_dia.config(text=f"${dev_parc_otro_dia:,.2f}")
            
            # Total Ventas Netas = Total Eleventa + Bugs + Cancel Otro Día + Dev Parc Otro Día - Dev Parc No Reg
            ventas_netas = total_eleventa + bug_dev_parc + cancel_no_form + bug_dev_parc_ot + cancel_otro_dia + dev_parc_otro_dia - dev_parc_no_reg
            self.lbl_corte_ventas_netas.config(text=f"${ventas_netas:,.2f}")
            
            # NOTA: Ya no sincronizamos Monto Facturas con Total Ventas Netas de Eleventa
            # porque Monto Facturas = facturas NO canceladas del día (sin considerar cancelaciones de otro día)
            
            self.lbl_corte_ganancia.config(text=f"${corte.ganancia:,.2f}")
            
            # --- ACTUALIZAR EXPLICACIÓN DE DIFERENCIA EN CANCELACIONES ---
            self.lbl_corte_exp_dev_ef.config(text=f"${devs.get('efectivo', 0):,.2f}")
            self.lbl_corte_exp_dev_cr.config(text=f"${devs.get('credito', 0):,.2f}")
            self.lbl_corte_exp_dev_tar.config(text=f"${devs.get('tarjeta', 0):,.2f}")
            self.lbl_corte_exp_total_dev.config(text=f"${corte.ventas.devoluciones_ventas:,.2f}")
            
        except FileNotFoundError as e:
            # Si no se encuentra la base de datos o isql
            self._limpiar_corte_cajero()
            print(f"⚠️ No se pudo cargar corte cajero: {e}")
        except Exception as e:
            # Cualquier otro error
            self._limpiar_corte_cajero()
            print(f"⚠️ Error al actualizar corte cajero: {e}")
    
    def _limpiar_corte_cajero(self):
        """Limpia todos los labels del corte cajero (pone en $0)."""
        labels_dinero = [
            'lbl_corte_fondo_caja', 'lbl_corte_ventas_efectivo', 'lbl_corte_abonos_efectivo',
            'lbl_corte_entradas', 'lbl_corte_total_efectivo', 'lbl_corte_salidas', 'lbl_corte_dev_efectivo', 'lbl_corte_total_dinero',
            'lbl_corte_canceladas_dia', 'lbl_corte_dev_parciales_dia'
        ]
        labels_ventas = [
            'lbl_corte_v_efectivo', 'lbl_corte_v_tarjeta', 'lbl_corte_v_credito',
            'lbl_corte_v_vales', 'lbl_corte_dev_ventas', 'lbl_corte_total_ventas', 
            'lbl_corte_ventas_netas', 'lbl_corte_ganancia', 'lbl_corte_canceladas', 'lbl_corte_dev_parciales',
            'lbl_corte_dev_parc_otro_dia', 'lbl_corte_bug_dev_parc', 'lbl_corte_bug_duplicados', 'lbl_corte_cancel_no_form',
            'lbl_corte_total_eleventa', 'lbl_corte_cancel_otro_dia', 'lbl_corte_dev_parc_no_reg'
        ]
        labels_exp = [
            'lbl_corte_exp_dev_ef', 'lbl_corte_exp_dev_cr', 'lbl_corte_exp_dev_tar', 'lbl_corte_exp_total_dev'
        ]
        # Labels del resumen en Asignación
        labels_asign = [
            'lbl_corte_asign_dinero', 'lbl_corte_asign_ventas', 'lbl_corte_asign_ganancia', 'lbl_corte_asign_devs'
        ]
        
        for lbl_name in labels_dinero + labels_ventas + labels_exp + labels_asign:
            if hasattr(self, lbl_name):
                getattr(self, lbl_name).config(text="$0.00")
        
        # Limpiar turno
        if hasattr(self, 'lbl_corte_asign_turno'):
            self.lbl_corte_asign_turno.config(text="#---")
        
        # Limpiar cancelaciones por usuario
        if hasattr(self, 'frame_cancel_usuarios'):
            for widget in self.frame_cancel_usuarios.winfo_children():
                widget.destroy()
        if hasattr(self, 'lbl_corte_total_cancelaciones'):
            self.lbl_corte_total_cancelaciones.config(text="$0.00")
        if hasattr(self, 'lbl_corte_num_cancelaciones'):
            self.lbl_corte_num_cancelaciones.config(text="0")
    
    def _mostrar_error_corte_cajero(self, error_msg: str):
        """Muestra un error relacionado con el Corte Cajero."""
        self._limpiar_corte_cajero()
        # Mostrar el error en el label de turno para que el usuario sepa
        if hasattr(self, 'lbl_corte_asign_turno'):
            if "No se encontró la base de datos" in error_msg:
                self.lbl_corte_asign_turno.config(text="#ERROR FDB")
            elif "No se encontró isql" in error_msg:
                self.lbl_corte_asign_turno.config(text="#NO ISQL")
            else:
                self.lbl_corte_asign_turno.config(text="#ERROR")

    def _mostrar_modal_dev_parciales_otro_dia(self):
        """Muestra un modal con los folios de facturas de otros días con devoluciones parciales."""
        if not hasattr(self, '_facturas_dev_parciales_otro_dia') or not self._facturas_dev_parciales_otro_dia:
            messagebox.showinfo("Sin datos", 
                               "No hay devoluciones parciales de facturas de otros días.",
                               parent=self.ventana)
            return
        
        # Crear ventana modal
        dialog = tk.Toplevel(self.ventana)
        dialog.title("📦 Devoluciones Parciales de Otro Día")
        dialog.geometry("500x350")
        dialog.transient(self.ventana)
        dialog.grab_set()
        
        # Centrar ventana
        dialog.update_idletasks()
        x = self.ventana.winfo_x() + (self.ventana.winfo_width() - dialog.winfo_width()) // 2
        y = self.ventana.winfo_y() + (self.ventana.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Frame principal
        frame = ttk.Frame(dialog, padding=15)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Título explicativo
        ttk.Label(frame, text="Estas son devoluciones de facturas de OTROS días\nque fueron procesadas en la fecha seleccionada:",
                  font=("Segoe UI", 10), foreground="#6a1b9a").pack(pady=(0, 10))
        
        # Treeview para mostrar los folios
        columns = ("folio", "fecha_factura", "total_devuelto")
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=10)
        
        tree.heading("folio", text="Folio")
        tree.heading("fecha_factura", text="Fecha Factura")
        tree.heading("total_devuelto", text="Total Devuelto")
        
        tree.column("folio", width=100, anchor=tk.CENTER)
        tree.column("fecha_factura", width=150, anchor=tk.CENTER)
        tree.column("total_devuelto", width=150, anchor=tk.E)
        
        # Agregar scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Llenar datos
        total_general = 0
        for factura in self._facturas_dev_parciales_otro_dia:
            tree.insert("", tk.END, values=(
                factura['folio'],
                factura['fecha_factura'],
                f"${factura['total_devuelto']:,.2f}"
            ))
            total_general += factura['total_devuelto']
        
        # Frame para total
        frame_total = ttk.Frame(dialog, padding=(15, 5))
        frame_total.pack(fill=tk.X)
        
        ttk.Label(frame_total, text="TOTAL:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        ttk.Label(frame_total, text=f"${total_general:,.2f}", 
                  font=("Segoe UI", 11, "bold"), foreground="#6a1b9a").pack(side=tk.RIGHT)
        
        # Nota explicativa
        frame_nota = ttk.Frame(dialog, padding=(15, 5))
        frame_nota.pack(fill=tk.X)
        ttk.Label(frame_nota, 
                  text="⚠️ Este monto NO está incluido en 'Monto Facturas' porque las facturas son de otros días.",
                  font=("Segoe UI", 8), foreground="#795548").pack()
        
        # Botón cerrar
        ttk.Button(dialog, text="Cerrar", command=dialog.destroy).pack(pady=10)

    def _mostrar_detalle_bugs_eleventa(self):
        """Muestra un modal con los detalles de bugs de Eleventa detectados."""
        if not hasattr(self, 'ds') or not self.ds.fecha:
            messagebox.showinfo("Sin datos", 
                               "No hay fecha seleccionada.",
                               parent=self.ventana)
            return
        
        try:
            import database_local as db_local
            
            # Obtener bugs guardados en SQLite
            bugs = db_local.obtener_bugs_eleventa_fecha(self.ds.fecha)
            
            if not bugs:
                messagebox.showinfo("Sin bugs", 
                                   f"No se detectaron bugs de Eleventa para {self.ds.fecha}",
                                   parent=self.ventana)
                return
            
            # Crear ventana modal
            dialog = tk.Toplevel(self.ventana)
            dialog.title("🐛 Bugs Eleventa Detectados")
            dialog.geometry("650x450")
            dialog.transient(self.ventana)
            dialog.grab_set()
            
            # Centrar ventana
            dialog.update_idletasks()
            x = self.ventana.winfo_x() + (self.ventana.winfo_width() - dialog.winfo_width()) // 2
            y = self.ventana.winfo_y() + (self.ventana.winfo_height() - dialog.winfo_height()) // 2
            dialog.geometry(f"+{x}+{y}")
            
            # Frame principal
            frame = ttk.Frame(dialog, padding=15)
            frame.pack(fill=tk.BOTH, expand=True)
            
            # Título explicativo
            ttk.Label(frame, 
                      text="Estos son errores detectados en Eleventa que causan\ndiscrepancias en los cálculos de devoluciones:",
                      font=("Segoe UI", 10), foreground="#c62828").pack(pady=(0, 10))
            
            # Treeview para mostrar los bugs
            columns = ("turno", "tipo", "descripcion", "monto")
            tree = ttk.Treeview(frame, columns=columns, show="headings", height=12)
            
            tree.heading("turno", text="Turno")
            tree.heading("tipo", text="Tipo Bug")
            tree.heading("descripcion", text="Descripción")
            tree.heading("monto", text="Monto Bug")
            
            tree.column("turno", width=60, anchor=tk.CENTER)
            tree.column("tipo", width=130, anchor=tk.CENTER)
            tree.column("descripcion", width=300, anchor=tk.W)
            tree.column("monto", width=100, anchor=tk.E)
            
            # Agregar scrollbar
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Llenar datos
            total_bugs = 0
            for bug in bugs:
                tipo_display = bug['tipo_bug'].replace('_', ' ').title()
                tree.insert("", tk.END, values=(
                    f"#{bug['turno_id']}" if bug['turno_id'] else "N/A",
                    tipo_display,
                    bug['descripcion'][:50] + "..." if len(bug['descripcion']) > 50 else bug['descripcion'],
                    f"${bug['monto_bug']:,.2f}"
                ))
                total_bugs += bug['monto_bug']
            
            # Frame para total
            frame_total = ttk.Frame(dialog, padding=(15, 5))
            frame_total.pack(fill=tk.X)
            
            ttk.Label(frame_total, text="TOTAL BUGS:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
            ttk.Label(frame_total, text=f"${total_bugs:,.2f}", 
                      font=("Segoe UI", 11, "bold"), foreground="#c62828").pack(side=tk.RIGHT)
            
            # Nota explicativa
            frame_nota = ttk.Frame(dialog, padding=(15, 5))
            frame_nota.pack(fill=tk.X)
            ttk.Label(frame_nota, 
                      text="⚠️ Este monto representa errores de Eleventa. El valor real de devoluciones\nes menor al reportado por TURNOS debido a estos bugs.",
                      font=("Segoe UI", 8), foreground="#795548").pack()
            
            # Botón cerrar
            ttk.Button(dialog, text="Cerrar", command=dialog.destroy).pack(pady=10)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", 
                                f"Error al cargar bugs: {e}",
                                parent=self.ventana)

    def _agregar_pago_proveedor(self):
        """Agrega un nuevo pago a proveedor (deshabilitada - funcionalidad movida a tab Gastos)."""
        # La funcionalidad de pago a proveedores ahora está en el tab de Gastos
        pass
    
    def _eliminar_pago_proveedor(self):
        """Elimina el pago a proveedor seleccionado (deshabilitada)."""
        pass
    
    def _agregar_prestamo(self):
        """Agrega un nuevo préstamo (deshabilitada - funcionalidad movida a tab Préstamos)."""
        # La funcionalidad de préstamos ahora está en el tab de Préstamos
        pass
    
    def _eliminar_prestamo(self):
        """Elimina el préstamo seleccionado (deshabilitada)."""
        pass

    def _actualizar_diferencia(self, neto: float):
        """Calcula diferencia usando totales de dinero del DataStore."""
        filtro = self.repartidor_filtro_var.get()
        total_dinero = self.ds.get_total_dinero(
            filtro if filtro and filtro != "(Todos)" else ''
        )
        diferencia = total_dinero - neto

        self.lbl_dinero_contado.config(text=f"${total_dinero:,.2f}")
        
        # Actualizar label de diferencia en el resumen
        if abs(diferencia) < 0.01:
            self.lbl_diferencia.config(text="$0.00  ✓", foreground="#2e7d32")
            self.lbl_diferencia_global.config(text="$0.00 ✓", foreground="#2e7d32")
        elif diferencia > 0:
            self.lbl_diferencia.config(text=f"+${diferencia:,.2f}", foreground="#1565c0")
            self.lbl_diferencia_global.config(text=f"+${diferencia:,.2f}", foreground="#1565c0")
        else:
            self.lbl_diferencia.config(text=f"${diferencia:,.2f}", foreground="#c62828")
            self.lbl_diferencia_global.config(text=f"${diferencia:,.2f}", foreground="#c62828")

    # ==================================================================
    # PESTAÑA 2 – DESCUENTOS POR FACTURA
    # ==================================================================
    def _crear_tab_descuentos(self):
        # Contenedor principal con PanedWindow horizontal
        paned_main = ttk.PanedWindow(self.tab_descuentos, orient=tk.HORIZONTAL)
        paned_main.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # === PANEL IZQUIERDO: Búsqueda y Datos Factura ===
        frame_izq = ttk.Frame(paned_main)
        paned_main.add(frame_izq, weight=1)
        
        # --- BÚSQUEDA ---
        frame_buscar = ttk.LabelFrame(frame_izq, text="🔍 BUSCAR FACTURA", padding=8)
        frame_buscar.pack(fill=tk.X, padx=5, pady=(5, 3))

        # Fila búsqueda
        frame_row1 = ttk.Frame(frame_buscar)
        frame_row1.pack(fill=tk.X)
        
        ttk.Label(frame_row1, text="Buscar:").pack(side=tk.LEFT)
        self.buscar_var = tk.StringVar()
        self.entry_buscar = ttk.Entry(frame_row1, textvariable=self.buscar_var, width=20)
        self.entry_buscar.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
        ttk.Button(frame_row1, text="✕", width=3,
                   command=self._limpiar_buscar).pack(side=tk.LEFT)

        # Fila filtros
        frame_row2 = ttk.Frame(frame_buscar)
        frame_row2.pack(fill=tk.X, pady=(4, 0))
        
        ttk.Label(frame_row2, text="Folio:").pack(side=tk.LEFT)
        self.folio_var = tk.StringVar()
        self.folio_combo = ttk.Combobox(frame_row2, textvariable=self.folio_var,
                                        width=8, state="readonly")
        self.folio_combo.pack(side=tk.LEFT, padx=4)
        self.folio_combo.bind("<<ComboboxSelected>>", self._on_folio_combo_seleccionado)
        
        ttk.Label(frame_row2, text="Rep:").pack(side=tk.LEFT, padx=(8, 0))
        self.filtro_rep_var = tk.StringVar(value="Todos")
        self.combo_filtro_rep = ttk.Combobox(frame_row2, textvariable=self.filtro_rep_var,
                                             width=12, state="readonly")
        self.combo_filtro_rep.pack(side=tk.LEFT, padx=4)
        self.combo_filtro_rep.bind("<<ComboboxSelected>>", lambda e: self._filtrar_resultados())

        # Treeview resultados búsqueda
        frame_resultados = ttk.Frame(frame_buscar)
        frame_resultados.pack(fill=tk.BOTH, pady=(6, 0), expand=True)
        
        tree_bus_container = ttk.Frame(frame_resultados)
        tree_bus_container.pack(fill=tk.BOTH, expand=True)

        self.tree_buscar = ttk.Treeview(
            tree_bus_container,
            columns=("folio", "nombre", "subtotal", "repartidor"),
            selectmode="extended", height=4, show="headings"
        )
        self.tree_buscar.column("folio",      anchor=tk.CENTER, width=60,  minwidth=50)
        self.tree_buscar.column("nombre",     anchor=tk.W,      width=150, minwidth=100)
        self.tree_buscar.column("subtotal",   anchor=tk.E,      width=80,  minwidth=60)
        self.tree_buscar.column("repartidor", anchor=tk.CENTER, width=80,  minwidth=60)

        self.tree_buscar.heading("folio",      text="Folio")
        self.tree_buscar.heading("nombre",     text="Cliente")
        self.tree_buscar.heading("subtotal",   text="Subtotal")
        self.tree_buscar.heading("repartidor", text="Rep.")

        self.tree_buscar.tag_configure("par",   background="#2d2d2d", foreground="#ffffff")
        self.tree_buscar.tag_configure("impar", background="#3d3d3d", foreground="#ffffff")

        scroll_bus_y = ttk.Scrollbar(tree_bus_container, orient=tk.VERTICAL, command=self.tree_buscar.yview)
        self.tree_buscar.configure(yscrollcommand=scroll_bus_y.set)
        
        self.tree_buscar.grid(row=0, column=0, sticky="nsew")
        scroll_bus_y.grid(row=0, column=1, sticky="ns")
        tree_bus_container.grid_rowconfigure(0, weight=1)
        tree_bus_container.grid_columnconfigure(0, weight=1)
        
        self.tree_buscar.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(self.tree_buscar))
        self.tree_buscar.bind("<<TreeviewSelect>>", self._on_tree_buscar_select)
        self.buscar_var.trace_add("write", lambda *a: self._filtrar_resultados())

        # --- DATOS FACTURA ---
        frame_fac = ttk.LabelFrame(frame_izq, text="🧾 FACTURA SELECCIONADA", padding=8)
        frame_fac.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        # Info básica en grid compacto
        info_frame = ttk.Frame(frame_fac)
        info_frame.pack(fill=tk.X)
        
        ttk.Label(info_frame, text="Cliente:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky=tk.W)
        self.lbl_cliente_desc = ttk.Label(info_frame, text="—", wraplength=200)
        self.lbl_cliente_desc.grid(row=0, column=1, sticky=tk.W, padx=5)

        ttk.Label(info_frame, text="Subtotal:", font=("Segoe UI", 9, "bold")).grid(row=1, column=0, sticky=tk.W)
        self.lbl_total_desc_fac = ttk.Label(info_frame, text="$0", foreground="#4CAF50")
        self.lbl_total_desc_fac.grid(row=1, column=1, sticky=tk.W, padx=5)

        ttk.Label(info_frame, text="Rep:", font=("Segoe UI", 9, "bold")).grid(row=2, column=0, sticky=tk.W)
        self.lbl_rep_desc = ttk.Label(info_frame, text="—")
        self.lbl_rep_desc.grid(row=2, column=1, sticky=tk.W, padx=5)
        
        # Productos de la factura
        ttk.Label(frame_fac, text="📦 Productos (clic para seleccionar):", 
                  font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(8, 4))
        
        tree_prod_container = ttk.Frame(frame_fac)
        tree_prod_container.pack(fill=tk.BOTH, expand=True)
        
        self.tree_productos = ttk.Treeview(tree_prod_container,
                                          columns=("producto", "cantidad", "precio"),
                                          selectmode="browse", height=5, show="headings")
        self.tree_productos.column("producto", anchor=tk.W, width=150, minwidth=100)
        self.tree_productos.column("cantidad", anchor=tk.CENTER, width=50, minwidth=40)
        self.tree_productos.column("precio", anchor=tk.E, width=80, minwidth=60)
        
        self.tree_productos.heading("producto", text="Producto")
        self.tree_productos.heading("cantidad", text="Cant")
        self.tree_productos.heading("precio", text="Precio")
        
        scroll_prod_y = ttk.Scrollbar(tree_prod_container, orient=tk.VERTICAL, command=self.tree_productos.yview)
        self.tree_productos.configure(yscrollcommand=scroll_prod_y.set)
        
        self.tree_productos.grid(row=0, column=0, sticky="nsew")
        scroll_prod_y.grid(row=0, column=1, sticky="ns")
        tree_prod_container.grid_rowconfigure(0, weight=1)
        tree_prod_container.grid_columnconfigure(0, weight=1)
        
        self.tree_productos.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(self.tree_productos))
        self.tree_productos.bind("<<TreeviewSelect>>", self._on_producto_seleccionado)

        # === PANEL DERECHO: Formulario + Lista ===
        frame_der = ttk.Frame(paned_main)
        paned_main.add(frame_der, weight=2)

        # --- FORMULARIO AJUSTE ---
        frame_agg = ttk.LabelFrame(frame_der, text="➕ AGREGAR AJUSTE DE PRECIO", padding=10)
        frame_agg.pack(fill=tk.X, padx=5, pady=(5, 3))
        
        # Grid responsive para el formulario
        frame_agg.columnconfigure(1, weight=1)
        frame_agg.columnconfigure(3, weight=1)

        # Fila 0: Artículo
        ttk.Label(frame_agg, text="Artículo:").grid(row=0, column=0, sticky=tk.W)
        self.articulo_desc_var = tk.StringVar()
        self.entry_articulo = ttk.Entry(frame_agg, textvariable=self.articulo_desc_var, state="readonly")
        self.entry_articulo.grid(row=0, column=1, columnspan=3, sticky=tk.EW, padx=5, pady=2)
        
        # Fila 1: Cantidad + Precio Original
        ttk.Label(frame_agg, text="Cantidad:").grid(row=1, column=0, sticky=tk.W)
        self.cantidad_ajuste_var = tk.StringVar(value="")
        self.entry_cantidad = ttk.Entry(frame_agg, textvariable=self.cantidad_ajuste_var, width=10, state="readonly")
        self.entry_cantidad.grid(row=1, column=1, sticky=tk.W, padx=5, pady=2)
        
        ttk.Label(frame_agg, text="Precio:").grid(row=1, column=2, sticky=tk.W)
        self.precio_original_var = tk.StringVar(value="")
        self.entry_precio_orig = ttk.Entry(frame_agg, textvariable=self.precio_original_var, width=12, state="readonly")
        self.entry_precio_orig.grid(row=1, column=3, sticky=tk.W, padx=5, pady=2)
        
        # Fila 2: Nuevo Precio + Total Diferencia
        ttk.Label(frame_agg, text="Nuevo Precio:", foreground="#ffc107", font=("Segoe UI", 9, "bold")).grid(row=2, column=0, sticky=tk.W)
        self.precio_nuevo_var = tk.StringVar(value="")
        self.entry_precio_nuevo = ttk.Entry(frame_agg, textvariable=self.precio_nuevo_var, width=12)
        self.entry_precio_nuevo.grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        self.entry_precio_nuevo.bind("<KeyRelease>", self._calcular_diferencia_ajuste)
        
        ttk.Label(frame_agg, text="Total Dif:", font=("Segoe UI", 9, "bold")).grid(row=2, column=2, sticky=tk.W)
        self.monto_desc_var = tk.StringVar(value="$0")
        self.lbl_total_dif = ttk.Label(frame_agg, textvariable=self.monto_desc_var, 
                                        font=("Segoe UI", 11, "bold"), foreground="#ff5722")
        self.lbl_total_dif.grid(row=2, column=3, sticky=tk.W, padx=5, pady=2)
        
        # Fila 3: Observación
        ttk.Label(frame_agg, text="Observación:").grid(row=3, column=0, sticky=tk.W)
        self.observacion_ajuste_var = tk.StringVar(value="")
        self.entry_observacion = ttk.Entry(frame_agg, textvariable=self.observacion_ajuste_var)
        self.entry_observacion.grid(row=3, column=1, columnspan=3, sticky=tk.EW, padx=5, pady=2)
        
        # Botón agregar
        ttk.Button(frame_agg, text="✚ Agregar Ajuste",
                   command=self._agregar_descuento).grid(row=4, column=0, columnspan=4, pady=(10, 5))

        # --- RESUMEN DE FACTURA (Total, Ajustes, Devoluciones) ---
        frame_resumen = ttk.LabelFrame(frame_der, text="📊 RESUMEN FACTURA SELECCIONADA", padding=8)
        frame_resumen.pack(fill=tk.X, padx=5, pady=3)
        
        # Grid para el resumen
        frame_resumen.columnconfigure(1, weight=1)
        frame_resumen.columnconfigure(3, weight=1)
        
        ttk.Label(frame_resumen, text="Total Factura:", font=("Segoe UI", 9)).grid(row=0, column=0, sticky=tk.W)
        self.lbl_resumen_total_fac = ttk.Label(frame_resumen, text="$0", font=("Segoe UI", 10, "bold"), foreground="#4CAF50")
        self.lbl_resumen_total_fac.grid(row=0, column=1, sticky=tk.W, padx=5)
        
        ttk.Label(frame_resumen, text="(-) Ajustes:", font=("Segoe UI", 9)).grid(row=0, column=2, sticky=tk.W)
        self.lbl_resumen_ajustes = ttk.Label(frame_resumen, text="$0", font=("Segoe UI", 10, "bold"), foreground="#ff5722")
        self.lbl_resumen_ajustes.grid(row=0, column=3, sticky=tk.W, padx=5)
        
        ttk.Label(frame_resumen, text="(-) Devoluciones:", font=("Segoe UI", 9)).grid(row=1, column=0, sticky=tk.W)
        self.lbl_resumen_devol = ttk.Label(frame_resumen, text="$0", font=("Segoe UI", 10, "bold"), foreground="#9c27b0")
        self.lbl_resumen_devol.grid(row=1, column=1, sticky=tk.W, padx=5)
        
        ttk.Label(frame_resumen, text="= Nuevo Total:", font=("Segoe UI", 9, "bold")).grid(row=1, column=2, sticky=tk.W)
        self.lbl_resumen_nuevo_total = ttk.Label(frame_resumen, text="$0", font=("Segoe UI", 11, "bold"), foreground="#1565c0")
        self.lbl_resumen_nuevo_total.grid(row=1, column=3, sticky=tk.W, padx=5)

        # --- LISTA AJUSTES REGISTRADOS ---
        frame_lista = ttk.LabelFrame(frame_der, text="📝 AJUSTES REGISTRADOS", padding=5)
        frame_lista.pack(fill=tk.BOTH, expand=True, padx=5, pady=3)

        tree_desc_container = ttk.Frame(frame_lista)
        tree_desc_container.pack(fill=tk.BOTH, expand=True)
        
        self.tree_desc = ttk.Treeview(tree_desc_container,
                                      columns=("folio", "cliente", "articulo", "cant", "precio", "nuevo", "dif", "obs"),
                                      selectmode="extended", height=8, show="headings")
        self.tree_desc.column("#0",       width=0, stretch=tk.NO)
        self.tree_desc.column("folio",    anchor=tk.CENTER, width=50,  minwidth=40)
        self.tree_desc.column("cliente",  anchor=tk.W,      width=100, minwidth=70)
        self.tree_desc.column("articulo", anchor=tk.W,      width=110, minwidth=70)
        self.tree_desc.column("cant",     anchor=tk.CENTER, width=40,  minwidth=35)
        self.tree_desc.column("precio",   anchor=tk.E,      width=60,  minwidth=50)
        self.tree_desc.column("nuevo",    anchor=tk.E,      width=60,  minwidth=50)
        self.tree_desc.column("dif",      anchor=tk.E,      width=70,  minwidth=55)
        self.tree_desc.column("obs",      anchor=tk.W,      width=100, minwidth=60)

        self.tree_desc.heading("folio",    text="Folio")
        self.tree_desc.heading("cliente",  text="Cliente")
        self.tree_desc.heading("articulo", text="Artículo")
        self.tree_desc.heading("cant",     text="Cant")
        self.tree_desc.heading("precio",   text="Precio")
        self.tree_desc.heading("nuevo",    text="Nuevo")
        self.tree_desc.heading("dif",      text="Diferencia")
        self.tree_desc.heading("obs",      text="Obs.")
        
        self.tree_desc.tag_configure("par",   background="#2d2d2d", foreground="#ffffff")
        self.tree_desc.tag_configure("impar", background="#3d3d3d", foreground="#ffffff")

        scroll_desc_y = ttk.Scrollbar(tree_desc_container, orient=tk.VERTICAL, command=self.tree_desc.yview)
        scroll_desc_x = ttk.Scrollbar(tree_desc_container, orient=tk.HORIZONTAL, command=self.tree_desc.xview)
        self.tree_desc.configure(yscrollcommand=scroll_desc_y.set, xscrollcommand=scroll_desc_x.set)
        
        self.tree_desc.grid(row=0, column=0, sticky="nsew")
        scroll_desc_y.grid(row=0, column=1, sticky="ns")
        scroll_desc_x.grid(row=1, column=0, sticky="ew")
        
        tree_desc_container.grid_rowconfigure(0, weight=1)
        tree_desc_container.grid_columnconfigure(0, weight=1)
        
        self.tree_desc.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(self.tree_desc))
        self.tree_desc.bind("<Button-3>", self._mostrar_menu_descuentos)

        # Total de ajustes a la derecha
        frame_total = ttk.Frame(frame_lista)
        frame_total.pack(fill=tk.X, pady=(6, 2))
        
        self.lbl_total_desc_factura = ttk.Label(frame_total, text="Total ajustes del día: $0",
                                                font=("Segoe UI", 11, "bold"), foreground="#ffc107")
        self.lbl_total_desc_factura.pack(side=tk.RIGHT, padx=10)

    # --- refrescar combo de folios (desde DataStore) ---
    def _refrescar_folio_combo_descuentos(self):
        folios = [str(v['folio']) for v in self.ds.get_ventas()]
        self.folio_combo['values'] = folios
        
        # Actualizar combo de repartidores
        reps = sorted(set(v['repartidor'] for v in self.ds.get_ventas() if v['repartidor']))
        self.combo_filtro_rep['values'] = ["Todos"] + reps
        
        # También refrescar el treeview de búsqueda si hay texto activo
        self._filtrar_resultados()

    # --- filtrar resultados en tiempo real mientras el usuario escribe ---
    def _filtrar_resultados(self):
        texto = self.buscar_var.get().strip().lower()
        rep_filtro = self.filtro_rep_var.get()
        self.tree_buscar.delete(*self.tree_buscar.get_children())

        ventas = self.ds.get_ventas()

        # Aplicar filtros
        if rep_filtro != "Todos":
            ventas = [v for v in ventas if v['repartidor'] == rep_filtro]
        
        # Si no hay texto de búsqueda mostrar todas las ventas (filtradas por repartidor)
        if not texto:
            resultados = ventas
        else:
            # Buscar en nombre Y en folio (así "1234" matchea el folio también)
            resultados = [
                v for v in ventas
                if texto in v['nombre'].lower() or texto in str(v['folio'])
            ]

        # Poblar treeview con colores alternados
        for i, v in enumerate(resultados):
            tag = "par" if i % 2 == 0 else "impar"
            self.tree_buscar.insert(
                "", tk.END,
                iid=str(v['folio']),
                values=(v['folio'], v['nombre'],
                        f"${v['subtotal']:,.2f}",
                        v['repartidor'] or "—"),
                tags=(tag,)
            )

    # --- limpiar el campo de búsqueda ---
    def _limpiar_buscar(self):
        self.buscar_var.set("")
        self.entry_buscar.focus_set()

    # --- cuando se selecciona una fila en el tree de búsqueda ---
    def _on_tree_buscar_select(self, event=None):
        """Se ejecuta cuando cambia la selección en el tree de búsqueda."""
        seleccion = self.tree_buscar.selection()
        if not seleccion:
            return
        # El iid es el folio
        folio_id = seleccion[0]
        self._seleccionar_factura_desde_tree(folio_id)

    # --- cuando el usuario elige un folio del combo ---
    def _on_folio_combo_seleccionado(self, event=None):
        folio_s = self.folio_var.get()
        if folio_s:
            self._seleccionar_factura_desde_tree(folio_s)

    # --- núcleo: dado un folio (como string), colorea la fila en el tree y carga datos ---
    def _seleccionar_factura_desde_tree(self, folio_id: str):
        # Desmarcar todas las filas a su color alternado original
        for i, item in enumerate(self.tree_buscar.get_children()):
            tag = "par" if i % 2 == 0 else "impar"
            self.tree_buscar.item(item, tags=(tag,))

        # Marcar la fila seleccionada
        try:
            self.tree_buscar.item(folio_id, tags=("selec",))
        except tk.TclError:
            pass  # la fila no existe en el tree actual (búsqueda activa)

        # Sincronizar combo de folio
        self.folio_var.set(folio_id)

        # Cargar datos de esa factura en la parte inferior
        self._cargar_factura_desc()

    # --- cargar datos de la factura seleccionada ---
    def _cargar_factura_desc(self):
        folio_s = self.folio_var.get()
        if not folio_s:
            return
        folio = int(folio_s)
        self.folio_actual_desc = folio

        # buscar en DataStore
        for v in self.ds.get_ventas():
            if v['folio'] == folio:
                self.lbl_cliente_desc.config(text=v['nombre'])
                self.lbl_total_desc_fac.config(text=f"${v['subtotal']:,.2f}")
                self.lbl_rep_desc.config(text=v['repartidor'] or "— Sin asignar")
                break

        # cargar productos de la factura desde la BD
        self._cargar_productos_factura(folio)
        
        # cargar todos los ajustes del día
        self._refrescar_lista_descuentos()
    
    def _refrescar_lista_descuentos(self):
        """Carga todos los ajustes de precio registrados del día."""
        self.tree_desc.delete(*self.tree_desc.get_children())
        
        if not USE_SQLITE or not self.ds.fecha:
            self.lbl_total_desc_factura.config(text="Total ajustes del día: $0")
            return
        
        # Obtener solo ajustes del día (tipo='ajuste')
        desc_lista = db_local.obtener_descuentos_fecha(self.ds.fecha)
        ajustes = [d for d in desc_lista if d.get('tipo') == 'ajuste']
        
        total_d = 0.0
        
        for i, d in enumerate(ajustes):
            folio = d.get('folio', 0)
            monto = d.get('monto', 0)
            cliente = d.get('cliente', '')
            desc_id = d.get('id', 0)  # ID del descuento para eliminar
            # El artículo puede estar en 'articulo' o 'observacion' (datos anteriores)
            articulo = d.get('articulo', '') or d.get('observacion', '')
            precio_original = d.get('precio_facturado', 0)
            precio_nuevo = d.get('precio_nuevo', 0)
            cantidad = d.get('cantidad', 0)
            observacion = d.get('observacion', '') if d.get('articulo') else ''
            
            # Si no tiene cliente guardado, buscar en DataStore
            if not cliente:
                for v in self.ds.get_ventas():
                    if v['folio'] == folio:
                        cliente = v['nombre']
                        break
            
            # Tag alternado
            tag = 'par' if i % 2 == 0 else 'impar'
            
            # Usar el ID del descuento como iid para poder eliminarlo
            self.tree_desc.insert("", tk.END,
                                  iid=f"desc_{desc_id}",
                                  values=(
                                      folio,
                                      cliente[:15] if cliente else "—",
                                      articulo[:15] if articulo else "—",
                                      f"{cantidad:.0f}" if cantidad else "—",
                                      f"${precio_original:,.0f}" if precio_original else "—",
                                      f"${precio_nuevo:,.0f}" if precio_nuevo else "—",
                                      f"${monto:,.0f}",
                                      observacion[:12] if observacion else ""
                                  ),
                                  tags=(tag,))
            total_d += monto
        
        self.lbl_total_desc_factura.config(text=f"Total ajustes del día: ${total_d:,.0f}")
        
        # Actualizar resumen de la factura seleccionada
        self._actualizar_resumen_factura()
    
    def _actualizar_resumen_factura(self):
        """Actualiza el resumen de la factura seleccionada (total, ajustes, devoluciones)."""
        if not hasattr(self, 'folio_actual_desc'):
            self.lbl_resumen_total_fac.config(text="$0")
            self.lbl_resumen_ajustes.config(text="$0")
            self.lbl_resumen_devol.config(text="$0")
            self.lbl_resumen_nuevo_total.config(text="$0")
            return
        
        folio = self.folio_actual_desc
        
        # Obtener total de la factura
        total_factura = 0.0
        for v in self.ds.get_ventas():
            if v['folio'] == folio:
                total_factura = v['subtotal']
                break
        
        # Obtener ajustes de esta factura
        total_ajustes = 0.0
        if USE_SQLITE and self.ds.fecha:
            desc_lista = db_local.obtener_descuentos_fecha(self.ds.fecha)
            for d in desc_lista:
                if d.get('folio') == folio and d.get('tipo') == 'ajuste':
                    total_ajustes += d.get('monto', 0)
        
        # Obtener devoluciones parciales de esta factura
        total_devol = 0.0
        if USE_SQLITE:
            devol_lista = db_local.obtener_devoluciones_parciales_folio(folio)
            for dev in devol_lista:
                # El campo se llama dinero_devuelto en la tabla devoluciones_parciales
                total_devol += dev.get('dinero_devuelto', 0)
        
        # Calcular nuevo total = Total Factura - Ajustes - Devoluciones
        nuevo_total = total_factura - total_ajustes - total_devol
        
        # Actualizar labels
        self.lbl_resumen_total_fac.config(text=f"${total_factura:,.0f}")
        self.lbl_resumen_ajustes.config(text=f"${total_ajustes:,.0f}")
        self.lbl_resumen_devol.config(text=f"${total_devol:,.0f}")
        self.lbl_resumen_nuevo_total.config(text=f"${nuevo_total:,.0f}")
    
    def _mostrar_menu_descuentos(self, event):
        """Muestra menú contextual para descuentos con opciones de copiar y eliminar."""
        # Seleccionar el item bajo el cursor
        item = self.tree_desc.identify_row(event.y)
        if item:
            self.tree_desc.selection_set(item)
        
        # Crear menú contextual
        menu = tk.Menu(self.ventana, tearoff=0)
        menu.add_command(label="📋 Copiar", command=lambda: self._copiar_seleccion_tree(self.tree_desc))
        menu.add_separator()
        menu.add_command(label="🗑️ Eliminar Descuento", command=self._eliminar_descuento_seleccionado,
                         foreground="#c62828")
        
        # Mostrar menú
        menu.tk_popup(event.x_root, event.y_root)
    
    def _eliminar_descuento_seleccionado(self):
        """Elimina el descuento seleccionado en tree_desc."""
        seleccion = self.tree_desc.selection()
        if not seleccion:
            messagebox.showwarning("Sin Selección", "Selecciona un descuento para eliminar.")
            return
        
        # Obtener el ID del descuento desde el iid (formato: desc_ID)
        item_id = seleccion[0]
        if not item_id.startswith("desc_"):
            messagebox.showerror("Error", "No se puede identificar el descuento.")
            return
        
        try:
            desc_id = int(item_id.replace("desc_", ""))
        except ValueError:
            messagebox.showerror("Error", "ID de descuento inválido.")
            return
        
        # Obtener datos para mostrar confirmación
        values = self.tree_desc.item(item_id, "values")
        folio = values[0] if values else "?"
        articulo = values[2] if len(values) > 2 else "?"
        monto = values[6] if len(values) > 6 else "?"
        
        # Confirmar eliminación
        if not messagebox.askyesno("Confirmar Eliminación", 
                                    f"¿Eliminar descuento del folio {folio}?\n\n"
                                    f"Artículo: {articulo}\n"
                                    f"Monto: {monto}"):
            return
        
        # Eliminar de la base de datos
        if USE_SQLITE and db_local.eliminar_descuento(desc_id):
            messagebox.showinfo("Éxito", "Descuento eliminado correctamente.")
            # Refrescar la lista
            self._refrescar_lista_descuentos()
            # Notificar cambio global
            self.ds.notify()
        else:
            messagebox.showerror("Error", "No se pudo eliminar el descuento.")

    def _cargar_productos_factura(self, folio: int):
        """Carga los productos de una factura desde la BD."""
        self.tree_productos.delete(*self.tree_productos.get_children())
        
        # Consulta correcta usando VENTATICKETS_ARTICULOS
        sql = (
            "SET HEADING ON;\n"
            "SELECT VA.PRODUCTO_NOMBRE, VA.CANTIDAD, VA.PRECIO_FINAL\n"
            "FROM VENTATICKETS_ARTICULOS VA\n"
            "INNER JOIN VENTATICKETS V ON VA.TICKET_ID = V.ID\n"
            f"WHERE V.FOLIO = {folio}\n"
            "ORDER BY VA.PRODUCTO_NOMBRE;\n"
        )
        
        ok, stdout, stderr = self._ejecutar_sql(sql)
        if not ok:
            return
        
        header_visto = False
        for linea in stdout.split('\n'):
            linea = linea.strip()
            if not linea or linea.startswith('='):
                continue
            if 'PRODUCTO_NOMBRE' in linea and 'CANTIDAD' in linea:
                header_visto = True
                continue
            if not header_visto:
                continue
            
            partes = linea.split()
            if len(partes) < 3:
                continue
            
            try:
                # Precio en pesos colombianos (NO dividir por 100)
                precio = float(partes[-1])
                cantidad = float(partes[-2])
                nombre = ' '.join(partes[:-2]).replace('<null>', '').strip()
                
                if nombre:
                    self.tree_productos.insert("", tk.END,
                                              values=(nombre, 
                                                     f"{cantidad:.0f}",
                                                     f"${precio:,.0f}"))
            except (ValueError, IndexError):
                continue

    def _on_producto_seleccionado(self, event=None):
        """Al seleccionar un producto de la lista, llena los campos del formulario de ajuste."""
        seleccion = self.tree_productos.selection()
        if not seleccion:
            return
        
        item = seleccion[0]
        valores = self.tree_productos.item(item, 'values')
        if not valores or len(valores) < 3:
            return
        
        # valores = (producto, cantidad, precio)
        producto = valores[0]
        cantidad_str = valores[1]
        precio_str = valores[2].replace('$', '').replace(',', '')
        
        try:
            cantidad = float(cantidad_str)
            precio = float(precio_str)
        except ValueError:
            return
        
        # Llenar los campos del formulario
        self.articulo_desc_var.set(producto)
        self.cantidad_ajuste_var.set(f"{cantidad:.0f}")
        self.precio_original_var.set(f"{precio:.0f}")
        self.precio_nuevo_var.set("")  # Limpiar para que el usuario ingrese
        self.monto_desc_var.set("$0")
        
        # Enfocar el campo de nuevo precio
        self.entry_precio_nuevo.focus_set()
    
    def _calcular_diferencia_ajuste(self, event=None):
        """Calcula el total diferencia = (Precio - Nuevo Precio) * Cantidad."""
        try:
            precio_orig = float(self.precio_original_var.get() or 0)
            precio_nuevo = float(self.precio_nuevo_var.get() or 0)
            cantidad = float(self.cantidad_ajuste_var.get() or 0)
            
            diferencia = (precio_orig - precio_nuevo) * cantidad
            self.monto_desc_var.set(f"${diferencia:,.0f}")
        except ValueError:
            self.monto_desc_var.set("$0")

    def _agregar_descuento(self):
        """Agrega un ajuste de precio a la factura seleccionada."""
        if not hasattr(self, 'folio_actual_desc'):
            messagebox.showwarning("Advertencia", "Selecciona una factura primero.")
            return
        try:
            articulo = self.articulo_desc_var.get().strip()
            precio_orig = float(self.precio_original_var.get() or 0)
            precio_nuevo = float(self.precio_nuevo_var.get() or 0)
            cantidad = float(self.cantidad_ajuste_var.get() or 0)
            observacion = self.observacion_ajuste_var.get().strip()
            monto = (precio_orig - precio_nuevo) * cantidad

            if not articulo:
                messagebox.showwarning("Artículo", "Selecciona un producto de la lista.")
                return
                
            if monto <= 0:
                messagebox.showwarning("Valor", "El nuevo precio debe ser menor al precio original.")
                return

            # buscar repartidor y cliente de la factura desde DataStore
            rep = ''
            cliente = ''
            for v in self.ds.get_ventas():
                if v['folio'] == self.folio_actual_desc:
                    rep = v['repartidor']
                    cliente = v['nombre']
                    break

            # Guardar en BD como ajuste
            if USE_SQLITE:
                db_local.agregar_descuento(
                    self.ds.fecha,
                    self.folio_actual_desc,
                    'ajuste',  # tipo fijo: ajuste
                    monto,
                    rep,
                    observacion,  # observacion
                    cliente,
                    articulo,
                    precio_orig,   # precio_facturado
                    precio_nuevo,  # precio_nuevo
                    cantidad       # cantidad
                )

            # reset campos
            self.precio_nuevo_var.set("")
            self.articulo_desc_var.set("")
            self.precio_original_var.set("")
            self.cantidad_ajuste_var.set("")
            self.observacion_ajuste_var.set("")
            self.monto_desc_var.set("$0")

            # recargar lista
            self._refrescar_lista_descuentos()

            # notificar al DataStore para que Liquidación se actualice
            self.ds._notificar()

            messagebox.showinfo("Listo", "Ajuste agregado.")
        except ValueError:
            messagebox.showerror("Error", "Verifica que el valor sea un número válido.")

    # ==================================================================
    # PESTAÑA 3 – GASTOS ADICIONALES
    # ==================================================================
    def _crear_tab_gastos(self):
        # --- barra de filtros ---
        frame_filtros = ttk.Frame(self.tab_gastos)
        frame_filtros.pack(fill=tk.X, padx=10, pady=(10, 4))
        
        ttk.Label(frame_filtros, text="🔍 Filtrar:", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(frame_filtros, text="Repartidor:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.filtro_gastos_rep_var = tk.StringVar(value="Todos")
        self.filtro_gastos_rep_combo = ttk.Combobox(frame_filtros, textvariable=self.filtro_gastos_rep_var,
                                                     width=15, state="readonly")
        self.filtro_gastos_rep_combo['values'] = ["Todos"]
        self.filtro_gastos_rep_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.filtro_gastos_rep_combo.bind("<<ComboboxSelected>>", lambda e: self._refrescar_gastos())
        
        ttk.Label(frame_filtros, text="Tipo:", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 5))
        self.filtro_gastos_tipo_var = tk.StringVar(value="Todos")
        self.filtro_gastos_tipo_combo = ttk.Combobox(frame_filtros, textvariable=self.filtro_gastos_tipo_var,
                                                      width=15, state="readonly")
        self.filtro_gastos_tipo_combo['values'] = ["Todos", "GASTO", "PAGO PROVEEDOR", "PRÉSTAMO", "NÓMINA", "SOCIO", "TRANSFERENCIA"]
        self.filtro_gastos_tipo_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.filtro_gastos_tipo_combo.bind("<<ComboboxSelected>>", lambda e: self._refrescar_gastos())
        
        ttk.Button(frame_filtros, text="🔄 Refrescar", command=self._refrescar_gastos).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_filtros, text="Limpiar", command=self._limpiar_filtro_gastos).pack(side=tk.LEFT, padx=2)
        
        # Contador de registros filtrados
        self.lbl_gastos_filtrados = ttk.Label(frame_filtros, text="", font=("Segoe UI", 9))
        self.lbl_gastos_filtrados.pack(side=tk.RIGHT, padx=10)
        
        # --- zona de entrada ---
        frame_entrada = ttk.LabelFrame(self.tab_gastos, text="➕ AÑADIR GASTO / PRÉSTAMO / PAGO", padding=(10, 8))
        frame_entrada.pack(fill=tk.X, padx=10, pady=(4, 4))

        # Fila 0: Tipo de registro (Combobox readonly con tipos fijos)
        ttk.Label(frame_entrada, text="Tipo:").grid(row=0, column=0, sticky=tk.W, pady=(0, 4))
        self.gasto_tipo_var = tk.StringVar(value="GASTO")
        self.gasto_tipo_combo = ttk.Combobox(frame_entrada, textvariable=self.gasto_tipo_var,
                                              width=18, state="readonly")
        self.gasto_tipo_combo.grid(row=0, column=1, sticky=tk.W, padx=(4, 20), pady=(0, 4))
        
        # Cargar tipos existentes
        self._cargar_tipos_gasto()

        # Fila 1: Repartidor
        ttk.Label(frame_entrada, text="Repartidor:").grid(row=1, column=0, sticky=tk.W, pady=(0, 4))
        self.gasto_rep_var = tk.StringVar()
        self.gasto_rep_combo = ttk.Combobox(frame_entrada, textvariable=self.gasto_rep_var,
                                            width=18, state="readonly")
        self.gasto_rep_combo.grid(row=1, column=1, sticky=tk.W, padx=(4, 20), pady=(0, 4))

        # Fila 2: Concepto (Entry libre)
        ttk.Label(frame_entrada, text="Concepto:").grid(row=2, column=0, sticky=tk.W, pady=(0, 4))
        self.gasto_concepto_var = tk.StringVar()
        ttk.Entry(frame_entrada, textvariable=self.gasto_concepto_var, width=50).grid(
            row=2, column=1, columnspan=4, sticky=tk.W, padx=(4, 0), pady=(0, 4))

        # Fila 3: Monto
        ttk.Label(frame_entrada, text="Monto:").grid(row=3, column=0, sticky=tk.W, pady=(0, 4))
        frame_monto_gasto = ttk.Frame(frame_entrada)
        frame_monto_gasto.grid(row=3, column=1, sticky=tk.W, padx=(4, 0), pady=(0, 4))
        ttk.Label(frame_monto_gasto, text="$").pack(side=tk.LEFT)
        self.gasto_monto_var = tk.StringVar(value="0.00")
        ttk.Entry(frame_monto_gasto, textvariable=self.gasto_monto_var,
                  width=12, justify=tk.RIGHT).pack(side=tk.LEFT)
        
        # Fila 4: Observaciones
        ttk.Label(frame_entrada, text="Observaciones:").grid(row=4, column=0, sticky=tk.W, pady=(0, 4))
        self.gasto_observ_var = tk.StringVar()
        ttk.Entry(frame_entrada, textvariable=self.gasto_observ_var, width=50).grid(
            row=4, column=1, columnspan=4, sticky=tk.W, padx=(4, 0), pady=(0, 4))
        
        # Fila 5: Botón
        ttk.Button(frame_entrada, text="＋  Añadir Registro",
                   command=self._añadir_gasto).grid(row=5, column=1, sticky=tk.W, pady=(8, 0))

        # --- tabla de gastos registrados ---
        frame_tabla = ttk.LabelFrame(self.tab_gastos, text="💸 GASTOS REGISTRADOS (Doble clic para editar)", padding=(5, 5))
        frame_tabla.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)

        # Contenedor con scrollbars
        tree_gastos_container = ttk.Frame(frame_tabla)
        tree_gastos_container.pack(fill=tk.BOTH, expand=True)

        self.tree_gastos = ttk.Treeview(
            tree_gastos_container,
            columns=("id", "tipo", "repartidor", "concepto", "monto", "observaciones"),
            selectmode="extended", height=16, show="headings"
        )
        self.tree_gastos.column("id",            anchor=tk.CENTER, width=40,  minwidth=40)
        self.tree_gastos.column("tipo",          anchor=tk.CENTER, width=110, minwidth=80)
        self.tree_gastos.column("repartidor",    anchor=tk.W,      width=100, minwidth=80)
        self.tree_gastos.column("concepto",      anchor=tk.W,      width=300, minwidth=150)
        self.tree_gastos.column("monto",         anchor=tk.E,      width=110, minwidth=80)
        self.tree_gastos.column("observaciones", anchor=tk.W,      width=200, minwidth=100)

        self.tree_gastos.heading("id",            text="🔢 ID")
        self.tree_gastos.heading("tipo",          text="📋 Tipo")
        self.tree_gastos.heading("repartidor",    text="👤 Repartidor")
        self.tree_gastos.heading("concepto",      text="📝 Concepto")
        self.tree_gastos.heading("monto",         text="💵 Monto")
        self.tree_gastos.heading("observaciones", text="📄 Observaciones")

        # Tags con colores para modo oscuro
        self.tree_gastos.tag_configure("gasto",     background="#2d2d2d", foreground="#ffffff")
        self.tree_gastos.tag_configure("proveedor", background="#1a237e", foreground="#90caf9")
        self.tree_gastos.tag_configure("prestamo",  background="#004d40", foreground="#80cbc4")
        self.tree_gastos.tag_configure("nomina",    background="#4a148c", foreground="#ce93d8")
        self.tree_gastos.tag_configure("socios",    background="#bf360c", foreground="#ffab91")

        # Scrollbars
        scroll_gastos_y = ttk.Scrollbar(tree_gastos_container, orient=tk.VERTICAL, command=self.tree_gastos.yview)
        scroll_gastos_x = ttk.Scrollbar(tree_gastos_container, orient=tk.HORIZONTAL, command=self.tree_gastos.xview)
        self.tree_gastos.configure(yscrollcommand=scroll_gastos_y.set, xscrollcommand=scroll_gastos_x.set)
        
        # Grid layout
        self.tree_gastos.grid(row=0, column=0, sticky="nsew")
        scroll_gastos_y.grid(row=0, column=1, sticky="ns")
        scroll_gastos_x.grid(row=1, column=0, sticky="ew")
        
        tree_gastos_container.grid_rowconfigure(0, weight=1)
        tree_gastos_container.grid_columnconfigure(0, weight=1)
        
        # Bindings
        self.tree_gastos.bind("<Control-c>", lambda e: self._copiar_seleccion_tree(self.tree_gastos))
        self.tree_gastos.bind("<Control-C>", lambda e: self._copiar_seleccion_tree(self.tree_gastos))
        self.tree_gastos.bind("<Double-1>", self._editar_gasto_doble_clic)  # Doble clic para editar
        self.tree_gastos.bind("<Button-3>", self._on_gastos_right_click)    # Clic derecho menú

        # --- barra inferior: totales por repartidor + total global (BOTTOM para visibilidad) ---
        frame_inf = ttk.LabelFrame(self.tab_gastos, text="📊 TOTALES", padding=(8, 6))
        frame_inf.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 8))

        # Label que se actualiza dinámicamente con totales desglosados
        self.lbl_totales_gastos = ttk.Label(frame_inf, text="—",
                                            font=("Segoe UI", 9), justify=tk.LEFT)
        self.lbl_totales_gastos.pack(anchor=tk.W)

        # Total global grande
        frame_total_g = ttk.Frame(frame_inf)
        frame_total_g.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(frame_total_g, text="TOTAL GASTOS:",
                  font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        self.lbl_total_gastos_global = ttk.Label(frame_total_g, text="$0.00",
                                                  font=("Segoe UI", 12, "bold"), foreground="#e65100")
        self.lbl_total_gastos_global.pack(side=tk.LEFT, padx=8)
        
        ttk.Label(frame_total_g, text="    |    PAGO PROVEEDORES:",
                  font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        self.lbl_total_pago_prov_gastos = ttk.Label(frame_total_g, text="$0.00",
                                                     font=("Segoe UI", 12, "bold"), foreground="#1565c0")
        self.lbl_total_pago_prov_gastos.pack(side=tk.LEFT, padx=8)
        
        ttk.Label(frame_total_g, text="    |    PRÉSTAMOS:",
                  font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        self.lbl_total_prestamos_gastos = ttk.Label(frame_total_g, text="$0.00",
                                                     font=("Segoe UI", 12, "bold"), foreground="#00695c")
        self.lbl_total_prestamos_gastos.pack(side=tk.LEFT, padx=8)
        
        ttk.Label(frame_total_g, text="    |    NÓMINA:",
                  font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        self.lbl_total_nomina_gastos = ttk.Label(frame_total_g, text="$0.00",
                                                  font=("Segoe UI", 12, "bold"), foreground="#7b1fa2")
        self.lbl_total_nomina_gastos.pack(side=tk.LEFT, padx=8)
        
        ttk.Label(frame_total_g, text="    |    SOCIOS:",
                  font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        self.lbl_total_socios_gastos = ttk.Label(frame_total_g, text="$0.00",
                                                  font=("Segoe UI", 12, "bold"), foreground="#e64a19")
        self.lbl_total_socios_gastos.pack(side=tk.LEFT, padx=8)
        
        ttk.Label(frame_total_g, text="    |    TRANSFERENCIAS:",
                  font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT)
        self.lbl_total_transferencias_gastos = ttk.Label(frame_total_g, text="$0.00",
                                                  font=("Segoe UI", 12, "bold"), foreground="#0288d1")
        self.lbl_total_transferencias_gastos.pack(side=tk.LEFT, padx=8)

    def _cargar_tipos_gasto(self):
        """Carga los tipos de gasto en el combobox."""
        # Tipos disponibles para registrar movimientos
        tipos = ["GASTO", "PAGO PROVEEDOR", "PRÉSTAMO", "NÓMINA", "SOCIO", "TRANSFERENCIA"]
        self.gasto_tipo_combo['values'] = tipos
        self.gasto_tipo_combo.config(state="readonly")  # No permitir tipos personalizados
    
    def _agregar_tipo_gasto(self):
        """Agrega un nuevo tipo de gasto."""
        nuevo_tipo = self.gasto_tipo_var.get().strip().upper()
        if not nuevo_tipo:
            messagebox.showwarning("Advertencia", "Ingrese un tipo de gasto")
            return
        
        # Verificar si ya existe
        tipos_actuales = list(self.gasto_tipo_combo['values'])
        if nuevo_tipo in tipos_actuales:
            messagebox.showinfo("Info", f"El tipo '{nuevo_tipo}' ya existe")
            return
        
        # Agregar a la BD
        if self.ds.agregar_concepto_gasto(nuevo_tipo):
            self._cargar_tipos_gasto()
            messagebox.showinfo("Éxito", f"Tipo '{nuevo_tipo}' agregado")
        else:
            # Si ya existe en BD, solo recargar
            self._cargar_tipos_gasto()

    def _crear_switch_toggle(self, parent, variable, texto, switch_id="default"):
        """Crea un switch toggle estilo iOS/Material Design."""
        frame = ttk.Frame(parent)
        
        # Inicializar diccionario de switches si no existe
        if not hasattr(self, '_switches'):
            self._switches = {}
        
        # Canvas para el switch
        canvas_width = 44
        canvas_height = 24
        switch_canvas = tk.Canvas(
            frame, width=canvas_width, height=canvas_height,
            bg='#3d3d3d', highlightthickness=0, cursor="hand2"
        )
        switch_canvas.pack(side=tk.LEFT, padx=(0, 8))
        
        # Colores
        switch_off_bg = '#555555'
        switch_on_bg = '#4caf50'
        switch_knob_color = '#ffffff'
        
        # Redondear esquinas con óvalos
        r = 10
        switch_canvas.create_oval(2, 2, 2+r*2, canvas_height-2, fill=switch_off_bg, outline='', tags=f'bg_left_{switch_id}')
        switch_canvas.create_oval(canvas_width-2-r*2, 2, canvas_width-2, canvas_height-2, fill=switch_off_bg, outline='', tags=f'bg_right_{switch_id}')
        switch_canvas.create_rectangle(2+r, 2, canvas_width-2-r, canvas_height-2, fill=switch_off_bg, outline='', tags=f'bg_center_{switch_id}')
        
        # Knob (círculo blanco)
        knob_r = 8
        knob_x = 12 if not variable.get() else canvas_width - 12
        knob_id = switch_canvas.create_oval(
            knob_x - knob_r, (canvas_height//2) - knob_r,
            knob_x + knob_r, (canvas_height//2) + knob_r,
            fill=switch_knob_color, outline='#cccccc'
        )
        
        # Guardar referencia en diccionario
        self._switches[switch_id] = {
            'canvas': switch_canvas,
            'variable': variable,
            'knob_id': knob_id,
            'off_bg': switch_off_bg,
            'on_bg': switch_on_bg
        }
        
        # Binding para toggle
        def toggle_switch(event=None):
            current = variable.get()
            variable.set(not current)
            self._actualizar_switch_visual(switch_id)
            # Si se activa préstamo, desactivar proveedor y viceversa
            if switch_id == "prestamo" and variable.get():
                self.es_pago_proveedor_var.set(False)
                self._actualizar_switch_visual("proveedor")
            elif switch_id == "proveedor" and variable.get():
                self.es_prestamo_var.set(False)
                self._actualizar_switch_visual("prestamo")
        
        switch_canvas.bind("<Button-1>", toggle_switch)
        
        # Label del texto
        ttk.Label(frame, text=texto, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        
        return frame
    
    def _actualizar_switch_visual(self, switch_id="default"):
        """Actualiza la apariencia visual de un switch específico."""
        if not hasattr(self, '_switches') or switch_id not in self._switches:
            return
        
        sw = self._switches[switch_id]
        canvas = sw['canvas']
        is_on = sw['variable'].get()
        
        # Colores según estado
        bg_color = sw['on_bg'] if is_on else sw['off_bg']
        
        # Actualizar colores del fondo
        canvas.itemconfig(f'bg_left_{switch_id}', fill=bg_color)
        canvas.itemconfig(f'bg_right_{switch_id}', fill=bg_color)
        canvas.itemconfig(f'bg_center_{switch_id}', fill=bg_color)
        
        # Mover el knob
        canvas_width = 44
        canvas_height = 24
        knob_r = 8
        knob_x = canvas_width - 12 if is_on else 12
        
        canvas.coords(
            sw['knob_id'],
            knob_x - knob_r, (canvas_height//2) - knob_r,
            knob_x + knob_r, (canvas_height//2) + knob_r
        )

    # --- refrescar tabla y totales de gastos ---
    def _refrescar_tab_gastos(self):
        # actualizar combo de repartidores
        reps = self.ds.get_repartidores()
        self.gasto_rep_combo['values'] = reps
        if self.gasto_rep_var.get() not in reps and reps:
            self.gasto_rep_var.set(reps[0])
        
        # Actualizar filtro de repartidores
        if hasattr(self, 'filtro_gastos_rep_combo'):
            self.filtro_gastos_rep_combo['values'] = ["Todos"] + list(reps)
        
        # Obtener filtros
        filtro_rep = self.filtro_gastos_rep_var.get() if hasattr(self, 'filtro_gastos_rep_var') else "Todos"
        filtro_tipo = self.filtro_gastos_tipo_var.get() if hasattr(self, 'filtro_gastos_tipo_var') else "Todos"

        # poblar treeview con gastos Y pagos a proveedores
        self.tree_gastos.delete(*self.tree_gastos.get_children())
        
        # Contadores para totales filtrados
        count_registros = 0
        total_filtrado = 0
        
        # Obtener gastos normales
        gastos = self.ds.get_gastos()
        for g in gastos:
            rep = g['repartidor']
            concepto_upper = g.get('concepto', '').upper()
            
            # Determinar tipo de gasto según el concepto guardado
            if 'NOMINA' in concepto_upper or 'NÓMINA' in concepto_upper:
                tipo_texto = "💰 Pago Nómina"
                tipo_interno = "NÓMINA"
            elif 'SOCIO' in concepto_upper:
                tipo_texto = "🤝 Pago Socios"
                tipo_interno = "SOCIO"
            elif 'PROVEEDOR' in concepto_upper:
                tipo_texto = "💼 Pago Proveedor"
                tipo_interno = "PAGO PROVEEDOR"
            elif rep.lower() in ('cajero', 'caja', 'cajera'):
                tipo_texto = "🏪 Gasto Cajero"
                tipo_interno = "GASTO"
            else:
                tipo_texto = "🔧 Gasto Rep."
                tipo_interno = "GASTO"
            
            # Aplicar filtros
            if filtro_rep != "Todos" and rep != filtro_rep:
                continue
            if filtro_tipo != "Todos" and tipo_interno != filtro_tipo:
                continue
            
            monto = g['monto']
            self.tree_gastos.insert("", tk.END, 
                                    iid=f"gasto_{g.get('id', 0)}",
                                    values=(g.get('id', ''),
                                            tipo_texto,
                                            rep,
                                            g['concepto'],
                                            f"${monto:,.2f}",
                                            g.get('observaciones', '') or ''),
                                    tags=("gasto",))
            count_registros += 1
            total_filtrado += monto
        
        # Obtener pagos a proveedores
        pagos_prov = self.ds.get_pagos_proveedores()
        for p in pagos_prov:
            rep = p.get('repartidor', '') or '—'
            
            # Aplicar filtros
            if filtro_rep != "Todos" and rep != filtro_rep and rep != '—':
                continue
            if filtro_tipo != "Todos" and filtro_tipo != "PAGO PROVEEDOR":
                continue
            
            monto = p.get('monto', 0)
            self.tree_gastos.insert("", tk.END,
                                    iid=f"prov_{p.get('id', 0)}",
                                    values=(p.get('id', ''),
                                            "💼 Pago Proveedor",
                                            rep,
                                            p.get('proveedor', ''),
                                            f"${monto:,.2f}",
                                            p.get('observaciones', '') or ''),
                                    tags=("proveedor",))
            count_registros += 1
            total_filtrado += monto

        # Obtener préstamos
        prestamos = self.ds.get_prestamos()
        for pr in prestamos:
            rep = pr.get('repartidor', '')
            
            # Aplicar filtros
            if filtro_rep != "Todos" and rep != filtro_rep:
                continue
            if filtro_tipo != "Todos" and filtro_tipo != "PRÉSTAMO":
                continue
            
            monto = pr.get('monto', 0)
            self.tree_gastos.insert("", tk.END,
                                    iid=f"prest_{pr.get('id', 0)}",
                                    values=(pr.get('id', ''),
                                            "💵 Préstamo",
                                            rep,
                                            pr.get('concepto', ''),
                                            f"${monto:,.2f}",
                                            pr.get('observaciones', '') or ''),
                                    tags=("prestamo",))
            count_registros += 1
            total_filtrado += monto
        
        # Obtener pagos de nómina (desde tabla dedicada)
        pagos_nomina = self.ds.get_pagos_nomina()
        for pn in pagos_nomina:
            rep = pn.get('empleado', '')
            
            # Aplicar filtros
            if filtro_rep != "Todos" and rep != filtro_rep:
                continue
            if filtro_tipo != "Todos" and filtro_tipo != "NÓMINA":
                continue
            
            monto = pn.get('monto', 0)
            self.tree_gastos.insert("", tk.END,
                                    iid=f"nomina_{pn.get('id', 0)}",
                                    values=(pn.get('id', ''),
                                            "💰 Nómina",
                                            rep,
                                            pn.get('concepto', ''),
                                            f"${monto:,.2f}",
                                            pn.get('observaciones', '') or ''),
                                    tags=("nomina",))
            count_registros += 1
            total_filtrado += monto
        
        # Obtener pagos a socios (desde tabla dedicada)
        pagos_socios = self.ds.get_pagos_socios()
        for ps in pagos_socios:
            rep = ps.get('socio', '')
            
            # Aplicar filtros
            if filtro_rep != "Todos" and rep != filtro_rep:
                continue
            if filtro_tipo != "Todos" and filtro_tipo != "SOCIO":
                continue
            
            monto = ps.get('monto', 0)
            self.tree_gastos.insert("", tk.END,
                                    iid=f"socios_{ps.get('id', 0)}",
                                    values=(ps.get('id', ''),
                                            "🤝 Socios",
                                            rep,
                                            ps.get('concepto', ''),
                                            f"${monto:,.2f}",
                                            ps.get('observaciones', '') or ''),
                                    tags=("socios",))
            count_registros += 1
            total_filtrado += monto
        
        # Obtener transferencias (desde tabla dedicada)
        transferencias = self.ds.get_transferencias()
        for tr in transferencias:
            rep = tr.get('destinatario', '')
            
            # Aplicar filtros
            if filtro_rep != "Todos" and rep != filtro_rep:
                continue
            if filtro_tipo != "Todos" and filtro_tipo != "TRANSFERENCIA":
                continue
            
            monto = tr.get('monto', 0)
            self.tree_gastos.insert("", tk.END,
                                    iid=f"transf_{tr.get('id', 0)}",
                                    values=(tr.get('id', ''),
                                            "💸 Transferencia",
                                            rep,
                                            tr.get('concepto', ''),
                                            f"${monto:,.2f}",
                                            tr.get('observaciones', '') or ''),
                                    tags=("transferencia",))
            count_registros += 1
            total_filtrado += monto
        
        # Actualizar contador de registros filtrados
        if hasattr(self, 'lbl_gastos_filtrados'):
            if filtro_rep != "Todos" or filtro_tipo != "Todos":
                self.lbl_gastos_filtrados.config(text=f"📊 Filtrado: {count_registros} registros | ${total_filtrado:,.2f}")
            else:
                self.lbl_gastos_filtrados.config(text=f"📊 Total: {count_registros} registros")

        # totales desglosados por repartidor (solo gastos)
        reps_activos = sorted({g['repartidor'] for g in gastos})
        lineas = []
        for r in reps_activos:
            t = self.ds.get_total_gastos(r)
            lineas.append(f"{r}:  ${t:,.2f}")
        self.lbl_totales_gastos.config(
            text=("  |  ".join(lineas)) if lineas else "Sin gastos registrados."
        )

        total_global = self.ds.get_total_gastos()
        self.lbl_total_gastos_global.config(text=f"${total_global:,.2f}")
        
        # Total pagos proveedores
        total_prov = self.ds.get_total_pagos_proveedores()
        self.lbl_total_pago_prov_gastos.config(text=f"${total_prov:,.2f}")
        
        # Total préstamos
        total_prest = self.ds.get_total_prestamos()
        self.lbl_total_prestamos_gastos.config(text=f"${total_prest:,.2f}")
        
        # Total nómina
        total_nomina = self.ds.get_total_pagos_nomina()
        self.lbl_total_nomina_gastos.config(text=f"${total_nomina:,.2f}")
        
        # Total socios
        total_socios = self.ds.get_total_pagos_socios()
        self.lbl_total_socios_gastos.config(text=f"${total_socios:,.2f}")
        
        # Total transferencias
        total_transferencias = self.ds.get_total_transferencias()
        self.lbl_total_transferencias_gastos.config(text=f"${total_transferencias:,.2f}")

    # --- Refrescar gastos (alias para filtros) ---
    def _refrescar_gastos(self):
        self._refrescar_tab_gastos()
    
    # --- Limpiar filtro de gastos ---
    def _limpiar_filtro_gastos(self):
        if hasattr(self, 'filtro_gastos_rep_var'):
            self.filtro_gastos_rep_var.set("Todos")
        if hasattr(self, 'filtro_gastos_tipo_var'):
            self.filtro_gastos_tipo_var.set("Todos")
        self._refrescar_tab_gastos()

    # --- añadir un gasto nuevo ---
    def _añadir_gasto(self):
        tipo = self.gasto_tipo_var.get()
        rep = self.gasto_rep_var.get().strip()
        concepto = self.gasto_concepto_var.get().strip()
        observaciones = self.gasto_observ_var.get().strip()
        try:
            monto = float(self.gasto_monto_var.get() or 0)
        except ValueError:
            messagebox.showerror("Error", "El monto debe ser un número válido.")
            return

        if not rep:
            messagebox.showwarning("Repartidor", "Selecciona un repartidor primero.")
            return
        if not concepto:
            messagebox.showwarning("Concepto", "Ingresa un concepto.")
            return
        if monto <= 0:
            messagebox.showwarning("Monto", "El monto debe ser mayor a 0.")
            return

        # Verificar tipo de registro según combobox
        if tipo == "PRÉSTAMO":
            self.ds.agregar_prestamo(rep, concepto, monto, observaciones)
        elif tipo == "PAGO PROVEEDOR":
            self.ds.agregar_pago_proveedor(proveedor=concepto, concepto=f"Pago por {rep}", monto=monto, repartidor=rep, observaciones=observaciones)
        elif tipo == "NÓMINA":
            # Guardar en tabla dedicada pago_nomina
            self.ds.agregar_pago_nomina(rep, concepto, monto, observaciones)
        elif tipo == "SOCIO":
            # Guardar en tabla dedicada pago_socios
            self.ds.agregar_pago_socios(rep, concepto, monto, observaciones)
        elif tipo == "TRANSFERENCIA":
            # Guardar en tabla dedicada transferencias
            self.ds.agregar_transferencia(rep, concepto, monto, observaciones)
        else:  # GASTO normal
            self.ds.agregar_gasto(rep, concepto, monto, observaciones)

        # reset campos (excepto repartidor que se mantiene)
        self.gasto_concepto_var.set("")
        self.gasto_monto_var.set("0.00")
        self.gasto_observ_var.set("")
        self.gasto_tipo_var.set("GASTO")
        
        # Refrescar
        self._refrescar_tab_gastos()
        self._refrescar_liquidacion()
        
        # Si fue un préstamo, refrescar también la pestaña de Préstamos
        if tipo == "PRÉSTAMO":
            self._refrescar_prestamos_tab()

    # --- editar gasto con doble clic ---
    def _editar_gasto_doble_clic(self, event):
        """Abre diálogo para editar el gasto seleccionado."""
        sel = self.tree_gastos.selection()
        if not sel:
            return
        
        item_id = sel[0]
        values = self.tree_gastos.item(item_id, "values")
        if not values:
            return
        
        # Determinar tipo y obtener datos
        es_proveedor = item_id.startswith("prov_")
        es_prestamo = item_id.startswith("prest_")
        es_nomina = item_id.startswith("nomina_")
        es_socios = item_id.startswith("socios_")
        es_transferencia = item_id.startswith("transf_")
        registro_id = int(values[0])
        
        self._mostrar_dialogo_editar_gasto(item_id, registro_id, es_proveedor, es_prestamo, es_nomina, es_socios, es_transferencia, values)
    
    def _mostrar_dialogo_editar_gasto(self, item_id, registro_id, es_proveedor, es_prestamo, es_nomina, es_socios, es_transferencia, values):
        """Muestra un diálogo para editar un gasto, pago a proveedor, préstamo, nómina, socios o transferencia."""
        dialog = tk.Toplevel(self.ventana)
        dialog.title("✏️ Editar Registro")
        dialog.geometry("500x380")
        dialog.resizable(False, False)
        dialog.transient(self.ventana)
        
        # Centrar diálogo
        dialog.update_idletasks()
        x = self.ventana.winfo_x() + (self.ventana.winfo_width() - 500) // 2
        y = self.ventana.winfo_y() + (self.ventana.winfo_height() - 380) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Esperar a que la ventana sea visible antes de grab_set
        dialog.wait_visibility()
        dialog.grab_set()
        
        # Frame principal
        frame = ttk.Frame(dialog, padding=20)
        frame.pack(fill=tk.BOTH, expand=True)
        
        # Detectar tipo inicial basado en el origen del registro
        if es_nomina:
            tipo_texto = "💰 NÓMINA"
            tipo_inicial = "NÓMINA"
        elif es_socios:
            tipo_texto = "🤝 SOCIO"
            tipo_inicial = "SOCIO"
        elif es_transferencia:
            tipo_texto = "💸 TRANSFERENCIA"
            tipo_inicial = "TRANSFERENCIA"
        elif es_prestamo:
            tipo_texto = "💵 PRÉSTAMO"
            tipo_inicial = "PRÉSTAMO"
        elif es_proveedor:
            tipo_texto = "💼 PAGO PROVEEDOR"
            tipo_inicial = "PAGO PROVEEDOR"
        else:
            tipo_texto = "🔧 GASTO"
            tipo_inicial = "GASTO"
        
        ttk.Label(frame, text=f"Editando {tipo_texto}", 
                  font=("Segoe UI", 11, "bold")).pack(anchor=tk.W, pady=(0, 15))
        
        # Tipo de registro (Combobox) - mismos tipos que creación
        frame_tipo = ttk.Frame(frame)
        frame_tipo.pack(fill=tk.X, pady=5)
        ttk.Label(frame_tipo, text="Tipo:", width=12).pack(side=tk.LEFT)
        tipo_var = tk.StringVar(value=tipo_inicial)
        tipos_disponibles = ["GASTO", "PAGO PROVEEDOR", "PRÉSTAMO", "NÓMINA", "SOCIO", "TRANSFERENCIA"]
        tipo_combo = ttk.Combobox(frame_tipo, textvariable=tipo_var, width=20,
                                   values=tipos_disponibles, state="readonly")
        tipo_combo.pack(side=tk.LEFT, padx=5)
        
        # Repartidor/Destinatario
        frame_rep = ttk.Frame(frame)
        frame_rep.pack(fill=tk.X, pady=5)
        ttk.Label(frame_rep, text="Repartidor:", width=12).pack(side=tk.LEFT)
        rep_var = tk.StringVar(value=values[2])
        rep_combo = ttk.Combobox(frame_rep, textvariable=rep_var, width=30, 
                                  values=self.ds.get_repartidores())
        rep_combo.pack(side=tk.LEFT, padx=5)
        
        # Concepto (Combobox editable)
        frame_conc = ttk.Frame(frame)
        frame_conc.pack(fill=tk.X, pady=5)
        ttk.Label(frame_conc, text="Concepto:", width=12).pack(side=tk.LEFT)
        conc_var = tk.StringVar(value=values[3])
        conc_combo = ttk.Combobox(frame_conc, textvariable=conc_var, width=32,
                                   values=self.ds.get_conceptos_gastos())
        conc_combo.pack(side=tk.LEFT, padx=5)
        
        # Monto
        frame_monto = ttk.Frame(frame)
        frame_monto.pack(fill=tk.X, pady=5)
        ttk.Label(frame_monto, text="Monto:", width=12).pack(side=tk.LEFT)
        monto_str = values[4].replace("$", "").replace(",", "")
        monto_var = tk.StringVar(value=monto_str)
        ttk.Entry(frame_monto, textvariable=monto_var, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Label(frame_monto, text="$").pack(side=tk.LEFT)
        
        # Observaciones
        frame_obs = ttk.Frame(frame)
        frame_obs.pack(fill=tk.X, pady=5)
        ttk.Label(frame_obs, text="Observaciones:", width=12).pack(side=tk.LEFT)
        obs_var = tk.StringVar(value=values[5] if len(values) > 5 else '')
        ttk.Entry(frame_obs, textvariable=obs_var, width=35).pack(side=tk.LEFT, padx=5)
        
        # Botones
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill=tk.X, pady=(20, 0))
        
        def guardar():
            try:
                nuevo_monto = float(monto_var.get().replace(",", ""))
            except:
                messagebox.showerror("Error", "Monto inválido", parent=dialog)
                return
            
            nuevo_rep = rep_var.get().strip()
            nuevo_conc = conc_var.get().strip()
            nuevo_tipo = tipo_var.get()
            nuevo_obs = obs_var.get().strip()
            
            if not nuevo_rep or not nuevo_conc or nuevo_monto <= 0:
                messagebox.showwarning("Advertencia", "Complete todos los campos", parent=dialog)
                return
            
            # Si el tipo cambió, eliminar el original y crear nuevo
            tipo_cambio = nuevo_tipo != tipo_inicial
            
            if tipo_cambio:
                # Eliminar registro original
                if es_nomina:
                    self.ds.eliminar_pago_nomina(registro_id)
                elif es_socios:
                    self.ds.eliminar_pago_socios(registro_id)
                elif es_transferencia:
                    self.ds.eliminar_transferencia(registro_id)
                elif es_prestamo:
                    self.ds.eliminar_prestamo(registro_id)
                elif es_proveedor:
                    self.ds.eliminar_pago_proveedor(registro_id)
                else:
                    self.ds.eliminar_gasto(registro_id)
                
                # Crear nuevo registro del tipo seleccionado
                if nuevo_tipo == "PRÉSTAMO":
                    self.ds.agregar_prestamo(nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif nuevo_tipo == "PAGO PROVEEDOR":
                    self.ds.agregar_pago_proveedor(proveedor=nuevo_conc, concepto=f"Pago por {nuevo_rep}", 
                                                   monto=nuevo_monto, repartidor=nuevo_rep, observaciones=nuevo_obs)
                elif nuevo_tipo == "NÓMINA":
                    self.ds.agregar_pago_nomina(nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif nuevo_tipo == "SOCIO":
                    self.ds.agregar_pago_socios(nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif nuevo_tipo == "TRANSFERENCIA":
                    self.ds.agregar_transferencia(nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                else:  # GASTO
                    self.ds.agregar_gasto(nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
            else:
                # Actualizar registro existente del mismo tipo
                if es_nomina:
                    self.ds.actualizar_pago_nomina(registro_id, nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif es_socios:
                    self.ds.actualizar_pago_socios(registro_id, nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif es_transferencia:
                    self.ds.actualizar_transferencia(registro_id, nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif es_prestamo:
                    self.ds.actualizar_prestamo(registro_id, nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
                elif es_proveedor:
                    self.ds.actualizar_pago_proveedor(registro_id, nuevo_conc, f"Pago por {nuevo_rep}", 
                                                       nuevo_monto, nuevo_rep, nuevo_obs)
                else:
                    self.ds.actualizar_gasto(registro_id, nuevo_rep, nuevo_conc, nuevo_monto, nuevo_obs)
            
            dialog.destroy()
            self._refrescar_tab_gastos()
            self._refrescar_liquidacion()
            # Si involucra préstamos, refrescar pestaña Préstamos
            if es_prestamo or nuevo_tipo == "PRÉSTAMO" or tipo_inicial == "PRÉSTAMO":
                self._refrescar_prestamos_tab()
        
        def eliminar():
            if messagebox.askyesno("Confirmar", "¿Eliminar este registro?", parent=dialog):
                if es_nomina:
                    self.ds.eliminar_pago_nomina(registro_id)
                elif es_socios:
                    self.ds.eliminar_pago_socios(registro_id)
                elif es_transferencia:
                    self.ds.eliminar_transferencia(registro_id)
                elif es_prestamo:
                    self.ds.eliminar_prestamo(registro_id)
                elif es_proveedor:
                    self.ds.eliminar_pago_proveedor(registro_id)
                else:
                    self.ds.eliminar_gasto(registro_id)
                dialog.destroy()
                self._refrescar_tab_gastos()
                self._refrescar_liquidacion()
                # Si era préstamo, refrescar pestaña Préstamos
                if es_prestamo:
                    self._refrescar_prestamos_tab()
        
        ttk.Button(frame_btns, text="💾 Guardar", command=guardar).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_btns, text="🗑️ Eliminar", command=eliminar).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_btns, text="❌ Cancelar", command=dialog.destroy).pack(side=tk.RIGHT, padx=5)

    # --- clic derecho en tabla de gastos → eliminar ---
    def _on_gastos_right_click(self, event):
        row = self.tree_gastos.identify_row(event.y)
        
        menu = tk.Menu(self.ventana, tearoff=0)
        
        # Opciones de copiar siempre disponibles
        menu.add_command(
            label="📋 Copiar selección (Ctrl+C)",
            command=lambda: self._copiar_seleccion_tree(self.tree_gastos)
        )
        menu.add_command(
            label="📄 Copiar toda la tabla",
            command=lambda: self._copiar_toda_tabla(self.tree_gastos)
        )
        
        # Si hay fila seleccionada, agregar opciones
        if row:
            self.tree_gastos.selection_set(row)
            values = self.tree_gastos.item(row, "values")
            if values:
                menu.add_separator()
                menu.add_command(
                    label="✏️ Editar registro",
                    command=lambda: self._editar_gasto_doble_clic(None)
                )
                
                # Determinar tipo de registro
                es_proveedor = row.startswith("prov_")
                es_prestamo = row.startswith("prest_")
                es_nomina = row.startswith("nomina_")
                es_socios = row.startswith("socios_")
                es_transferencia = row.startswith("transf_")
                registro_id = int(values[0])
                
                menu.add_command(
                    label="🗑️ Eliminar registro",
                    command=lambda: self._eliminar_registro_gastos(es_proveedor, es_prestamo, es_nomina, es_socios, es_transferencia, registro_id, values)
                )

        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
    
    def _eliminar_registro_gastos(self, es_proveedor, es_prestamo, es_nomina, es_socios, es_transferencia, registro_id, values):
        """Elimina un gasto, pago a proveedor, préstamo, nómina, socios o transferencia."""
        if es_nomina:
            tipo = "pago de nómina"
        elif es_socios:
            tipo = "pago a socios"
        elif es_transferencia:
            tipo = "transferencia"
        elif es_prestamo:
            tipo = "préstamo"
        elif es_proveedor:
            tipo = "pago a proveedor"
        else:
            tipo = "gasto"
            
        if messagebox.askyesno("Confirmar",
                               f"¿Eliminar {tipo}?\n\n"
                               f"Concepto: {values[3]}\n"
                               f"Monto: {values[4]}"):
            if es_nomina:
                self.ds.eliminar_pago_nomina(registro_id)
            elif es_socios:
                self.ds.eliminar_pago_socios(registro_id)
            elif es_transferencia:
                self.ds.eliminar_transferencia(registro_id)
            elif es_prestamo:
                self.ds.eliminar_prestamo(registro_id)
            elif es_proveedor:
                self.ds.eliminar_pago_proveedor(registro_id)
            else:
                self.ds.eliminar_gasto(registro_id)
            # Refrescar tabla y liquidación
            self._refrescar_tab_gastos()
            self._refrescar_liquidacion()
            # Si era préstamo, refrescar pestaña Préstamos
            if es_prestamo:
                self._refrescar_prestamos_tab()

    def _eliminar_gasto(self, idx: int):
        gastos = self.ds.get_gastos()
        if 0 <= idx < len(gastos):
            g = gastos[idx]
            if messagebox.askyesno("Confirmar",
                                   f"¿Eliminar gasto?\n\n"
                                   f"Repartidor: {g['repartidor']}\n"
                                   f"Concepto:   {g['concepto']}\n"
                                   f"Monto:      ${g['monto']:,.2f}"):
                self.ds.eliminar_gasto(g.get('id', idx))
                # Refrescar
                self._refrescar_tab_gastos()
                self._refrescar_liquidacion()

    # ==================================================================
    # PESTAÑA 4 – CONTEO DE DINERO (por repartidor)
    # ==================================================================
    DENOMINACIONES = [
        ("BILLETES", [
            ("$100.000", 100000), ("$50.000", 50000), ("$20.000", 20000),
            ("$10.000", 10000),   ("$5.000", 5000),   ("$2.000", 2000),
        ]),
        ("MONEDAS", [
            ("$1.000", 1000), ("$500", 500), ("$200", 200),
            ("$100", 100),    ("$50", 50),
        ])
    ]
    # Lista plana de valores para iterar en orden
    _VALORES_ORDEN = [100000, 50000, 20000, 10000, 5000, 2000,
                      1000, 500, 200, 100, 50]

    def _crear_tab_dinero(self):
        # Colores del modo oscuro
        BG_DARK = '#1e1e1e'
        BG_CARD = '#2d2d2d'
        TEXT_PRIMARY = '#ffffff'
        PRIMARY = '#2196f3'
        
        # Variable para rastrear la sesión de conteo actual
        self._sesion_conteo_actual = None
        
        # --- barra superior: selección de repartidor y botones ---
        frame_top = ttk.Frame(self.tab_dinero)
        frame_top.pack(fill=tk.X, padx=10, pady=(10, 4))

        ttk.Label(frame_top, text="Repartidor:").pack(side=tk.LEFT, padx=(0, 4))
        self.dinero_rep_var = tk.StringVar()
        self.dinero_rep_combo = ttk.Combobox(frame_top, textvariable=self.dinero_rep_var,
                                             width=20, state="readonly")
        self.dinero_rep_combo.pack(side=tk.LEFT, padx=(0, 8))
        self.dinero_rep_combo.bind("<<ComboboxSelected>>", self._on_dinero_rep_cambio)
        
        # Botones de acción para conteos múltiples
        ttk.Button(frame_top, text="➕ Nuevo Conteo", 
                   command=self._nuevo_conteo_sesion, style="Success.TButton").pack(side=tk.LEFT, padx=(10, 5))
        ttk.Button(frame_top, text="💾 Guardar", 
                   command=self._guardar_conteo_sesion_actual).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(frame_top, text="🗑️ Eliminar", 
                   command=self._eliminar_conteo_sesion_actual, style="Danger.TButton").pack(side=tk.LEFT)

        # ============================================================
        # CONTENEDOR PRINCIPAL CON DOS PANELES (IZQUIERDA Y DERECHA)
        # ============================================================
        frame_principal = ttk.Frame(self.tab_dinero)
        frame_principal.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        frame_principal.columnconfigure(0, weight=1)
        frame_principal.columnconfigure(1, weight=1)
        frame_principal.rowconfigure(0, weight=1)
        
        # ============================================================
        # PANEL IZQUIERDO: CONTEO DE DINERO
        # ============================================================
        frame_outer = ttk.LabelFrame(frame_principal, text="💰 CONTEO DE DINERO", padding=(8, 6))
        frame_outer.grid(row=0, column=0, sticky="nsew", padx=(0, 5))

        canvas = tk.Canvas(frame_outer, highlightthickness=0, bg=BG_CARD)
        scroll = ttk.Scrollbar(frame_outer, orient=tk.VERTICAL, command=canvas.yview)
        self._dinero_frame_scroll = tk.Frame(canvas, bg=BG_CARD)

        self._dinero_frame_scroll.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self._dinero_frame_scroll, anchor=tk.NW)
        canvas.configure(yscrollcommand=scroll.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # vars de cantidad y subtotal por denominación
        self.denom_vars = {}   # valor_int → StringVar (cantidad)
        self.denom_sub  = {}   # valor_int → StringVar (subtotal mostrado)
        self._dinero_trace_ids = []  # para desconectar traces al cambiar rep

        for grupo, items in self.DENOMINACIONES:
            lbl_grupo = tk.Label(self._dinero_frame_scroll, text=grupo,
                      font=("Segoe UI", 10, "bold"), bg=BG_CARD, fg=PRIMARY)
            lbl_grupo.pack(fill=tk.X, pady=(12, 4), padx=4)
            
            sep = tk.Frame(self._dinero_frame_scroll, height=1, bg='#404040')
            sep.pack(fill=tk.X, padx=4)

            for nombre, valor in items:
                row = tk.Frame(self._dinero_frame_scroll, bg=BG_CARD)
                row.pack(fill=tk.X, pady=2, padx=4)

                tk.Label(row, text=nombre, width=11, anchor=tk.E, 
                        bg=BG_CARD, fg=TEXT_PRIMARY, font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 6))

                cant_var = tk.StringVar(value="0")
                self.denom_vars[valor] = cant_var

                tk.Label(row, text="×", bg=BG_CARD, fg=TEXT_PRIMARY).pack(side=tk.LEFT)
                
                entry = tk.Entry(row, textvariable=cant_var, width=7,
                          justify=tk.CENTER, bg='#3d3d3d', fg=TEXT_PRIMARY,
                          insertbackground=TEXT_PRIMARY, relief='flat',
                          font=("Segoe UI", 9))
                entry.pack(side=tk.LEFT, padx=(4, 10))

                sub_var = tk.StringVar(value="$0")
                self.denom_sub[valor] = sub_var
                tk.Label(row, textvariable=sub_var, width=14,
                          anchor=tk.E, bg=BG_CARD, fg='#4caf50', 
                          font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT)

                tid = cant_var.trace_add("write", lambda *a: self._recalcular_dinero())
                self._dinero_trace_ids.append((cant_var, tid))

        # total inferior
        sep_total = tk.Frame(self._dinero_frame_scroll, height=2, bg=PRIMARY)
        sep_total.pack(fill=tk.X, pady=8, padx=4)
        
        frame_total = tk.Frame(self._dinero_frame_scroll, bg=BG_CARD)
        frame_total.pack(fill=tk.X, pady=4, padx=4)
        
        tk.Label(frame_total, text="TOTAL:", font=("Segoe UI", 11, "bold"),
                  width=24, anchor=tk.E, bg=BG_CARD, fg=TEXT_PRIMARY).pack(side=tk.LEFT)
        self.lbl_dinero_total = tk.Label(frame_total, text="$0.00",
                                          font=("Segoe UI", 14, "bold"), 
                                          bg=BG_CARD, fg=PRIMARY)
        self.lbl_dinero_total.pack(side=tk.LEFT, padx=6)

        # ============================================================
        # PANEL DERECHO: RESUMEN DE CONTEOS POR REPARTIDOR
        # ============================================================
        # PANEL DERECHO: DIVIDIDO EN DOS SECCIONES
        # ============================================================
        frame_derecho = ttk.Frame(frame_principal)
        frame_derecho.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        frame_derecho.rowconfigure(0, weight=1)
        frame_derecho.rowconfigure(1, weight=1)
        frame_derecho.columnconfigure(0, weight=1)
        
        # ============================================================
        # SECCIÓN SUPERIOR: CONTEOS DEL REPARTIDOR SELECCIONADO
        # ============================================================
        frame_conteos_rep = ttk.LabelFrame(frame_derecho, text="📋 CONTEOS DEL REPARTIDOR", padding=(8, 6))
        frame_conteos_rep.grid(row=0, column=0, sticky="nsew", pady=(0, 5))
        
        # Tabla de conteos del repartidor
        tree_sesiones_container = ttk.Frame(frame_conteos_rep)
        tree_sesiones_container.pack(fill=tk.BOTH, expand=True)
        
        self.tree_sesiones = ttk.Treeview(
            tree_sesiones_container,
            columns=("id", "hora", "descripcion", "total"),
            selectmode="browse", height=5
        )
        self.tree_sesiones.column("#0", width=0, stretch=tk.NO)
        self.tree_sesiones.column("id", width=0, stretch=tk.NO)  # Oculto
        self.tree_sesiones.column("hora", anchor=tk.CENTER, width=60)
        self.tree_sesiones.column("descripcion", anchor=tk.W, width=120)
        self.tree_sesiones.column("total", anchor=tk.E, width=100)
        
        self.tree_sesiones.heading("hora", text="🕐 Hora")
        self.tree_sesiones.heading("descripcion", text="📝 Descripción")
        self.tree_sesiones.heading("total", text="💵 Total")
        
        self.tree_sesiones.tag_configure("par", background="#2d2d2d", foreground="#ffffff")
        self.tree_sesiones.tag_configure("impar", background="#3d3d3d", foreground="#ffffff")
        self.tree_sesiones.tag_configure("seleccionado", background="#1565c0", foreground="#ffffff")
        
        scroll_sesiones = ttk.Scrollbar(tree_sesiones_container, orient=tk.VERTICAL, 
                                         command=self.tree_sesiones.yview)
        self.tree_sesiones.configure(yscrollcommand=scroll_sesiones.set)
        
        self.tree_sesiones.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_sesiones.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Binding para cargar conteo al seleccionar
        self.tree_sesiones.bind("<<TreeviewSelect>>", self._on_sesion_seleccionada)
        
        # Total del repartidor
        frame_total_rep = ttk.Frame(frame_conteos_rep)
        frame_total_rep.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Label(frame_total_rep, text="💰 Total Repartidor:", 
                  font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        self.lbl_total_rep_conteos = ttk.Label(frame_total_rep, text="$0.00", 
                                                font=("Segoe UI", 11, "bold"),
                                                foreground="#4caf50")
        self.lbl_total_rep_conteos.pack(side=tk.LEFT)
        
        # ============================================================
        # SECCIÓN INFERIOR: RESUMEN GENERAL DE TODOS LOS REPARTIDORES
        # ============================================================
        frame_resumen = ttk.LabelFrame(frame_derecho, text="📊 RESUMEN GENERAL DEL DÍA", padding=(8, 6))
        frame_resumen.grid(row=1, column=0, sticky="nsew", pady=(5, 0))
        
        # Tabla de resumen
        tree_conteo_container = ttk.Frame(frame_resumen)
        tree_conteo_container.pack(fill=tk.BOTH, expand=True)
        
        self.tree_conteos = ttk.Treeview(
            tree_conteo_container,
            columns=("repartidor", "num_conteos", "total"),
            selectmode="browse", height=5
        )
        self.tree_conteos.column("#0", width=0, stretch=tk.NO)
        self.tree_conteos.column("repartidor", anchor=tk.W, width=120)
        self.tree_conteos.column("num_conteos", anchor=tk.CENTER, width=60)
        self.tree_conteos.column("total", anchor=tk.E, width=100)
        
        self.tree_conteos.heading("repartidor", text="Repartidor")
        self.tree_conteos.heading("num_conteos", text="# Conteos")
        self.tree_conteos.heading("total", text="Total")
        
        # Configurar tags para colores alternados (igual que otros treeviews)
        self.tree_conteos.tag_configure("par", background="#2d2d2d", foreground="#ffffff")
        self.tree_conteos.tag_configure("impar", background="#3d3d3d", foreground="#ffffff")
        
        scroll_conteos = ttk.Scrollbar(tree_conteo_container, orient=tk.VERTICAL, 
                                        command=self.tree_conteos.yview)
        self.tree_conteos.configure(yscrollcommand=scroll_conteos.set)
        
        self.tree_conteos.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_conteos.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Binding para ver detalle al hacer doble click
        self.tree_conteos.bind("<Double-1>", self._ver_detalle_conteo)
        # Binding para cargar el repartidor al seleccionar
        self.tree_conteos.bind("<<TreeviewSelect>>", self._on_conteo_seleccionado)
        
        # Total general de conteos
        frame_total_general = ttk.Frame(frame_resumen)
        frame_total_general.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Separator(frame_total_general, orient="horizontal").pack(fill=tk.X, pady=3)
        
        frame_suma = ttk.Frame(frame_total_general)
        frame_suma.pack(fill=tk.X)
        
        ttk.Label(frame_suma, text="💰 TOTAL GENERAL:", 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        self.lbl_total_general_conteos = ttk.Label(frame_suma, text="$0.00", 
                                                    font=("Segoe UI", 12, "bold"),
                                                    foreground="#2196f3")
        self.lbl_total_general_conteos.pack(side=tk.LEFT)

    # --- cuando el usuario cambia de repartidor en el combo ---
    def _on_dinero_rep_cambio(self, event=None):
        # El nuevo repartidor ya está en la variable
        nuevo_rep = self.dinero_rep_var.get()
        
        # Limpiar selección de sesión actual
        self._sesion_conteo_actual = None
        
        # Actualizar tabla de conteos del repartidor
        self._actualizar_tabla_sesiones_rep()
        
        # Limpiar campos de conteo para nuevo ingreso
        self._limpiar_campos_conteo()
        
        # Actualizar referencia del repartidor actual
        self._rep_dinero_anterior = nuevo_rep

    # --- limpiar todos los campos de conteo ---
    def _limpiar_campos_conteo(self):
        """Limpia todos los campos de denominación y el total."""
        for valor, var in self.denom_vars.items():
            var.set("0")
        self._recalcular_dinero_sin_guardar()

    # --- obtener el conteo actual de los campos ---
    def _obtener_conteo_actual(self) -> dict:
        """Obtiene el conteo actual de los campos como dict {denom: cant}."""
        conteo = {}
        for valor, var in self.denom_vars.items():
            try:
                cant = int(var.get() or 0)
                if cant > 0:
                    conteo[valor] = cant
            except ValueError:
                pass
        return conteo

    # --- nuevo conteo (sesión) ---
    def _nuevo_conteo_sesion(self):
        """Crea una nueva sesión de conteo para el repartidor actual."""
        rep = self.dinero_rep_var.get()
        if not rep:
            messagebox.showwarning("Advertencia", "Seleccione un repartidor primero")
            return
        
        # Limpiar campos para nuevo conteo
        self._limpiar_campos_conteo()
        self._sesion_conteo_actual = None  # Nueva sesión (sin ID aún)

    # --- guardar conteo sesión actual ---
    def _guardar_conteo_sesion_actual(self):
        """Guarda el conteo actual como una sesión (nueva o existente)."""
        rep = self.dinero_rep_var.get()
        if not rep:
            messagebox.showwarning("Advertencia", "Seleccione un repartidor primero")
            return
        
        conteo = self._obtener_conteo_actual()
        if not conteo or sum(conteo.values()) == 0:
            messagebox.showwarning("Advertencia", "Ingrese al menos una denominación")
            return
        
        # Pedir descripción opcional
        descripcion = simpledialog.askstring(
            "Descripción del Conteo",
            "Ingrese una descripción (opcional):",
            parent=self.ventana
        ) or ""
        
        if USE_SQLITE and self.ds.fecha:
            if self._sesion_conteo_actual:
                # Actualizar sesión existente
                ok = db_local.actualizar_conteo_sesion(
                    self._sesion_conteo_actual, conteo, descripcion
                )
                if ok:
                    messagebox.showinfo("Éxito", "Conteo actualizado correctamente")
            else:
                # Crear nueva sesión
                sesion_id = db_local.guardar_conteo_sesion(
                    self.ds.fecha, rep, conteo, descripcion
                )
                if sesion_id:
                    self._sesion_conteo_actual = sesion_id
                    messagebox.showinfo("Éxito", "Conteo guardado correctamente")
                else:
                    messagebox.showerror("Error", "No se pudo guardar el conteo")
                    return
        
        # Actualizar tablas
        self._actualizar_tabla_sesiones_rep()
        self._actualizar_tabla_conteos()
        self._refrescar_liquidacion()

    # --- eliminar conteo sesión actual ---
    def _eliminar_conteo_sesion_actual(self):
        """Elimina la sesión de conteo actualmente seleccionada."""
        if not self._sesion_conteo_actual:
            messagebox.showwarning("Advertencia", "Seleccione un conteo para eliminar")
            return
        
        if not messagebox.askyesno("Confirmar", "¿Está seguro de eliminar este conteo?"):
            return
        
        if USE_SQLITE:
            ok = db_local.eliminar_conteo_sesion(self._sesion_conteo_actual)
            if ok:
                messagebox.showinfo("Éxito", "Conteo eliminado")
                self._sesion_conteo_actual = None
                self._limpiar_campos_conteo()
                self._actualizar_tabla_sesiones_rep()
                self._actualizar_tabla_conteos()
                self._refrescar_liquidacion()
            else:
                messagebox.showerror("Error", "No se pudo eliminar el conteo")

    # --- cuando se selecciona una sesión de la tabla ---
    def _on_sesion_seleccionada(self, event=None):
        """Carga los datos de la sesión seleccionada en los campos."""
        sel = self.tree_sesiones.selection()
        if not sel:
            return
        
        values = self.tree_sesiones.item(sel[0], "values")
        if not values:
            return
        
        sesion_id = int(values[0])
        self._sesion_conteo_actual = sesion_id
        
        # Cargar detalle de la sesión
        if USE_SQLITE:
            detalle = db_local.obtener_detalle_conteo_sesion(sesion_id)
            
            # Actualizar campos
            for valor, var in self.denom_vars.items():
                var.set(str(detalle.get(valor, 0)))
            
            self._recalcular_dinero_sin_guardar()

    # --- actualizar tabla de sesiones del repartidor ---
    def _actualizar_tabla_sesiones_rep(self):
        """Actualiza la tabla de conteos (sesiones) del repartidor seleccionado."""
        if not hasattr(self, 'tree_sesiones'):
            return
        
        # Limpiar tabla
        for item in self.tree_sesiones.get_children():
            self.tree_sesiones.delete(item)
        
        rep = self.dinero_rep_var.get()
        total_rep = 0
        
        if USE_SQLITE and self.ds.fecha and rep:
            sesiones = db_local.obtener_conteos_sesion_repartidor(self.ds.fecha, rep)
            for i, sesion in enumerate(sesiones):
                sesion_id = sesion.get('id', 0)
                hora = sesion.get('hora', '')
                desc = sesion.get('descripcion', '') or f"Conteo #{i+1}"
                total = sesion.get('total', 0)
                total_rep += total
                
                tag = "par" if i % 2 == 0 else "impar"
                self.tree_sesiones.insert(
                    "", tk.END,
                    values=(sesion_id, hora, desc, f"${total:,.2f}"),
                    tags=(tag,)
                )
        
        # Actualizar total del repartidor
        if hasattr(self, 'lbl_total_rep_conteos'):
            self.lbl_total_rep_conteos.config(text=f"${total_rep:,.2f}")

    # --- guardar el conteo de un repartidor específico (compatibilidad) ---
    def _guardar_dinero_rep_especifico(self, repartidor: str):
        """Compatibilidad: no hace nada, los conteos se guardan explícitamente."""
        pass

    # --- guardar el estado actual de las entradas al DataStore (compatibilidad) ---
    def _guardar_dinero_actual(self):
        """Compatibilidad: no hace nada automáticamente."""
        pass

    # --- cargar conteo desde DataStore al grid de entradas ---
    def _cargar_dinero_rep(self, repartidor: str):
        """Carga los conteos del repartidor y muestra la lista."""
        # Limpiar campos
        self._limpiar_campos_conteo()
        self._sesion_conteo_actual = None
        
        # Actualizar tabla de sesiones
        self._actualizar_tabla_sesiones_rep()
        
        # Marcar este repartidor como el anterior para el próximo cambio
        self._rep_dinero_anterior = repartidor

    # --- recalcular subtotales y total SIN guardar (para carga inicial) ---
    def _recalcular_dinero_sin_guardar(self):
        total = 0
        for valor, var in self.denom_vars.items():
            try:
                cant = int(var.get() or 0)
            except ValueError:
                cant = 0
            sub = cant * valor
            total += sub
            self.denom_sub[valor].set(f"${sub:,}")
        self.lbl_dinero_total.config(text=f"${total:,.2f}")

    # --- recalcular subtotales y total al cambiar una cantidad ---
    def _recalcular_dinero(self):
        """Recalcula subtotales y total (no guarda automáticamente)."""
        total = 0
        for valor, var in self.denom_vars.items():
            try:
                cant = int(var.get() or 0)
            except ValueError:
                cant = 0
            sub = cant * valor
            total += sub
            self.denom_sub[valor].set(f"${sub:,}")

        self.lbl_dinero_total.config(text=f"${total:,.2f}")

    # --- refrescar combo de repartidores (llamado desde _on_data_changed) ---
    def _refrescar_tab_dinero(self):
        reps = self.ds.get_repartidores()
        self.dinero_rep_combo['values'] = reps
        actual = self.dinero_rep_var.get()
        if actual not in reps:
            # seleccionar el primero disponible (o quedarlo vacío)
            nuevo_rep = reps[0] if reps else ""
            self.dinero_rep_var.set(nuevo_rep)
            if nuevo_rep:
                self._cargar_dinero_rep(nuevo_rep)
                self._rep_dinero_anterior = nuevo_rep
        else:
            # Actualizar sesiones del repartidor actual
            self._actualizar_tabla_sesiones_rep()
        # Actualizar tabla de resumen de conteos
        self._actualizar_tabla_conteos()

    def _actualizar_tabla_conteos(self):
        """Actualiza la tabla de resumen de conteos de todos los repartidores."""
        if not hasattr(self, 'tree_conteos'):
            return
        
        # Limpiar tabla
        for item in self.tree_conteos.get_children():
            self.tree_conteos.delete(item)
        
        # Obtener resumen de conteos desde SQLite (múltiples)
        total_general = 0
        if USE_SQLITE and self.ds.fecha:
            resumen = db_local.obtener_resumen_conteos_multiples_fecha(self.ds.fecha)
            for i, conteo in enumerate(resumen):
                repartidor = conteo.get('repartidor', '')
                num_conteos = conteo.get('num_conteos', 0)
                total = conteo.get('total', 0)
                total_general += total
                # Usar tags alternados para filas pares/impares
                tag = "par" if i % 2 == 0 else "impar"
                self.tree_conteos.insert(
                    "", tk.END,
                    values=(repartidor, num_conteos, f"${total:,.2f}"),
                    tags=(tag,)
                )
        
        # Actualizar total general
        self.lbl_total_general_conteos.config(text=f"${total_general:,.2f}")

    def _on_conteo_seleccionado(self, event=None):
        """Carga el conteo del repartidor seleccionado en la tabla de resumen."""
        sel = self.tree_conteos.selection()
        if not sel:
            return
        
        values = self.tree_conteos.item(sel[0], "values")
        if not values:
            return
        
        repartidor = values[0]
        
        # Cambiar el combo al repartidor seleccionado
        self.dinero_rep_var.set(repartidor)
        
        # Cargar los datos del repartidor (sesiones)
        self._sesion_conteo_actual = None
        self._limpiar_campos_conteo()
        self._actualizar_tabla_sesiones_rep()
        self._rep_dinero_anterior = repartidor

    def _ver_detalle_conteo(self, event=None):
        """Muestra el detalle de todos los conteos de un repartidor al hacer doble click."""
        sel = self.tree_conteos.selection()
        if not sel:
            return
        
        values = self.tree_conteos.item(sel[0], "values")
        if not values:
            return
        
        repartidor = values[0]
        
        # Obtener todas las sesiones del repartidor
        if USE_SQLITE and self.ds.fecha:
            sesiones = db_local.obtener_conteos_sesion_repartidor(self.ds.fecha, repartidor)
            
            # Crear ventana de detalle
            ventana = tk.Toplevel(self.ventana)
            ventana.title(f"Detalle Conteos - {repartidor}")
            ventana.geometry("500x500")
            ventana.transient(self.ventana)
            
            # Frame principal
            frame = ttk.Frame(ventana, padding=10)
            frame.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(frame, text=f"💰 CONTEOS DE {repartidor}", 
                      font=("Segoe UI", 11, "bold")).pack(pady=(0, 10))
            
            # Crear un notebook para cada sesión
            total_general = 0
            
            for i, sesion in enumerate(sesiones):
                sesion_id = sesion.get('id', 0)
                hora = sesion.get('hora', '')
                desc = sesion.get('descripcion', '') or f"Conteo #{i+1}"
                total_sesion = sesion.get('total', 0)
                total_general += total_sesion
                
                # Frame para cada sesión
                frame_sesion = ttk.LabelFrame(frame, text=f"🕐 {hora} - {desc}", padding=5)
                frame_sesion.pack(fill=tk.X, pady=5)
                
                # Detalle de la sesión
                detalle = db_local.obtener_detalle_conteo_sesion(sesion_id)
                
                # Mostrar denominaciones en una línea
                denoms_texto = []
                for denom, cant in sorted(detalle.items(), reverse=True):
                    if cant > 0:
                        denoms_texto.append(f"${denom:,}×{cant}")
                
                texto_denoms = " | ".join(denoms_texto) if denoms_texto else "Sin detalle"
                ttk.Label(frame_sesion, text=texto_denoms, 
                          font=("Segoe UI", 9), foreground="#888888").pack(anchor=tk.W)
                
                ttk.Label(frame_sesion, text=f"Total: ${total_sesion:,.2f}", 
                          font=("Segoe UI", 10, "bold"), foreground="#4caf50").pack(anchor=tk.E)
            
            # Total general
            ttk.Separator(frame, orient="horizontal").pack(fill=tk.X, pady=10)
            ttk.Label(frame, text=f"💰 TOTAL GENERAL: ${total_general:,.2f}", 
                      font=("Segoe UI", 12, "bold"), foreground="#2196f3").pack()
            
            # Botón cerrar
            ttk.Button(frame, text="Cerrar", command=ventana.destroy).pack(pady=10)

    def _editar_repartidor_conteo(self):
        """Permite editar el nombre del repartidor en un conteo."""
        sel = self.tree_conteos.selection()
        if not sel:
            messagebox.showwarning("Selección", "Selecciona un conteo de la tabla para editar.")
            return
        
        values = self.tree_conteos.item(sel[0], "values")
        if not values:
            return
        
        repartidor_actual = values[0]
        
        # Crear ventana de edición
        ventana = tk.Toplevel(self)
        ventana.title("Editar Repartidor")
        ventana.geometry("350x150")
        ventana.transient(self)
        ventana.grab_set()
        
        frame = ttk.Frame(ventana, padding=15)
        frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(frame, text="✏️ EDITAR REPARTIDOR", 
                  font=("Segoe UI", 11, "bold")).pack(pady=(0, 15))
        
        # Repartidor actual
        frame_actual = ttk.Frame(frame)
        frame_actual.pack(fill=tk.X, pady=5)
        ttk.Label(frame_actual, text="Actual:").pack(side=tk.LEFT)
        ttk.Label(frame_actual, text=repartidor_actual, font=("Segoe UI", 9, "bold"),
                  foreground="#1565c0").pack(side=tk.LEFT, padx=(10, 0))
        
        # Nuevo nombre
        frame_nuevo = ttk.Frame(frame)
        frame_nuevo.pack(fill=tk.X, pady=5)
        ttk.Label(frame_nuevo, text="Nuevo:").pack(side=tk.LEFT)
        
        # Combo con repartidores existentes o entrada libre
        reps = self.ds.get_repartidores()
        nuevo_var = tk.StringVar(value=repartidor_actual)
        combo_nuevo = ttk.Combobox(frame_nuevo, textvariable=nuevo_var, values=reps, width=20)
        combo_nuevo.pack(side=tk.LEFT, padx=(10, 0))
        
        def guardar_cambio():
            nuevo_rep = nuevo_var.get().strip().upper()
            if not nuevo_rep:
                messagebox.showwarning("Error", "El nombre no puede estar vacío.")
                return
            if nuevo_rep == repartidor_actual:
                ventana.destroy()
                return
            
            # Actualizar en la BD
            if USE_SQLITE and self.ds.fecha:
                if db_local.actualizar_repartidor_conteo(self.ds.fecha, repartidor_actual, nuevo_rep):
                    messagebox.showinfo("Éxito", f"Repartidor cambiado de '{repartidor_actual}' a '{nuevo_rep}'")
                    ventana.destroy()
                    # Refrescar tabla y liquidación
                    self._actualizar_tabla_conteos()
                    self._refrescar_liquidacion()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar el repartidor.")
        
        # Botones
        frame_btns = ttk.Frame(frame)
        frame_btns.pack(fill=tk.X, pady=(15, 0))
        ttk.Button(frame_btns, text="💾 Guardar", command=guardar_cambio).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame_btns, text="Cancelar", command=ventana.destroy).pack(side=tk.LEFT)

    # ==================================================================
    # TAB AUDITORÍA DE CORTE
    # ==================================================================
    def _crear_tab_auditoria(self):
        """Crea la pestaña de auditoría para analizar el corte de caja"""
        tab = self.tab_auditoria
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(1, weight=1)

        # ══════════════════════════════════════════════════════════════
        # FRAME SUPERIOR - Controles
        # ══════════════════════════════════════════════════════════════
        frame_ctrl = ttk.LabelFrame(tab, text="  🔍 Parámetros de Auditoría  ", padding=10)
        frame_ctrl.grid(row=0, column=0, sticky="ew", padx=10, pady=(10,5))
        frame_ctrl.columnconfigure(5, weight=1)

        # Fecha
        ttk.Label(frame_ctrl, text="📅 Fecha:", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=(0,5))
        self.audit_fecha_var = tk.StringVar(value=self.ds.fecha)
        self.audit_fecha_entry = ttk.Entry(frame_ctrl, textvariable=self.audit_fecha_var, width=12)
        self.audit_fecha_entry.grid(row=0, column=1, padx=5)

        # Botón Analizar
        btn_analizar = ttk.Button(frame_ctrl, text="📊 Analizar Corte", command=self._ejecutar_auditoria)
        btn_analizar.grid(row=0, column=2, padx=15)

        # Separador
        ttk.Separator(frame_ctrl, orient="vertical").grid(row=0, column=3, sticky="ns", padx=10)

        # Buscar Folio específico
        ttk.Label(frame_ctrl, text="🔎 Buscar Folio:", font=("Segoe UI", 10, "bold")).grid(row=0, column=4, padx=(0,5))
        self.audit_folio_var = tk.StringVar()
        self.audit_folio_entry = ttk.Entry(frame_ctrl, textvariable=self.audit_folio_var, width=10)
        self.audit_folio_entry.grid(row=0, column=5, padx=5, sticky="w")
        btn_buscar = ttk.Button(frame_ctrl, text="🔍 Buscar", command=self._buscar_folio_auditoria)
        btn_buscar.grid(row=0, column=6, padx=5)

        # ══════════════════════════════════════════════════════════════
        # FRAME PRINCIPAL - Contenido con scroll
        # ══════════════════════════════════════════════════════════════
        frame_main = ttk.Frame(tab)
        frame_main.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
        frame_main.columnconfigure(0, weight=1)
        frame_main.rowconfigure(0, weight=1)

        # Canvas con scroll
        canvas = tk.Canvas(frame_main, highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame_main, orient="vertical", command=canvas.yview)
        self.audit_scroll_frame = ttk.Frame(canvas)

        self.audit_scroll_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.audit_scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # Scroll con rueda del mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # Para Linux
        canvas.bind_all("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind_all("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        self.audit_canvas = canvas
        self.audit_scroll_frame.columnconfigure(0, weight=1)

        # Inicializar contenido vacío
        self._crear_contenido_auditoria_vacio()

    def _crear_contenido_auditoria_vacio(self):
        """Muestra un mensaje inicial en la pestaña de auditoría"""
        for widget in self.audit_scroll_frame.winfo_children():
            widget.destroy()

        frame = ttk.Frame(self.audit_scroll_frame, padding=50)
        frame.grid(row=0, column=0, sticky="nsew")

        ttk.Label(frame, text="📊 Auditoría de Corte de Caja", 
                  font=("Segoe UI", 16, "bold")).pack(pady=10)
        ttk.Label(frame, text="Seleccione una fecha y presione 'Analizar Corte'\npara ver el desglose detallado de facturas.",
                  font=("Segoe UI", 11), justify="center").pack(pady=20)
        ttk.Label(frame, text="También puede buscar un folio específico\npara ver su información detallada.",
                  font=("Segoe UI", 10), foreground="gray", justify="center").pack(pady=10)

    def _ejecutar_auditoria(self):
        """Ejecuta el análisis de auditoría para la fecha seleccionada"""
        fecha = self.audit_fecha_var.get().strip()
        if not fecha:
            messagebox.showwarning("Fecha requerida", "Por favor ingrese una fecha.")
            return

        # Limpiar contenido anterior
        for widget in self.audit_scroll_frame.winfo_children():
            widget.destroy()

        # Obtener datos de la BD
        try:
            datos = self._obtener_datos_auditoria(fecha)
        except Exception as e:
            messagebox.showerror("Error", f"Error al obtener datos:\n{e}")
            return

        self._mostrar_resultados_auditoria(datos, fecha)

    def _obtener_datos_auditoria(self, fecha):
        """Obtiene todos los datos necesarios para la auditoría"""
        import subprocess
        import os
        import shutil

        isql_path = "/opt/firebird/bin/isql"
        original_db = os.path.join(os.path.dirname(__file__), "PDVDATA.FDB")
        user_id = os.getuid()
        tmp_db = f"/tmp/PDVDATA_{user_id}.FDB"

        # Copiar BD si no existe
        if not os.path.exists(tmp_db):
            shutil.copy2(original_db, tmp_db)
            os.chmod(tmp_db, 0o666)

        dsn = f"localhost:{tmp_db}"

        def ejecutar_sql(sql):
            cmd = [isql_path, "-u", "SYSDBA", "-p", "masterkey", "-ch", "UTF8", dsn]
            proc = subprocess.run(cmd, input=sql, capture_output=True, text=True)
            return proc.stdout

        # 1. Total de facturas del día (TODAS)
        sql_total = f"""
        SET HEADING ON;
        SELECT COUNT(*) AS TOTAL_FACTURAS, 
               COALESCE(SUM(TOTAL), 0) AS TOTAL_MONTO
        FROM FACTURAS 
        WHERE CAST(FECHA AS DATE) = '{fecha}';
        """

        # 2. Facturas canceladas del mismo día
        sql_canceladas_dia = f"""
        SET HEADING ON;
        SELECT COUNT(*) AS CANCELADAS_DIA,
               COALESCE(SUM(TOTAL), 0) AS MONTO_CANCELADAS
        FROM FACTURAS 
        WHERE CAST(FECHA AS DATE) = '{fecha}'
        AND CANCELADO = 1;
        """

        # 3. Facturas NO canceladas del día
        sql_no_canceladas = f"""
        SET HEADING ON;
        SELECT COUNT(*) AS NO_CANCELADAS,
               COALESCE(SUM(TOTAL), 0) AS MONTO_NO_CANCELADAS
        FROM FACTURAS 
        WHERE CAST(FECHA AS DATE) = '{fecha}'
        AND (CANCELADO = 0 OR CANCELADO IS NULL);
        """

        # 4. Facturas de OTRO día canceladas EN esta fecha
        sql_cancel_otro_dia = f"""
        SET HEADING ON;
        SELECT F.FOLIO, F.TOTAL, CAST(F.FECHA AS DATE) AS FECHA_CREACION,
               CAST(F.CANCELADO_FECHA AS DATE) AS FECHA_CANCELACION
        FROM FACTURAS F
        WHERE CAST(F.CANCELADO_FECHA AS DATE) = '{fecha}'
        AND CAST(F.FECHA AS DATE) <> '{fecha}'
        AND F.CANCELADO = 1
        ORDER BY F.FOLIO;
        """

        # 5. Detalle de facturas canceladas del día
        sql_detalle_cancel = f"""
        SET HEADING ON;
        SELECT F.FOLIO, F.TOTAL, F.CLIENTE, F.FORMAPAGO
        FROM FACTURAS F
        WHERE CAST(F.FECHA AS DATE) = '{fecha}'
        AND F.CANCELADO = 1
        ORDER BY F.FOLIO;
        """

        # 6. Resumen por forma de pago (NO canceladas)
        sql_por_formapago = f"""
        SET HEADING ON;
        SELECT FORMAPAGO, COUNT(*) AS CANTIDAD, COALESCE(SUM(TOTAL), 0) AS MONTO
        FROM FACTURAS
        WHERE CAST(FECHA AS DATE) = '{fecha}'
        AND (CANCELADO = 0 OR CANCELADO IS NULL)
        GROUP BY FORMAPAGO
        ORDER BY FORMAPAGO;
        """

        # Ejecutar todas las consultas
        result_total = ejecutar_sql(sql_total)
        result_cancel_dia = ejecutar_sql(sql_canceladas_dia)
        result_no_cancel = ejecutar_sql(sql_no_canceladas)
        result_otro_dia = ejecutar_sql(sql_cancel_otro_dia)
        result_detalle = ejecutar_sql(sql_detalle_cancel)
        result_formapago = ejecutar_sql(sql_por_formapago)

        # Parsear resultados
        def parsear_numeros(texto, num_valores=2):
            lineas = [l.strip() for l in texto.strip().split('\n') if l.strip() and not l.startswith('=')]
            for linea in lineas:
                if any(c.isdigit() for c in linea):
                    partes = linea.split()
                    valores = []
                    for p in partes:
                        try:
                            valores.append(float(p.replace(',', '')))
                        except:
                            pass
                    if len(valores) >= num_valores:
                        return valores[:num_valores]
            return [0] * num_valores

        def parsear_tabla(texto):
            lineas = [l for l in texto.strip().split('\n') if l.strip() and not l.startswith('=')]
            filas = []
            for linea in lineas:
                if any(c.isdigit() for c in linea):
                    filas.append(linea)
            return filas

        total_facturas, monto_total = parsear_numeros(result_total)
        cancel_dia, monto_cancel = parsear_numeros(result_cancel_dia)
        no_cancel, monto_no_cancel = parsear_numeros(result_no_cancel)

        return {
            'total_facturas': int(total_facturas),
            'monto_total': monto_total,
            'canceladas_dia': int(cancel_dia),
            'monto_canceladas': monto_cancel,
            'no_canceladas': int(no_cancel),
            'monto_no_canceladas': monto_no_cancel,
            'facturas_otro_dia': parsear_tabla(result_otro_dia),
            'detalle_canceladas': parsear_tabla(result_detalle),
            'por_formapago': parsear_tabla(result_formapago),
            'raw_otro_dia': result_otro_dia,
            'raw_detalle': result_detalle,
            'raw_formapago': result_formapago
        }

    def _mostrar_resultados_auditoria(self, datos, fecha):
        """Muestra los resultados del análisis de auditoría"""
        parent = self.audit_scroll_frame
        parent.columnconfigure(0, weight=1)

        row = 0

        # ══════════════════════════════════════════════════════════════
        # TÍTULO
        # ══════════════════════════════════════════════════════════════
        titulo = ttk.Label(parent, text=f"📊 Análisis de Corte - {fecha}", 
                           font=("Segoe UI", 14, "bold"))
        titulo.grid(row=row, column=0, sticky="w", padx=10, pady=(10,15))
        row += 1

        # ══════════════════════════════════════════════════════════════
        # RESUMEN EJECUTIVO
        # ══════════════════════════════════════════════════════════════
        frame_resumen = ttk.LabelFrame(parent, text="  📋 RESUMEN EJECUTIVO  ", padding=15)
        frame_resumen.grid(row=row, column=0, sticky="ew", padx=10, pady=5)
        frame_resumen.columnconfigure((0,1,2,3), weight=1)
        row += 1

        # Tarjetas de resumen
        self._crear_tarjeta_resumen(frame_resumen, 0, "📄 Total Facturas", 
                                     f"{datos['total_facturas']}", f"${datos['monto_total']:,.2f}", "#3498db")
        self._crear_tarjeta_resumen(frame_resumen, 1, "❌ Canceladas (Día)", 
                                     f"{datos['canceladas_dia']}", f"${datos['monto_canceladas']:,.2f}", "#e74c3c")
        self._crear_tarjeta_resumen(frame_resumen, 2, "✅ NO Canceladas", 
                                     f"{datos['no_canceladas']}", f"${datos['monto_no_canceladas']:,.2f}", "#27ae60")
        
        # Facturas de otro día canceladas hoy
        cant_otro_dia = len(datos['facturas_otro_dia'])
        self._crear_tarjeta_resumen(frame_resumen, 3, "📅 Canceladas Otro Día", 
                                     f"{cant_otro_dia}", "Ver detalles ↓", "#9b59b6")

        # ══════════════════════════════════════════════════════════════
        # EXPLICACIÓN DEL CÁLCULO
        # ══════════════════════════════════════════════════════════════
        frame_calculo = ttk.LabelFrame(parent, text="  🧮 ANÁLISIS DEL CORTE  ", padding=15)
        frame_calculo.grid(row=row, column=0, sticky="ew", padx=10, pady=10)
        row += 1

        texto_calculo = f"""
┌─────────────────────────────────────────────────────────────────────────────┐
│  CÁLCULO DE VENTAS EN EFECTIVO (Corte de Caja)                              │
├─────────────────────────────────────────────────────────────────────────────┤
│                                                                             │
│  📊 Según la Base de Datos:                                                 │
│     • Total facturas del día:     {datos['total_facturas']:>5}  →  ${datos['monto_total']:>15,.2f}   │
│     • Facturas canceladas:        {datos['canceladas_dia']:>5}  →  ${datos['monto_canceladas']:>15,.2f}   │
│     • Facturas válidas:           {datos['no_canceladas']:>5}  →  ${datos['monto_no_canceladas']:>15,.2f}   │
│                                                                             │
│  ✅ El programa muestra: ${datos['monto_no_canceladas']:>15,.2f}                              │
│     (Solo facturas NO canceladas del día)                                   │
│                                                                             │
│  📝 NOTA: Las facturas canceladas de OTROS días que se cancelan HOY         │
│     NO afectan el total del corte de hoy.                                   │
│                                                                             │
└─────────────────────────────────────────────────────────────────────────────┘
"""
        lbl_calculo = ttk.Label(frame_calculo, text=texto_calculo, font=("Consolas", 10), 
                                 justify="left", anchor="w")
        lbl_calculo.pack(fill="x", expand=True)

        # ══════════════════════════════════════════════════════════════
        # DETALLE CANCELADAS DEL DÍA
        # ══════════════════════════════════════════════════════════════
        if datos['canceladas_dia'] > 0:
            frame_cancel = ttk.LabelFrame(parent, text=f"  ❌ FACTURAS CANCELADAS DEL DÍA ({datos['canceladas_dia']})  ", padding=10)
            frame_cancel.grid(row=row, column=0, sticky="ew", padx=10, pady=5)
            row += 1

            # Mostrar texto raw formateado
            txt_cancel = tk.Text(frame_cancel, height=min(10, datos['canceladas_dia'] + 3), 
                                  font=("Consolas", 9), wrap="none")
            txt_cancel.insert("1.0", datos['raw_detalle'])
            txt_cancel.config(state="disabled")
            txt_cancel.pack(fill="x", expand=True)

        # ══════════════════════════════════════════════════════════════
        # FACTURAS DE OTRO DÍA CANCELADAS HOY
        # ══════════════════════════════════════════════════════════════
        if len(datos['facturas_otro_dia']) > 0:
            frame_otro = ttk.LabelFrame(parent, 
                text=f"  📅 FACTURAS DE OTRO DÍA CANCELADAS EN {fecha} ({len(datos['facturas_otro_dia'])})  ", 
                padding=10)
            frame_otro.grid(row=row, column=0, sticky="ew", padx=10, pady=5)
            row += 1

            txt_otro = tk.Text(frame_otro, height=min(8, len(datos['facturas_otro_dia']) + 3), 
                                font=("Consolas", 9), wrap="none")
            txt_otro.insert("1.0", datos['raw_otro_dia'])
            txt_otro.config(state="disabled")
            txt_otro.pack(fill="x", expand=True)

            # Nota explicativa
            nota = ttk.Label(frame_otro, 
                text="⚠️ Estas facturas fueron CREADAS en otro día pero CANCELADAS hoy.\n"
                     "   NO se descuentan del corte de hoy (ya se descontaron cuando se crearon).",
                font=("Segoe UI", 9, "italic"), foreground="#e67e22")
            nota.pack(pady=(10,0), anchor="w")

        # ══════════════════════════════════════════════════════════════
        # DESGLOSE POR FORMA DE PAGO
        # ══════════════════════════════════════════════════════════════
        frame_fp = ttk.LabelFrame(parent, text="  💳 DESGLOSE POR FORMA DE PAGO (No Canceladas)  ", padding=10)
        frame_fp.grid(row=row, column=0, sticky="ew", padx=10, pady=5)
        row += 1

        txt_fp = tk.Text(frame_fp, height=8, font=("Consolas", 9), wrap="none")
        txt_fp.insert("1.0", datos['raw_formapago'])
        txt_fp.config(state="disabled")
        txt_fp.pack(fill="x", expand=True)

        # ══════════════════════════════════════════════════════════════
        # VERIFICACIÓN CRUZADA
        # ══════════════════════════════════════════════════════════════
        frame_verif = ttk.LabelFrame(parent, text="  ✅ VERIFICACIÓN CRUZADA  ", padding=15)
        frame_verif.grid(row=row, column=0, sticky="ew", padx=10, pady=10)
        row += 1

        diferencia = datos['monto_total'] - datos['monto_canceladas'] - datos['monto_no_canceladas']
        if abs(diferencia) < 0.01:
            estado = "✅ CORRECTO"
            color = "#27ae60"
        else:
            estado = f"⚠️ DIFERENCIA: ${diferencia:,.2f}"
            color = "#e74c3c"

        verif_texto = f"""
  Total del día:           ${datos['monto_total']:>15,.2f}
- Canceladas del día:      ${datos['monto_canceladas']:>15,.2f}
─────────────────────────────────────────
= Facturas válidas:        ${datos['monto_no_canceladas']:>15,.2f}

Estado: {estado}
"""
        lbl_verif = ttk.Label(frame_verif, text=verif_texto, font=("Consolas", 11), justify="left")
        lbl_verif.pack(anchor="w")

        # Actualizar scroll region
        self.audit_scroll_frame.update_idletasks()
        self.audit_canvas.configure(scrollregion=self.audit_canvas.bbox("all"))

    def _crear_tarjeta_resumen(self, parent, col, titulo, valor, subtitulo, color):
        """Crea una tarjeta de resumen visual"""
        frame = ttk.Frame(parent, padding=10)
        frame.grid(row=0, column=col, sticky="nsew", padx=5, pady=5)

        ttk.Label(frame, text=titulo, font=("Segoe UI", 9)).pack()
        ttk.Label(frame, text=valor, font=("Segoe UI", 18, "bold")).pack(pady=5)
        ttk.Label(frame, text=subtitulo, font=("Segoe UI", 10)).pack()

    def _buscar_folio_auditoria(self):
        """Busca información detallada de un folio específico"""
        folio = self.audit_folio_var.get().strip()
        if not folio:
            messagebox.showwarning("Folio requerido", "Por favor ingrese un número de folio.")
            return

        try:
            folio_int = int(folio)
        except ValueError:
            messagebox.showerror("Error", "El folio debe ser un número.")
            return

        # Buscar el folio
        import subprocess
        import os
        import shutil

        isql_path = "/opt/firebird/bin/isql"
        original_db = os.path.join(os.path.dirname(__file__), "PDVDATA.FDB")
        user_id = os.getuid()
        tmp_db = f"/tmp/PDVDATA_{user_id}.FDB"

        if not os.path.exists(tmp_db):
            shutil.copy2(original_db, tmp_db)
            os.chmod(tmp_db, 0o666)

        dsn = f"localhost:{tmp_db}"

        sql = f"""
SET HEADING ON;
SELECT 
    F.FOLIO,
    F.TOTAL,
    F.SUBTOTAL,
    F.DESCUENTO,
    F.CLIENTE,
    F.FORMAPAGO,
    CAST(F.FECHA AS DATE) AS FECHA_CREACION,
    CAST(F.FECHA AS TIME) AS HORA,
    F.CANCELADO,
    CAST(F.CANCELADO_FECHA AS DATE) AS FECHA_CANCELACION,
    CAST(F.CANCELADO_FECHA AS TIME) AS HORA_CANCELACION,
    F.VENDEDOR
FROM FACTURAS F
WHERE F.FOLIO = {folio_int};
"""

        cmd = [isql_path, "-u", "SYSDBA", "-p", "masterkey", "-ch", "UTF8", dsn]
        proc = subprocess.run(cmd, input=sql, capture_output=True, text=True)
        resultado = proc.stdout

        if "FOLIO" not in resultado or str(folio_int) not in resultado:
            messagebox.showinfo("No encontrado", f"No se encontró la factura con folio {folio_int}")
            return

        # Mostrar en ventana emergente
        ventana = tk.Toplevel(self.ventana)
        ventana.title(f"📄 Detalle Factura #{folio_int}")
        ventana.geometry("700x400")
        ventana.transient(self.ventana)

        frame = ttk.Frame(ventana, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=f"📄 Información de Factura #{folio_int}", 
                  font=("Segoe UI", 14, "bold")).pack(pady=(0,15))

        txt = tk.Text(frame, font=("Consolas", 10), wrap="none")
        txt.insert("1.0", resultado)
        txt.config(state="disabled")
        txt.pack(fill="both", expand=True)

        ttk.Button(frame, text="Cerrar", command=ventana.destroy).pack(pady=(15,0))

    # ==================================================================
    # GUARDAR LIQUIDACIÓN EN BD
    # ==================================================================
    def _guardar_liquidacion(self):
        """Guarda la liquidación actual en la base de datos SQLite."""
        fecha = self.ds.fecha
        rep_filtro = self.repartidor_filtro_var.get()
        
        if not rep_filtro or rep_filtro == "(Todos)":
            messagebox.showwarning("Seleccionar Repartidor", 
                                   "Debes seleccionar un repartidor específico para guardar la liquidación.")
            return
        
        # Obtener datos del resumen financiero
        try:
            # Total vendido
            total_vendido = float(self.lbl_total_subtotal.cget("text").replace("$", "").replace(",", ""))
            # Total descuentos
            total_descuentos = float(self.lbl_total_descuentos.cget("text").replace("$", "").replace(",", ""))
            # Total gastos
            total_gastos = float(self.lbl_total_gastos.cget("text").replace("$", "").replace(",", ""))
            # Neto a entregar
            neto = float(self.lbl_neto.cget("text").replace("$", "").replace(",", ""))
            # Dinero contado
            dinero_contado = float(self.lbl_dinero_contado.cget("text").replace("$", "").replace(",", ""))
            # Diferencia
            diferencia_text = self.lbl_diferencia.cget("text").replace("$", "").replace(",", "")
            diferencia = float(diferencia_text) if diferencia_text and diferencia_text != "N/A" else 0
            
            # Datos adicionales
            total_creditos = float(self.lbl_total_creditos.cget("text").replace("$", "").replace(",", ""))
            total_devoluciones = float(self.lbl_total_devoluciones.cget("text").replace("$", "").replace(",", ""))
            total_ajustes = float(self.lbl_total_ajustes.cget("text").replace("$", "").replace(",", ""))
            total_canceladas = float(self.lbl_total_canceladas.cget("text").replace("$", "").replace(",", ""))
            
        except (ValueError, AttributeError) as e:
            messagebox.showerror("Error", f"Error al obtener datos de liquidación: {e}")
            return
        
        # Preparar datos para guardar
        datos = {
            'total_ventas': total_vendido,
            'total_descuentos': total_descuentos,
            'total_creditos': total_creditos,
            'total_devoluciones': total_devoluciones,
            'total_ajustes': total_ajustes,
            'total_gastos': total_gastos,
            'total_canceladas': total_canceladas,
            'total_dinero': dinero_contado,
            'neto': neto,
            'diferencia': diferencia,
            'fecha_guardado': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Confirmar
        msg = f"¿Guardar liquidación de {rep_filtro} para {fecha}?\n\n"
        msg += f"Total Vendido: ${total_vendido:,.2f}\n"
        msg += f"Total Descuentos: ${total_descuentos:,.2f}\n"
        msg += f"Total Gastos: ${total_gastos:,.2f}\n"
        msg += f"Neto a Entregar: ${neto:,.2f}\n"
        msg += f"Dinero Contado: ${dinero_contado:,.2f}\n"
        msg += f"Diferencia: ${diferencia:,.2f}"
        
        if not messagebox.askyesno("Confirmar Guardado", msg):
            return
        
        # Guardar en SQLite
        liq_id = db_local.guardar_liquidacion(fecha, rep_filtro, datos)
        
        if liq_id > 0:
            messagebox.showinfo("Guardado", 
                f"✅ Liquidación guardada correctamente.\n\n"
                f"ID: {liq_id}\n"
                f"Repartidor: {rep_filtro}\n"
                f"Fecha: {fecha}")
        else:
            messagebox.showerror("Error", "No se pudo guardar la liquidación.")

    # ==================================================================
    # REPORTE DE LIQUIDACIÓN
    # ==================================================================
    def _generar_reporte(self):
        fecha = self.ds.fecha
        reps  = self.ds.get_repartidores()

        if not reps:
            messagebox.showwarning("Sin datos", "No hay repartidores asignados.")
            return

        tipo_map = {"credito": "Crédito", "devolucion": "Devolución", "ajuste": "Ajuste"}
        desc_todos = cargar_descuentos()

        # ── construir datos por repartidor ──────────────────────────────
        datos_por_rep = {}   # rep → { ventas, descuentos, gastos, dinero, totales }

        for rep in reps:
            ventas_rep = [v for v in self.ds.get_ventas() if v['repartidor'] == rep]

            # descuentos filtrados
            desc_rep = []
            for fk, datos in desc_todos.items():
                for d in datos.get("descuentos", []):
                    if d.get("fecha", "").startswith(fecha) and d.get("repartidor") == rep:
                        desc_rep.append({
                            'folio': fk,
                            'tipo': tipo_map.get(d['tipo'], d['tipo']),
                            'monto': d.get('monto', 0),
                            'observacion': d.get('observacion', '')
                        })

            gastos_rep  = self.ds.get_gastos(rep)
            dinero_rep  = self.ds.get_dinero(rep)   # {valor_int: cantidad}

            total_sub   = sum(v['subtotal'] for v in ventas_rep)
            total_desc  = sum(d['monto'] for d in desc_rep)
            total_gasto = sum(g['monto'] for g in gastos_rep)
            total_din   = self.ds.get_total_dinero(rep)
            neto        = total_sub - total_desc - total_gasto
            diferencia  = total_din - neto

            datos_por_rep[rep] = {
                'ventas':      ventas_rep,
                'descuentos':  desc_rep,
                'gastos':      gastos_rep,
                'dinero':      dinero_rep,
                'total_sub':   total_sub,
                'total_desc':  total_desc,
                'total_gasto': total_gasto,
                'total_din':   total_din,
                'neto':        neto,
                'diferencia':  diferencia,
            }

        # ── texto preview (todos los reps concatenados) ─────────────────
        sep  = "═" * 64
        thin = "─" * 64
        texto_total = ""

        for rep, d in datos_por_rep.items():
            texto_total += (
                f"\n{sep}\n"
                f"  LIQUIDACIÓN DE REPARTIDOR\n"
                f"  Repartidor : {rep}\n"
                f"  Fecha      : {fecha}\n"
                f"{sep}\n\n"
                f"VENTAS DEL DÍA\n"
            )
            for v in d['ventas']:
                cancelada_txt = " [CANCELADA]" if v.get('cancelada', False) else ""
                texto_total += f"  Folio {v['folio']:>6}  {v['nombre']:<40} ${v['subtotal']:>12,.2f}{cancelada_txt}\n"
            texto_total += f"  {thin}\n  Total Subtotal          ${d['total_sub']:>18,.2f}\n\n"

            texto_total += "DESCUENTOS\n"
            if d['descuentos']:
                for desc in d['descuentos']:
                    texto_total += (f"  Folio {desc['folio']:>6}: {desc['tipo']:<12} "
                                    f"${desc['monto']:>12,.2f}   {desc['observacion']}\n")
            else:
                texto_total += "  Ninguno\n"
            texto_total += f"  {thin}\n  Total Descuentos        ${d['total_desc']:>18,.2f}\n\n"

            texto_total += "GASTOS ADICIONALES\n"
            if d['gastos']:
                for g in d['gastos']:
                    texto_total += f"  {g['concepto']:<40} ${g['monto']:>12,.2f}\n"
            else:
                texto_total += "  Ninguno\n"
            texto_total += f"  {thin}\n  Total Gastos            ${d['total_gasto']:>18,.2f}\n"
            texto_total += f"  {thin}\n  Total a Descontar       ${d['total_desc']+d['total_gasto']:>18,.2f}\n\n"

            texto_total += (
                f"{sep}\n"
                f"  TOTAL NETO A PAGAR      ${d['neto']:>18,.2f}\n"
                f"{sep}\n\n"
                f"CONTEO DE DINERO\n"
            )
            for valor in self._VALORES_ORDEN:
                cant = d['dinero'].get(valor, 0)
                if cant > 0:
                    texto_total += f"  ${valor:>7,} × {cant:>4} = ${cant*valor:>13,}\n"
            texto_total += f"  {thin}\n  TOTAL DINERO            ${d['total_din']:>18,.2f}\n\n"
            texto_total += f"  DIFERENCIA              ${d['diferencia']:>18,.2f}\n"
            if abs(d['diferencia']) < 0.01:
                texto_total += "\n  ✓  LIQUIDACIÓN CUADRADA\n"
            texto_total += "\n"

        # ── ventana de preview ───────────────────────────────────────────
        win = tk.Toplevel(self.ventana)
        win.title("Reporte de Liquidación")
        win.geometry("780x680")
        win.transient(self.ventana)

        txt = tk.Text(win, font=("Consolas", 9), wrap=tk.NONE)
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=(8, 0))
        txt.insert(tk.END, texto_total)
        txt.config(state=tk.DISABLED)

        frame_btn = ttk.Frame(win)
        frame_btn.pack(pady=8)

        ttk.Button(frame_btn, text="📊  Exportar Excel",
                   command=lambda: self._exportar_excel(datos_por_rep, fecha)).pack(side=tk.LEFT, padx=4)
        def _guardar_txt():
            nombre = f"Liquidacion_todos_{fecha}.txt"
            with open(nombre, 'w', encoding='utf-8') as f:
                f.write(texto_total)
            messagebox.showinfo("Guardado", f"Reporte guardado:\n{nombre}", parent=win)
        ttk.Button(frame_btn, text="💾  Guardar .txt",
                   command=_guardar_txt).pack(side=tk.LEFT, padx=4)

    # ── exportar a Excel ─────────────────────────────────────────────────
    def _exportar_excel(self, datos_por_rep: dict, fecha: str):
        # ─ estilos reutilizables ─
        font_title   = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        font_header  = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_seccion = Font(name='Arial', size=11, bold=True)
        font_normal  = Font(name='Arial', size=10)
        font_monto   = Font(name='Arial', size=10)
        font_neto    = Font(name='Arial', size=12, bold=True, color='2E7D32')
        font_dif_ok  = Font(name='Arial', size=11, bold=True, color='2E7D32')
        font_dif_no  = Font(name='Arial', size=11, bold=True, color='C62828')

        fill_title   = PatternFill('solid', fgColor='1565C0')
        fill_header  = PatternFill('solid', fgColor='1976D2')
        fill_seccion = PatternFill('solid', fgColor='E3F2FD')
        fill_resumen = PatternFill('solid', fgColor='FFF3E0')
        fill_neto    = PatternFill('solid', fgColor='E8F5E9')
        fill_dinero  = PatternFill('solid', fgColor='F3E5F5')
        fill_par     = PatternFill('solid', fgColor='FFFFFF')
        fill_impar   = PatternFill('solid', fgColor='F5F5F5')

        align_c = Alignment(horizontal='center', vertical='center')
        align_l = Alignment(horizontal='left',   vertical='center')
        align_r = Alignment(horizontal='right',  vertical='center')

        borde_fino = Border(
            left=Side(style='thin', color='BDBDBD'),
            right=Side(style='thin', color='BDBDBD'),
            top=Side(style='thin', color='BDBDBD'),
            bottom=Side(style='thin', color='BDBDBD')
        )

        fmt_moneda = '#,##0.00'

        wb = Workbook()
        wb.remove(wb.active)   # eliminar hoja vacía por defecto

        for rep, d in datos_por_rep.items():
            # nombre de hoja: máximo 31 chars, sin caracteres ilegales
            nombre_hoja = rep[:31].replace('/', '-').replace('\\', '-').replace('[', '').replace(']', '')
            ws = wb.create_sheet(title=nombre_hoja)

            ws.column_dimensions['A'].width = 12   # Folio
            ws.column_dimensions['B'].width = 35   # Cliente
            ws.column_dimensions['C'].width = 16   # Subtotal
            ws.column_dimensions['D'].width = 12   # Cancelada
            ws.column_dimensions['E'].width = 22   # Observación (descuentos)
            ws.column_dimensions['F'].width = 18   # Extra

            r = 1   # cursor de fila

            # ══ TÍTULO ═══════════════════════════════════════════════════
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.cell(r, 1, value=f"LIQUIDACIÓN DE REPARTIDOR — {rep}")
            ws.cell(r, 1).font = font_title
            ws.cell(r, 1).fill = fill_title
            ws.cell(r, 1).alignment = align_c
            for c in range(1, 6):
                ws.cell(r, c).fill = fill_title
            r += 1

            ws.cell(r, 1, value="Fecha:")
            ws.cell(r, 1).font = font_seccion
            ws.cell(r, 2, value=fecha)
            ws.cell(r, 2).font = font_normal
            r += 2

            # ══ VENTAS DEL DÍA ═══════════════════════════════════════════
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.cell(r, 1, value="VENTAS DEL DÍA")
            ws.cell(r, 1).font = font_seccion
            ws.cell(r, 1).fill = fill_seccion
            for c in range(1, 5):
                ws.cell(r, c).fill = fill_seccion
                ws.cell(r, c).border = borde_fino
            r += 1

            # headers ventas
            hdrs_ventas = ["Folio", "Cliente", "Subtotal", "Cancelada"]
            fila_hdr_ventas = r
            for i, h in enumerate(hdrs_ventas, 1):
                cell = ws.cell(r, i, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_c
                cell.border = borde_fino
            r += 1

            fila_inicio_ventas = r
            for idx, v in enumerate(d['ventas']):
                fill_f = fill_par if idx % 2 == 0 else fill_impar
                cancelada = v.get('cancelada', False)
                ws.cell(r, 1, value=v['folio']);      ws.cell(r, 1).alignment = align_c
                ws.cell(r, 2, value=v['nombre']);     ws.cell(r, 2).alignment = align_l
                ws.cell(r, 3, value=v['subtotal']);   ws.cell(r, 3).number_format = fmt_moneda
                ws.cell(r, 3).alignment = align_r
                ws.cell(r, 4, value="SÍ" if cancelada else "NO"); ws.cell(r, 4).alignment = align_c
                for c in range(1, 5):
                    ws.cell(r, c).font  = font_normal
                    ws.cell(r, c).fill  = fill_f
                    ws.cell(r, c).border = borde_fino
                r += 1
            fila_fin_ventas = r - 1

            # fila Total Subtotal con fórmula SUM
            ws.cell(r, 2, value="Total Subtotal:")
            ws.cell(r, 2).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 2).alignment = align_r
            if fila_inicio_ventas <= fila_fin_ventas:
                ws.cell(r, 3, value=f"=SUM(C{fila_inicio_ventas}:C{fila_fin_ventas})")
            else:
                ws.cell(r, 3, value=0)
            ws.cell(r, 3).number_format = fmt_moneda
            ws.cell(r, 3).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 3).alignment = align_r
            for c in range(1, 5):
                ws.cell(r, c).fill  = fill_resumen
                ws.cell(r, c).border = borde_fino
            r += 2

            # ══ DESCUENTOS ════════════════════════════════════════════════
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            ws.cell(r, 1, value="DESCUENTOS")
            ws.cell(r, 1).font = font_seccion
            ws.cell(r, 1).fill = fill_seccion
            for c in range(1, 5):
                ws.cell(r, c).fill = fill_seccion
                ws.cell(r, c).border = borde_fino
            r += 1

            hdrs_desc = ["Folio", "Tipo", "Monto", "Observación"]
            for i, h in enumerate(hdrs_desc, 1):
                cell = ws.cell(r, i, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_c
                cell.border = borde_fino
            r += 1

            fila_inicio_desc = r
            for idx, desc in enumerate(d['descuentos']):
                fill_f = fill_par if idx % 2 == 0 else fill_impar
                ws.cell(r, 1, value=desc['folio']);       ws.cell(r, 1).alignment = align_c
                ws.cell(r, 2, value=desc['tipo']);        ws.cell(r, 2).alignment = align_c
                ws.cell(r, 3, value=desc['monto']);       ws.cell(r, 3).number_format = fmt_moneda
                ws.cell(r, 3).alignment = align_r
                ws.cell(r, 4, value=desc['observacion']); ws.cell(r, 4).alignment = align_l
                for c in range(1, 5):
                    ws.cell(r, c).font  = font_normal
                    ws.cell(r, c).fill  = fill_f
                    ws.cell(r, c).border = borde_fino
                r += 1
            fila_fin_desc = r - 1

            ws.cell(r, 3, value="Total Descuentos:")
            ws.cell(r, 3).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 3).alignment = align_r
            if fila_inicio_desc <= fila_fin_desc:
                ws.cell(r, 4, value=f"=SUM(C{fila_inicio_desc}:C{fila_fin_desc})")
            else:
                ws.cell(r, 4, value=0)
            ws.cell(r, 4).number_format = fmt_moneda
            ws.cell(r, 4).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 4).alignment = align_r
            for c in range(1, 5):
                ws.cell(r, c).fill  = fill_resumen
                ws.cell(r, c).border = borde_fino
            r += 2

            # ══ GASTOS ADICIONALES ════════════════════════════════════════
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r, 1, value="GASTOS ADICIONALES")
            ws.cell(r, 1).font = font_seccion
            ws.cell(r, 1).fill = fill_seccion
            for c in range(1, 4):
                ws.cell(r, c).fill = fill_seccion
                ws.cell(r, c).border = borde_fino
            r += 1

            hdrs_gasto = ["#", "Concepto", "Monto"]
            for i, h in enumerate(hdrs_gasto, 1):
                cell = ws.cell(r, i, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_c
                cell.border = borde_fino
            r += 1

            fila_inicio_gasto = r
            for idx, g in enumerate(d['gastos']):
                fill_f = fill_par if idx % 2 == 0 else fill_impar
                ws.cell(r, 1, value=idx + 1);           ws.cell(r, 1).alignment = align_c
                ws.cell(r, 2, value=g['concepto']);      ws.cell(r, 2).alignment = align_l
                ws.cell(r, 3, value=g['monto']);         ws.cell(r, 3).number_format = fmt_moneda
                ws.cell(r, 3).alignment = align_r
                for c in range(1, 4):
                    ws.cell(r, c).font  = font_normal
                    ws.cell(r, c).fill  = fill_f
                    ws.cell(r, c).border = borde_fino
                r += 1
            fila_fin_gasto = r - 1

            ws.cell(r, 2, value="Total Gastos:")
            ws.cell(r, 2).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 2).alignment = align_r
            if fila_inicio_gasto <= fila_fin_gasto:
                ws.cell(r, 3, value=f"=SUM(C{fila_inicio_gasto}:C{fila_fin_gasto})")
            else:
                ws.cell(r, 3, value=0)
            ws.cell(r, 3).number_format = fmt_moneda
            ws.cell(r, 3).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 3).alignment = align_r
            for c in range(1, 4):
                ws.cell(r, c).fill  = fill_resumen
                ws.cell(r, c).border = borde_fino
            r += 2

            # ══ RESUMEN FINANCIERO ════════════════════════════════════════
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r, 1, value="RESUMEN FINANCIERO")
            ws.cell(r, 1).font = font_seccion
            ws.cell(r, 1).fill = fill_seccion
            for c in range(1, 4):
                ws.cell(r, c).fill = fill_seccion
                ws.cell(r, c).border = borde_fino
            r += 1

            resumen_items = [
                ("Total Subtotal",     d['total_sub']),
                ("Total Descuentos",   d['total_desc']),
                ("Total Gastos",       d['total_gasto']),
                ("Total a Descontar",  d['total_desc'] + d['total_gasto']),
            ]
            for label, valor in resumen_items:
                ws.cell(r, 1, value=label)
                ws.cell(r, 1).font = font_normal
                ws.cell(r, 1).alignment = align_l
                ws.cell(r, 2, value=valor)
                ws.cell(r, 2).number_format = fmt_moneda
                ws.cell(r, 2).font = font_normal
                ws.cell(r, 2).alignment = align_r
                for c in range(1, 3):
                    ws.cell(r, c).fill  = fill_resumen
                    ws.cell(r, c).border = borde_fino
                r += 1

            # NETO
            ws.cell(r, 1, value="TOTAL NETO A PAGAR")
            ws.cell(r, 1).font = font_neto
            ws.cell(r, 1).fill = fill_neto
            ws.cell(r, 1).alignment = align_l
            ws.cell(r, 1).border = borde_fino
            ws.cell(r, 2, value=d['neto'])
            ws.cell(r, 2).number_format = fmt_moneda
            ws.cell(r, 2).font = font_neto
            ws.cell(r, 2).fill = fill_neto
            ws.cell(r, 2).alignment = align_r
            ws.cell(r, 2).border = borde_fino
            r += 2

            # ══ CONTEO DE DINERO ══════════════════════════════════════════
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.cell(r, 1, value="CONTEO DE DINERO")
            ws.cell(r, 1).font = font_seccion
            ws.cell(r, 1).fill = fill_dinero
            for c in range(1, 4):
                ws.cell(r, c).fill = fill_dinero
                ws.cell(r, c).border = borde_fino
            r += 1

            hdrs_din = ["Denominación", "Cantidad", "Subtotal"]
            for i, h in enumerate(hdrs_din, 1):
                cell = ws.cell(r, i, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_c
                cell.border = borde_fino
            r += 1

            fila_inicio_din = r
            for idx, valor in enumerate(self._VALORES_ORDEN):
                cant = d['dinero'].get(valor, 0)
                fill_f = fill_par if idx % 2 == 0 else fill_impar
                ws.cell(r, 1, value=f"${valor:,}");   ws.cell(r, 1).alignment = align_l
                ws.cell(r, 2, value=cant);             ws.cell(r, 2).alignment = align_c
                ws.cell(r, 3, value=cant * valor);     ws.cell(r, 3).number_format = fmt_moneda
                ws.cell(r, 3).alignment = align_r
                for c in range(1, 4):
                    ws.cell(r, c).font  = font_normal
                    ws.cell(r, c).fill  = fill_f
                    ws.cell(r, c).border = borde_fino
                r += 1
            fila_fin_din = r - 1

            ws.cell(r, 2, value="TOTAL DINERO:")
            ws.cell(r, 2).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 2).alignment = align_r
            ws.cell(r, 3, value=f"=SUM(C{fila_inicio_din}:C{fila_fin_din})")
            ws.cell(r, 3).number_format = fmt_moneda
            ws.cell(r, 3).font = Font(name='Arial', size=10, bold=True)
            ws.cell(r, 3).alignment = align_r
            for c in range(1, 4):
                ws.cell(r, c).fill  = fill_dinero
                ws.cell(r, c).border = borde_fino
            r += 2

            # ══ DIFERENCIA ════════════════════════════════════════════════
            ws.cell(r, 1, value="DIFERENCIA")
            ws.cell(r, 1).font = Font(name='Arial', size=11, bold=True)
            ws.cell(r, 1).alignment = align_l
            ws.cell(r, 1).border = borde_fino
            ws.cell(r, 2, value=d['diferencia'])
            ws.cell(r, 2).number_format = fmt_moneda
            ws.cell(r, 2).font = font_dif_ok if abs(d['diferencia']) < 0.01 else font_dif_no
            ws.cell(r, 2).alignment = align_r
            ws.cell(r, 2).border = borde_fino
            if abs(d['diferencia']) < 0.01:
                ws.cell(r, 3, value="✓ CUADRADA")
                ws.cell(r, 3).font = font_dif_ok
            r += 1

        # ── guardar archivo ──────────────────────────────────────────────
        nombre_archivo = f"Liquidacion_{fecha}.xlsx"
        try:
            wb.save(nombre_archivo)
            messagebox.showinfo("Exportado",
                                f"Archivo Excel generado exitosamente:\n\n{nombre_archivo}\n\n"
                                f"Contiene {len(datos_por_rep)} hoja(s): "
                                f"{', '.join(datos_por_rep.keys())}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el archivo:\n{e}")

    # ==================================================================
    # EJECUCIÓN SQL (Firebird via isql)
    # ==================================================================
    def _ejecutar_sql(self, sql: str):
        try:
            if not os.path.exists(self.ruta_fdb):
                return False, "", f"Archivo no encontrado: {self.ruta_fdb}"
            
            # Detectar sistema operativo
            es_windows = sys.platform.startswith('win')
            
            if es_windows:
                # En Windows, buscar isql.exe en rutas comunes de instalación
                posibles_isql = [
                    r"C:\Program Files\Firebird\Firebird_5_0\isql.exe",  # Firebird 5.0
                    r"C:\Program Files\Firebird\Firebird_4_0\isql.exe",
                    r"C:\Program Files (x86)\Firebird\Firebird_4_0\bin\isql.exe",
                    r"C:\Program Files\Firebird\Firebird_3_0\isql.exe",
                    r"C:\Program Files (x86)\Firebird\Firebird_3_0\bin\isql.exe",
                    r"C:\Program Files\Firebird\Firebird_2_5\bin\isql.exe",
                    r"C:\Program Files (x86)\Firebird\Firebird_2_5\bin\isql.exe",
                    r"C:\Program Files\Firebird\bin\isql.exe",
                    r"C:\Program Files (x86)\Firebird\bin\isql.exe",
                ]
                
                # Buscar isql en rutas estándar
                isql_path = None
                for ruta in posibles_isql:
                    if os.path.exists(ruta):
                        isql_path = ruta
                        break
                
                # Si no encuentra, intentar ejecutar 'isql' directamente (puede estar en PATH)
                if not isql_path:
                    try:
                        # Verificar si isql está en PATH
                        resultado_test = subprocess.run(
                            ['isql', '-version'],
                            capture_output=True,
                            timeout=2,
                            encoding='utf-8',
                            creationflags=subprocess.CREATE_NO_WINDOW
                        )
                        if resultado_test.returncode == 0:
                            isql_path = 'isql'
                    except:
                        pass
                
                if not isql_path:
                    return False, "", (
                        "No se encontró isql de Firebird.\n\n"
                        "Firebird no parece estar instalado correctamente.\n"
                        "Verifica la instalación en:\n"
                        "https://www.firebirdsql.org/download/\n\n"
                        "O agrega Firebird\\bin al PATH de tu sistema."
                    )
                
                # En Windows: NO usar sudo, ejecutar directamente
                cmd = [isql_path, '-u', 'SYSDBA', '-p', 'masterkey', self.ruta_fdb]
            else:
                # En Linux/Unix: conectar via TCP/IP al servidor Firebird
                # Firebird necesita acceso al archivo, así que copiamos a /tmp
                import shutil
                import tempfile
                
                # Usar un nombre único en /tmp para evitar conflictos de permisos
                tmp_fdb = f'/tmp/PDVDATA_{os.getuid()}.FDB'
                
                # Intentar copiar el archivo
                try:
                    # Si el archivo temporal existe y podemos escribir, lo eliminamos primero
                    if os.path.exists(tmp_fdb):
                        try:
                            os.remove(tmp_fdb)
                        except:
                            pass
                    
                    if os.path.exists(self.ruta_fdb):
                        shutil.copy2(self.ruta_fdb, tmp_fdb)
                        # Dar permisos completos al archivo
                        os.chmod(tmp_fdb, 0o666)
                except Exception as e:
                    return False, "", f"Error copiando archivo a /tmp: {str(e)}"
                
                # Usar siempre /tmp para la conexión TCP/IP
                fdb_path = tmp_fdb
                
                # El comando de isql para conexión TCP/IP
                cmd = [self.isql_path]
                # El SQL debe incluir el CONNECT con localhost
                sql = f"CONNECT 'localhost:{fdb_path}' USER 'SYSDBA' PASSWORD 'masterkey';\n" + sql
            
            # Agregar QUIT al final del SQL para que isql termine correctamente
            sql_completo = sql.strip()
            if not sql_completo.endswith(';'):
                sql_completo += ';'
            sql_completo += '\nQUIT;'
            
            # En Windows usar cp1252 (Windows-1252) en lugar de utf-8
            # porque Firebird a menudo devuelve datos en esa codificación
            encoding_usar = 'cp1252' if es_windows else 'utf-8'
            
            # Configurar kwargs para subprocess
            run_kwargs = {
                'input': sql_completo,
                'capture_output': True,
                'text': True,
                'timeout': 30,
                'encoding': encoding_usar,
                'errors': 'ignore'
            }
            # En Windows, ocultar ventana de CMD
            if es_windows:
                run_kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW
            
            resultado = subprocess.run(cmd, **run_kwargs)
            
            # En Linux con conexión TCP/IP, isql siempre muestra "Use CONNECT..." en stderr
            # pero eso no es un error si hay datos válidos en stdout
            stdout = resultado.stdout or ""
            stderr = resultado.stderr or ""
            
            # Verificar si hay datos válidos en la salida
            # Buscar líneas de separación (===) o datos numéricos típicos de resultados
            tiene_datos = '===' in stdout or any(
                keyword in stdout.upper() for keyword in ['COUNT', 'FOLIO', 'NOMBRE', 'TOTAL', 'SUBTOTAL']
            )
            
            # También verificar si hay líneas con datos numéricos (típico de resultados)
            if not tiene_datos:
                for linea in stdout.split('\n'):
                    linea = linea.strip()
                    # Si hay una línea con números que parece ser un resultado de consulta
                    if linea and any(c.isdigit() for c in linea) and not linea.startswith('Use '):
                        tiene_datos = True
                        break
            
            # Es exitoso si: returncode es 0, O si hay datos válidos en stdout
            exito = resultado.returncode == 0 or tiene_datos
            
            # Si hay datos válidos, limpiar el stderr del mensaje "Use CONNECT..."
            # ya que ese mensaje no es un error real cuando hay datos
            if tiene_datos and 'Use CONNECT' in stderr:
                stderr = ""
            
            return exito, stdout, stderr
            
        except subprocess.TimeoutExpired:
            return False, "", "Timeout: La consulta SQL tardó demasiado (>30s)"
        except FileNotFoundError:
            return False, "", "No se pudo ejecutar isql. Verifica que Firebird esté instalado."
        except Exception as e:
            return False, "", str(e)


# ===========================================================================
#  ENTRY POINT
# ===========================================================================
def main():
    ventana = tk.Tk()
    ventana.configure(bg="#1e1e1e")  # Fondo oscuro por defecto

    # Estilo global - Modo oscuro
    estilo = ttk.Style()
    estilo.theme_use('clam')  # tema que permite mayor personalización

    app = LiquidadorRepartidores(ventana)
    ventana.mainloop()


if __name__ == '__main__':
    main()