#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXPORTADOR DE LISTA DE PRECIOS
==============================
Exporta la lista de productos con precios de venta desde PDVDATA.FDB a Excel.
"""

import os
import sys
from datetime import datetime

# Agregar el directorio actual al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Verificar dependencias
try:
    import pandas as pd
except ImportError:
    import subprocess
    print("Instalando pandas...")
    subprocess.run([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"], check=True)
    import pandas as pd

try:
    import fdb
except ImportError:
    import subprocess
    print("Instalando fdb...")
    subprocess.run([sys.executable, "-m", "pip", "install", "fdb"], check=True)
    import fdb

from core.firebird_setup import get_firebird_setup, get_fdb_connection


def obtener_productos():
    """Obtiene la lista de productos con precios usando fdb."""
    
    setup = get_firebird_setup()
    print(f"📁 Base de datos: {setup.get_default_db_path()}")
    print(f"🔧 Biblioteca Firebird: {setup.fb_library_name}")
    
    try:
        conn = get_fdb_connection()
        cursor = conn.cursor()
        
        # Consulta SQL para obtener productos (columnas reales de Eleventa)
        sql = """
            SELECT 
                CODIGO,
                DESCRIPCION,
                UMEDIDA,
                PVENTA,
                PMAYOREOFINAL,
                DINVENTARIO
            FROM PRODUCTOS
            WHERE ELIMINADO_EN IS NULL
            ORDER BY DESCRIPCION
        """
        
        print("🔍 Consultando productos...")
        cursor.execute(sql)
        
        productos = []
        for row in cursor.fetchall():
            productos.append({
                'Código': row[0] or '',
                'Descripción': (row[1] or '').strip(),
                'Unidad': row[2] or 'PZA',
                'Precio Venta': float(row[3] or 0),
                'Precio Mayoreo': float(row[4] or 0),
                'Existencia': int(row[5] or 0)
            })
        
        conn.close()
        return productos
        
    except Exception as e:
        print(f"❌ Error al conectar: {e}")
        return []


def exportar_excel(productos, archivo_salida):
    """Exporta los productos a un archivo Excel."""
    if not productos:
        print("❌ No hay productos para exportar")
        return False
    
    df = pd.DataFrame(productos)
    
    # Exportar a Excel con formato
    with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Lista de Precios', index=False)
        
        # Obtener la hoja
        worksheet = writer.sheets['Lista de Precios']
        
        # Ajustar anchos de columna
        anchos = {'A': 15, 'B': 50, 'C': 10, 'D': 15, 'E': 15, 'F': 12}
        for col, ancho in anchos.items():
            worksheet.column_dimensions[col].width = ancho
        
        # Formato de moneda para Precio Venta (columna D)
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Estilo para encabezados
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Formato para datos
        for row in range(2, len(df) + 2):
            # Precio Venta como moneda
            cell = worksheet.cell(row=row, column=4)
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            
            # Precio Mayoreo como moneda
            cell = worksheet.cell(row=row, column=5)
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            
            # Existencia como número
            cell = worksheet.cell(row=row, column=6)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='center')
    
    return True


def main():
    print("=" * 60)
    print("📋 EXPORTADOR DE LISTA DE PRECIOS")
    print("=" * 60)
    
    # Obtener productos
    productos = obtener_productos()
    
    if not productos:
        print("❌ No se pudieron obtener los productos")
        return 1
    
    print(f"✅ Se encontraron {len(productos)} productos")
    
    # Mostrar algunos ejemplos
    print("\n📋 Primeros 5 productos:")
    for p in productos[:5]:
        print(f"   {p['Código']}: {p['Descripción'][:40]} - ${p['Precio Venta']:,.2f}")
    
    # Generar nombre de archivo con fecha
    base_dir = os.path.dirname(os.path.abspath(__file__))
    fecha = datetime.now().strftime('%Y%m%d_%H%M')
    archivo_salida = os.path.join(base_dir, f'Lista_Precios_{fecha}.xlsx')
    
    # Exportar
    if exportar_excel(productos, archivo_salida):
        print(f"\n✅ Archivo exportado exitosamente:")
        print(f"   📄 {archivo_salida}")
        print(f"   📊 {len(productos)} productos")
        return 0
    
    return 1


if __name__ == '__main__':
    sys.exit(main())
